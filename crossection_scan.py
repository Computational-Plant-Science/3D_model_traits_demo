"""
Version: 1.5

Summary: compute the segmentaiton and label of cross section image sequence 

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 crossection_scan.py -p ~/ply_data/cross_section_scan/ -th 2.35


argument:
("-p", "--path", required = True,    help = "path to image file")
("-ft", "--filetype", required = False, default = 'jpg',   help = "Image filetype")
("-th", "--threshold", required = False, default = '2.35', type = float, help = "threshold to remove outliers")

"""

#!/usr/bin/python
# Standard Libraries

import matplotlib 
matplotlib.use('Agg')

import glob
import os,fnmatch,os.path
import argparse
import shutil
import cv2

#import morphsnakes
import math

from numpy import NaN, Inf, arange, isscalar, asarray, array
import numpy as np

from matplotlib import pyplot as plt

from sklearn import linear_model
from itertools import compress

from skimage.measure import regionprops, label
from skimage.morphology import watershed, convex_hull_image
from skimage.color import label2rgb
from skimage.util import invert

import time

from scipy import ndimage

#import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from rdp import rdp

from openpyxl import load_workbook
from openpyxl import Workbook

import csv

import warnings
warnings.filterwarnings("ignore")

def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False


def get_median_filtered(signal, threshold=3):
    """construct median filter"""
    
    signal = signal.copy()
    
    difference = np.abs(signal - np.median(signal))
    
    median_difference = np.median(difference)
    
    if median_difference == 0:
        s = 0
    else:
        s = difference / float(median_difference)
        
    mask = s > threshold
    
    signal[mask] = np.median(signal)
    
    return signal


# get middle value of a list
def findMiddle(input_list):
    middle = float(len(input_list))/2
    if middle % 2 != 0:
        return input_list[int(middle - .5)]
    else:
        #return (input_list[int(middle)], input_list[int(middle-1)])
        return (input_list[int(middle)])

# get average of a list 
def Average(lst): 
    return sum(lst) / len(lst)

# compute radius from area
def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    return radius 

#compute external contour traits
def comp_external_contour(orig,thresh):
    
    img_height, img_width, img_channels = orig.shape
    
    #Convert the mean shift image to grayscale, then apply Otsu's thresholding
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    convexhull = convex_hull_image(gray)
    
    img_convexhull = np.uint8(convexhull)*255
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh_hull = cv2.threshold(img_convexhull, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #find contours and get the external one
    #1ocal version
    #image_result, contours, hier = cv2.findContours(img_convexhull, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #container version
    contours, hier = cv2.findContours(img_convexhull, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    #print("len(contours)")
    #print(len(contours))
    
    #label image regions
    #label_image_convexhull = label(convexhull)
    
    #Measure properties 
    regions = regionprops(img_convexhull)
    
    #eccentricity = regions[0].eccentricity
    
    
    d_major = regions[0].major_axis_length
    d_minor = regions[0].minor_axis_length
    
    eccentricity =  d_minor/d_major
    
    #eccentricity = 1

    #center location of region
    y_cvh, x_cvh = regions[0].centroid
    #print("Convexhull center of root system: {0}, {1} \n".format(int(x_cvh),int(y_cvh)))
    
    convexhull_diameter = regions[0].equivalent_diameter 
    
    return img_convexhull, convexhull_diameter, y_cvh, x_cvh, eccentricity, d_major, d_minor


    
def root_area_label(image_file):
    """compute the segmentaiton and label of cross section sequence"""
    
    #Parse image path  and create result image path
    path, filename = os.path.split(image_file)
    
    print("processing image : {0} \n".format(str(filename)))
    
    #load the image and perform pyramid mean shift filtering to aid the thresholding step
    imgcolor = cv2.imread(image_file)
    
    imgcolor_copy = imgcolor
    
    #imgcolor = ~imgcolor
    
    #accquire image dimensions 
    height, width, channels = imgcolor.shape
    
    #print(height, width, channels)
    
    #shifted = cv2.pyrMeanShiftFiltering(imgcolor, 5, 5)
    
    #define image morphology operation kernel
    kernel = cv2.getStructuringElement(cv2.MORPH_CROSS,(3,3))
    
    #perfrom dilation/closing to connect nearby contours
    dilation = cv2.dilate(imgcolor, kernel, iterations = 3)
    closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel, iterations = 3)
    erode = cv2.erode(closing, kernel, iterations = 1)
    
    #Image binarization by apltying otsu threshold
    gray = cv2.cvtColor(closing, cv2.COLOR_BGR2GRAY)
    
    # Convert BGR to GRAY
    #img_lab = cv2.cvtColor(erode, cv2.COLOR_BGR2LAB)
    
    gray = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2GRAY)
    

    #Obtain the threshold image using OTSU adaptive filter
    ret, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
 

    #find contours
    #container version
    contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #local version
    #_, contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    #result_img_path = save_path_ac + str(filename[0:-4]) + '_ac.png'
    #cv2.imwrite(result_img_path,binary)
    
    
    #draw all the filled contours
    for c in contours:
        
        #fill the connected contours
        contours_img = cv2.drawContours(binary, [c], -1, (255, 255, 255), cv2.FILLED)

    # define kernel
    connectivity = 8
    
    #find connected components 
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(contours_img, connectivity , cv2.CV_32S)
    
    #find the component with largest area 
    largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA]) 
    
    #collection of all component area
    areas = [s[4] for s in stats]
    
    #sorted_idx = np.argsort(areas)
    
    #sum of all component area
    area_sum = sum(areas)
    
    # average of component area
    area_avg = sum(areas)/len(np.unique(labels))
    
    #area_avg = sum(areas)/len(labels)
    #print("Area of components: {0} \n".format(int(area_sum)))
    
    #unique values in label results
    print("Length of region is: {0} \n".format(len(np.unique(labels))))
    
    
    #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))
    label_hue[labels == largest_label] = np.uint8(15)
    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
    
    #define result path for labeled images
    result_img_path = save_path_label + str(filename[0:-4]) + '_label.png'
    
    # save results
    cv2.imwrite(result_img_path, labeled_img)
    
    
    #Creat background image to display the location of detected roots
    image_background = np.zeros((height, width, 3), dtype = "uint8")
    
    # assign image values
    image_background[:] = (0, 0, 0)
    
    radius_scale = 1
    
    for i in range(1, len(centroids)):
        
        (x, y) = centroids[i]
        
        #radius of each contour
        #r = int(math.sqrt(0.5*areas[i]/math.pi))
        r = 5
        
        if r > 0:
            if i == largest_label:
                cv2.circle(image_background, (int(x), int(y)), int(r*radius_scale), (0, 128, 255), cv2.FILLED)
            else:
                cv2.circle(image_background, (int(x), int(y)), int(r*1), (0, 255, 255), cv2.FILLED)
    
    #define result path for simplified segmentation result
    result_img_path = save_path_ac + str(filename[0:-4]) + '.png'
    
    #write out results
    cv2.imwrite(result_img_path,image_background)
    
    
    return area_avg, area_sum, len(np.unique(labels))



def find_outlier_index(properity_list, threshold):
    """detect outlier index"""
 
    point = np.asarray(properity_list)

    point_filtered = get_median_filtered(point, threshold)

    outlier_idx = np.where(point_filtered != point)[0]

    return outlier_idx


def draw_properity(ax, point, outlier_idx, color):
    """visualize results"""

    y = point.flatten()

    x = np.asarray(range(1, len(point)+1)).flatten()

    ax.plot(x, y, label='root properity')

    ax.scatter(x, y, c = 'b', marker = 'o')

    ax.scatter(x[outlier_idx], point[outlier_idx], c = color, marker = '*')
    
    

def outlier_visualization(result, thresh_value):
    """visualize outlier frame"""
    
    outlier_index = [[] for i in range(3)]

    fig = plt.figure(1)

    ax = fig.gca()
    
    for i in range(0,1):
        
        outlier_index[i] = find_outlier_index( list(zip(*result)[i]), threshold = args["threshold"])
        
        draw_properity(ax, np.asarray(list(zip(*result)[i])), outlier_index[i], 'r')

    plt.legend()
   

    point = np.asarray(list(zip(*result)[0]))
    
    y = point.flatten()

    X = np.asarray(range(1, len(point)+1)).reshape(-1, 1) 
    
    # Fit line using all data
    lr  =  linear_model.LinearRegression()
    lr.fit(X, y)

    # Robustly fit linear model with RANSAC algorithm
    ransac  =  linear_model.RANSACRegressor()
    ransac.fit(X, y)
    inlier_mask  =  ransac.inlier_mask_
    outlier_mask  =  np.logical_not(inlier_mask)

    # Predict data of estimated models
    line_X  =  np.arange(X.min(), X.max())[:, np.newaxis]
    line_y  =  lr.predict(line_X)
    line_y_ransac  =  ransac.predict(line_X)

    lw  =  2
    
    ax.scatter(X[inlier_mask], y[inlier_mask], color = 'yellowgreen', marker = '.', label = 'Inliers')
    ax.scatter(X[outlier_mask], y[outlier_mask], color = 'gold', marker = '.', label = 'Outliers')

    ax.plot(line_X, line_y, color = 'navy', linewidth = lw, label = 'Linear regressor')
    ax.plot(line_X, line_y_ransac, color = 'cornflowerblue', linewidth = lw, label = 'RANSAC regressor')

    ax.legend(loc = 'lower right')
    plt.xlabel("Input")
    plt.ylabel("Response")
    plt.savefig('outlier.png')
    #plt.show()
    plt.close(fig)
    
    outlier_index = list(compress(xrange(len(outlier_mask)), outlier_mask))
    
    return outlier_index


def angle(directions):
    """Return the angle between vectors"""
    vec2 = directions[1:]
    vec1 = directions[:-1]

    norm1 = np.sqrt((vec1 ** 2).sum(axis=1))
    norm2 = np.sqrt((vec2 ** 2).sum(axis=1))
    cos = (vec1 * vec2).sum(axis=1) / (norm1 * norm2)   
    return np.arccos(cos)


def first_derivative(x) :
    """first derivative function"""
    return x[2:] - x[0:-2]

def second_derivative(x) :
    """second derivative function"""
    return x[2:] - 2 * x[1:-1] + x[:-2]

def curvature(x, y) :
    """compute curvature"""
    x_1 = first_derivative(x)
    x_2 = second_derivative(x)
    y_1 = first_derivative(y)
    y_2 = second_derivative(y)
    return np.abs(x_1 * y_2 - y_1 * x_2) / np.sqrt((x_1**2 + y_1**2)**3)


def turning_points(x, y, turning_points, smoothing_radius,cluster_radius):
    """define angle computation for turing points detection"""
    
    if smoothing_radius:
        weights = np.ones(2 * smoothing_radius + 1)
        new_x = ndimage.convolve1d(x, weights, mode='constant', cval=0.0)
        new_x = new_x[smoothing_radius:-smoothing_radius] / np.sum(weights)
        new_y = ndimage.convolve1d(y, weights, mode='constant', cval=0.0)
        new_y = new_y[smoothing_radius:-smoothing_radius] / np.sum(weights)
    else :
        new_x, new_y = x, y
        
    k = curvature(new_x, new_y)
    turn_point_idx = np.argsort(k)[::-1]
    t_points = []
    
    while len(t_points) < turning_points and len(turn_point_idx) > 0:
        t_points += [turn_point_idx[0]]
        idx = np.abs(turn_point_idx - turn_point_idx[0]) > cluster_radius
        turn_point_idx = turn_point_idx[idx]
        
    t_points = np.array(t_points)
    t_points += smoothing_radius + 1
    
    return t_points.astype(int)


def ecdf(a):
    x, counts = np.unique(a, return_counts=True)
    cusum = np.cumsum(counts)
    return x, cusum / cusum[-1]


def CDF_visualization(result):
    """visualize CDF"""
    
    ###################################################################
    #print(list(zip(*result)[0]))
     #write measured parameters as excel file 
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    trait_file = (save_path_excel + '/CDF.xlsx')
    
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active
    
    data = list(zip(*result))[0]

    for row in enumerate(data):
    
        sheet.append(row)

    #save the csv file
    wb.save(trait_file)
    ####################################################################
    
    '''
    x, y = ecdf(list(zip(*result))[0])
    x = np.insert(x, 0, x[0])
    y = np.insert(y, 0, 0.)
    
    fig = plt.figure(1)
    plt.plot(x, y, drawstyle='steps-post')
    plt.grid(True)
    plt.savefig('ecdf.png')

    result_file_CDF = save_path_excel + '/'  + 'ecdf.png'
    #result_file_CDF = result_file_CDF.replace('.txt','_cdf.png')
    plt.savefig(result_file_CDF)
    plt.close()
    '''
    
    num_bins = 10
    
    #counts, bin_edges = np.histogram(list(zip(*result)[0]), bins = num_bins, normed = True)
    counts, bin_edges = np.histogram(list(zip(*result))[0], bins = num_bins)
    
    # compute CDF curve
    cdf = np.cumsum(counts)
    
    #cdf = cdf / cdf[-1] #normalize
    
    x = bin_edges[1:]
    y = cdf
    
    # assembly points of CDF curve 
    trajectory = np.vstack((x, y)).T

    index_turning_pt = turning_points(x, y, turning_points = 4, smoothing_radius = 2, cluster_radius = 2)
    

    #Ramer-Douglas-Peucker Algorithm
    #simplify points et using rdp library 
    simplified_trajectory = rdp(trajectory, epsilon = 0.00200)
    #simplified_trajectory = rdp(trajectory)
    sx, sy = simplified_trajectory.T
    
    #print(sx)
    #print(sy)

    #compute plateau in curve
    dis_sy = [j-i for i, j in zip(sy[:-1], sy[1:])]
    
    #get index of plateau location
    index_sy = [i for i in range(len(dis_sy)) if dis_sy[i] <= 1.3]
    
    dis_index_sy = [j-i for i, j in zip(index_sy[:-1], index_sy[1:])]
    
    for idx, value in enumerate(dis_index_sy):
        
        if idx < len(index_sy)-2:
        
            if value == dis_index_sy[idx+1]:
            
                index_sy.remove(index_sy[idx+1])
    
    # Define a minimum angle to treat change in direction
    # as significant (valuable turning point).
    #min_angle = np.pi / 36.0
    min_angle = np.pi / 180.0
    #min_angle = np.pi /1800.0
    
    # Compute the direction vectors on the simplified_trajectory.
    directions = np.diff(simplified_trajectory, axis = 0)
    theta = angle(directions)

    # Select the index of the points with the greatest theta.
    # Large theta is associated with greatest change in direction.
    idx = np.where(theta > min_angle)[0] + 1
    
    '''
    # Visualize trajectory and its simplified version.
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(x, y, 'r--', label='trajectory')
    ax.plot(sx, sy, 'b-', label='simplified trajectory')
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.legend(loc='best')
    plt.savefig('trajectory.png')
    plt.close()
    
    # Visualize valuable turning points on the simplified trjectory.
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(sx, sy, 'gx-', label='simplified trajectory')
    ax.plot(sx[idx], sy[idx], 'ro', markersize = 7, label='turning points')
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.legend(loc='best')

    plt.savefig('turning.png')
    plt.close()
    '''
    
    
    index_turning_pt = sorted(idx)
    
    Turing_points =  np.unique(sy[idx].astype(int))
    
    #max_idx = max(max_idx)
    #print("Turing points: {0} \n".format(Turing_points))
    
    # plot CDF 
    #fig = plt.plot(bin_edges[1:], cdf, '-r', label = 'CDF')
    fig = plt.figure(1)
    plt.grid(True)
    plt.legend(loc='right')
    plt.title('CDF curve ')
    plt.xlabel('Root area, unit:pixel')
    plt.ylabel('Depth of level-set, unit:pixel')
    
    plt.plot(sx, sy, 'gx-', label='simplified trajectory')
    plt.plot(x, y, '-b', label = 'CDF')
    #plt.plot(sx[idx], sy[idx], 'ro', markersize = 7, label='turning points')
    
    plt.plot(sx[index_sy], sy[index_sy], 'ro', markersize = 7, label='plateau points')
    
    #plt.plot(sx[index_turning_pt], sy[index_turning_pt], 'bo', markersize = 7, label='turning points')
    #plt.vlines(sx[index_turning_pt], sy[index_turning_pt]-100, sy[index_turning_pt]+100, color='b', linewidth = 2, alpha = 0.3)
    
    plt.legend(loc='best')
    
    result_file_CDF = save_path_excel + '/'  + 'cdf.png'
    #result_file_CDF = result_file_CDF.replace('.txt','_cdf.png')
    plt.savefig(result_file_CDF)
    plt.close()
    
    #return Turing_points, sy
    return sy
    


# detect root convexhull area based on regionprops method
def root_system_trait(image_file):
    
    path, filename = os.path.split(image_file)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    #print("processing image : {0} \n".format(str(filename)))
    
    #result_img_path = file_path + str(filename[0:-4]) + '_lab.png'
       
    #print(result_img_path)
    
     # load the image and perform pyramid mean shift filtering to aid the thresholding step
    imgcolor = cv2.imread(image_file)
    
    # if cross scan images are white background and black foreground
    #imgcolor = ~imgcolor
    
    # accquire image dimensions 
    height, width, channels = imgcolor.shape
    #shifted = cv2.pyrMeanShiftFiltering(image, 5, 5)

    #Image binarization by apltying otsu threshold
    img = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2GRAY)
    
    # Convert BGR to GRAY
    img_lab = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2LAB)
    
    gray = cv2.cvtColor(img_lab, cv2.COLOR_BGR2GRAY)
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Compute the gemetrical shape in convexhull
    (img_convexhull, convexhull_diameter, y_cvh, x_cvh, eccentricity, d_major, d_minor) = comp_external_contour(imgcolor.copy(),thresh)
    
    #print("convexhull_diameter: {0} \n".format(convexhull_diameter))
    
   
    #define result path for labeled images
    #result_img_path = save_path_excel + str(filename[0:-4]) + '_convex.png'
    
    # save results
    #cv2.imwrite(result_img_path,img_convexhull)
    
    
    #print("eccentricity : {0} \n".format(str(eccentricity)))
    #print("d_major : {0} \n".format(str(d_major)))
    #print("d_minor : {0} \n".format(str(d_minor)))
    
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
   
    connectivity = 3
    
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary , connectivity , cv2.CV_32S)
   
    largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA]) 
    
    areas = [s[4] for s in stats]
    
    areas.remove(max(areas))
    
    radius = [area_radius(x) for x in areas]
    
    min_radius = 1
    
    radius = [x for x in radius if x > min_radius]
    
    #print(radius)
    #print(type(radius))
    
    sorted_idx_as = np.argsort(radius)
    
    sorted_idx_de = sorted_idx_as[::-1]
    
    #print(sorted(radius, reverse=True))
    
    radius_max = radius[sorted_idx_de[0]]

    
    if pattern_id == 1:
        
        radius = [x for x in radius if x == radius_max]
        
        num_primary_root = 1
        num_lateral_root = 0
        num_fine_root = 0
    
    elif (pattern_id == 2):
        
        num_primary_root = 1
        num_lateral_root = len(radius) - 1
        num_fine_root = 0

    elif (pattern_id == 3):
        
        #radius = [x for x in radius if x != radius_max]
        
        num_primary_root = 0
        num_lateral_root = 0
        num_fine_root = len(radius)

    '''
    print("num_primary_root is : {0}\n".format(str(num_primary_root)))
    print("num_lateral_root is : {0}\n".format(str(num_lateral_root)))
    print("num_fine_root is : {0}\n".format(str(num_fine_root)))
    '''
    return base_name, convexhull_diameter, len(radius), num_primary_root, num_lateral_root, num_fine_root, eccentricity, d_major, d_minor, sorted(radius, reverse=True)
    



def parallel_root_system_trait(images):
    
    
    # parallel processing 
    agents = multiprocessing.cpu_count() - 2
    chunksize = 3
    
    with closing(Pool(processes = agents)) as pool:
        
        result = pool.map(root_system_trait, images, chunksize)
        
        pool.terminate()
    
    base_name_rec = list(zip(*result))[0]
    convexhull_diameter_rec = list(zip(*result))[1]
    len_radius_rec = list(zip(*result))[2]
    num_primary_root_rec = list(zip(*result))[3]
    num_lateral_root_rec = list(zip(*result))[4]
    num_fine_root_rec = list(zip(*result))[5]
    eccentricity_rec = list(zip(*result))[6]
    d_major_rec = list(zip(*result))[7]
    d_minor_rec = list(zip(*result))[8]
    radius_rec = list(zip(*result))[9]
    
    '''
    #loop to all tracked trace files
    for file_idx, fname in enumerate(images):

        (base_name_rec, convexhull_diameter_rec, len_radius_rec, num_primary_root_rec, num_lateral_root_rec, num_fine_root_rec, eccentricity_rec, d_major_rec, d_minor_rec, radius_rec) = root_system_trait(file_idx)
        
    '''
    eccentricity_avg = Average(eccentricity_rec)
    d_major_avg = Average(d_major_rec)
    d_minor_avg = Average(d_minor_rec)
    
    print("eccentricity_avg : {0} \n".format(str(eccentricity_avg)))
    print("d_major_avg : {0} \n".format(str(d_major_avg)))
    print("d_minor_avg : {0} \n".format(str(d_minor_avg)))
    
    #num_primary_root_avg = Average(num_primary_root_rec)
    #num_lateral_root_avg = Average(num_lateral_root_rec)
    #num_fine_root_avg = Average(num_fine_root_rec)
    
    #print(type(num_primary_root_rec.sort()))
    
    Root_system_diameter = Average(convexhull_diameter_rec)
    
    num_primary_root_avg = math.floor(findMiddle((num_primary_root_rec)))
    num_lateral_root_avg = math.floor(max((num_lateral_root_rec)))
    num_fine_root_avg = math.floor(max((num_fine_root_rec)))
    
    
    print("num_primary_root : {0} \n".format(str(num_primary_root_avg)))
    print("num_lateral_root : {0} \n".format(str(num_lateral_root_avg)))
    print("num_fine_root : {0} \n".format(str(num_fine_root_avg)))
   
    #print(radius_rec[0], len(radius_rec[0]))
    avr_len = Average(len_radius_rec)
    
    #print(avr_len)
    
    #print(radius_rec)
    #print(type(radius_rec))
    
    #write measured parameters as excel file 
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #trait_file = (parent_path + '/system_traits_' + str(pattern_id) + '.xlsx')
    trait_file = (save_path_excel + '/system_traits_detail.xlsx')
    
   
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'Root system diameter'
        sheet.cell(row = 1, column = 3).value = 'Number of roots'
        sheet.cell(row = 1, column = 4).value = 'Number of primary root'
        sheet.cell(row = 1, column = 5).value = 'Number of lateral root'
        sheet.cell(row = 1, column = 6).value = 'Number of fine root'
        sheet.cell(row = 1, column = 7).value = 'Each root radius'
    
   
    for idx, row in enumerate(radius_rec):
        
        row.insert(0, base_name_rec[idx])
        row.insert(1, convexhull_diameter_rec[idx])
        row.insert(2, len_radius_rec[idx])
        row.insert(3, num_primary_root_rec[idx])
        row.insert(4, num_lateral_root_rec[idx])
        row.insert(5, num_fine_root_rec[idx])
        
        sheet.append(row)
    

    #save the xlsx file
    wb.save(trait_file)
    
    
    #write measured parameters as excel file 
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #trait_file = (parent_path + '/system_traits_' + str(pattern_id) + '.xlsx')
    trait_file = (save_path_excel + '/system_traits.xlsx')
    
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

    else:
        # Keep presets
        wb = Workbook()
        
        sheet = wb.active
        
        sheet.title = "system_traits" 

        sheet.cell(row = 1, column = 1).value = 'Root system diameter'
        sheet.cell(row = 1, column = 2).value = 'Number of primary root'
        sheet.cell(row = 1, column = 3).value = 'Number of lateral root'
        sheet.cell(row = 1, column = 4).value = 'Number of fine root'
        sheet.cell(row = 1, column = 5).value = 'Eccentricity'
        sheet.cell(row = 1, column = 6).value = 'System diameter max'
        sheet.cell(row = 1, column = 7).value = 'System diameter min'
        sheet.cell(row = 1, column = 8).value = 'Whorl number'
        sheet.cell(row = 1, column = 9).value = 'Whorl distance1'
        sheet.cell(row = 1, column = 10).value = 'Whorl distance2'
        sheet.cell(row = 1, column = 11).value = 'Whorl distance3'

    data = [ Root_system_diameter, num_primary_root_avg, num_lateral_root_avg, num_fine_root_avg, (eccentricity_avg), (d_major_avg), (d_minor_avg)]
    
    #print(data)
    
    sheet.append(data)
        

    sheet.cell(row = 2, column = 8).value = count
    sheet.cell(row = 2, column = 9).value = whorl_dis_array[0]
    
    if count > 2:
        sheet.cell(row = 2, column = 9).value = whorl_dis_array[1]
    else:
        sheet.cell(row = 2, column = 10).value = "NaN"
    
    
    #save the csv file
    wb.save(trait_file)
    
    

if __name__ == '__main__':

    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required = False, default = 'png',   help = "Image filetype")
    ap.add_argument("-th", "--threshold", required = False, default = '2.35', type = float, help = "threshold to remove outliers")
    args = vars(ap.parse_args())

    global file_path, save_path_ac, save_path_label, parent_path, pattern_id, count, whorl_dis_array, save_path_excel, n_images
    
    # setting path to cross section image files
    file_path = args["path"]
    ext = args['filetype']
    
    thresh_value = args["threshold"]
     
    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype

    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    n_images = len(imgList)
    
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    mkpath = parent_path + '/' + str('active_component')
    mkdir(mkpath)
    save_path_ac = mkpath + '/'
    
    mkpath = parent_path + '/' + str('lable')
    mkdir(mkpath)
    save_path_label = mkpath + '/'
    
    
    mkpath = parent_path + '/' + str('excel')
    mkdir(mkpath)
    save_path_excel = mkpath + '/'

    #print "results_folder: " + save_path_ac  
    
    # Run this with a pool of avaliable agents having a chunksize of 3 until finished
    # run image labeling fucntion to accquire segmentation for each cross section image
    agents = multiprocessing.cpu_count() - 2
    chunksize = 3
    with closing(Pool(processes = agents)) as pool:
        result = pool.map(root_area_label, imgList, chunksize)
        pool.terminate()
    
    
    #visualzie the CDF graph of first return value 
    list_thresh = sorted(CDF_visualization(result))
   
    #compute plateau in curve
    dis_array = [j-i for i, j in zip(list_thresh[:-1], list_thresh[1:])]
    
     
    #get index of plateau location
    index = [i for i in range(len(dis_array)) if dis_array[i] <= 1.3]
    
    dis_index = [j-i for i, j in zip(index[:-1], index[1:])]
    
    for idx, value in enumerate(dis_index):
        
        if idx < len(index)-2:
        
            if value == dis_index[idx+1]:
            
                index.remove(index[idx+1])
    
    
    reverse_index = sorted(index, reverse = True)
    
    #count = sum(1 for x in dis_array if float(x) <= 1.3)
    #get whorld number count 
    
    count = len(index)
    
    print("number of whorls is: {0} \n".format(count))
    
    #compute whorl location
    whorl_dis = []
    whorl_loc = []
    
    for idx, value in enumerate(reverse_index, start=1):
        
        #dis_value = list_thresh[value+1] - list_thresh[value-1]
        loc_value = int(len(imgList) - list_thresh[value+1])
        whorl_loc.append(loc_value)
        #print("adding value : {0} \n".format(str(loc_value)))
    
    
    #compute whorl distance
    whorl_dis_array = [j-i for i, j in zip(whorl_loc[:-1], whorl_loc[1:])]
   
    whorl_loc.extend([0, len(imgList)])
    
    whorl_loc = list(dict.fromkeys(whorl_loc))
    
    whorl_loc_ex = sorted(whorl_loc)
    

    print("list_thresh : {0} \n".format(str(list_thresh)))
    print("dis_array : {0} \n".format(str(dis_array)))
    print("index : {0} \n".format(str(index)))
    print("dis_index : {0} \n".format(str(dis_index)))
    print("reverse_index : {0} \n".format(str(reverse_index)))
    print("whorl_loc : {0} \n".format(str(whorl_loc)))
    print("whorl_dis_array : {0} \n".format(str(whorl_dis_array)))
    print("whorl_loc_ex : {0} \n".format(str(whorl_loc_ex)))

    
    
    #divide the image list into n chunks  
    list_part = [] 
    
    for idx, val in enumerate(whorl_loc_ex):
        
        print(idx, val)
        
        if idx < len(whorl_loc_ex)-1:
            
            sublist = imgList[val:whorl_loc_ex[idx+1]]
        
            list_part.append(sublist)                 
    
    #print(whorl_loc_ex)
    
    print(len(list_part))
    
    
 
    #for i in range(0, len(list_part)):
    
    for i in range(0, len(list_part)):
        if i == 0 :
            pattern_id = 1
        elif i == 1:
            pattern_id = 2
        else:
            pattern_id = 3
        
        print(pattern_id)
        
        parallel_root_system_trait(list_part[i])
    

    
    '''
    #convert excel to cvs file 
    
    trait_file = (save_path_excel + '/system_traits.xlsx')
    
    trait_file_csv = (save_path_excel + '/system_traits.csv')

    wb = load_workbook(trait_file)
    sh = wb.active

    #with open(trait_file_csv, 'wb') as f:  # open('test.csv', 'w', newline="") for python 3
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    
    '''
