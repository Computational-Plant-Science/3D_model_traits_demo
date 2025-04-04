"""
Version: 1.5

Summary: Analyze and visualize tracked traces

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 track_load_ori.py -p ~/Ptvpy_test/ -f trace_result.csv -v True


argument:
("-p", "--path", required = True, help="path to trace file")
("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")

default file format: *.csv 

"""

# Import python libraries
import numpy as np
from numpy import mean
from numpy import arctan2, sqrt
from numpy import matrix, average


import glob
import fnmatch
import os, os.path
import math
import sys
#import shutil

import matplotlib.pyplot as plt
import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
        
from scipy.spatial import distance
import scipy.linalg 

from mayavi import mlab

import imagesize 
import progressbar
from time import sleep

import itertools

#import warnings
#warnings.filterwarnings("ignore", category=FutureWarning)
#warnings.filterwarnings("ignore")

from tabulate import tabulate
import pandas as pd

import  moviepy.editor as mpy

'''
from skimage.morphology import skeletonize
import sknw
import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import axes3d, Axes3D 

from skimage.morphology import skeletonize_3d
from network_3d import skel2graph, plot_graph
from networkx import nx

import dask
import dask.array as da
'''

def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        #shutil.rmtree(dirpath)
        # if exists, return 
        #print path+' path exists!'
        return False

#colormap mapping
def get_cmap(n, name = 'BrBG'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)
    
# compute the path length along the trace
def pathlength(x,y,z):

    n = len(x)
    
    lv = [sqrt((x[i]-x[i-1])**2 + (y[i]-y[i-1])**2 + (z[i]-z[i-1])**2) for i in range(n)]
    
    return sum(lv)

# compute distance between consective point sets
def points_seg_length(coords):
    
    d = np.diff(coords, axis=0)
    
    segdists = np.sqrt((d ** 2).sum( axis = 1))
    
    # calculate length of line
    #l = np.sqrt( np.diff(X)**2 + np.diff(Y)**2 + np.diff(Z)**2 )
    
    return sum(segdists)

#compute distance between two point sets in a line
def line_length_2pt(coords):
    
    start = coords[0]
    end = coords[len(coords)-1]
    dst = distance.euclidean(start, end)
    #dist = numpy.linalg.norm(a-b)
    
    return dst


# compute angle between two 3D points
def points_angle(x, y, z, line_length):
    
    theta_offest = np.zeros(4)
    r_offest = np.zeros(4)
        
    #calculate angles
    for offest in range(0,4):
        
        interval = int(((offest+1)*0.25)*line_length)  
        
        if interval >= len(x):
            interval = len(x)-1
        
        cx = x[interval] - x[0]
        cy = y[interval] - y[0]  
        cz = z[interval] - z[0]
        
        (r,theta,phi) = asSpherical(cx, cy, cz)
        
        if theta > 90:
            theta = 180 -theta

        theta_offest[offest] = theta

        r_offest[offest] = r

    return r_offest[2], theta_offest[2], phi

#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    hxy = np.hypot(x, y)
    r = np.hypot(hxy, z)
    elevation = np.arctan2(z, hxy)*180/math.pi
    azimuth = np.arctan2(y, x)*180/math.pi
    return r[2], azimuth[2], elevation[2]
    '''
    if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
    '''
    

#coordinates transformation from cartesian coords to sphere coord system
def appendSpherical_np(xyz):
    
    ptsnew = np.hstack((xyz, np.zeros(xyz.shape)))
    xy = xyz[:,0]**2 + xyz[:,1]**2
    ptsnew[:,3] = np.sqrt(xy + xyz[:,2]**2)
    ptsnew[:,4] = np.arctan2(np.sqrt(xy), xyz[:,2]) # for elevation angle defined from Z-axis down
    #ptsnew[:,4] = np.arctan2(xyz[:,2], np.sqrt(xy)) # for elevation angle defined from XY-plane up
    ptsnew[:,5] = np.arctan2(xyz[:,1], xyz[:,0])
    return ptsnew[:,3],ptsnew[:,4],ptsnew[:,5]

# remove otliers
def reject_outliers(data, m = 2.):
    
    d = np.abs(data - np.median(data))
    mdev = np.median(d)
    s = d / (mdev if mdev else 1.)
    return data[s < m]


# visualize the traces and return their properities
def trace_visualize(trace_array, array_index_rec, fit_linepts_rec, index_pair_rec, connect_pts_rec):
    
    # properities initialization for traits computation
    index_rec = []
    
    X_rec = []
    Y_rec = []
    Z_rec = []
    
    connect_label_rec = []

    index_label_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius_rec = []
    
    color_rec = []
    
    #image_chunk = np.zeros((416, 414, 282))
    

    #color = np.arange(0, 1, 1/n_trace).tolist()
    
    index_pair_rec_arr = np.asarray(index_pair_rec)
    
    index_pair_rec_unique = np.unique(index_pair_rec_arr)
        
    #print("index_pair_rec_unique {}\n".format(index_pair_rec_unique))
    
    #print("array_index_rec {}\n".format(array_index_rec))
    
    #array_index_rec_part = np.delete(array_index_rec, index_pair_rec_unique, None)
    
    #print("array_index_rec_part {}\n".format(array_index_rec_part))
    
    #print("len(array_index_rec) {}, len(index_pair_rec_arr) {}, len(connect_pts_rec) {}\n".format(len(array_index_rec),len(index_pair_rec_arr),len(connect_pts_rec)))
    
    cmap = get_cmap(len(array_index_rec))
    
    for idx, index_value in enumerate(array_index_rec):
        
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        if index_value in index_pair_rec_unique:
        #if(len(list(filter (lambda x : x == index_value, index_pair_rec_unique))) > 0):
        
            index_loc = np.where(index_pair_rec_arr[:,0] == index_value)

            index_loc_list = [x for xs in index_loc for x in xs]

            for idx_pair_loc in index_loc_list:
                
                #print("current idx {0}: pair_value ({1},{2})".format(idx, index_value, index_pair_rec_arr[idx_pair_loc, 1]))
                
                coords_arr = np.asarray(connect_pts_rec[idx_pair_loc])

                x = coords_arr[:,0]
                y = coords_arr[:,1]
                z = coords_arr[:,2]

                radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_pair_rec_arr[idx_pair_loc, 0])][:,4])
                
                #compute line path length
                line_length = points_seg_length(coords_arr)
                
                (r, azimuth, elevation) = cart2sph(x, y, z)

                X_rec.append(x)
                Y_rec.append(y)
                Z_rec.append(z)

                
                index_label_rec.append(index_pair_rec_arr[idx_pair_loc, 0])
                connect_label_rec.append('1')
                
                index_rec.append(idx)
                length_rec.append(line_length)
                angle_rec.append(azimuth)
                diameter_rec.append(radius_mean)
                projection_radius_rec.append(r)
                color_rec.append(color_rgb)
                
        else:
            #print("current idx {0}: index value {1}".format(idx, index_value))
            
            X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]/X_scale
            Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]/Y_scale
            Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/Z_scale

            radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])

            trace_radius = trace_array[np.where(trace_array[:,0] == index_value)][:,4]

            scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
            
            #compute line path length
            line_length = points_seg_length(np.stack(( X, Y, Z ), axis = 1))
            
            (r, azimuth, elevation) = cart2sph(X, Y, Z)
            
            X_rec.append(X)
            Y_rec.append(Y)
            Z_rec.append(Z)

            index_label_rec.append(index_value)
            connect_label_rec.append('0')
            
            index_rec.append(idx)
            length_rec.append(line_length)
            angle_rec.append(azimuth)
            diameter_rec.append(radius_mean)
            projection_radius_rec.append(r)
            color_rec.append(color_rgb)
    
    #import random
    #color = list(random.sample(range(0, 1), len(array_index_rec)))
    
    if args["visualize"]:
    
        print("Visualizing tracked traces...\n")
        
        ###########################################################################################################3
        
        #cmap = get_cmap(len(X_rec))
        
        fig_trace = mlab.figure('Root_structure', bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (720,1080))
        
        for idx, (x, y, z, diameter_rec_value, connect_label_rec_value, index_label_rec_value, color_rec_value) in enumerate(zip(X_rec, Y_rec, Z_rec, diameter_rec, connect_label_rec, index_label_rec, color_rec)):
            
            #color_rgb = cmap(idx)[:len(cmap(idx))-1]
            
            #pts = mlab.points3d(x, y, z, color = color_rgb, mode = 'point')
            #pts.actor.property.set(point_size = 5.5)
            
            if (int(connect_label_rec_value) > 0):
                
                pts = mlab.plot3d(x, y, z, color = color_rec_value, opacity = 0.3, representation = 'wireframe', transparent = True, tube_radius = tube_scale*diameter_rec_value)
                
                #pts = mlab.points3d(x, y, z, color = color_rec_value, mode = 'point')
                #pts.actor.property.set(point_size = 5.5)
                
            else:
                
                pts = mlab.plot3d(x, y, z, color = color_rec_value, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = tube_scale*diameter_rec_value)

            #pts = mlab.text3d(x[0], y[0], z[0], str(index_label_rec_value), scale = (4, 4, 4), color = (1, 0.0, 0.0))
        
        
        
        #snapshot = (save_path_result + directory_name + '.png')
        
        #mlab.savefig(snapshot, sizesize = (720,1080))
        
        #obj_file = (save_path_result + 'model.obj')
        
        #mlab.savefig(obj_file)
       
        mlab.show()
        
        ############################################################################################################
        
        '''
        #############################################################################################################
        #animation display
        #fig_trace_animation = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (1920,1080))
        
        fig_trace_animation = mlab.figure('Root_structure', bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (720,1080))
        
        # duration of the animation in seconds (it will loop)
        duration = len(X_rec)  
        
        def make_frame(t):
            
            idx_t = int(t)
            #mlab.clf() # clear the figure (to reset the colors)
            #pts = mlab.points3d(X_rec[idx_t], Y_rec[idx_t], Z_rec[idx_t], color = cmap(idx_t)[:len(cmap(idx_t))-1], mode = 'point')
            
            #pts.actor.property.set(point_size = 2.5)

            #mlab.clf() # clear the figure (to reset the colors)
            #pts = mlab.points3d(X_rec[idx_t], Y_rec[idx_t], Z_rec[idx_t], color = color_rec[idx_t], mode = 'point')
            
            #pts.actor.property.set(point_size = 2.5)

            pts = mlab.plot3d(X_rec[idx_t], Y_rec[idx_t], Z_rec[idx_t], color = color_rec[idx_t], opacity = 0.1, representation = 'surface', transparent = True, tube_radius = tube_scale*diameter_rec[idx_t])

            #pts = mlab.plot3d(line_x, line_y, line_z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 1*radius_mean_rec[idx])

            pts = mlab.text3d(X_rec[idx_t][len(X_rec[idx_t])-1], Y_rec[idx_t][len(X_rec[idx_t])-1], Z_rec[idx_t][len(X_rec[idx_t])-1], str(idx_t+1), scale = (text_size, text_size, text_size), color = (1, 0.0, 0.0))
            
            mlab.view(camera_azimuth, camera_elevation, camera_distance, camera_focalpoint)
            
            print(mlab.view())
            
            f = mlab.gcf()
            f.scene._lift()
            
            return mlab.screenshot(antialiased = True)
        
        animation = mpy.VideoClip(make_frame, duration = duration)

        animation.write_gif((save_path_result + directory_name + '.gif'), fps = 5)
        
        #animation.write_videofile((save_path_result + 'structure.mp4'), fps = 5)
        
        #show model
        mlab.show()
        '''

        ####################################################################Pipeline Visualiztion
        '''
        N = len(X)
        
        # We create a list of positions and connections, each describing a line.
        # We will collapse them in one array before plotting.
        x = list()
        y = list()
        z = list()
        s = list()
        connections = list()

        # The index of the current point in the total amount of points
        index = 0

        # Create each line one after the other in a loop
        for i in range(N):
            x.append(X[i])
            y.append(Y[i])
            z.append(Z[i])
            s.append(scalars[idx])
            # This is the tricky part: in a line, each point is connected
            # to the one following it. We have to express this with the indices
            # of the final set of points once all lines have been combined
            # together, this is why we need to keep track of the total number of
            # points already created (index)
            connections.append(np.vstack(
                               [np.arange(index,   index + N - 1.5),
                                np.arange(index + 1, index + N - .5)]
                                    ).T)
        index += N
        
         # Now collapse all positions, scalars and connections in big arrays
        x = np.hstack(x)
        y = np.hstack(y)
        z = np.hstack(z)
        s = np.hstack(s)
        connections = np.vstack(connections)

        # Create the points
        src = mlab.pipeline.scalar_scatter(x, y, z, s)

        # Connect them
        src.mlab_source.dataset.lines = connections
        src.update()

        # The stripper filter cleans up connected lines
        #lines = mlab.pipeline.stripper(src)
        
        #lines = mlab.pipeline.tube(src, tube_radius=0.005, tube_sides=6)

        # Finally, display the set of lines
        mlab.pipeline.surface(src, colormap = 'Accent', line_width = 3, opacity =.4)
        '''
        #############################################################################
            
            
        #show model
        #mlab.show()
    
    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec, X_rec, Y_rec, Z_rec
    


# visualize the traces and return their properities
def trace_visualize_simple(trace_array, array_index_rec, fit_linepts_rec, index_pair_rec, connect_pts_rec):
    
    # properities initialization
    index_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius_rec = []
    
    index_label_rec = []
    
    X_rec = []
    Y_rec = []
    Z_rec = []
    
    color_rec = []

    cmap = get_cmap(len(array_index_rec))
    #color = np.arange(0, 1, 1/n_trace).tolist()
    
    # collect all the tracked traces and their properties
    for idx, index_value in enumerate(array_index_rec):

        #t = idx
        #print(idx, index_value)
        
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]/X_scale
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]/Y_scale
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/Z_scale
        
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        trace_radius = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
        
        ####################################################################
        

         #recompute the connected trace properities
        coords = np.stack(( X, Y, Z ), axis = 1)

        #compute line angle
        line_length = points_seg_length(coords)

        (r, azimuth, elevation) = cart2sph(X, Y, Z)
        '''
        if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
        '''
        #print("Trace {0} properities:".format(idx))
        #print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X), line_length, angle))

      
        # record all parameters
        X_rec.append(X)
        Y_rec.append(Y)
        Z_rec.append(Z)

        index_label_rec.append(index_value)

        index_rec.append(idx)
        length_rec.append(line_length)
        angle_rec.append(azimuth)
        diameter_rec.append(radius_mean)
        projection_radius_rec.append(r)
        color_rec.append(color_rgb)
        
        #index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, color_rec, index_label_rec, X_rec, Y_rec, Z_rec
    
    ##################################################visualize structure
    if args["visualize"]:
    
        print("Visualizing tracked traces...")
        
        #cmap = get_cmap(len(X_rec))
        
        fig_myv = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))
        
        for idx, (x, y, z, diameter_rec_value, index_label_rec_value, color_rec_value) in enumerate(zip(X_rec, Y_rec, Z_rec, diameter_rec, index_label_rec, color_rec)):
            
            #color_rgb = cmap(idx)[:len(cmap(idx))-1]
            
            pts = mlab.points3d(x, y, z, color = color_rec_value, mode = 'point')
            pts.actor.property.set(point_size = 5.5)
            
            pts = mlab.plot3d(x, y, z, color = color_rec_value, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = tube_scale*diameter_rec_value)
            pts = mlab.text3d(x[0], y[0], z[0], str(index_label_rec_value), scale = (4, 4, 4), color = (1, 0.0, 0.0))
        
        mlab.show()
        
        #########################################################################################
        '''
        #animation display
        #fig_myv = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (1920,1080))
        
        fig_myv = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (1280,720))
        
        cmap = get_cmap(len(X_rec))
        
        mlab.orientation_axes(True)
        # duration of the animation in seconds (it will loop)
        duration = len(array_index_rec) 
        
        def make_frame(t):
            
            idx_t = int(t)
           
            #mlab.clf() # clear the figure (to reset the colors)
            pts = mlab.points3d(X_rec[idx_t], Y_rec[idx_t], Z_rec[idx_t], color = cmap(idx_t)[:len(cmap(idx_t))-1], mode = 'point')
            
            pts.actor.property.set(point_size = 2.5)

            pts = mlab.plot3d(X_rec[idx_t], Y_rec[idx_t], Z_rec[idx_t], color = cmap(idx_t)[:len(cmap(idx_t))-1], opacity = 0.1, representation = 'surface', transparent = True, tube_radius = tube_scale*diameter_rec[idx_t])

            #pts = mlab.plot3d(line_x, line_y, line_z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 1*radius_mean_rec[idx])

            pts = mlab.text3d(X_rec[idx_t][len(X_rec[idx_t])-1], Y_rec[idx_t][len(X_rec[idx_t])-1], Z_rec[idx_t][len(X_rec[idx_t])-1], str(idx_t+1), scale = (text_size, text_size, text_size), color = (1, 0.0, 0.0))
            
            mlab.view(camera_azimuth, camera_elevation, camera_distance, camera_focalpoint)
            
            #print(mlab.view())
            
            f = mlab.gcf()
            f.scene._lift()
            
            return mlab.screenshot(antialiased=True)
        
        animation = mpy.VideoClip(make_frame, duration = duration)
        
        animation.write_gif((save_path_result + 'structure.gif'), fps = 5)
        
        #animation.write_videofile((save_path_result + 'structure.mp4'), fps = 5)
        
        #show model
        mlab.show()
        '''

    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec, X_rec, Y_rec, Z_rec



# SVD fiting lines to 3D points
def line_fiting_3D(data):

    # Calculate the mean of the points, i.e. the 'center' of the cloud
    datamean = data.mean(axis = 0)
    
    x_range = data[:,0].max() - data[:,0].min()
    y_range = data[:,1].max() - data[:,1].min()
    z_range = data[:,2].max() - data[:,2].min()
    
    
    data_range =(x_range, y_range, z_range)
    
    data_range_mean = sum(data_range) / len(data_range)
    
    # Do an SVD on the mean-centered data.
    uu, dd, vv = np.linalg.svd(data - datamean)

    # Now vv[0] contains the first principal component, i.e. the direction
    # vector of the 'best fit' line in the least squares sense.

    # Now generate some points along this best fit line, for plotting.

    # use -7, 7 since the spread of the data is roughly 14
    # and we want it to have mean 0 (like the points we did
    # the svd on). Also, it's a straight line, so we only need 2 points.
    linepts = vv[0] * np.mgrid[-1*data_range_mean:data_range_mean:2j][:, np.newaxis]

    # shift by the mean to get the line in the right place
    linepts += datamean
    
    return linepts


# compute trace properities
def trace_compute(trace_array, trace_index):
    
    #import scipy.optimize as optimize
    
    print("Processing tracked trace properities...")
    
    #initialize parameters
    index_rec = []
    length_euclidean_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius_rec = []
    fit_linepts_rec = []
    
    index_label_rec = []
    
    X_fit = []
    Y_fit = []
    Z_fit = []
    
    X_rec = []
    Y_rec = []
    Z_rec = []
    
    
    p = np.array([0, 0, 0])
    q = np.array([0, 0, 1])
    r = np.array([0, 1, 1])


    
    #progress bar display
    bar = progressbar.ProgressBar(maxval = len(trace_index))
    
    bar.start()

    #scan all traces
    for idx, index_value in enumerate(trace_index):

        #print(idx, index_value)
        bar.update(idx+1)
        
        sleep(0.1)
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]/X_scale
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]/Y_scale
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/Z_scale

        #traits measurement
        ##################################################################
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]

        #compute line length
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        #print(coords.shape)
        
        '''
        #global curve interpolation
        degree = 1  # cubic curve
        curve = fitting.interpolate_curve(coords.tolist(), degree)
        #curve = fitting.approximate_curve(coords.tolist(), degree)
        evalpts = np.array(curve.evalpts)
        x_curve = evalpts[:, 0]
        y_curve = evalpts[:, 1]
        z_curve = evalpts[:, 2]
        '''
        # fitting curve
        #(xLinespace, yLinespace, zLinespace) = cubic_spline_interpolate(X, Y, Z)
        #fit_linepts = np.stack(( xLinespace, yLinespace, zLinespace ), axis = 1)

        # fitting straight line
        fit_linepts = line_fiting_3D(coords)
        
        xLinespace = np.array((fit_linepts[0][0], fit_linepts[1][0]))
        yLinespace = np.array((fit_linepts[0][1], fit_linepts[1][1]))
        zLinespace = np.array((fit_linepts[0][2], fit_linepts[1][2]))
        
        #print(fit_linepts[0], fit_linepts[1])
        
        #line_length = pathlength(X, Y, Z)
        
        #compute line angle
        line_length = points_seg_length(coords)
        
        #compute line length bewteen start and end points
        line_length_euclidean = line_length_2pt(coords)
        
        #print("line_length {0} properities:".format(line_length))
        #print("line_length_2pt {0} properities:".format(line_length_euclidean))
        
        (r, azimuth, elevation) = cart2sph(X, Y, Z)
        '''
        if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
        '''
        #print("Trace {0} properities:".format(index_value))
        #print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X), line_length, angle))
        #print(popt)
        
        #print("Angle:{0} {1} {2}\n".format(r, theta, phi))

        # record all parameters
        index_rec.append(idx)
        
        length_euclidean_rec.append(line_length_euclidean)
        angle_rec.append(azimuth)
        #diameter_rec.append(radius_mean)
        #projection_radius_rec.append(r)
        
        fit_linepts_rec.append(fit_linepts)
        index_label_rec.append(index_value)
 
        X_fit.append(xLinespace)
        Y_fit.append(yLinespace)
        Z_fit.append(zLinespace)
        
        #X_fit.append(x_curve)
        #Y_fit.append(y_curve)
        #Z_fit.append(z_curve)
        
        X_rec.append(X)
        Y_rec.append(Y)
        Z_rec.append(Z)
    
    bar.finish()
    
    # remove outlier trace
    avg_length = mean(length_euclidean_rec) 
    
    indexes_length_remove = [idx for idx, element in enumerate(length_euclidean_rec) if element < (avg_length*length_thresh)]
    
    #indexes_angle_remove = [idx for idx, element in enumerate(angle_rec) if element > 70]
    
    #indexes_length_remove_label = [index_label_rec[i] for i in indexes_length_remove] 
    
    #print("avg_length is {0}".format(avg_length))
    #print("indexes_length_remove is {0}".format(indexes_length_remove))
    
    #print("indexes_angle_remove is {0}".format(indexes_angle_remove))
    
    '''
    fig_ori = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))
        
    for (x_fit, y_fit, z_fit, x_ori, y_ori, z_ori, index_value) in zip(X_fit, Y_fit, Z_fit, X_rec, Y_rec, Z_rec, index_label_rec):
        
        #print("connect_rec_value: {0} index {1}".format(connect_rec_value, index_rec_vis_value))
        
        #pts = mlab.points3d(x_fit, y_fit, z_fit, color = (0.2, 0.4, 0.5), mode = 'point')
        
        pts = mlab.points3d(x_ori, y_ori, z_ori, color = (0.8, 0.0, 0.0), mode = 'point')
        
        pts.actor.property.set(point_size = 5.5)
        
        #pts = mlab.plot3d(x_fit, y_fit, z_fit, color = (0.2, 0.4, 0.5))
        
        pts = mlab.plot3d(x_fit, y_fit, z_fit, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 0.5)

        pts = mlab.text3d(x_ori[0], y_ori[0], z_ori[0], str(index_value), scale = (4, 4, 4), color = (1, 0.0, 0.0))
        
    mlab.show()
    '''

    #return index_rec, length_euclidean_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec, fit_linepts_rec, indexes_length_remove 

    return index_rec, index_label_rec, fit_linepts_rec, indexes_length_remove, length_euclidean_rec




def angle_between(vector_1, vector_2):
    
    unit_vector_1 = vector_1 / np.linalg.norm(vector_1)

    unit_vector_2 = vector_2 / np.linalg.norm(vector_2)

    dot_product = np.dot(unit_vector_1, unit_vector_2)

    angle = np.arccos(dot_product)
 
    return np.rad2deg(angle)


def interpolate_pts_3D(data, data_range, z_range):

    # Calculate the mean of the points, i.e. the 'center' of the cloud
    datamean = data.mean(axis=0)

    # Do an SVD on the mean-centered data.
    uu, dd, vv = np.linalg.svd(data - datamean)

    # Now vv[0] contains the first principal component, i.e. the direction
    # vector of the 'best fit' line in the least squares sense.
       # Now generate some points along this best fit line, for plotting.

    # I use -7, 7 since the spread of the data is roughly 14
    # and we want it to have mean 0 (like the points we did
    # the svd on). Also, it's a straight line, so we only need 2 points.
    linepts = vv[0] * np.mgrid[(-1*data_range):data_range:complex(0,z_range)][:, np.newaxis]

    # shift by the mean to get the line in the right place
    linepts += datamean
    
    return linepts



def solveEquations(P,L,U,y):
    y1=np.dot(P,y)
    y2=y1
    m=0
    for m in range(0, len(y)):
        for n in range(0, m):
            y2[m] = y2[m] - y2[n] * L[m][n]
        y2[m] = y2[m] / L[m][m]
    y3 = y2
    for m in range(len(y) - 1,-1,-1):
        for n in range(len(y) - 1, m, -1):
            y3[m] = y3[m] - y3[n] * U[m][n]
        y3[m] = y3[m] / U[m][m]
    return y3

'''
    Scipy tool with high complexity.
    P stands for the permutation Matrix
    L stands for the lower-triangle Matrix
    U stands for the upper-triangle Matrix
    matrix·x = y
    P·matrix = L·U
    P·matrix·x = L·U·x = P·y
    L·U·x = y1
    U·x = y2
    x = y3
'''  
def doLUFactorization(matrix):    
    P, L, U = scipy.linalg.lu(matrix)
    return P, L, U   

#calculate each parameters of location.
def func(x1,x2,t,v1,v2,t1,t2):
    ft=((t2-t)**3*v1+(t-t1)**3*v2)/6+(t-t1)*(x2-v2/6)+(t2-t)*(x1-v1/6)
    return ft
   
def cubic_spline_interpolate(x_axis,y_axis,z_axis):
    '''
        prepare right-side vector
    '''
    dx=[]
    dy=[]
    dz=[]
    matrix=[]
    n=2
    while n<len(x_axis):
        dx.append(3*(x_axis[n]-2*x_axis[n-1]+x_axis[n-2]))
        dy.append(3*(y_axis[n]-2*y_axis[n-1]+y_axis[n-2]))
        dz.append(3*(z_axis[n]-2*z_axis[n-1]+z_axis[n-2]))
        n=n+1   
    '''
        produce square matrix looks like :
        [[2.0, 0.5, 0.0, 0.0], [0.5, 2.0, 0.5, 0.0], [0.0, 0.5, 2.0, 0.5], [0.0, 0.0, 2.0, 0.5]]
        the classes of the matrix depends on the length of x_axis(number of nodes)
    '''
    matrix.append([float(2), float(0.5)])
    for m in range(len(x_axis)-4):
        matrix[0].append(float(0))                
    n=2
    while n<len(x_axis)-2:
        matrix.append([])
        for m in range(n-2):
            matrix[n-1].append(float(0)) 
              
        matrix[n-1].append(float(0.5))
        matrix[n-1].append(float(2))
        matrix[n-1].append(float(0.5))
        
        for m in range(len(x_axis)-n-3):
            matrix[n-1].append(float(0)) 
        n=n+1
        
    matrix.append([])
    for m in range(n-2):
        matrix[n-1].append(float(0))    
    matrix[n-1].append(float(0.5))    
    matrix[n-1].append(float(2))
    '''
        LU Factorization may not be optimal method to solve this regular matrix. 
        If you guys have better idea to solve the Equation, please contact me.
        As the LU Factorization algorithm cost 2*n^3/3 + O(n^2) (e.g. Doolittle algorithm, Crout algorithm, etc).
        (How about Rx = Q'y using matrix = QR (Schmidt orthogonalization)?)
        If your application field requires interpolating into constant number nodes, 
        It is highly recommended to cache the P,L,U and reuse them to get O(n^2) complexity.
    '''
    P, L, U = doLUFactorization(matrix)
    u=solveEquations(P,L,U,dx)
    v=solveEquations(P,L,U,dy)
    w=solveEquations(P,L,U,dz)
    
    '''
        define gradient of start/end point
    '''
    m=0
    U=[0]
    V=[0]
    W=[0]
    while m<len(u):
        U.append(u[m])
        V.append(v[m])
        W.append(w[m])
        m=m+1
    U.append(0)
    V.append(0)
    W.append(0)
   
    #return U,V,W
    
    n_iteration = 100
    
    m = 1
    xLinespace=[]
    yLinespace=[]
    zLinespace=[]
    while m<len(x_axis):
        for t in np.arange(m-1,m,1/float(n_iteration)):
            xLinespace.append(func(x_axis[m-1],x_axis[m],t,U[m-1],U[m],m-1,m))
            yLinespace.append(func(y_axis[m-1],y_axis[m],t,V[m-1],V[m],m-1,m))
            zLinespace.append(func(z_axis[m-1],z_axis[m],t,W[m-1],W[m],m-1,m))
        m=m+1
    
    return xLinespace, yLinespace, zLinespace
    



'''
def pts_interpolation(U,V,W,x_axis,y_axis,z_axis):
    
    n_iteration = 10
    
    m = 1
    xLinespace=[]
    yLinespace=[]
    zLinespace=[]
    while m<len(x_axis):
        for t in np.arange(m-1,m,1/float(n_iteration)):
            xLinespace.append(func(x_axis[m-1],x_axis[m],t,U[m-1],U[m],m-1,m))
            yLinespace.append(func(y_axis[m-1],y_axis[m],t,V[m-1],V[m],m-1,m))
            zLinespace.append(func(z_axis[m-1],z_axis[m],t,W[m-1],W[m],m-1,m))
        m=m+1
    
    return xLinespace, yLinespace, zLinespace
'''


def dist_pt2line(p, q, r):
    
    x = p-q
    
    t = np.dot(r-q, x)/np.dot(x, x)
    
    return np.linalg.norm(t*(p-q)+q-r)



def trace_connect(trace_array, array_index_rec, fit_linepts_rec):
    
    #import similaritymeasures
    #from scipy.spatial import distance
    
    print("connecting tracked traces...")
    
    index_pair_rec = []
    connect_pts_rec = []
    index_label_rec = []
    color_rec = []

    cmap = get_cmap(len(array_index_rec))
    
    for idx, index_value in enumerate(array_index_rec):

        #print(idx, index_value)
                
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]/X_scale
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]/Y_scale
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/Z_scale
        
        #stack array
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
        
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        #extract coordinates of the fitted straight line of current trace
        arr_lines = np.asarray(fit_linepts_rec[idx])
        
        vector_x = arr_lines[0,0] - arr_lines[1,0]
        vector_y = arr_lines[0,1] - arr_lines[1,1]
        vector_z = arr_lines[0,2] - arr_lines[1,2]
        
        vector_line = np.array([vector_x, vector_y, vector_z])
        
        vector_vertical = (([0, 0, 1]))

        angle = angle_between(vector_line, vector_vertical)

        #print("coordinate {0} {1} {2}:\n".format(np.max()))
        
        #print("angle index {0} is {1}:".format(index_value, angle))

        
        for idx_next, index_value_next in enumerate(array_index_rec):
            
            if (idx != idx_next): 
                
                #extract coordinates of the next trace
                X_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,1]/X_scale
                Y_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,2]/Y_scale
                Z_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,3]/Z_scale
                
                #stack array
                coords_next = np.stack(( X_next, Y_next, Z_next ), axis = 1)
                
                #extract coordinates of the fitted line of next trace
                arr_lines_next = np.asarray(fit_linepts_rec[idx_next])

                vector_x_next = arr_lines_next[0,0] - arr_lines_next[1,0]
                vector_y_next = arr_lines_next[0,1] - arr_lines_next[1,1]
                vector_z_next = arr_lines_next[0,2] - arr_lines_next[1,2]
                
                vector_line_next = np.array([vector_x_next, vector_y_next, vector_z_next])
                
                #angle between two traces
                angle_diff_vector = angle_between(vector_line, vector_line_next)
                
                #angle between current trace and new trace to be connected
                angle_diff_vector_2connect = angle_between(vector_line, np.subtract(coords[len(coords)-1], coords_next[0]))

                '''
                #interpolate straight lines between connecting points
                connect_vector = np.vstack((coords[len(coords)-1],coords_next[0]))
                #print("coordinate current {0}".format(coords[len(coords)-1]))
                #print("coords_next {0}:".format(coords_next[0]))
                #print(connect_vector)
                #print("angle_diff_vector_2connect {0}:\n".format(angle_diff_vector_2connect))
               

                #compute the data range to be connected
                x_diff = np.diff(connect_vector[:,0])
                y_diff = np.diff(connect_vector[:,1])
                z_diff = np.diff(connect_vector[:,2])
                
                data_range = max(abs(x_diff), abs(y_diff), abs(z_diff))
                
                #print(max_connect_pts[:,2])
                #print(min_connect_pts[:,2])
                
                squared_dist = np.sum((coords[len(coords)-1]-coords_next[0])**2, axis=0)
                dist = np.sqrt(squared_dist)
                
                #z_range = abs(max_connect_pts[:,2] - min_connect_pts[:,2])
                
                #compute number of points to be connected
                z_range = dist
                '''
                
                # quantify the difference between the two curves using PCM
                #pcm_diff = similaritymeasures.pcm(coords, coords_next)
                
                #distance between connection points
                dst_connection_pts = distance.euclidean(coords[len(coords)-1], coords_next[0])
                
                #distance between connection points
                #dst_pts_line = distance.euclidean(coords[len(coords)-1], coords_next[0])
                
                #p = arr_lines[0,:]
                #q = arr_lines[1,:]
                #r = coords_next[0]
                
                #dst_pts_line = dist_pt2line(arr_lines[0,:], arr_lines[1,:], coords_next[0])
                
                #dst_pts_line = distance.euclidean((coords[0] + coords[len(coords)-1])*0.5, coords_next[0])
                
                dst_pts_line = distance.euclidean((arr_lines[0,:] + arr_lines[1,:])*0.5, coords_next[0])
                
                dst_current_line =  distance.euclidean(arr_lines[0,:], arr_lines[1,:])
                
                dst_next_line =  distance.euclidean(arr_lines_next[0,:], arr_lines_next[1,:])

                #print("dst_connection_pts {0}, dst_pts_line {1}:".format(dst_connection_pts, dst_pts_line))
                
                '''
                p = np.array([0, 0, 0])
                q = np.array([0, 0, 1])
                r = np.array([0, 1, 1])
                
                dst_pts_line = dist_pt2line(p, q, r)
                
                print()
                '''
                
                X_combine = np.hstack([X, X_next])
                Y_combine = np.hstack([Y, Y_next])
                Z_combine = np.hstack([Z, Z_next])
                

                #if (angle_diff_vector < 20) and (pcm_diff < 105) and (dst_diff < 100):
                if (angle_diff_vector < angle_connect_thresh) and (angle_diff_vector_2connect < angle_connect_thresh) and (dst_connection_pts > 0.5*dst_pts_line) and(dst_connection_pts < (dst_current_line + dst_next_line)):

                    index_pair = (index_value, index_value_next)
                    index_pair_rec.append(index_pair)
                    
                    #interpolate the points to connect two traces
                    #connect_pts = interpolate_pts_3D(connect_vector, data_range*interpolation_range_factor, int(z_range*interpolation_range_factor))
                    #connect_pts_rec.append(connect_pts)
                    
                    #print("connect_pts type {0} \n:".format(type(connect_pts)))
                    #print("connect_pts size {0} \n:".format(str(connect_pts.shape)))
                    
                    (xLinespace, yLinespace, zLinespace) = cubic_spline_interpolate(X_combine, Y_combine, Z_combine)
                    
                    coords_connect_pts = np.stack(( xLinespace, yLinespace, zLinespace ), axis = 1)
                    
                    #print("X size {0}, X combine size {1}, xLinespace size {2}, coords_connect_pts size {3}\n".format(len(X), len(X_combine), len(xLinespace), coords_connect_pts.shape))
                    
                    connect_pts_rec.append(coords_connect_pts)
                    
                    
                 
                    #print(connect_vector)
                    #print(connect_pts)
                    #print("Trace index {0} are connected by {1} points in 3D space...".format(index_pair, len(xLinespace)))
                    print("Trace index pair {} are connected in 3D space...".format(index_pair))
                    #print("dst_connection_pts {0}, dst_pts_line {1}\n".format(dst_connection_pts, dst_pts_line))
                    
                    #print("Angle difference between lines: {0}, Angle difference between connected lines: {1}.\n".format(angle_diff_vector, angle_diff_vector_2connect))
            

 
    
    #print(index_pair_rec)
    
    #print(connect_pts_rec)
    print("connect tracked traces finished...\n")
    
     
    return index_pair_rec, connect_pts_rec
    
    

if __name__ == '__main__':
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "Path to slice & trace file")
    ap.add_argument("-f", "--file", required = True, help = "Trace file")
    ap.add_argument("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")
    args = vars(ap.parse_args())

    global angle_connect_thresh, X_scale, Y_scale, Z_scale, text_size, \
    length_thresh, save_path_result, camera_azimuth, camera_elevation, \
    camera_distance, camera_focalpoint, tube_scale, interpolation_range_factor, directory_name
    
    #extract csv trace result file 
    #trace_filetype = '*.csv' 
    file_path = args["path"]
    #accquire file list
    #trace_file_list = sorted(fnmatch.filter(os.listdir(args["path"]), trace_filetype))
    
    directory_name = os.path.basename(os.path.dirname(file_path)) 
    
    #print("folder_name: {0} ...\n".format(directory_name))
    
    fname = file_path + args["file"]
    
    #print(trace_file_list)  
    
   
    
    length_thresh = 0.4
    
    angle_connect_thresh = 30
    
    X_scale = 2.0
    Y_scale = 2.0
    Z_scale = -2.5
    
    text_size = 6
    
    tube_scale = 0.8
    
    interpolation_range_factor = 0.55
    
    
    
    camera_azimuth = 45.0
    camera_elevation = 54.0
    camera_distance = 1148.0
    camera_focalpoint = np.array([218.0, 216.0, -143.0]) 
    
    #45.00000000000001, 54.73561031724528, 1148.0687014896587, array([ 218.14042502,  216.51101153, -143.00329134]))
    
    '''
    #accquire slice image file list
    image_filetype = '*.png'
    image_file_path = file_path + image_filetype
  
    imgList = sorted(glob.glob(image_file_path))
    
    #print(imgList)

    #get image size and number
    n_slices = len(imgList)
    
    if n_slices > 0 :
        width, height = imagesize.get(imgList[0])
        #print(width, height, n_slices)
    else:
        print("Empty image folder, abort!")
        sys.exit(0)
    '''

    # Create the folder to save the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #mkpath = parent_path + '/' + str('analysis_result')
    mkpath = file_path + str('analysis_result')
    mkdir(mkpath)
    save_path_result = mkpath + '/'
  
    
    #load tracked trace file
    df = pd.read_csv(fname)
    trace_pd = df[['particle', 'x', 'y', 'frame', 'size', 'mass', 'raw_mass']]
    trace_index = trace_pd["particle"].unique()
    trace_number = len(trace_index)

    # convert pd format to nump array
    trace_array = trace_pd.to_numpy()
    
    # trait computation 
    (index_rec, index_label_rec, fit_linepts_rec, indexes_length_remove,length_euclidean_rec) = trace_compute(trace_array, trace_index)

    #print("indexes_length_remove: {}".format(indexes_length_remove))
    
    #remove outliers based on euclidean length
    for index in sorted(indexes_length_remove, reverse = True):
        del index_rec[index]
        del index_label_rec[index]
        del fit_linepts_rec[index]
        del length_euclidean_rec[index]
    
    
    
    index_rec_arr = np.asarray(index_label_rec)
    
    #connect traces with similar angle and location
    (index_pair_rec, connect_pts_rec) = trace_connect(trace_array, index_rec_arr, fit_linepts_rec)
    
    #print("index_pair_rec {}".format(index_pair_rec))
    

    '''
    remove all index of connected and small length traces 
    #######################################################################
    index_pair_rec_arr = np.asarray(index_pair_rec)
    #remove traces based on trace length outliers and connected parts
    index_pair_remove_arr = np.asarray(np.unique(index_pair_rec_arr[:,1]))
    indexes_length_remove_arr = np.asarray(indexes_length_remove)
    
    #print(index_pair_remove_arr)
    
    #print(indexes_length_remove_arr)
    
    remove_indexes_combine = np.unique(np.concatenate([index_pair_remove_arr,indexes_length_remove_arr]))
    
    #print(np.asarray(index_rec))
    #print(remove_indexes_combine)

    indexes_delete = np.asarray(np.where(np.isin(np.asarray(index_rec), remove_indexes_combine)))
    
    #print(indexes_delete.flatten())
    
    index_rec_new = [j for i, j in enumerate(index_rec) if i not in indexes_delete]
    
    print(index_rec_new)
    '''
    #######################################################################################
    
    
    if(len(index_pair_rec) > 0):
        
        #print("Trace connected!")
        
        (index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec, X_rec, Y_rec, Z_rec) = trace_visualize(trace_array, index_rec_arr, fit_linepts_rec, index_pair_rec, connect_pts_rec)
    
    else:
        print("No connection parts were detected...\n")
        
        (index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec, X_rec, Y_rec, Z_rec) = trace_visualize_simple(trace_array, index_rec_arr, fit_linepts_rec, index_pair_rec, connect_pts_rec)
    
    #############################################################################################
    '''
    skel = skeletonize(image_chunk)
    
    skel = skel.astype(np.bool) #data needs to be bool
    
    G = skel2graph(skel) #create graph

    
    #compute edge properities 
    numer_total = G.size()
    
    #plot the graph, use the z component to colorcode both the edges and the nodes, scale nodes according to their degree
    #plot_graph(G,node_color_keyword='z',edge_color_keyword='z',scale_node_keyword='degree')
    
    edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'x', edge_color_keyword = 'x')
    
    #tube_surf, pts, edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'z', edge_color_keyword = 'z')
    
    #print("edge_node_unique: {0}\n".format(edge_node_unique))

    index = []
    edgecount = len(edge_node_n1_select)
    
    for i in range(edgecount):
        
        n1 = edge_node_n1_select[i]
        n2 = edge_node_n2_select[i]
        
        try:
            print("Properities of root index {0}:".format(i))
            print("node1 = {0}, node2 = {1} ".format(n1, n2))
            print("angle = {0}, length = {1} ".format(angle_select[i], length_select[i]))
            print("projection_radius = {0}\n ".format(projection_select[i]))
            
            index.append(i+1)
            
        except IndexError:
            pass
    #show the binary data
    #mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show()
    '''
    ########################################################################################
    
    #output total number of traces detected
    print("Summary: {0} unique root trajectories were detected...\n".format(len(index_rec)))
    
    #output in command window in a sum table
    trait_sum = []

    for row in zip(range(len(index_rec)), length_rec, angle_rec, diameter_rec, projection_radius_rec, index_label_rec):
       
       trait_sum.append(row)

    table = tabulate(trait_sum, headers = ['Root trace index', 'Length', 'Angle', 'Diameter', 'Projection radius' ,'Index_group'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    
    ##################################################################
    #Start of writing measured parameters as excel file 

    #parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #base_folder = os.path.basename(file_path[:-1])
    #trait_file = (parent_path + '/' + base_folder + 'root_trace_measure' + '.xlsx')
    #trait_file_csv = (parent_path + '/' + base_folder + 'root_trace_measure' + '.csv')
    
    trait_file = (save_path_result + 'root_trace_measure' + '.xlsx')
    
    trait_file_csv = (save_path_result + 'root_trace_measure' + '.csv')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

        os.remove(trait_file)
        
    else:
        # Keep presents
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'Root trace index'
        sheet.cell(row = 1, column = 2).value = 'Root trace length'
        sheet.cell(row = 1, column = 3).value = 'Root trace angle'
        sheet.cell(row = 1, column = 4).value = 'Root trace diameter'
        sheet.cell(row = 1, column = 5).value = 'Root trace projection radius'
        sheet.cell(row = 1, column = 6).value = 'Root number in total'
    
    sheet = wb.active

    for row in zip(index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec):
        sheet.append(row)
    
    sheet.cell(row = 2, column = 6).value = trace_number
    
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Trait result was saved in: " + trait_file + "\n")
    
    
    '''
    import openpyxl
    import csv

    wb = load_workbook(trait_file)
    sh = wb.get_active_sheet()
    
    #with open(trait_file_csv, 'wb') as f:
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    '''
    
    ##################################################################
    #End of writing measured parameters as excel file 
    
    
   






