"""
Version: 1.5

Summary: compute the cross section plane based on 3d model

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE

    python3 skeleton_analyze.py -p ~/example/ -m1 test_skeleton.ply -m2 test_aligned.ply -m3 ~/example/slices/ -v 0
    
    python3 skeleton_analyze.py -p /srv/test/ -m1 test_skeleton.ply -m2 test_aligned.ply -m3 /srv/test/slices/ -v 0


argument:
("-p", "--path", required=True,    help="path to *.ply model file")
("-m", "--model", required=True,    help="file name")

"""
#!/usr/bin/env python



import math

# import the necessary packages
from plyfile import PlyData, PlyElement
import numpy as np 
from numpy import interp

from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
from operator import itemgetter

from skimage.segmentation import watershed
from skimage.feature import peak_local_max
from skimage.morphology import convex_hull_image
from skimage.measure import regionprops

#from scipy.spatial import KDTree
from scipy import ndimage
import cv2


from pathlib import Path

import glob
import os
import sys
import open3d as o3d
import copy
import shutil
import argparse
from dev_code import par_config
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

from findpeaks import findpeaks

import graph_tool.all as gt

import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import math
import itertools

#from tabulate import tabulate
from rdp import rdp

# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)



# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False




#calculate length of a 3D path or curve
def path_length(X, Y, Z):

    n = len(X)
     
    lv = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1,n)]
    
    L = sum(lv)
    
    return L


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        #return angle
        
        
        if angle > 0 and angle < 45:
            return (90 - angle)
        elif angle < 90:
            return angle
        else:
            return (180- angle)
        

#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    
    hxy = np.hypot(x, y)
    
    r = np.hypot(hxy, z)
    
    elevation = np.arctan2(z, hxy)*180/math.pi
    
    azimuth = np.arctan2(y, x)*180/math.pi
    
    return r[2], azimuth[2], elevation[2]
    '''
    if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
    '''


# median-absolute-deviation (MAD) based outlier detection
def mad_based_outlier(points, thresh):
    
    if len(points.shape) == 1:
        
        points = points[:,None]
    
    median = np.median(points, axis=0)
    
    diff = np.sum((points - median)**2, axis=-1)
    
    diff = np.sqrt(diff)
    
    med_abs_deviation = np.median(diff)
    
    if med_abs_deviation == 0:
        
        modified_z_score = 0.6745 * diff / 1
    
    else:
        modified_z_score = 0.6745 * diff / med_abs_deviation
    
    return modified_z_score > thresh
    

'''
# compute nearest neighbors of the anchor_pt_idx in point cloud by building KDTree
def get_neighbors(Data_array_pt, anchor_pt_idx, search_radius):
    
    pcd = o3d.geometry.PointCloud()
    
    pcd.points = o3d.utility.Vector3dVector(Data_array_pt)
    
    #pcd.paint_uniform_color([0.5, 0.5, 0.5])
    
    #o3d.visualization.draw_geometries([pcd])
    
    # Build KDTree from point cloud for fast retrieval of nearest neighbors
    pcd_tree = o3d.geometry.KDTreeFlann(pcd)
    
    #print("Paint the 00th point red.")
    
    #pcd.colors[anchor_pt_idx] = [1, 0, 0]
    
    #print("Find its 50 nearest neighbors, paint blue.")
    
    [k, idx, _] = pcd_tree.search_knn_vector_3d(pcd.points[anchor_pt_idx], search_radius)
    
    #print("nearest neighbors = {}\n".format(sorted(np.asarray(idx[1:]))))

    return idx
'''
           


def get_pt_sel(Data_array_pt):
    
    ####################################################################
    
    # load skeleton coordinates and radius 
    Z_pt_sorted = np.sort(Data_array_pt[:,2])
    
    idx_sel = int(len(Z_pt_sorted)*0.08) 
    
    Z_mid = Z_pt_sorted[idx_sel]

    # mask
    Z_mask = (Data_array_pt[:,2] <= Z_mid) & (Data_array_pt[:,2] >= Z_pt_sorted[0]) 
    
    Z_pt_sel = Data_array_pt[Z_mask]
    
    '''
    ############################################################
    pcd_Z_mask = o3d.geometry.PointCloud()
    
    pcd_Z_mask.points = o3d.utility.Vector3dVector(Z_pt_sel)
    
    Z_mask_ply = result_path + "Z_mask.ply"
    
    o3d.visualization.draw_geometries([pcd_Z_mask])
    
    o3d.io.write_point_cloud(Z_mask_ply, pcd_Z_mask)
    ############################################################
    '''
    
    return Z_pt_sel
    
    
    

# compute dimensions of point cloud and nearest neighbors by KDTree
def get_pt_parameter(Data_array_pt, n_paths):
    
    
    ####################################################################
    pcd = o3d.geometry.PointCloud()
    
    pcd.points = o3d.utility.Vector3dVector(Data_array_pt)
    
    #pcd.paint_uniform_color([0.5, 0.5, 0.5])
    
   
    # get convex hull of a point cloud is the smallest convex set that contains all points.
    #hull, _ = pcd.compute_convex_hull()
    #hull_ls = o3d.geometry.LineSet.create_from_triangle_mesh(hull)
    #hull_ls.paint_uniform_color((1, 0, 0))
    
    # get AxisAlignedBoundingBox
    aabb = pcd.get_axis_aligned_bounding_box()
    #aabb.color = (0, 1, 0)
    
    #Get the extent/length of the bounding box in x, y, and z dimension.
    aabb_extent = aabb.get_extent()
    
    aabb_extent_half = aabb.get_half_extent()
    
    # get OrientedBoundingBox
    #obb = pcd.get_oriented_bounding_box()
    
    #obb.color = (1, 0, 0)

    #visualize the convex hull as a red LineSet
    #o3d.visualization.draw_geometries([pcd, aabb, obb, hull_ls])
    pt_diameter_max = max(aabb_extent[0], aabb_extent[1])

    pt_diameter_min = min(aabb_extent_half[0], aabb_extent_half[1])


    pt_length = (aabb_extent[2])


    pt_volume = np.pi * ((pt_diameter_max + pt_diameter_min)*0.5) ** 2 * pt_length


    pt_density = n_paths/(pt_diameter_max)**2
    
        
    pt_diameter = (pt_diameter_max + pt_diameter_min)*0.5
    
    pt_diameter_max*=par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).distance_tracking_max
    pt_diameter_min*=par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).distance_tracking_min
    
    if n_paths > 0:
        pt_diameter*=par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).distance_tracking_avg
    else:
        pt_diameter*=par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).thresh_distance_ratio
        
    return pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume, pt_density
    
    
    
#find the closest points from a points sets to a fix point using Kdtree, O(log n) 
def closest_point(point_set, anchor_point):
    
    kdtree = KDTree(point_set)
    
    (d, i) = kdtree.query(anchor_point)
    
    #print("closest point:", point_set[i])
    
    return  i, point_set[i]

    

def find_nearest(array, value):
    
    array = np.asarray(array)
    
    idx = (np.abs(array - value)).argmin()
    
    #return array[idx]
    return idx


#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.get_cmap(name,n+1)




# smooth 1d list
def smooth_data_convolve_average(arr, span):
    re = np.convolve(arr, np.ones(span * 2 + 1) / (span * 2 + 1), mode="same")

    # The "my_average" part: shrinks the averaging window on the side that 
    # reaches beyond the data, keeps the other side the same size as given 
    # by "span"
    re[0] = np.average(arr[:span])
    for i in range(1, span + 1):
        re[i] = np.average(arr[:i + span])
        re[-i] = np.average(arr[-i - span:])
    return re    



# Find closest number to k in given list
def closest(lst, K):
     
    v_closet = lst[min(range(len(lst)), key = lambda i: abs(lst[i]-K))]
    
    idx_list = [i for i, value in enumerate(lst) if value == v_closet]
    
    return idx_list, v_closet




#cluster 1D list using Kmeans
def cluster_list(list_array, n_clusters):
    
    data = np.array(list_array)
    
    if data.ndim == 1:
        
        data = data.reshape(-1,1)
    
    #kmeans = KMeans(n_clusters).fit(data.reshape(-1,1))
        
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(data)
    
    #kmeans = KMeans(n_clusters).fit(data)
    
    labels = kmeans.labels_
    
    centers = kmeans.cluster_centers_
    
    center_labels = kmeans.predict(centers)
    
    #print(kmeans.cluster_centers_)
    '''
    #visualzie data clustering 
    ######################################################
    y_kmeans = kmeans.predict(data)
    
    plt.scatter(data[:, 0], data[:, 1], c=y_kmeans, s=50, cmap='viridis')

    centers = kmeans.cluster_centers_
    
    plt.scatter(centers[:, 0], centers[:, 1], c='black', s=200, alpha=0.5)
    
    #plt.legend()

    plt.show()
    
    '''
    
    
    '''
    from sklearn.metrics import silhouette_score
    
    range_n_clusters = [2, 3, 4, 5, 6, 7, 8]
    silhouette_avg = []
    for num_clusters in range_n_clusters:
     
         # initialize kmeans
         kmeans = KMeans(n_clusters=num_clusters)
         kmeans.fit(data)
         cluster_labels = kmeans.labels_
         
         # silhouette score
         silhouette_avg.append(silhouette_score(data, cluster_labels))
    
    plt.plot(range_n_clusters,silhouette_avg,'bx-')
    plt.xlabel('Values of K')
    plt.ylabel('Silhouette score')
    plt.title('Silhouette analysis For Optimal k')
    plt.show()
    
    
    Sum_of_squared_distances = []
    K = range(2,8)
    for num_clusters in K :
        kmeans = KMeans(n_clusters=num_clusters)
        kmeans.fit(data)
        Sum_of_squared_distances.append(kmeans.inertia_)
        
    plt.plot(K,Sum_of_squared_distances,'bx-')
    plt.xlabel('Values of K') 
    plt.ylabel('Sum of squared distances/Inertia') 
    plt.title('Elbow Method For Optimal k')
    plt.show()
    '''
    ######################################################
    
    
    return labels, centers, center_labels


#remove outliers
def outlier_remove(data_list):

    #find index of k biggest elements in list
    ####################################################
    
    k = int(len(data_list) * 0.8)
    
    #print(k)
    
    #k biggest
    idx_dominant = np.argsort(data_list)[-k:]
    
    #k smallest
    #idx_dominant_dis_closest_pts = np.argsort(dis_closest_pts)[:k]
    
    #print("idx_dominant_dis_closest_pts = {}".format(idx_dominant_dis_closest_pts))
    
    #print(idx_dominant_dis_closest_pts)
    
    outlier_remove_list = [data_list[index] for index in idx_dominant] 
    
    #print("outlier_remove_list = {}".format(outlier_remove_list))
    
    return outlier_remove_list, idx_dominant
    ####################################################


# save point cloud data from numpy array as ply file, open3d compatiable format
def write_ply(path, data_numpy_array):
    
    #data_range = 100
    
    #Normalize data range for generate cross section level set scan
    #min_max_scaler = preprocessing.MinMaxScaler(feature_range = (0, data_range))

    #point_normalized = min_max_scaler.fit_transform(data_numpy_array)
    
    #initialize pcd object for open3d 
    pcd = o3d.geometry.PointCloud()
     
    pcd.points = o3d.utility.Vector3dVector(data_numpy_array)
    
    # get the model center postion
    model_center = pcd.get_center()
    
    # geometry points are translated directly to the model_center position
    pcd.translate(-1*(model_center))
    
    #write out point cloud file
    o3d.io.write_point_cloud(path, pcd, write_ascii = True)
    
    
    # check saved file
    if os.path.exists(path):
        print("Converted 3d model was saved at {0}".format(path))
        return True
    else:
        return False
        print("Model file converter failed !")
        #sys.exit(0)


# compute diameter from area
def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    
    #note: return diameter instead of radius
    return 2*radius 



# segmentation of overlapping components 
def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
    
    #localMax = peak_local_max(D,  min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    #print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels
    


# analyze cross section paramters
def crosssection_analysis(image_file):
    
    # load the image 
    imgcolor = cv2.imread(image_file)
    
    # if cross scan images are white background and black foreground
    #imgcolor = ~imgcolor
    
    # accquire image dimensions 
    height, width, channels = imgcolor.shape
    #shifted = cv2.pyrMeanShiftFiltering(image, 5, 5)

    #Image binarization by apltying otsu threshold
    img = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2GRAY)
    
    # Convert BGR to GRAY
    img_lab = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2LAB)
    
    gray = cv2.cvtColor(img_lab, cv2.COLOR_BGR2GRAY)
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    
    
    #find contours and fill contours 
    ####################################################################
    #container version
    contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #local version
    #_, contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contours_img = []
    
    #define image morphology operation kernel
    #kernel = cv2.getStructuringElement(cv2.MORPH_CROSS,(3,3))
    
    #draw all the filled contours
    for c in contours:
        
        #fill the connected contours
        contours_img = cv2.drawContours(binary, [c], -1, (255, 255, 255), cv2.FILLED)
        
        #contours_img = cv2.erode(contours_img, kernel, iterations = 5)

    #Obtain the threshold image using OTSU adaptive filter
    thresh_filled = cv2.threshold(contours_img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary_filled = cv2.threshold(contours_img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    
    
    # process filled contour images to extract connected Components
    ####################################################################
    
    contours, hier = cv2.findContours(binary_filled, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #local version
    #_, contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #draw all the filled contours
    for c in contours:
        
        #fill the connected contours
        contours_img = cv2.drawContours(binary, [c], -1, (255, 255, 255), cv2.FILLED)

    # define kernel
    connectivity = 8
    
    #find connected components 
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(contours_img, connectivity , cv2.CV_32S)
    
    #find the component with largest area 
    largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA]) 
    
    #collection of all component area
    areas = [s[4] for s in stats]
    
    # average of component area
    area_avg = sum(areas)/len(np.unique(labels))
    
    #area_avg = sum(areas)
    
    ###################################################################
    # segment overlapping components
    #make backup image
    orig = imgcolor.copy()

    #watershed based segmentaiton 
    labels = watershed_seg(contours_img, thresh_filled, min_distance_value)
    
    #labels = watershed_seg(contours_img, thresh_filled, 20)
    
    
    N_seg = len(np.unique(labels))-1
    
    
    
    #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))
    #label_hue[labels == largest_label] = np.uint8(15)
    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
    
    # save label results
    #result_file = (save_path_label + base_name + '_label.png')
    
    #print(result_file)

    #cv2.imwrite(result_file, labeled_img)
    

    
    return N_seg, labeled_img


#compute angle
def angle(directions):
    """Return the angle between vectors"""
    vec2 = directions[1:]
    vec1 = directions[:-1]

    norm1 = np.sqrt((vec1 ** 2).sum(axis=1))
    norm2 = np.sqrt((vec2 ** 2).sum(axis=1))
    cos = (vec1 * vec2).sum(axis=1) / (norm1 * norm2)   
    return np.arccos(cos)


#first derivative function
def first_derivative(x) :

    return x[2:] - x[0:-2]


#second derivative function
def second_derivative(x) :
    
    return x[2:] - 2 * x[1:-1] + x[:-2]


#compute curvature
def curvature(x, y) :

    x_1 = first_derivative(x)
    x_2 = second_derivative(x)
    y_1 = first_derivative(y)
    y_2 = second_derivative(y)
    return np.abs(x_1 * y_2 - y_1 * x_2) / np.sqrt((x_1**2 + y_1**2)**3)


#define angle computation for turing points detection
def turning_points(x, y, turning_points, smoothing_radius,cluster_radius):

    
    if smoothing_radius:
        weights = np.ones(2 * smoothing_radius + 1)
        new_x = ndimage.convolve1d(x, weights, mode='constant', cval=0.0)
        new_x = new_x[smoothing_radius:-smoothing_radius] / np.sum(weights)
        new_y = ndimage.convolve1d(y, weights, mode='constant', cval=0.0)
        new_y = new_y[smoothing_radius:-smoothing_radius] / np.sum(weights)
    else :
        new_x, new_y = x, y
        
    k = curvature(new_x, new_y)
    turn_point_idx = np.argsort(k)[::-1]
    t_points = []
    
    while len(t_points) < turning_points and len(turn_point_idx) > 0:
        t_points += [turn_point_idx[0]]
        idx = np.abs(turn_point_idx - turn_point_idx[0]) > cluster_radius
        turn_point_idx = turn_point_idx[idx]
        
    t_points = np.array(t_points)
    t_points += smoothing_radius + 1
    
    return t_points.astype(int)
    

# visualize CDF curve
def CDF_visualization(radius_avg_rec):
    
    trait_file = (label_path + '/CDF.xlsx')
    '''
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active
    '''
    if os.path.isfile(trait_file):
        # update values

        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(1, sheet.max_row+1) # for entire sheet

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active
    
    for row in enumerate(radius_avg_rec):
    
        sheet.append(row)

    #save the csv file
    wb.save(trait_file)
    
    num_bins = 10
    
    #counts, bin_edges = np.histogram(list(zip(*result)[0]), bins = num_bins, normed = True)
    counts, bin_edges = np.histogram(radius_avg_rec, bins = num_bins)
    
    # compute CDF curve
    cdf = np.cumsum(counts)
    
    #cdf = cdf / cdf[-1] #normalize
    
    x = bin_edges[1:]
    y = cdf
    
    # assembly points of CDF curve 
    trajectory = np.vstack((x, y)).T

    index_turning_pt = turning_points(x, y, turning_points = 4, smoothing_radius = 2, cluster_radius = 2)
    

    #Ramer-Douglas-Peucker Algorithm
    #simplify points et using rdp library 
    simplified_trajectory = rdp(trajectory, epsilon = 0.00200)
    #simplified_trajectory = rdp(trajectory)
    sx, sy = simplified_trajectory.T
    
    #print(sx)
    #print(sy)

    #compute plateau in curve
    dis_sy = [j-i for i, j in zip(sy[:-1], sy[1:])]
    
    #get index of plateau location
    index_sy = [i for i in range(len(dis_sy)) if dis_sy[i] <= 1.3]
    
    dis_index_sy = [j-i for i, j in zip(index_sy[:-1], index_sy[1:])]
    
    for idx, value in enumerate(dis_index_sy):
        
        if idx < len(index_sy)-2:
        
            if value == dis_index_sy[idx+1]:
            
                index_sy.remove(index_sy[idx+1])
    
    # Define a minimum angle to treat change in direction
    # as significant (valuable turning point).
    #min_angle = np.pi / 36.0
    min_angle = np.pi / 180.0
    #min_angle = np.pi /1800.0
    
    # Compute the direction vectors on the simplified_trajectory.
    directions = np.diff(simplified_trajectory, axis = 0)
    theta = angle(directions)

    # Select the index of the points with the greatest theta.
    # Large theta is associated with greatest change in direction.
    idx = np.where(theta > min_angle)[0] + 1
    
    index_turning_pt = sorted(idx)
    
    Turing_points =  np.unique(sy[idx].astype(int))
    
    #max_idx = max(max_idx)
    #print("Turing points: {0} \n".format(Turing_points))
    
    # plot CDF 
    #fig = plt.plot(bin_edges[1:], cdf, '-r', label = 'CDF')
    fig = plt.figure(1)
    #ax = fig.add_subplot(111)
    plt.grid(True)
    #plt.legend(loc = 'right')
    plt.title('CDF curve')
    plt.xlabel('Root area, unit:pixel')
    plt.ylabel('Depth of level-set, unit:pixel')
    
    plt.plot(sx, sy, 'gx-', label = 'simplified trajectory')
    plt.plot(bin_edges[1:], cdf, '-b', label = 'CDF')
    #plt.plot(sx[idx], sy[idx], 'ro', markersize = 7, label='turning points')
    
    plt.plot(sx[index_sy], sy[index_sy], 'ro', markersize = 7, label = 'plateau points')
    
    #plt.plot(sx[index_turning_pt], sy[index_turning_pt], 'bo', markersize = 7, label='turning points')
    #plt.vlines(sx[index_turning_pt], sy[index_turning_pt]-100, sy[index_turning_pt]+100, color='b', linewidth = 2, alpha = 0.3)
    #plt.legend(loc='best')
    
    result_file_CDF = label_path + '/'  + 'cdf.png'
    plt.savefig(result_file_CDF)
    plt.close()
    
    return sy


# compute number of whorls
def wholr_number_count(imgList):
    
    area_avg_rec = []
    
    density_rec = []
    
    for img in imgList:

        #(area_avg, area_sum, n_unique_labels) = root_area_label(img)
        
        (radius_avg, area_avg, density) = crosssection_analysis(img)
        
        area_avg_rec.append(area_avg)
        
        density_rec.append(density)
    
    #visualzie the CDF graph of first return value 
    list_thresh = sorted(CDF_visualization(area_avg_rec))

    #compute plateau in curve
    dis_array = [j-i for i, j in zip(list_thresh[:-1], list_thresh[1:])]
    
    #get index of plateau location
    index = [i for i in range(len(dis_array)) if dis_array[i] <= 1.3]
    
    dis_index = [j-i for i, j in zip(index[:-1], index[1:])]
    
    for idx, value in enumerate(dis_index):
        
        if idx < len(index)-2:
        
            if value == dis_index[idx+1]:
            
                index.remove(index[idx+1])
    
    
    reverse_index = sorted(index, reverse = True)
    
    #count = sum(1 for x in dis_array if float(x) <= 1.3)
    #get whorl number count 
    
    count_wholrs = int(math.ceil(len(index)))

    
    #compute wholr location
     #compute whorl location
    whorl_dis = []
    whorl_loc = []
    
    for idx, value in enumerate(reverse_index, start=1):
        
        #dis_value = list_thresh[value+1] - list_thresh[value-1]
        loc_value = int(len(imgList) - list_thresh[value+1])
        whorl_loc.append(loc_value)
        #print("adding value : {0} \n".format(str(loc_value)))
    
    
    #compute whorl distance
    whorl_dis_array = [j-i for i, j in zip(whorl_loc[:-1], whorl_loc[1:])]
    whorl_loc.extend([0, len(imgList)])
    whorl_loc = list(dict.fromkeys(whorl_loc))
    whorl_loc_ex = sorted(whorl_loc)
    
    #print("list_thresh : {0} \n".format(str(list_thresh)))
    
    
    return count_wholrs, whorl_loc_ex, sum(density_rec)/len(density_rec)




# compute parameters from cross section scan
def crosssection_scan(imgList, result_path):
    
    
    List_N_seg = []
    
    ####################################################################
    for image_file in imgList:
        
        path, filename = os.path.split(image_file)
    
        base_name = os.path.splitext(os.path.basename(filename))[0]
        
        (n_seg, labeled_img) = crosssection_analysis(image_file)
        
        print("{} N_labels = {}".format(base_name, n_seg))
        
        #result_file = (result_path + base_name + '_label.png')
        
        #cv2.imwrite(result_file, labeled_img)
        
        List_N_seg.append(n_seg)
        
    #List_N_seg_ori = List_N_seg
    
    #########################################################################
    # find the first 1 in list 
    idx_f = List_N_seg.index(1)
    
    print("first index {}\n".format(idx_f))
    
    # adjust the list values 
    if idx_f > 0:
        
        List_N_seg[0:idx_f] = [1 for x in List_N_seg[0:idx_f]]
    
    
    # find the index of peak max value
    if np.argmax(List_N_seg) > 0:
        
        List_N_seg = List_N_seg[0: np.argmax(List_N_seg)]
    
    ###################################################################
    #span = 3
  
    List_N_seg_smooth = smooth_data_convolve_average(np.array(List_N_seg), span)
    
    
    #List_N_seg_smooth = List_N_seg

    ####################################################################
    # peak detection
    # Data
    X = List_N_seg_smooth
    
    # Initialize
    fp = findpeaks(method='peakdetect', lookahead = 1)
    
    # return dictionary object
    results = fp.fit(X)

    # Plot
    fp.plot()
    
    result_file = result_path + 'N_seg.png'
    
    plt.savefig(result_file)
    
    plt.close()
    #####################################################################

    # parse the dictionary object and ge the pd.DataFrame
    df = results.get("df")
 
    # print df names
    #print(list(df.columns))      #x    y  labx  valley   peak

    # Use boolean indexing to extract peak locations to a new pd frame
    df_peak = df.loc[df['peak'] == True]

    #print(df_peak)
    
    # convert x, y values into list
    peak_y = df_peak['y'].tolist()
    peak_x = df_peak['x'].tolist()
    
    
    
    ###############################################################################
    N_cluster = 4
    
    # cluster list to find the center values
    (labels, centers, center_labels) = cluster_list(peak_y, n_clusters = N_cluster)
    
    sorted_idx = np.argsort(centers[:,0])
    
    centers = centers[sorted_idx]
    
    print("centers = {}\n".format(centers))
    
    #print("sorted_idx = {}\n".format(sorted_idx))
    
    
    #print("centers = {}\n".format(centers[sorted_idx]))
    

    
    '''
    plt.scatter([i for i in range(len(peak_y))], peak_y, c=labels)
    
    result_file = file_path + 'scatter.png'
    
    plt.savefig(result_file)
    
    plt.close()
    '''
    
    N_count = []
    idx_N_count = []
    
    # get idx of the cluster center in "y" series 
    for idx, (center_value) in enumerate(centers):
        
        (idx_list, v_closet) = closest(peak_y, center_value)
        
        
        if len(idx_list) > 0:
        
            print("idx_list = {}, v_closet = {}\n".format(peak_x[idx_list[0]], v_closet))
            
            N_count.append(v_closet)
            idx_N_count.append(peak_x[idx_list[0]])
    
    #print(N_count)
    #print(idx_N_count)
    
    if len(N_count) > 1:
        
        N_count[1] = N_count[1] if  N_count[1] > 10 else (N_count[1] + N_count[0])
    

    if len(N_count) > 3:
    
        N_1 = int(N_count[1])
        
        N_2 = int(N_count[2])
        
        R_1 = abs((idx_N_count[2] - idx_N_count[1])/len(imgList))
            
        R_2 = abs((idx_N_count[3] - idx_N_count[2])/len(imgList))
        

    elif len(N_count) > 2:
        
        N_1 = int(N_count[1])
        
        N_2 = int(N_count[2])
        
        R_1 = 0.104
        R_2 = R_1*0.62
    
    else:
        N_1 = int(N_count[0])
        
        N_2 = int(N_count[0])
        
        R_1 = 0.104
        R_2 = R_1*0.62

    n_whorl = int(len(N_count)/2)
    

    
    if N_2 > 40 and N_1 > 22:
            N_1*= 0.57
            
    if N_2> 50: 
            N_2*= 0.75
    
    print("R_1 = {}, R_2 = {}\n".format(R_1,R_2))
    
    
    if min_distance_value == 34:
        
        m = N_2
        N_1 = m
        #N_2 = m
    
    
    if R_1 > 0.19:
        R_1 = interp(R_1,[0.19,1],[0.0,0.19])
    if R_2 > 0.19:
        R_2 = interp(R_2,[0.19,1],[0.0,0.19])

    return int(N_1), int(N_2), R_1, R_2, n_whorl


def dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, select):
    
    loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(loc_start), min(loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(loc_end), min(loc_end)))
    
    max_dimension_length = abs(max(max(loc_start), max(loc_end) - min(min(loc_start), min(loc_end))))

    min_dimension_length = abs(min(min(loc_start), min(loc_end) - max(max(loc_start), max(loc_end))))
    
    #print("max_length = {}\n".format(max_length))
    
    if select == 1:
        return max_dimension_length
    else:
        return min_dimension_length



def calculate_wcss(data):
    wcss = []
    for n in range(2, 10):
        kmeans = KMeans(n_clusters=n)
        kmeans.fit(X=data)
        wcss.append(kmeans.inertia_)

    return wcss


def get_radius(radius_arr):
    
    #radius_arr = np.array([avg_radius_stem, avg_first_diameter, avg_second_diameter, avg_third_diameter]) 

    radius_arr = np.sort(radius_arr, axis = None)     

    avg_third_diameter = radius_arr[0]*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).snake_speed_min
    avg_second_diameter = radius_arr[1]*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).snake_speed_min
    avg_first_diameter = radius_arr[2]*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).snake_speed_max
    
    return avg_first_diameter, avg_second_diameter, avg_third_diameter
    
    


def optimal_number_of_clusters(wcss):
    x1, y1 = 2, wcss[0]
    x2, y2 = 20, wcss[len(wcss)-1]

    distances = []
    for i in range(len(wcss)):
        x0 = i+2
        y0 = wcss[i]
        numerator = abs((y2-y1)*x0 - (x2-x1)*y0 + x2*y1 - y2*x1)
        denominator = math.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        distances.append(numerator/denominator)

    return distances.index(max(distances)) + 0





# Skeleton analysis
def analyze_skeleton(current_path, filename_skeleton, filename_ptcloud, imgList):
    
    model_skeleton = current_path + filename_skeleton
    print("Loading 3D skeleton file {}...\n".format(filename_skeleton))
    #model_skeleton_basename = os.path.splitext(model_skeleton)[0]
    
    model_basename = Path(model_skeleton).stem
    

    
    #load the ply format skeleton file 
    try:
        with open(model_skeleton, 'rb') as f:
            plydata_skeleton = PlyData.read(f)
            num_vertex_skeleton = plydata_skeleton.elements[0].count
            N_edges_skeleton = len(plydata_skeleton['edge'].data['vertex_indices'])
            array_edges_skeleton = plydata_skeleton['edge'].data['vertex_indices']
            
            #print("Ply data structure: \n")
            #print(plydata_skeleton)
            #print("\n")
            
            print("Number of 3D points in skeleton model: {0} \n".format(num_vertex_skeleton))
            print("Number of edges: {0} \n".format(N_edges_skeleton))

        
    except:
        sys.exit("Model skeleton file does not exist!")
    
    
    #Parse ply format skeleton file and Extract the data
    Data_array_skeleton = np.zeros((num_vertex_skeleton, 4))
    
    # load skeleton coordinates and radius 
    Data_array_skeleton[:,0] = plydata_skeleton['vertex'].data['x']
    Data_array_skeleton[:,1] = plydata_skeleton['vertex'].data['y']
    Data_array_skeleton[:,2] = plydata_skeleton['vertex'].data['z']
    Data_array_skeleton[:,3] = plydata_skeleton['vertex'].data['radius']
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    radius_vtx = Data_array_skeleton[:,3]
    
    ####################################################################


    # build directed graph from skeleton/structure data
    ####################################################################
    print("Building directed graph from 3D skeleton/structure ...\n")
    G_unordered = gt.Graph(directed = True)
    
    # assert directed graph
    #print(G.is_directed())
    
    nodes = G_unordered.add_vertex(num_vertex_skeleton)
    
    G_unordered.add_edge_list(array_edges_skeleton.tolist()) 
    
    #gt.graph_draw(G_unordered, vertex_text = G_unordered.vertex_index, output = current_path + "graph_view.pdf")
    
    
    # find all end vertices by fast iteration of all vertices
    end_vlist = []
    
    end_vlist_offset = []
    
    for v in G_unordered.iter_vertices():
        
        #print(G.vertex(v).out_degree(), G.vertex(v).in_degree())
        
        if G_unordered.vertex(v).out_degree() == 0 and G_unordered.vertex(v).in_degree() == 1:
        
            end_vlist.append(v)
            
            if (v+1) == num_vertex_skeleton:
                end_vlist_offset.append(v)
            else:
                end_vlist_offset.append(v+1)
            
    #print("end_vlist = {} \n".format(end_vlist))
    #print("end_vlist_offset = {} \n".format(end_vlist_offset))
    # save graph
    ##########################################
    graph_file = (result_path + model_basename + '_graph.gt.gz')
    
    G_unordered.save(graph_file)

    #test angle calculation
    #vector1 = [0,0,1]
    # [1,0,0]
    #vector2 = [0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1]
    #print(dot_product_angle(vector1,vector2))
    #print(cart2sph(0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1)) 
    
    
    # parse all the sub branches edges and vetices, start, end vetices
    #####################################################################################
    sub_branch_list = []
    sub_branch_length_rec = []
    sub_branch_angle_rec = []
    sub_branch_start_rec = []
    sub_branch_end_rec = []
    sub_branch_projection_rec = []
    sub_branch_radius_rec = []
    
    sub_branch_xs_rec = []
    sub_branch_ys_rec = []
    sub_branch_zs_rec = []
    
    #factor = 0.77
    
    #if len(end_vlist) == len(end_vlist_offset):
        
    for idx, v_end in enumerate(end_vlist):
        
        #print(idx, v_end)
        #construct list of vertices in sub branches
        if idx == 0:
            v_list = [*range(0, int(end_vlist[idx])+1)]
        else:
            v_list = [*range(int(end_vlist[idx-1])+1, int(end_vlist[idx])+1)]
            
        # change type to interger 
        int_v_list = [int(i) for i in v_list]
        
        # current sub branch length
        sub_branch_length = path_length(X_skeleton[int_v_list], Y_skeleton[int_v_list], Z_skeleton[int_v_list])
        
        # current sub branch start and end points 
        start_v = [X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]]
        
        end_v = [X_skeleton[int_v_list[0]] - X_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Y_skeleton[int_v_list[0]] - Y_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Z_skeleton[int_v_list[0]] - Z_skeleton[int_v_list[int(len(int_v_list)-1.0)]]]
        
        # angle of current branch vs Z direction
        angle_sub_branch = dot_product_angle(start_v, end_v)*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).dis_tracking_ratio
        
        # projection radius of current branch length
        p0 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]])
        
        p1 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[-1]]])
        
        projection_radius = np.linalg.norm(p0 - p1)
        
        #radius value from fitted point cloud contours
        radius_edge = float(radius_vtx[int_v_list[0]])*1
        
        sub_branch_xs = X_skeleton[int_v_list[0]]
        sub_branch_ys = Y_skeleton[int_v_list[0]]
        sub_branch_zs = Z_skeleton[int_v_list[0]]
        
        
        # save computed parameters for each branch
        sub_branch_list.append(v_list)
        sub_branch_length_rec.append(sub_branch_length)
        sub_branch_angle_rec.append(angle_sub_branch)
        sub_branch_start_rec.append(int_v_list[0])
        sub_branch_end_rec.append(int_v_list[-1])
        sub_branch_projection_rec.append(projection_radius)
        sub_branch_radius_rec.append(radius_edge)
        
        sub_branch_xs_rec.append(sub_branch_xs)
        sub_branch_ys_rec.append(sub_branch_ys)
        sub_branch_zs_rec.append(sub_branch_zs)
    
    #print(min(sub_branch_angle_rec))
    #print(max(sub_branch_angle_rec))
    

    
    '''
    # sort branches according to the start vertex location(Z value) 
    ####################################################################
    Z_loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    
    sorted_idx_Z_loc = np.argsort(Z_loc_value)
    
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_Z_loc]
    
    
    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_Z_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_Z_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_Z_loc]

    #print("sub_branch_length_rec = {}\n".format(sub_branch_length_rec[0:20]))
    
    '''
    # sort branches according to length feature in descending order
    ####################################################################
    sorted_idx_len = np.argsort(sub_branch_length_rec)
    
    #reverse the order from accending to descending
    sorted_idx_len_loc = sorted_idx_len[::-1]

    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_len_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_len_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_len_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_len_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_len_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_len_loc]
    sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_len_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_len_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_len_loc]

    ####################################################################

    
    '''
    Z_loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    Z_loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(Z_loc_start), min(Z_loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(Z_loc_end), min(Z_loc_end)))
    
    max_length = abs(max(max(Z_loc_start), max(Z_loc_end) - min(min(Z_loc_start), min(Z_loc_end))))
    '''
    max_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    
    min_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    
    print("max_length = {} {} {}\n".format(max_length_x, max_length_y, max_length_z))
    
    print("min_length = {} {} {}\n".format(min_length_x, min_length_y, min_length_z))
    
            
    
    # construct sub branches with length and radius feature 
    ####################################################################
    combined_list = np.array(list(zip(sub_branch_length_rec, sub_branch_radius_rec))).reshape(len(sub_branch_length_rec), 2)
    
    # calculating the within clusters sum-of-squares 
    sum_of_squares = calculate_wcss(combined_list)
    
    # calculating the optimal number of clusters
    n_optimal = optimal_number_of_clusters(sum_of_squares)
    
    
    
    
    print("optimal_number_of_clusters = {}\n".format(n_optimal))
    
    n_optimal = 4
    
    # find sub branches cluster with length and radius feature 
    ####################################################################
    cluster_number = n_optimal + 4
    
    (labels, centers, center_labels) = cluster_list(combined_list, n_clusters = cluster_number)
    
    sorted_idx = np.argsort(centers[:,0])[::-1]

    print("sorted_idx = {}\n".format(sorted_idx))
    
    
    indices_level = []
    sub_branch_level = []
    sub_branch_start_level = []
    sub_branch_startZ_level = []
    sub_branch_startY_level = []
    sub_branch_startX_level = []
    
    radius_level = []
    length_level = []
    angle_level = []
    projection_level = []
    
    for idx, (idx_value) in enumerate(sorted_idx):
        
        #print(labels_length_rec.tolist().index(idx_value))
        
        #print("cluster {}, center value {}".format(idx, idx_value))
        indices = [i for i, x in enumerate(labels.tolist()) if x == idx_value]
        
        #print(indices)
        
        sub_branch_start_rec_selected = [sub_branch_start_rec[i] for i in indices]
        Z_loc = [Z_skeleton[index] for index in sub_branch_start_rec_selected]
        Y_loc = [Y_skeleton[index] for index in sub_branch_start_rec_selected]
        X_loc = [X_skeleton[index] for index in sub_branch_start_rec_selected]
        
        
        sub_loc = [sub_branch_list[index] for index in indices]
        radius_loc = [sub_branch_radius_rec[index] for index in indices]
        length_loc = [sub_branch_length_rec[index] for index in indices]
        angle_loc = [sub_branch_angle_rec[index] for index in indices]
        projection_loc = [sub_branch_projection_rec[index] for index in indices]
        
        '''
        print("max = {} min = {} ".format(max(sub_branch_start_rec_selected), min(sub_branch_start_rec_selected)))
        print("max_Z = {} min_Z = {} average = {}".format(max(Z_loc), min(Z_loc), np.mean(Z_loc)))
        print("max_radius = {} min_radius = {} average = {}".format(max(radius_loc), min(radius_loc), np.mean(radius_loc)))
        print("max_length = {} min_length = {} average = {}".format(max(length_loc), min(length_loc), np.mean(length_loc)))
        print("max_angle = {} min_angle = {} average = {}".format(max(angle_loc), min(angle_loc), np.mean(angle_loc)))
        print("max_projection = {} min_projection = {}".format(max(projection_loc), min(projection_loc), np.mean(projection_loc)))
        print("number of roots = {} {} {}\n".format(len(indices), len(Z_loc), len(radius_loc)))
        '''
        
        indices_level.append(indices)
        sub_branch_level.append(sub_loc)
        sub_branch_start_level.append(sub_branch_start_rec_selected)
        sub_branch_startZ_level.append(Z_loc)
        sub_branch_startY_level.append(Y_loc)
        sub_branch_startX_level.append(X_loc)
        
        radius_level.append(radius_loc)
        length_level.append(length_loc)
        angle_level.append(angle_loc)
        projection_level.append(projection_loc)
    
    
    
    
    ###################################################################
    
    for idx in range(cluster_number):
        
        print("sub_branch_level[{}] = {}\n".format(idx, len(sub_branch_level[idx])))
    

    

    #compute paramters

    #N_1 = int((len(indices_level[2]) + len(indices_level[1]) + len(indices_level[0]))*0.5)
    avg_first_length = np.mean(length_level[1])
    avg_first_angle = np.mean(angle_level[0])*1.12
    avg_first_diameter = np.mean(radius_level[1])*2
    avg_first_projection = np.mean(projection_level[1])
    

    
    #N_2 = int((len(indices_level[2]) + len(indices_level[1]) + len(indices_level[0]))*0.5)
    avg_second_length = np.mean(length_level[2])
    avg_second_angle = np.mean(angle_level[1])*0.88
    avg_second_diameter = np.mean(radius_level[2])*2
    avg_second_projection = np.mean(projection_level[2])
    
    avg_third_diameter = np.mean(radius_level[3])


  

    ####################################################################
    n_paths = 0
    
    for i in range(4):
        
        n_paths+= len(indices_level[i])
    
    print("paths = {}".format(n_paths))

    
    ################################################################################################################################
    (N_1, N_2, R_1, R_2, n_whorl) = crosssection_scan(imgList, current_path)
        
    print("Found cross-section image in {}\n".format(slice_image_path))

    print("N1 = {}, N2 = {}\n".format(N_1, N_2))

    print("R1 = {}, R2 = {}\n".format(R_1, R_2))
    ###################################################################
    #initialize parameters
    pt_diameter_max = pt_diameter_min = pt_length = pt_diameter = pt_eccentricity = pt_stem_diameter = pt_density = 0
        
    #load aligned ply point cloud file
    if not (filename_ptcloud is None):
        
        model_pcloud = current_path + filename_ptcloud
        
        print("Loading 3D point cloud {}...\n".format(filename_ptcloud))
        
        model_pcloud_name_base = os.path.splitext(model_pcloud)[0]
        
        pcd = o3d.io.read_point_cloud(model_pcloud)
        
        Data_array_pcloud = np.asarray(pcd.points)
        
        #print(Data_array_pcloud.shape)
        
        if pcd.has_colors():
            
            print("Render colored point cloud\n")
            
            pcd_color = np.asarray(pcd.colors)
            
            if len(pcd_color) > 0: 
                
                pcd_color = np.rint(pcd_color * 255.0)
            
            #pcd_color = tuple(map(tuple, pcd_color))
        else:
            
            print("Generate random color\n")
        
            pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
        


        #compute dimensions of point cloud data
        (pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume, pt_density) = get_pt_parameter(Data_array_pcloud, 1)
        
        s_diameter_max = pt_diameter_max
        s_diameter_min = pt_diameter_min
        s_diameter = pt_diameter
        s_length = pt_length
        
        avg_density = pt_density
        
        pt_eccentricity = (pt_diameter_min/pt_diameter_max)
        
        avg_volume = pt_volume
        
        wdis_1 = R_1*pt_length*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).kernel_max_radius
        
        wdis_2 = R_2*pt_length*par_config.match_par(min_distance_value, par_config.md_list, par_config.List_Par).kernel_min_radius
        
        #print("pt_diameter_max = {} pt_diameter_min = {} pt_diameter = {} pt_length = {} pt_volume = {}\n".format(pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume))
        
        #print("W1 = {}, W2 = {}\n".format(wdis_1, wdis_2))
        
        ###############################################
        
        pt_stem = get_pt_sel(Data_array_pcloud)
        
        (stem_diameter_max, stem_diameter_min, stem_diameter, stem_length, stem_volume, stem_density) = get_pt_parameter(pt_stem, 0)
        
        #print("setm_diameter = {} stem_length = {} \n".format(stem_diameter, stem_length))
        
        if stem_diameter >  pt_diameter*0.5:
            stem_diameter*=0.3
        
        avg_radius_stem = stem_diameter
        
        ##############################################

                
        radius_arr = np.array([avg_radius_stem, avg_first_diameter, avg_second_diameter, avg_third_diameter]) 
        radius_arr = np.sort(radius_arr, axis = None) 
        
        
        (avg_first_diameter, avg_second_diameter, avg_third_diameter) = get_radius(radius_arr)


        avg_first_projection = abs(avg_first_length*np.cos(np.pi*avg_first_angle/180))
        
        #print("avg_second_length = {} avg_second_angle = {} np.sin(avg_second_angle) = {}\n".format(avg_second_length,avg_second_angle, np.cos(np.pi*avg_second_angle/180)))
        avg_second_projection = abs(avg_second_length*np.cos(np.pi*avg_second_angle/180))
        

        
    #Skeleton Visualization pipeline
    ####################################################################
    # The number of points per line
    
    if args["visualize_model"] == 1:
    
        from mayavi import mlab
        from tvtk.api import tvtk

        N = 2
        
        mlab.figure("Structure_graph", size = (800, 800), bgcolor = (0, 0, 0))
        
        mlab.clf()
       
        # visualize 3d points
        #pts = mlab.points3d(X_skeleton[0], Y_skeleton[0], Z_skeleton[0], color = (0.58, 0.29, 0), mode = 'sphere', scale_factor = 0.15)
        #pts = mlab.points3d(X_skeleton[neighbors_idx], Y_skeleton[neighbors_idx], Z_skeleton[neighbors_idx], mode = 'sphere', color=(0,0,1), scale_factor = 0.05)
        #pts = mlab.points3d(X_skeleton[end_vlist_offset], Y_skeleton[end_vlist_offset], Z_skeleton[end_vlist_offset], color = (1,1,1), mode = 'sphere', scale_factor = 0.03)
        #pts = mlab.points3d(X_skeleton[sub_branch_start_rec_selected], Y_skeleton[sub_branch_start_rec_selected], Z_skeleton[sub_branch_start_rec_selected], color = (1,0,0), mode = 'sphere', scale_factor = 0.08)
        
        

        
        N_sublist = 4
        
        cmap = get_cmap(N_sublist)
        
        for idx in range(N_sublist):
            
            color_rgb = cmap(idx)[:len(cmap(idx))-1]
            
            print("sub_branch_level[{}] = {}\n".format(idx, len(sub_branch_level[idx])))

            
            pts = mlab.points3d(sub_branch_startX_level[idx], sub_branch_startY_level[idx], sub_branch_startZ_level[idx], color = color_rgb, mode = 'sphere', scale_factor = 0.08)
            
        
            
            for i, (sub_branch, sub_branch_start, sub_branch_radius) in enumerate(zip(sub_branch_level[idx], sub_branch_start_level[idx], radius_level[idx])):

                pts = mlab.points3d(X_skeleton[sub_branch], Y_skeleton[sub_branch], Z_skeleton[sub_branch], color = color_rgb, mode = 'sphere', scale_factor = 0.02)
                
                pts = mlab.text3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start]-0.05, str("{:.2f}".format(sub_branch_radius)), color = (0,1,0), scale = (0.01, 0.01, 0.01))
        

        #pts = mlab.points3d(Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2], mode = 'point')
        
        #mlab.show()
        
        
        #visualize point cloud model with color
        ####################################################################
        '''
        if not (filename_ptcloud is None):
            
            x, y, z = Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2] 
            
            
            pts = mlab.points3d(x,y,z, mode = 'point')
            
            sc = tvtk.UnsignedCharArray()
            
            sc.from_array(pcd_color)

            pts.mlab_source.dataset.point_data.scalars = sc
            
            pts.mlab_source.dataset.modified()
            
        '''
        
        
        #visualize skeleton model, edge, nodes
        ####################################################################
        if args["visualize_model"] == 1:
        
            x = list()
            y = list()
            z = list()
            s = list()
            connections = list()
            
            # The index of the current point in the total amount of points
            index = 0
            
            # Create each line one after the other in a loop
            #for i in range(10000):
            for i in range(N_edges_skeleton):
            #for val in vlist_path:

                if True:
                    #i = int(val)
                    #print("Edges {0} has nodes {1}, {2}\n".format(i, array_edges[i][0], array_edges[i][1]))
                  
                    x.append(X_skeleton[array_edges_skeleton[i][0]])
                    y.append(Y_skeleton[array_edges_skeleton[i][0]])
                    z.append(Z_skeleton[array_edges_skeleton[i][0]])
                    
                    x.append(X_skeleton[array_edges_skeleton[i][1]])
                    y.append(Y_skeleton[array_edges_skeleton[i][1]])
                    z.append(Z_skeleton[array_edges_skeleton[i][1]])
                    
                    # The scalar parameter for each line
                    s.append(array_edges_skeleton[i])
                    
                    # This is the tricky part: in a line, each point is connected
                    # to the one following it. We have to express this with the indices
                    # of the final set of points once all lines have been combined
                    # together, this is why we need to keep track of the total number of
                    # points already created (index)
                    #connections.append(np.vstack(array_edges[i]).T)
                    
                    connections.append(np.vstack(
                                   [np.arange(index,   index + N - 1.5),
                                    np.arange(index + 1, index + N - .5)]
                                        ).T)
                    index += 2
            
            
            # Now collapse all positions, scalars and connections in big arrays
            x = np.hstack(x)
            y = np.hstack(y)
            z = np.hstack(z)
            s = np.hstack(s)
            #connections = np.vstack(connections)

            # Create the points
            src = mlab.pipeline.scalar_scatter(x, y, z, s)

            # Connect them
            src.mlab_source.dataset.lines = connections
            src.update()

            # display the set of lines
            mlab.pipeline.surface(src, colormap = 'Accent', line_width = 5, opacity = 0.7)

            # And choose a nice view
            #mlab.view(33.6, 106, 5.5, [0, 0, .05])
            #mlab.roll(125)
        
        mlab.show()
    

    return s_diameter_max, s_diameter_min, s_diameter, avg_radius_stem, \
        N_1, avg_first_angle, avg_first_diameter, \
        N_2, avg_second_angle, avg_second_diameter, \
        wdis_1, wdis_2






if __name__ == '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", dest = "path", type = str, required = True, help = "path to *.ply model file")
    ap.add_argument("-m1", "--model_skeleton", dest = "model_skeleton", type = str, required = False, default = None, help = "skeleton file name")
    ap.add_argument("-m2", "--model_pcloud", dest = "model_pcloud", type = str, required = False, default = None, help = "point cloud model file name, same path with ply model")
    ap.add_argument("-m3", "--slice_path", dest = "slice_path", type = str, required = False,  help = "Cross section/slices image folder path in png format")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    ap.add_argument("-md", "--min_dis", required = False, type = int, default = 35,   help = "min distance for watershed segmentation")
    ap.add_argument("-dr", "--dis_tracking_ratio", required = False, type = float, default = 0.71,   help = "ratio for tracking")
    ap.add_argument("-dm", "--distance_tracking_max", required = False, type = float, default = 1.12,   help = "max distance for tracking segs")
    ap.add_argument("-di", "--distance_tracking_min", required = False, type = float, default = 0.50,   help = "min distance for tracking segs")
    ap.add_argument("-dv", "--distance_tracking_avg", required = False, type = float, default = 1.41,   help = "avg distance for tracking segs")
    ap.add_argument("-td", "--thresh_distance_ratio", required = False, type = float, default = 1.41,   help = "paramter to smooth the 1d curve")
    ap.add_argument("-sm", "--snake_speed_max", required = False, type = float, default = 0.32,   help = "min distance for watershed segmentation")
    ap.add_argument("-si", "--snake_speed_min", required = False, type = float, default = 0.90,   help = "paramter to smooth the 1d curve")
    ap.add_argument("-ki", "--kernel_max_radius", required = False, type = float, default = 1.12,   help = "max convolution kernel radius")
    ap.add_argument("-km", "--kernel_min_radius", required = False, type = float, default = 0.77,   help = "min convolution kernel radius")
    ap.add_argument("-sp", "--span", required = False, type = float, default = 3,   help = "paramter to smooth the 1d curve")    
    ap.add_argument("-v", "--visualize_model", dest = "visualize_model", required = False, type = int, default = 0, help = "Display model or not, deafult not display")
    args = vars(ap.parse_args())



    
    # setting input path to model file 
    current_path = args["path"]
    #current_path = os.path.join(current_path, '')
    
    folder_name = os.path.basename(os.path.dirname(current_path))
    
    ###################################################################
    # check file name input and default file name
    if args["model_skeleton"] is None:

        # search for file with default name
        filename_skeleton = current_path + folder_name + '_skeleton.ply'
        
        print(filename_skeleton)
        
        if os.path.isfile(filename_skeleton):
            print("Default skeleton file: {}\n".format(filename_skeleton))
            filename_skeleton = folder_name + '_skeleton.ply'
        else:
            print("Skeleton model is not found!\n")
            sys.exit()
    else:
        filename_skeleton = args["model_skeleton"]
    

    if args["model_pcloud"] is None:
        
        # search for file with default name
        filename_ptcloud = current_path + folder_name + '_aligned.ply'
        
        if os.path.isfile(filename_ptcloud):
            print("Default model file: {}\n".format(filename_ptcloud))
            filename_ptcloud = folder_name + '_aligned.ply'
        else:
            print("Aligned pointclod model is not found!\n")
            sys.exit()

    else:
        filename_ptcloud = args["model_pcloud"]
    
    
    ####################################################################
   
    
    # output path
    #result_path = args["output_path"] if args["output_path"] is not None else os.getcwd()
    
    result_path = args["output_path"] if args["output_path"] is not None else current_path
    
    result_path = os.path.join(result_path, '')
    
    # result path
    print ("results_folder: {}\n".format(result_path))
    

    #create label result file folder
    #label_path = result_path + '/label'

    # slice image path
    if args["slice_path"] is None:
        
        slice_image_path = args["path"] + '/slices/*.png'
        
    else:
        
        slice_image_path = args["slice_path"] + '*.png'
     
    # obtain image file list
    imgList = sorted(glob.glob(slice_image_path))

    n_images = len(imgList)
    
    if len(imgList) > 0:
        print("Processing {} slices of cross section images\n".format(n_images))
        
    else:
        print("Could not load cross-section image in {}\n".format(slice_image_path))
        sys.exit()

    ############################################################################################################

    
    (idx_list, min_distance_value) = (closest(par_config.md_list, args['min_dis'] ))
    
    span = args['span']
    
    print("Matched Min dis = {}\n".format(min_distance_value))
 

    
    #############################################################################################
    
    (s_diameter_max, s_diameter_min, s_diameter, avg_radius_stem, \
        N_1, avg_first_angle, avg_first_diameter, \
        N_2, avg_second_angle, avg_second_diameter, \
        wdis_1, wdis_2) = analyze_skeleton(current_path, filename_skeleton, filename_ptcloud, imgList)
    

    trait_sum = []
    
    
    trait_sum.append([s_diameter_max, s_diameter_min, s_diameter, avg_radius_stem, \
        N_1, avg_first_angle, avg_first_diameter, \
        N_2, avg_second_angle, avg_second_diameter, \
        wdis_1, wdis_2])
    

    #save reuslt file
    ####################################################################
    
    
    #trait_path = os.path.dirname(current_path + filename_skeleton)
    
    #folder_name = os.path.basename(trait_path)
    
    #print("current_path folder ={}".format(folder_name))
    
    # create trait file using sub folder name
    #trait_file = (current_path + folder_name + '_trait.xlsx')
    
    
    trait_file = (result_path + folder_name + '_trait.xlsx')
    
    #trait_file_csv = (current_path + 'trait.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        
        
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'root system diameter max'
        sheet.cell(row = 1, column = 2).value = 'root system diameter min'
        sheet.cell(row = 1, column = 3).value = 'root system diameter'
        sheet.cell(row = 1, column = 4).value = 'stem diameter'
        sheet.cell(row = 1, column = 5).value = 'number of youngest nodal root'
        sheet.cell(row = 1, column = 6).value = 'youngest nodal root angle'
        sheet.cell(row = 1, column = 7).value = 'youngest nodal root diameter'
        sheet.cell(row = 1, column = 8).value = 'number of 2nd youngest nodal root'
        sheet.cell(row = 1, column = 9).value = '2nd youngest nodal root angle'
        sheet.cell(row = 1, column = 10).value = '2nd youngest nodal root diameter'
        sheet.cell(row = 1, column = 11).value = 'youngest - 2nd youngest whorl distance'
        sheet.cell(row = 1, column = 12).value = '2nd youngest - 3rd youngest whorl distance'
        

        
    for row in trait_sum:
        sheet.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error in saving Result file\n")


    

