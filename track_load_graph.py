"""
Version: 1.5

Summary: Analyze and visualize tracked traces

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 track_load_graph.py -p /home/suxingliu/Ptvpy_test/ -f result.csv -v True


argument:
("-p", "--path", required = True, help="path to trace file")
("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")

default file format: *.csv 

"""

# Import python libraries
import numpy as np
from numpy import arctan2, sqrt
#import numexpr as ne

import glob
import fnmatch
import os, os.path
import math
import sys

import matplotlib.pyplot as plt
import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
        
from scipy.spatial import distance

from mayavi import mlab

import imagesize 
import itertools

#import warnings
#warnings.filterwarnings("ignore", category=FutureWarning)
#warnings.filterwarnings("ignore")

from tabulate import tabulate
import pandas as pd



from skimage.morphology import skeletonize
import sknw
import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import axes3d, Axes3D 

from skimage.morphology import skeletonize_3d
from network_3d import skel2graph, plot_graph
from networkx import nx

import dask
import dask.array as da


def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False

#colormap mapping
def get_cmap(n, name = 'viridis'):
    """get the color mapping""" 
    #viridis
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)
    
# compute the path length along the trace
def pathlength(x,y,z):

    n = len(x)
    
    lv = [sqrt((x[i]-x[i-1])**2 + (y[i]-y[i-1])**2 + (z[i]-z[i-1])**2) for i in range(n)]
    
    return sum(lv)

# compute distance between consective point sets
def points_seg_length(coords):
    
    d = np.diff(coords, axis=0)
    
    segdists = np.sqrt((d ** 2).sum( axis = 1))
    
    # calculate length of line
    #l = np.sqrt( np.diff(X)**2 + np.diff(Y)**2 + np.diff(Z)**2 )
    
    return sum(segdists)


# compute angle between two 3D points
def points_angle(x, y, z, line_length):
    
    theta_offest = np.zeros(4)
    r_offest = np.zeros(4)
        
    #calculate angles
    for offest in range(0,4):
        
        interval = int(((offest+1)*0.25)*line_length)  
        
        if interval >= len(x):
            interval = len(x)-1
        
        cx = x[interval] - x[0]
        cy = y[interval] - y[0]  
        cz = z[interval] - z[0]
        
        (r,theta,phi) = asSpherical(cx, cy, cz)
        
        if theta > 90:
            theta = 180 -theta

        theta_offest[offest] = theta

        r_offest[offest] = r

    return r_offest[2], theta_offest[2], phi

#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    hxy = np.hypot(x, y)
    r = np.hypot(hxy, z)
    elevation = np.arctan2(z, hxy)*180/math.pi
    azimuth = np.arctan2(y, x)*180/math.pi
    return r[2], azimuth[2], elevation[2]

#coordinates transformation from cartesian coords to sphere coord system
def appendSpherical_np(xyz):
    
    ptsnew = np.hstack((xyz, np.zeros(xyz.shape)))
    xy = xyz[:,0]**2 + xyz[:,1]**2
    ptsnew[:,3] = np.sqrt(xy + xyz[:,2]**2)
    ptsnew[:,4] = np.arctan2(np.sqrt(xy), xyz[:,2]) # for elevation angle defined from Z-axis down
    #ptsnew[:,4] = np.arctan2(xyz[:,2], np.sqrt(xy)) # for elevation angle defined from XY-plane up
    ptsnew[:,5] = np.arctan2(xyz[:,1], xyz[:,0])
    return ptsnew[:,3],ptsnew[:,4],ptsnew[:,5]

# remove otliers
def reject_outliers(data, m = 2.):
    
    d = np.abs(data - np.median(data))
    mdev = np.median(d)
    s = d / (mdev if mdev else 1.)
    return data[s < m]

# visualize the traces and return their properities
def trace_visualize(trace_array, array_index_rec, n_trace, fit_linepts_rec, index_pair_rec, connect_pts_rec):
    
    # properities initialization
    index_rec_new = []
    length_rec_new = []
    angle_rec_new = []
    diameter_rec_new = []
    projection_radius_rec_new = []
    index_label_rec_new = []
    color_rec = []
    
    
    image_chunk = np.zeros((416, 414, 282))
    
    if args["visualize"]:
    
        print("Visualizing tracked traces...")
     
        f = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))

        cmap = get_cmap(n_trace)
        
        #color = np.arange(0, 1, 1/n_trace).tolist()
        
        index_pair_rec_arr = np.asarray(index_pair_rec)
        
        #print(len(index_pair_rec_arr))
        
        #import random
        #color = list(random.sample(range(0, 1), n_trace))
 
        for idx, index_value in enumerate(array_index_rec):

            #print(idx, index_value)
            
            X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]
            Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]
            Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/3
            
            radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
            
            trace_radius = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
            
            #print(type(trace_radius))
            
            scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
            
            #print("fitting lines")
            #print(fit_linepts_rec[idx][0], fit_linepts_rec[idx][1])
            
            arr_lines = np.asarray(fit_linepts_rec[idx])
            
            
            line_x = arr_lines[:,0]
            line_y = arr_lines[:,1]
            line_z = arr_lines[:,2]
            
            ####################################################################Visualiztion
            
            #generate different colors for different traces
            color_rgb = cmap(idx)[:len(cmap(idx))-1]
            
            color_rec.append(color_rgb)
            
          
            index_pair = -1
            
            if(len(index_pair_rec_arr) > 0):
                
                #print("index_pair_rec_arr not empty")
            # draw connection parts and pair trace
                if(len(list(filter (lambda x : x == index_value, index_pair_rec_arr[:,0]))) > 0):
                   
                   #print("found index {}".format(index_value))

                   index_match = np.where(np.unique(index_pair_rec_arr[:,0]) == index_value)
                   
                   index_pair = int(index_pair_rec_arr[int(index_match[0]), 1])
                   #print("found index pair {}".format(index_pair))
                   
                   #print(index_match)
                   
                   index_connect_pts_rec = int(index_match[0])
                   
                   #print("index value {}".format(index_connect_pts_rec))
                   
                   coords_arr = np.asarray(connect_pts_rec[index_connect_pts_rec])
                   
                  
                   x = coords_arr[:,0]
                   y = coords_arr[:,1]
                   z = coords_arr[:,2]
                  
                   
                   x_pair = trace_array[np.where(trace_array[:,0] == index_pair)][:,1]
                   y_pair = trace_array[np.where(trace_array[:,0] == index_pair)][:,2]
                   z_pair = trace_array[np.where(trace_array[:,0] == index_pair)][:,3]/3
                   
                   X_combine = np.hstack([X, x, x_pair])
                   Y_combine = np.hstack([Y, y, y_pair])
                   Z_combine = np.hstack([Z, z, z_pair])
                   
                   #print("shape {} {} {} {} \n".format(X_combine.shape, X.shape, x.shape, x_pair.shape))
                   
                   radius_mean_pair = np.mean(trace_array[np.where(trace_array[:,0] == index_pair)][:,4])
                   
                   
                   #Draw 3d points and lines 
                   pts = mlab.points3d(x, y, z, color = color_rgb, mode = 'point')
                   pts = mlab.points3d(x_pair, y_pair, z_pair, color = color_rgb, mode = 'point')
                   pts.actor.property.set(point_size = 5.5)
                   
                   pts = mlab.plot3d(x, y, z, color = color_rgb, opacity = 0.3, representation = 'wireframe', transparent = True, tube_radius = 1*radius_mean)
                   pts = mlab.plot3d(x_pair, y_pair, z_pair, color = color_rgb, opacity = 0.3, representation = 'wireframe', transparent = True, tube_radius = 1*radius_mean_pair)
                  
                   
                else:
                   X_combine = X
                   Y_combine = Y
                   Z_combine = Z

                # Skip draw connection pair parts
                if (len(list(filter (lambda x : x == index_value, index_pair_rec_arr[:,1]))) > 0):
                    print("Trace {} was connected...".format(idx))
                else:
                    #Draw 3d points and lines 
                    pts = mlab.points3d(X, Y, Z, color = color_rgb, mode = 'point')
                    pts.actor.property.set(point_size = 5.5)

                    #Draw 3d points and lines 
                    #pts = mlab.points3d(X, Y, Z, color = color_rgb, mode = 'point')
                    #pts.actor.property.set(point_size = 5.5)
                    #pts.glyph.color_mode = 'color_by_scalar'

                    pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 1*radius_mean)
                    
                    #pts = mlab.plot3d(line_x, line_y, line_z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 1*radius_mean)

                    pts = mlab.text3d(X[0], Y[0], Z[0], str(idx+1), scale = (4, 4, 4), color = (1, 0.0, 0.0))
                
                ############################################################################################
                
                #recompute the connected trace properities
                coords = np.stack(( X_combine, Y_combine, Z_combine ), axis = 1)
        
                #print(coords.shape)

                 #compute line angle
                line_length = points_seg_length(coords)
                
                (r, azimuth, elevation) = cart2sph(X_combine, Y_combine, Z_combine)
                
                if azimuth > 90:
                    angle = 180 - azimuth
                elif azimuth < 0:
                    angle = 90 + azimuth
                else:
                    angle = azimuth

                print("Trace {0} properities:".format(idx))
                print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X_combine), line_length, angle))
                

                '''
                # record all parameters
                index_rec_new.append(idx)
                length_rec_new.append(line_length)
                angle_rec_new.append(angle)
                diameter_rec_new.append(radius_mean)
                projection_radius_rec_new.append(r)
                index_label_rec_new.append(index_value)
                '''
                #########################################################################################

            else:
                
                pts = mlab.plot3d(X, Y, Z, color = color_rgb, opacity = 0.3, representation = 'surface', transparent = True, tube_radius = 1*radius_mean)
            
            
            image_chunk[X_combine.astype(int), Y_combine.astype(int), Z_combine.astype(int)] = 1
            
            
            # record all parameters
            index_rec_new.append(idx)
            length_rec_new.append(line_length)
            angle_rec_new.append(angle)
            diameter_rec_new.append(radius_mean)
            projection_radius_rec_new.append(r)
            index_label_rec_new.append(index_value)
            
            ####################################################################Pipeline Visualiztion
            '''
            N = len(X)
            
            # We create a list of positions and connections, each describing a line.
            # We will collapse them in one array before plotting.
            x = list()
            y = list()
            z = list()
            s = list()
            connections = list()

            # The index of the current point in the total amount of points
            index = 0

            # Create each line one after the other in a loop
            for i in range(N):
                x.append(X[i])
                y.append(Y[i])
                z.append(Z[i])
                s.append(scalars[idx])
                # This is the tricky part: in a line, each point is connected
                # to the one following it. We have to express this with the indices
                # of the final set of points once all lines have been combined
                # together, this is why we need to keep track of the total number of
                # points already created (index)
                connections.append(np.vstack(
                                   [np.arange(index,   index + N - 1.5),
                                    np.arange(index + 1, index + N - .5)]
                                        ).T)
            index += N
            
             # Now collapse all positions, scalars and connections in big arrays
            x = np.hstack(x)
            y = np.hstack(y)
            z = np.hstack(z)
            s = np.hstack(s)
            connections = np.vstack(connections)

            # Create the points
            src = mlab.pipeline.scalar_scatter(x, y, z, s)

            # Connect them
            src.mlab_source.dataset.lines = connections
            src.update()

            # The stripper filter cleans up connected lines
            #lines = mlab.pipeline.stripper(src)
            
            #lines = mlab.pipeline.tube(src, tube_radius=0.005, tube_sides=6)

            # Finally, display the set of lines
            mlab.pipeline.surface(src, colormap = 'Accent', line_width = 3, opacity =.4)
            '''
            #############################################################################
            
            
        #show model
        mlab.show()
    
               
    return index_rec_new, length_rec_new, angle_rec_new, diameter_rec_new, projection_radius_rec_new, index_label_rec_new, image_chunk
    


# SVD fiting lines to 3D points
def line_fiting_3D(data):

    # Calculate the mean of the points, i.e. the 'center' of the cloud
    datamean = data.mean(axis = 0)

    # Do an SVD on the mean-centered data.
    uu, dd, vv = np.linalg.svd(data - datamean)

    # Now vv[0] contains the first principal component, i.e. the direction
    # vector of the 'best fit' line in the least squares sense.

    # Now generate some points along this best fit line, for plotting.

    # use -7, 7 since the spread of the data is roughly 14
    # and we want it to have mean 0 (like the points we did
    # the svd on). Also, it's a straight line, so we only need 2 points.
    linepts = vv[0] * np.mgrid[-7:7:2j][:, np.newaxis]

    # shift by the mean to get the line in the right place
    linepts += datamean
    
    return linepts


# compute trace properities
def trace_compute(trace_array, trace_index, trace_number):
    
    #import scipy.optimize as optimize
    
    print("Processing tracked trace properities...")
    
    #initialize parameters
    index_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius_rec = []
    fit_linepts_rec = []


    for idx, index_value in enumerate(trace_index):

        #print(idx, index_value)
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/3

        #traits measurement
        ##################################################################
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]

        #compute line length
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        #print(coords.shape)
        
        #A = np.array([(19,20,24), (10,40,28), (10,50,31)])

        #guess = (1,1)
        
        #Optimal values for the parameters so that the sum of the squared residuals of f(xdata, *popt) - ydata is minimized.
        #popt, pcov = optimize.curve_fit(func, coords[:,:2], coords[:,2])
        
        fit_linepts = line_fiting_3D(coords)
        
        #print(fit_linepts[0], fit_linepts[1])
        
        #line_length = pathlength(X, Y, Z)
        
         #compute line angle
        line_length = points_seg_length(coords)
        
        (r, azimuth, elevation) = cart2sph(X, Y, Z)
        
        if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth

        #print("Trace {0} properities:".format(index_value))
        #print("Number of points:{}, Length:{:.2f}, Angle:{:.2f} \n".format(len(X), line_length, angle))
        #print(popt)
        
        #print("Angle:{0} {1} {2}\n".format(r, theta, phi))

        # record all parameters
        index_rec.append(index_value)
        length_rec.append(line_length)
        angle_rec.append(angle)
        diameter_rec.append(radius_mean)
        projection_radius_rec.append(r)
        fit_linepts_rec.append(fit_linepts)
 
    # remove outlier trace
    avg_length = sum(length_rec) / len(length_rec) 
    
    indexes_length_remove = [idx for idx, element in enumerate(length_rec) if element < (avg_length*0.8)]
    
    
    for index in sorted(indexes_length_remove, reverse = True):
        del index_rec[index]
        del length_rec[index]
        del angle_rec[index]
        del diameter_rec[index]
        del projection_radius_rec[index]
        del fit_linepts_rec[index]

    #print("avg_length is {0}".format(avg_length))
    #print("indexes is {0}".format(indexes))

    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, fit_linepts_rec, indexes_length_remove

def angle_between(vector_1, vector_2):
    
    unit_vector_1 = vector_1 / np.linalg.norm(vector_1)

    unit_vector_2 = vector_2 / np.linalg.norm(vector_2)

    dot_product = np.dot(unit_vector_1, unit_vector_2)

    angle = np.arccos(dot_product)
 
    return np.rad2deg(angle)


def interpolate_pts_3D(data, data_range, z_range):

    # Calculate the mean of the points, i.e. the 'center' of the cloud
    datamean = data.mean(axis=0)

    # Do an SVD on the mean-centered data.
    uu, dd, vv = np.linalg.svd(data - datamean)

    # Now vv[0] contains the first principal component, i.e. the direction
    # vector of the 'best fit' line in the least squares sense.
       # Now generate some points along this best fit line, for plotting.

    # I use -7, 7 since the spread of the data is roughly 14
    # and we want it to have mean 0 (like the points we did
    # the svd on). Also, it's a straight line, so we only need 2 points.
    linepts = vv[0] * np.mgrid[(-1*data_range):data_range:complex(0,z_range)][:, np.newaxis]

    # shift by the mean to get the line in the right place
    linepts += datamean
    
    return linepts


def connect_trace(trace_array, array_index_rec, n_trace, fit_linepts_rec):
    
    #import similaritymeasures
    #from scipy.spatial import distance
    
    print("connecting tracked traces...")
    
    index_pair_rec = []
    connect_pts_rec = []
    
    for idx, index_value in enumerate(array_index_rec):

        #print(idx, index_value)
        
        X = trace_array[np.where(trace_array[:,0] == index_value)][:,1]
        Y = trace_array[np.where(trace_array[:,0] == index_value)][:,2]
        Z = trace_array[np.where(trace_array[:,0] == index_value)][:,3]/3
        
        #stack array
        coords = np.stack(( X, Y, Z ), axis = 1)
        
        radius_mean = np.mean(trace_array[np.where(trace_array[:,0] == index_value)][:,4])
        
        scalars = trace_array[np.where(trace_array[:,0] == index_value)][:,4]
        
        #extract coordinates of the fitted line of current trace
        arr_lines = np.asarray(fit_linepts_rec[idx])
        
        vector_x = arr_lines[0,0] - arr_lines[1,0]
        vector_y = arr_lines[0,1] - arr_lines[1,1]
        vector_z = arr_lines[0,2] - arr_lines[1,2]
        
        vector_line = np.array([vector_x, vector_y, vector_z])
        
        vector_vertical = (([0, 0, 1]))

        angle = angle_between(vector_line, vector_vertical)

        #print("coordinate {0} {1} {2}:\n".format(np.max()))
        
        #print("angle index {0} is {1}:".format(index_value, angle))

        
        for idx_next, index_value_next in enumerate(array_index_rec):
            
            if (idx != idx_next): 
                
                #extract coordinates of the next trace
                X_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,1]
                Y_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,2]
                Z_next = trace_array[np.where(trace_array[:,0] == index_value_next)][:,3]/3
                
                #stack array
                coords_next = np.stack(( X_next, Y_next, Z_next ), axis = 1)
                
                #extract coordinates of the fitted line of next trace
                arr_lines_next = np.asarray(fit_linepts_rec[idx_next])

                vector_x_next = arr_lines_next[0,0] - arr_lines_next[1,0]
                vector_y_next = arr_lines_next[0,1] - arr_lines_next[1,1]
                vector_z_next = arr_lines_next[0,2] - arr_lines_next[1,2]
                
                vector_line_next = np.array([vector_x_next, vector_y_next, vector_z_next])
                
                #angle between two traces
                angle_diff_vector = angle_between(vector_line, vector_line_next)
                
                #angle between current trace and new trace to be connected
                angle_diff_vector_2connect = angle_between(vector_line, np.subtract(coords[len(coords)-1], coords_next[0]))

                connect_vector = np.vstack((coords[len(coords)-1],coords_next[0]))
                #print("coordinate current {0}".format(coords[len(coords)-1]))
                #print("coords_next {0}:".format(coords_next[0]))
                #print(connect_vector)
                #print("angle_diff_vector_2connect {0}:\n".format(angle_diff_vector_2connect))
               

                #compute the data range to be connected
                x_diff = np.diff(connect_vector[:,0])
                y_diff = np.diff(connect_vector[:,1])
                z_diff = np.diff(connect_vector[:,2])
                
                data_range = max(abs(x_diff), abs(y_diff), abs(z_diff))
                
                #print(max_connect_pts[:,2])
                #print(min_connect_pts[:,2])
                
                squared_dist = np.sum((coords[len(coords)-1]-coords_next[0])**2, axis=0)
                dist = np.sqrt(squared_dist)
                
                #z_range = abs(max_connect_pts[:,2] - min_connect_pts[:,2])
                
                #compute number of points to be connected
                z_range = dist
                
                # quantify the difference between the two curves using PCM
                #pcm_diff = similaritymeasures.pcm(coords, coords_next)
                #dst_diff = distance.euclidean(coords[len(coords)-1], coords_next[0])
                
                #if (angle_diff_vector < 20) and (pcm_diff < 105) and (dst_diff < 100):
                if (angle_diff_vector < angle_connect_thresh) and (angle_diff_vector_2connect < angle_connect_thresh) :

                    index_pair = (index_value, index_value_next)
                    index_pair_rec.append(index_pair)
                    
                    #interpolate the points to connect two traces
                    connect_pts = interpolate_pts_3D(connect_vector, data_range*0.8, int(z_range*0.8))
                    connect_pts_rec.append(connect_pts)
                   
                    #print(connect_vector)
                    #print(connect_pts)
                    print("Trace index {0} are connected by {1} points in 3D space...".format(index_pair, int(z_range)))
                    print("Angle difference between lines: {0}, Angle difference between connected lines: {1}.\n".format(angle_diff_vector, angle_diff_vector_2connect))
                    
                
    #print(index_pair_rec)
    #print(connect_pts_rec)
    print("connect tracked traces fnished...\n")
    
    return index_pair_rec, connect_pts_rec
    
    

if __name__ == '__main__':
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "Path to slice & trace file")
    ap.add_argument("-f", "--file", required = True, help = "Trace file")
    ap.add_argument("-v", "--visualize", required = False, default = False, type = bool, help = "Visualize result or not")
    args = vars(ap.parse_args())

    #extract csv trace result file 
    #trace_filetype = '*.csv' 
    file_path = args["path"]
    #accquire file list
    #trace_file_list = sorted(fnmatch.filter(os.listdir(args["path"]), trace_filetype))
    
    fname = file_path + args["file"]
    
    #print(trace_file_list)  
    
    global angle_connect_thresh 
    
    angle_connect_thresh = 60
    
    '''
    #accquire slice image file list
    image_filetype = '*.png'
    image_file_path = file_path + image_filetype
  
    imgList = sorted(glob.glob(image_file_path))
    
    #print(imgList)

    #get image size and number
    n_slices = len(imgList)
    
    if n_slices > 0 :
        width, height = imagesize.get(imgList[0])
        #print(width, height, n_slices)
    else:
        print("Empty image folder, abort!")
        sys.exit(0)
    '''

    # Create the folder to save the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    mkpath = parent_path + '/' + str('analysis_result')
    mkdir(mkpath)
    save_path_result = mkpath + '/'
  
    
    #load tracked trace file
    df = pd.read_csv(fname)
    trace_pd = df[['particle', 'x', 'y', 'frame', 'size', 'mass', 'raw_mass']]
    trace_index = trace_pd["particle"].unique()
    trace_number = len(trace_index)

    # convert pd format to nump array
    trace_array = trace_pd.to_numpy()
    
    # trait computation 
    (index_rec, length_rec, angle_rec, diameter_rec, projection_radius_rec, fit_linepts_rec, indexes_length_remove) = trace_compute(trace_array, trace_index, trace_number)
    
    index_rec_arr = np.asarray(index_rec)

    #connect traces with similar angle and location
    (index_pair_rec, connect_pts_rec) = connect_trace(trace_array, index_rec_arr, len(index_rec_arr), fit_linepts_rec)
    
    '''
    remove all index of connected and small length traces 
    #######################################################################
    index_pair_rec_arr = np.asarray(index_pair_rec)
    #remove traces based on trace length outliers and connected parts
    index_pair_remove_arr = np.asarray(np.unique(index_pair_rec_arr[:,1]))
    indexes_length_remove_arr = np.asarray(indexes_length_remove)
    
    #print(index_pair_remove_arr)
    
    #print(indexes_length_remove_arr)
    
    remove_indexes_combine = np.unique(np.concatenate([index_pair_remove_arr,indexes_length_remove_arr]))
    
    #print(np.asarray(index_rec))
    #print(remove_indexes_combine)

    indexes_delete = np.asarray(np.where(np.isin(np.asarray(index_rec), remove_indexes_combine)))
    
    #print(indexes_delete.flatten())
    
    index_rec_new = [j for i, j in enumerate(index_rec) if i not in indexes_delete]
    
    print(index_rec_new)
    #######################################################################################
    '''
    
    
    (index_rec_new, length_rec_new, angle_rec_new, diameter_rec_new, projection_radius_rec_new, index_label_rec_new, image_chunk) = trace_visualize(trace_array, index_rec_arr, len(index_rec_arr), fit_linepts_rec, index_pair_rec, connect_pts_rec)
    
    
    '''
    skel = skeletonize(image_chunk)
    
    skel = skel.astype(np.bool) #data needs to be bool
    
    G = skel2graph(skel) #create graph

    
    #compute edge properities 
    numer_total = G.size()
    
    #plot the graph, use the z component to colorcode both the edges and the nodes, scale nodes according to their degree
    #plot_graph(G,node_color_keyword='z',edge_color_keyword='z',scale_node_keyword='degree')
    
    edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'x', edge_color_keyword = 'x')
    
    #tube_surf, pts, edge_node_n1_select, edge_node_n2_select, angle_select, length_select, projection_select = plot_graph(G, node_color_keyword = 'z', edge_color_keyword = 'z')
    
    #print("edge_node_unique: {0}\n".format(edge_node_unique))

    index = []
    edgecount = len(edge_node_n1_select)
    
    for i in range(edgecount):
        
        n1 = edge_node_n1_select[i]
        n2 = edge_node_n2_select[i]
        
        try:
            print("Properities of root index {0}:".format(i))
            print("node1 = {0}, node2 = {1} ".format(n1, n2))
            print("angle = {0}, length = {1} ".format(angle_select[i], length_select[i]))
            print("projection_radius = {0}\n ".format(projection_select[i]))
            
            index.append(i+1)
            
        except IndexError:
            pass
        
    
    
                
    #show the binary data
    #mlab.contour3d(image_chunk.astype(np.float),contours=[.5],opacity=.5,color=(1,1,1))
    
    mlab.show()
    
    
    '''
    #output total number of traces detected
    print("Summary: {0} unique root trajectories were detected...\n".format(len(index_rec_new)))
    
    #output in command window in a sum table
    trait_sum = []
    
    for row in zip(range(len(index_rec_new)), length_rec_new, angle_rec_new, diameter_rec_new, projection_radius_rec_new, index_label_rec_new):
       trait_sum.append(row)

    table = tabulate(trait_sum, headers = ['Root trace index', 'Length', 'Angle', 'Diameter', 'Projection radius' ,'Index_group'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    
    
    '''
    ##################################################################
    #Start of writing measured parameters as excel file 

    #parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    #base_folder = os.path.basename(file_path[:-1])
    #trait_file = (parent_path + '/' + base_folder + 'root_trace_measure' + '.xlsx')
    #trait_file_csv = (parent_path + '/' + base_folder + 'root_trace_measure' + '.csv')
    
    trait_file = (save_path_result + 'root_trace_measure' + '.xlsx')
    
    trait_file_csv = (save_path_result + 'root_trace_measure' + '.csv')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

        os.remove(trait_file)
        
    else:
        # Keep presents
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'Root trace index'
        sheet.cell(row = 1, column = 2).value = 'Root trace length'
        sheet.cell(row = 1, column = 3).value = 'Root trace angle'
        sheet.cell(row = 1, column = 4).value = 'Root trace diameter'
        sheet.cell(row = 1, column = 5).value = 'Root trace projection radius'
        sheet.cell(row = 1, column = 6).value = 'Root number in total'
    
    sheet = wb.active

    for row in zip(index_rec, length_rec, angle_rec, diameter_rec, projection_radius):
        sheet.append(row)
    
    sheet.cell(row = 2, column = 6).value = trace_number
    
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Trait result was saved in: " + trait_file + "\n")
    
    import openpyxl
    import csv

    wb = load_workbook(trait_file)
    sh = wb.get_active_sheet()
    
    #with open(trait_file_csv, 'wb') as f:
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    
    
    ##################################################################
    #End of writing measured parameters as excel file 
    '''
    
   






