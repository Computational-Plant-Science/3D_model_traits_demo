"""
Version: 1.5

Summary: Analyze the 3d model using Graph representation and Quaternions  

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 skeleton_graph.py -p ~/example/quaternion/tiny/ -m1 tiny_skeleton.ply -v 1 -r 3 -tq 0

    python3 skeleton_graph.py -p ~/example/quaternion/species_comp/Maize_B101/ -m1 B101_skeleton.ply -r 50 -tq 0


"""
#!/usr/bin/env python

from mayavi import mlab
from tvtk.api import tvtk

# import the necessary packages
from plyfile import PlyData, PlyElement
import numpy as np 
from numpy import interp


from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from operator import itemgetter
import argparse
    

from skimage.segmentation import watershed
from skimage.feature import peak_local_max
from skimage.morphology import convex_hull_image
from skimage.measure import regionprops

from scipy.spatial import KDTree
from scipy import ndimage
from scipy.spatial.transform import Rotation as R
from scipy.spatial import Delaunay


import random

import cv2

import glob
import os
import sys
import open3d as o3d
import copy
import shutil

import graph_tool.all as gt

from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import math
import itertools

#from tabulate import tabulate
from rdp import rdp

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import csv


import pandas as pd
import plotly
import plotly.graph_objs as go

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
from matplotlib import colors
from matplotlib.ticker import PercentFormatter

from pyquaternion import Quaternion

#validation for average_quaternions
from sksurgerycore.algorithms.averagequaternions import average_quaternions


# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)



from sklearn import metrics

from spherecluster import SphericalKMeans
from spherecluster import VonMisesFisherMixture
from spherecluster import sample_vMF


def graph_plot(x, y, z, start_idx, end_idx, color_rgb):
    """ Show the graph edges using Mayavi

        Parameters
        -----------
        x: ndarray
            x coordinates of the points
        y: ndarray
            y coordinates of the points
        z: ndarray
            z coordinates of the points
        edge_scalars: ndarray, optional
            optional data to give the color of the edges.
        kwargs:
            extra keyword arguments are passed to quiver3d.
    """
    vec = mlab.quiver3d(x[start_idx],
                        y[start_idx],
                        z[start_idx],
                        x[end_idx] - x[start_idx],
                        y[end_idx] - y[start_idx],
                        z[end_idx] - z[start_idx],
                        color = color_rgb,
                        mode = '2darrow',
                        line_width = 1, 
                        scale_factor = 1,
                        opacity = 1.0)
                        
    #if edge_scalars is not None:
        #vec.glyph.color_mode = 'color_by_scalar'
    
    return vec




# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False




#calculate length of a 3D path or curve
def path_length(X, Y, Z):

    n = len(X)
     
    lv = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1,n)]
    
    L = sum(lv)
    
    return L


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        #return angle
        
        
        if angle > 0 and angle < 45:
            return (90 - angle)
        elif angle < 90:
            return angle
        else:
            return (180- angle)
        
    
#find the closest points from a points sets to a fix point using Kdtree, O(log n) 
def closest_point(point_set, anchor_point):
    
    kdtree = KDTree(point_set)
    
    (d, i) = kdtree.query(anchor_point)
    
    #print("closest point:", point_set[i])
    
    return  (i, point_set[i])


#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)


#compute muti-dimension distance between two points
def multiDimenDist(point1,point2):
   #find the difference between the two points, its really the same as below
   deltaVals = [point2[dimension]-point1[dimension] for dimension in range(len(point1))]
   runningSquared = 0
   #because the pythagarom theorm works for any dimension we can just use that
   for coOrd in deltaVals:
       runningSquared += coOrd**2
   return runningSquared**(1/2)
   

#compute two vectors form two points in muti-dimension space
def findVec(point1,point2,unitSphere = False):
  #setting unitSphere to True will make the vector scaled down to a sphere with a radius one, instead of it's orginal length
  finalVector = [0 for coOrd in point1]
  for dimension, coOrd in enumerate(point1):
      #finding total differnce for that co-ordinate(x,y,z...)
      deltaCoOrd = point2[dimension]-coOrd
      #adding total difference
      finalVector[dimension] = deltaCoOrd
  if unitSphere:
      totalDist = multiDimenDist(point1,point2)
      unitVector =[]
      for dimen in finalVector:
          unitVector.append( dimen/totalDist)
      return unitVector
  else:
      return finalVector



#get rotation matrix between two vectors using scipy
def get_rotation_matrix(vec2, vec1):
    """ Find the rotation matrix that aligns vec1 to vec2
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    
    
    vec1 = np.reshape(vec1, (1, -1))
    
    vec2 = np.reshape(vec2, (1, -1))
    
    r = R.align_vectors(vec2, vec1)
        
    return r[0].as_matrix()
    
    '''
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return numpy.eye(3) #cross of all zeros only occurs on identical directions
    '''

    


#Find the rotation matrix that aligns vec1 to vec2
def rotation_matrix_from_vectors(vec1, vec2):
    """ 
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return np.eye(3) #cross of all zeros only occurs on identical directions


# Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
# The quaternions are arranged as (w,x,y,z), with w being the scalar
# The result will be the average quaternion of the input. Note that the signs
# of the output quaternion can be reversed, since q and -q describe the same orientation
def averageQuaternions(Q):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))

    for i in range(0,M):
        q = Q[i,:]
        # multiply q with its transposed version q' and add A
        A = np.outer(q,q) + A

    # scale
    A = (1.0/M)*A
    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)
    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]
    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0])
    return np.ravel(eigenVectors[:,0])


# Average multiple quaternions with specific weights
# The weight vector w must be of the same length as the number of rows in the
# quaternion maxtrix Q
def weightedAverageQuaternions(Q, w):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))
    weightSum = 0

    for i in range(0,M):
        q = Q[i,:]
        A = w[i] * np.outer(q,q) + A
        weightSum += w[i]

    # scale
    A = (1.0/weightSum) * A

    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)

    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]

    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0].A1)
    return np.ravel(eigenVectors[:,0])


def euler_to_rotMat(yaw, pitch, roll):
    Rz_yaw = np.array([
        [np.cos(yaw), -np.sin(yaw), 0],
        [np.sin(yaw),  np.cos(yaw), 0],
        [          0,            0, 1]])
    Ry_pitch = np.array([
        [ np.cos(pitch), 0, np.sin(pitch)],
        [             0, 1,             0],
        [-np.sin(pitch), 0, np.cos(pitch)]])
    Rx_roll = np.array([
        [1,            0,             0],
        [0, np.cos(roll), -np.sin(roll)],
        [0, np.sin(roll),  np.cos(roll)]])
    # R = RzRyRx
    rotMat = np.dot(Rz_yaw, np.dot(Ry_pitch, Rx_roll))
    
    return rotMat


# RPY/Euler angles to Rotation Vector
def euler_to_rotVec(yaw, pitch, roll):

    # compute the rotation matrix
    Rmat = euler_to_rotMat(yaw, pitch, roll)
    
    theta = math.acos(((Rmat[0, 0] + Rmat[1, 1] + Rmat[2, 2]) - 1) / 2)
    sin_theta = math.sin(theta)
    if sin_theta == 0:
        rx, ry, rz = 0.0, 0.0, 0.0
    else:
        multi = 1 / (2 * math.sin(theta))
        rx = multi * (Rmat[2, 1] - Rmat[1, 2]) * theta
        ry = multi * (Rmat[0, 2] - Rmat[2, 0]) * theta
        rz = multi * (Rmat[1, 0] - Rmat[0, 1]) * theta
    return rx, ry, rz





def product_Quaternions(quaternion_arr):
    
    #A cool feature of quaternions is that they can be intuitively chained together to form a composite rotation from a sequence of discrete rotations:
    product_quaternion = quaternion_arr[0,:]
    
    #print(product_quaternion)

    #Quaternion(numpy.array([a, b, c, d]))

    for idx, q_r in enumerate(quaternion_arr):

        if idx>0:
            
            product_quaternion = Quaternion(product_quaternion)*Quaternion(q_r)
        
    return product_quaternion.elements
            


#find shortest path between start and end vertex
def short_path_finder(G_unordered, start_v, end_v):
    
    #find shortest path between start and end vertex
    ####################################################################
    
    #define start and end vertex index
    #start_v = 0
    #end_v = 1559
    #end_v = 608
    
    
    #print(X_skeleton[start_v], Y_skeleton[start_v], Z_skeleton[start_v])
    
    # find shortest path in the graph between start and end vertices 
    
    #vlist, elist = gt.all_shortest_paths(G_unordered, G_unordered.vertex(start_v), G_unordered.vertex(end_v))
    
    vlist, elist = gt.shortest_path(G_unordered, G_unordered.vertex(start_v), G_unordered.vertex(end_v))
    
    vlist_path = [str(v) for v in vlist]
    
    elist_path = [str(e) for e in elist]
    
    #change format form str to int
    int_vlist_path = [int(i) for i in vlist_path]
    
    #print(int_vlist_path)
    
    if len(vlist_path) > 0: 
        
        return int_vlist_path
    else:
        print("No shortest path found in graph...\n")
        
        return 0


def dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, select):
    
    loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(loc_start), min(loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(loc_end), min(loc_end)))
    
    max_dimension_length = abs(max(max(loc_start), max(loc_end) - min(min(loc_start), min(loc_end))))

    min_dimension_length = abs(min(min(loc_start), min(loc_end) - max(max(loc_start), max(loc_end))))
    
    #print("max_length = {}\n".format(max_length))
    
    if select == 1:
        return max_dimension_length
    else:
        return min_dimension_length



def calculate_wcss(data):
    wcss = []
    for n in range(2, 10):
        kmeans = KMeans(n_clusters=n)
        kmeans.fit(X=data)
        wcss.append(kmeans.inertia_)

    return wcss



def optimal_number_of_clusters(wcss):
    x1, y1 = 2, wcss[0]
    x2, y2 = 20, wcss[len(wcss)-1]

    distances = []
    for i in range(len(wcss)):
        x0 = i+2
        y0 = wcss[i]
        numerator = abs((y2-y1)*x0 - (x2-x1)*y0 + x2*y1 - y2*x1)
        denominator = math.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        distances.append(numerator/denominator)

    return distances.index(max(distances)) + 0



def his_plot(path_length_rec, current_path, folder_name):
    
  
    legend = ['Path length histogram distribution']

    N_points = len(path_length_rec)
    x = path_length_rec
    n_bins = 20
    
    
    # Creating histogram
    fig, axs = plt.subplots(1, 1, figsize =(10, 7), tight_layout = True)


    # Remove axes splines
    for s in ['top', 'bottom', 'left', 'right']:
        axs.spines[s].set_visible(False)

    # Remove x, y ticks
    axs.xaxis.set_ticks_position('none')
    axs.yaxis.set_ticks_position('none')

    # Add padding between axes and labels
    axs.xaxis.set_tick_params(pad = 5)
    axs.yaxis.set_tick_params(pad = 10)

    # Add x, y gridlines
    axs.grid(visible = True, color ='grey', linestyle ='-.', linewidth = 0.5, alpha = 0.6)

    bin_size = 0.1
    min_edge = 0.1
    max_edge = 5.0
    Nplus1 = (max_edge-min_edge)/bin_size + 1
    bin_list = np.linspace(min_edge, max_edge, int(Nplus1))


    #fixed_bins = [0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]
    
    # Creating histogram
    #N, bins, patches = axs.hist(x, bins = fixed_bins)
    
    #N, bins, patches = axs.hist(x, bins = bin_list)
    
    N, bins, patches = axs.hist(x, bins = n_bins)
    
    axs.set_ylim([0, 150])

    # Setting color
    fracs = ((N**(1 / 5)) / N.max())
    norm = colors.Normalize(fracs.min(), fracs.max())

    for thisfrac, thispatch in zip(fracs, patches):
        color = plt.cm.viridis(norm(thisfrac))
        thispatch.set_facecolor(color)

    # Adding extra features   
    plt.xlabel("Node based path length in 3D model space")
    plt.ylabel("Counts")
    plt.legend(legend)
    plt.title('Node based path length distribution')
    
    
    # create trait file using sub folder name
    path_histogram = (current_path + folder_name + '_his.png')
    
    plt.savefig(path_histogram)




#cluster 1D list using Kmeans
def cluster_list(list_array, n_clusters):
    
    data = np.array(list_array)
    
    if data.ndim == 1:
        
        data = data.reshape(-1,1)
    
    #kmeans = KMeans(n_clusters).fit(data.reshape(-1,1))
        
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(data)
    
    #kmeans = KMeans(n_clusters).fit(data)
    
    labels = kmeans.labels_
    
    centers = kmeans.cluster_centers_
    
    center_labels = kmeans.predict(centers)
    
    #print(kmeans.cluster_centers_)
    '''
    #visualzie data clustering 
    ######################################################
    y_kmeans = kmeans.predict(data)
    
    plt.scatter(data[:, 0], data[:, 1], c=y_kmeans, s=50, cmap='viridis')

    centers = kmeans.cluster_centers_
    
    plt.scatter(centers[:, 0], centers[:, 1], c='black', s=200, alpha=0.5)
    
    #plt.legend()

    plt.show()
    
    '''
    
    
    '''
    from sklearn.metrics import silhouette_score
    
    range_n_clusters = [2, 3, 4, 5, 6, 7, 8]
    silhouette_avg = []
    for num_clusters in range_n_clusters:
     
         # initialize kmeans
         kmeans = KMeans(n_clusters=num_clusters)
         kmeans.fit(data)
         cluster_labels = kmeans.labels_
         
         # silhouette score
         silhouette_avg.append(silhouette_score(data, cluster_labels))
    
    plt.plot(range_n_clusters,silhouette_avg,'bx-')
    plt.xlabel('Values of K')
    plt.ylabel('Silhouette score')
    plt.title('Silhouette analysis For Optimal k')
    plt.show()
    
    
    Sum_of_squared_distances = []
    K = range(2,8)
    for num_clusters in K :
        kmeans = KMeans(n_clusters=num_clusters)
        kmeans.fit(data)
        Sum_of_squared_distances.append(kmeans.inertia_)
        
    plt.plot(K,Sum_of_squared_distances,'bx-')
    plt.xlabel('Values of K') 
    plt.ylabel('Sum of squared distances/Inertia') 
    plt.title('Elbow Method For Optimal k')
    plt.show()
    '''
    ######################################################
    
    
    return labels, centers, center_labels



# get consecutive elements pairing in list:
def pairwise(q_list):
 
    # use itertools.tee to create two iterators from the list
    a, b = itertools.tee(q_list)
     
    # advance the iterator by one element
    next(b, None)
     
    # use zip to pair the elements from the two iterators
    res = list(zip(a, b))  
    
    return res
    


# compute the distance between two quaternions accounting for the sign ambiguity.
def quaternion_list_distance(q_list):

    # get the pair of the elements from the list
    res = pairwise(q_list)
    
    # Create a 4 element array containing the final quaternion mutltipy results
    q_distance = []
    
    # make the column of the same length for easy operate
    q_distance.append([0,0,0])
     
    #compute mutltipy of adjacent pair of quaternions and then loop all elements 
    for idx, q_pair_value in enumerate(res):
        
        #Quaternion(numpy.array([a, b, c, d]))
        
        Q_Current = Quaternion(np.array(q_pair_value[0]))
        Q_Next = Quaternion(np.array(q_pair_value[1]))
        
        #This function does not measure the distance on the hypersphere, 
        #but it takes into account the fact that q and -q encode the same rotation. 
        #It is thus a good indicator for rotation similarities.
        # Quaternion absolute distance.
        Q_D_absolute = Quaternion.absolute_distance(Q_Current, Q_Next)
        
        # Quaternion intrinsic distance.
        #Although q0^(-1)*q1 != q1^(-1)*q0, the length of the path joining them is given by the logarithm of those product quaternions, the norm of which is the same.
        Q_D_intrinsic = Quaternion.distance(Q_Current, Q_Next)
        
        # Quaternion symmetrized distance.
        #Find the intrinsic symmetrized geodesic distance between q0 and q1.
        Q_D_symmetrized = Quaternion.sym_distance(Q_Current, Q_Next)
        
        q_distance.append([Q_D_absolute, Q_D_intrinsic, Q_D_symmetrized])


    #print(q_distance)
    
    return q_distance



# compute mutltipy two quaternions in a pair:
def quaternion_multiply(Q0, Q1):
    """
    Multiplies two quaternions.
 
    Input
    :param Q0: A 4 element array containing the first quaternion (q01,q11,q21,q31) 
    :param Q1: A 4 element array containing the second quaternion (q02,q12,q22,q32) 
 
    Output
    :return: A 4 element array containing the final quaternion (q03,q13,q23,q33) 
 
    """
    
    
    # Extract the values from Q0
    w0 = Q0[0]
    x0 = Q0[1]
    y0 = Q0[2]
    z0 = Q0[3]
     
    # Extract the values from Q1
    w1 = Q1[0]
    x1 = Q1[1]
    y1 = Q1[2]
    z1 = Q1[3]
     
    # Computer the product of the two quaternions, term by term
    Q0Q1_w = w0 * w1 - x0 * x1 - y0 * y1 - z0 * z1
    Q0Q1_x = w0 * x1 + x0 * w1 + y0 * z1 - z0 * y1
    Q0Q1_y = w0 * y1 - x0 * z1 + y0 * w1 + z0 * x1
    Q0Q1_z = w0 * z1 + x0 * y1 - y0 * x1 + z0 * w1
     
    # Create a 4 element array containing the final quaternion
    final_quaternion = np.array([Q0Q1_w, Q0Q1_x, Q0Q1_y, Q0Q1_z])
     
    # Return a 4 element array containing the final quaternion (q02,q12,q22,q32) 
    return final_quaternion



    


# compute mutltipy of a list of quaternions:
def quaternion_list_multiply(q_list):

    # get the pair of the elements from the list
    res = pairwise(q_list)
    
    # Create a 4 element array containing the final quaternion mutltipy results
    q_mutiply = np.array([0, 0, 0, 0])
     
    #compute mutltipy of adjacent pair of quaternions and then loop all elements 
    for idx, q_pair_value in enumerate(res):
        
        if idx < 1: 
            q_mutiply = quaternion_multiply(res[0][0], res[0][1])
        else:
            q_mutiply = quaternion_multiply(q_mutiply, res[idx][1])

    
    return q_mutiply



#def quaternion_list_addition(q_list):




# Get the Differential quaternion representing of a path, using substraction the quaternion formed by element-wise difference between adjacent quaternions
# a new Quaternion object representing the difference of the inputs. The difference is not guaranteed to be a unit quaternion.
def quaternion_list_differential(q_list):

    # get the pair of the elements from the list
    res = pairwise(q_list)
    
    # Create a 4 element array containing the final quaternion mutltipy results
    q_diff = []
    
    # make the column of the same length for easy operate
    q_diff.append([0,0,0,0])
     
    #compute mutltipy of adjacent pair of quaternions and then loop all elements 
    for idx, q_pair_value in enumerate(res):
        
        Q_Current = Quaternion(np.array(q_pair_value[0]))
        Q_Next = Quaternion(np.array(q_pair_value[1]))
        
        diff_elements = (Q_Current - Q_Next).elements
        
        q_diff.append(diff_elements)
        
    
    q_diff = np.array(q_diff)
    
    return q_diff.sum(axis = 0)



# compute addition of a list of quaternions:
def quaternion_list_addition(q_list):

    # Create a 4 element array containing the final quaternion mutltipy results
    q_addition = ([0,0,0,0])
    
    
    #compute mutltipy of adjacent pair of quaternions and then loop all elements 
    for idx, q_value in enumerate(q_list):
        
        Q_Current = Quaternion(np.array(q_value))
        
        q_addition = Quaternion(q_addition) + Q_Current
        
    
    q_sum = np.array(q_addition.elements)
    
    return q_sum
    
    


# compute the distance between quaternions 
def quaternion_list_distance(q_list):

    # get the pair of the elements from the list
    res = pairwise(q_list)
    
    # Create a 4 element array containing the final quaternion  results
    q_distance = []
    
    # make the column of the same length for easy operate
    q_distance.append([0,0,0])
    
    
    #compute mutltipy of adjacent pair of quaternions and then loop all elements 
    for idx, q_pair_value in enumerate(res):
        
        #Quaternion(numpy.array([a, b, c, d]))
        
        Q_Current = Quaternion(np.array(q_pair_value[0]))
        Q_Next = Quaternion(np.array(q_pair_value[1]))
        
        #This function does not measure the distance on the hypersphere, 
        #but it takes into account the fact that q and -q encode the same rotation. 
        #It is thus a good indicator for rotation similarities.
        # Quaternion absolute distance.
        Q_D_absolute = Quaternion.absolute_distance(Q_Current, Q_Next)
        
        # Quaternion intrinsic distance.
        #Although q0^(-1)*q1 != q1^(-1)*q0, the length of the path joining them is given by the logarithm of those product quaternions, the norm of which is the same.
        Q_D_intrinsic = Quaternion.distance(Q_Current, Q_Next)
        
        # Quaternion symmetrized distance.
        #Find the intrinsic symmetrized geodesic distance between q0 and q1.
        Q_D_symmetrized = Quaternion.sym_distance(Q_Current, Q_Next)
        
        q_distance.append([Q_D_absolute, Q_D_intrinsic, Q_D_symmetrized])

    
    q_distance = np.array(q_distance)
    
    #print(q_distance.shape)
    
    (sum_Q_D_absolute, sum_Q_D_intrinsic, sum_Q_D_symmetrized)  = q_distance.sum(axis = 0)
    
    
    return sum_Q_D_absolute, sum_Q_D_intrinsic, sum_Q_D_symmetrized



# compute the rotation vector from a quaternion 
def rotVec_from_quaternion(quaternion_value):

    # get Rotation matrix from quaternion
    rot = R.from_quat(quaternion_value)

    # get the rotation vector
    rotVec = rot.as_rotvec()
    
    return rotVec
    


# compute two orthonormal vectors a and b such that the cross product of the two vectors equals another unit vector k
def orthonormal_vectors(k):

    # get unit vector
    k_unit = k / np.linalg.norm(k)
    
    # take a random vector
    x = np.random.randn(3)  

    # make it orthogonal to k
    x -= x.dot(k_unit) * k_unit      
    
    # normalize x
    x /= np.linalg.norm(x)  
    
    # cross product with k
    y = np.cross(k_unit, x)      
    
    return x, y, k_unit



# visualize all the connection points along the longest path
def draw_nodes_index(X_skeleton, Y_skeleton, Z_skeleton, node_idx_list, color_rgb_value, scale_factor):
    

    # show the index of connection points
    for i,  idx_pts in enumerate(node_idx_list):

        graph_vis = mlab.text3d(X_skeleton[idx_pts], Y_skeleton[idx_pts], Z_skeleton[idx_pts], \
                                str("{:.0f}".format(idx_pts)), color = color_rgb_value, scale = (scale_factor, scale_factor, scale_factor))

    # visualize all the connection points
    graph_vis = mlab.points3d(X_skeleton[node_idx_list], Y_skeleton[node_idx_list], Z_skeleton[node_idx_list], 
                            color = color_rgb_value, mode = 'sphere', scale_factor = scale_factor)
                            
    return graph_vis



# visualize rotation vectors
def draw_rotation_vectors(rotVec, color_vec, line_width_value):
    
    
    for i, vectors in enumerate(rotVec):
        
        zeros = np.zeros(len(vectors))

        vec_vis = mlab.quiver3d(zeros, zeros, zeros, np.asarray(vectors)[:,0], np.asarray(vectors)[:,1], np.asarray(vectors)[:,2], color = color_vec[i], mode = '2darrow', line_width = line_width_value)
    
                            
    return vec_vis
    
    



def optimize_n_clusters(data_list):

    
    ###############################################################################
    #Use Elbow Method methods to determine optimal value of number_cluster.
    

    silhouette_avg = []
    
    for num_clusters in list(range(2,20)):
        
        kmeans = KMeans(n_clusters = num_clusters, init = "k-means++", n_init = 10)
        
        kmeans.fit_predict(data_list)
        
        score = silhouette_score(data_list, kmeans.labels_)
        
        silhouette_avg.append(score)


    best_k = np.argmax(silhouette_avg) + 2
    
    print ("Best K: {}".format(best_k))
    
    return best_k




# using ploty to show the scatterplot of quaternion values
def plot_quaternion_result(quaternion_path_all, percent_all, file_output, type_quaternion, dimension):
    
        
    if dimension == 4:

        if type_quaternion == 0:
            cols_q = ['quaternion_a','quaternion_b','quaternion_c', 'quaternion_d']
        elif type_quaternion == 1:
            cols_q = ['composition_quaternion_a','composition_quaternion_b','composition_quaternion_c', 'composition_quaternion_d']
        elif type_quaternion == 2:
            cols_q = ['diff_quaternion_a','diff_quaternion_b','diff_quaternion_c', 'diff_quaternion_d']
        elif type_quaternion == 3:
            cols_q = ['distance_absolute','distance_intrinsic', 'distance_symmetrized']
        
        data = pd.DataFrame(quaternion_path_all, columns = cols_q)

        #Set marker properties
        #markercolor = data['quaternion_a']
        
        markercolor = percent_all
        
        #Make Plotly figure
        fig1 = go.Scatter3d(x=data[cols_q[1]],
                        y=data[cols_q[2]],
                        z=data[cols_q[3]],
                        marker=dict(color=markercolor,
                                    opacity=1,
                                    reversescale=True,
                                    colorscale='Viridis',
                                    colorbar=dict(thickness=10),
                                    size=5),
                        line=dict (width=0.02),
                        mode='markers')
        
                     
        #Make Plot.ly Layout
        mylayout = go.Layout(scene=dict(xaxis=dict( title=str(cols_q[1])), 
                                        yaxis=dict( title=str(cols_q[2])),
                                        zaxis=dict(title=str(cols_q[3]))),)
        
        #Plot and save html
        plotly.offline.plot({"data": [fig1],
                         "layout": mylayout},
                         auto_open=False,
                         filename=file_output)
    elif dimension == 3:
        
        
        if type_quaternion == 0:
            cols_q = ['rotVec_avg_0','rotVec_avg_1','rotVec_avg_2']
        elif type_quaternion == 1:
            cols_q = ['rotVec_composition_0','rotVec_composition_1','rotVec_composition_2']
        elif type_quaternion == 2:
            cols_q = ['rotVec_diff_0','rotVec_diff_1','rotVec_diff_2']
        elif type_quaternion == 3:
            cols_q = ['distance_absolute','distance_intrinsic', 'distance_symmetrized']

        data = pd.DataFrame(quaternion_path_all, columns = cols_q)

        #Set marker properties
        #markercolor = data['quaternion_a']
        
        markercolor = percent_all
        
        #Make Plotly figure
        fig1 = go.Scatter3d(x=data[cols_q[0]],
                        y=data[cols_q[1]],
                        z=data[cols_q[2]],
                        marker=dict(color=markercolor,
                                    opacity=1,
                                    reversescale=True,
                                    colorscale='Viridis',
                                    colorbar=dict(thickness=10),
                                    size=5),
                        line=dict (width=0.02),
                        mode='markers')
        
                     
        #Make Plot.ly Layout
        mylayout = go.Layout(scene=dict(xaxis=dict( title=str(cols_q[0])),
                                    yaxis=dict( title=str(cols_q[1])),
                                    zaxis=dict(title=str(cols_q[2]))),)
        
        #Plot and save html
        plotly.offline.plot({"data": [fig1],
                         "layout": mylayout},
                         auto_open=False,
                         filename=file_output)



def in_hull(p, hull):
    """
    Test if points in `p` are in `hull`

    `p` should be a `NxK` coordinates of `N` points in `K` dimensions
    `hull` is either a scipy.spatial.Delaunay object or the `MxK` array of the 
    coordinates of `M` points in `K`dimensions for which Delaunay triangulation
    will be computed
    """
    
    if not isinstance(hull,Delaunay):
        hull = Delaunay(hull)

    return hull.find_simplex(p)>=0

    



def analyze_path_traits(vlist_path, Data_array_skeleton):
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    
    # Compute current path length between nodes
    path_length_N2N = path_length(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path])
    
    rotVec_rec_avg = []
    rotVec_rec_composition = []
    rotVec_rec_diff = []
    
   
    index_inv = []
    
    list_quaternion = []
    
    #print("closest_pts = {}\n".format(vlist_path[i]))

    # get all nodes pairs
    idx_pair_arr = np.vstack(pairwise(vlist_path))

    start_idx = idx_pair_arr[:,0]
    end_idx = idx_pair_arr[:,1]

    # compute all connected vectors in the path
    vector_list = []
    for idx, (s, e) in enumerate(zip(start_idx, end_idx)):

        p1 = [X_skeleton[s], Y_skeleton[s], Z_skeleton[s]]
        p2 = [X_skeleton[e], Y_skeleton[e], Z_skeleton[e]]

        vector_in_pair = findVec(p1,p2)
        vector_list.append(vector_in_pair)

    # get index list pairs of adjacent vector 
    vector_pair = pairwise(range(len(vector_list)))


    sum_quaternion = np.zeros([len(vector_pair), 4])

    # loop over all adjacent vector pairs 
    for i, value in enumerate(vector_pair):
        
        start_vec_idx = list(value)[0]
        end_vec_idx = list(value)[1]

        # compoute rotation matrix between adjacent directed vectors
        mat = get_rotation_matrix(vec1 = vector_list[start_vec_idx], vec2 = vector_list[end_vec_idx])

        # compoute quaternion between adjacent directed vectors
        #The returned quaternion value is in scalar-last (x, y, z, w) format.
        quaternion_r = R.from_matrix(mat).as_quat()

        #compute rotation vector between adjacent directed vectors
        rotVec_r = R.from_matrix(mat).as_rotvec()

        # change the order of the quaternion_r value from (x, y, z, w)  to (w, x, y, z)
        quaternion_r_rearanged = np.hstack((quaternion_r[3], quaternion_r[0], quaternion_r[1], quaternion_r[2]))

        #euler_r = R.from_matrix(mat).as_euler('xyz', degrees = True)
           
        sum_quaternion[i,:] = quaternion_r_rearanged

        list_quaternion.append(list(quaternion_r_rearanged))
        
        ##################################################


    #compute quaternion traits of the current path
    ######################################################################

    # Average of quaternions
    ###############################################################
    # compute average of quaternions from Quaternion averaging functions from scikit-surgerycore, The quaternions input are arranged as (w,x,y,z),

    #sample_list = ([0,0,0,0], [1,2,3,4], [1,2,3,4])

    #avg_quaternion = quaternion_list_addition(list_quaternion)

    #print("sum_quaternion = {}\n".format(sum_quaternion))

    # use eigenvalues to compute average of quaternions, The quaternions input are arranged as (w,x,y,z) with w being the scalar
    avg_quaternion = average_quaternions(sum_quaternion)


    # Composition of quaternions
    #################################################################
    q_composition = quaternion_list_multiply(list_quaternion)

    q_composition = q_composition.flatten()

    #composition_path_rec.append(q_composition)

    # get the rotation vector
    #rotVec_rec_composition.append(rotVec_from_quaternion(q_composition))


    # Differential of quaternions
    #################################################################

    q_diff = quaternion_list_differential(list_quaternion)

    #diff_path_rec.append(q_diff)

    # get the rotation vector
    #rotVec_rec_diff.append(rotVec_from_quaternion(q_diff))



    # Distance of quaternions
    #################################################################

    (cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized) = quaternion_list_distance(list_quaternion)

    #distance_path_rec.append([cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized])


    return  avg_quaternion, q_composition, q_diff, cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized, path_length_N2N








#######################################################################################################
########################################################################################################
# Skeleton analysis
def analyze_skeleton(current_path, filename_skeleton, filename_pcloud):
    
    model_skeleton = current_path + filename_skeleton
    print("Loading 3D skeleton file {}...\n".format(filename_skeleton))
    model_skeleton_name_base = os.path.splitext(model_skeleton)[0]
    
    trait_path = os.path.dirname(current_path + filename_skeleton)
    
    folder_name = os.path.basename(trait_path)
    
    
    
    #load the ply format skeleton file 
    try:
        with open(model_skeleton, 'rb') as f:
            plydata_skeleton = PlyData.read(f)
            num_vertex_skeleton = plydata_skeleton.elements[0].count
            N_edges_skeleton = len(plydata_skeleton['edge'].data['vertex_indices'])
            array_edges_skeleton = plydata_skeleton['edge'].data['vertex_indices']

            print("Ply data structure: \n")
            print(plydata_skeleton)
            #print("\n")
            print("Number of 3D points in skeleton model: {0} \n".format(num_vertex_skeleton))
            print("Number of edges: {0} \n".format(N_edges_skeleton))
            

    except:
        sys.exit("Model skeleton file does not exist!")
    

    
    #Parse ply format skeleton file and Extract the data
    Data_array_skeleton = np.zeros((num_vertex_skeleton, 3))
    
    Data_array_skeleton[:,0] = plydata_skeleton['vertex'].data['x']
    Data_array_skeleton[:,1] = plydata_skeleton['vertex'].data['y']
    Data_array_skeleton[:,2] = plydata_skeleton['vertex'].data['z']
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    
    #radius of vertex
    #radius_skeleton = np.zeros((num_vertex_skeleton, 1))
    #radius_skeleton = plydata_skeleton['vertex'].data['radius']
    
    #print(radius_skeleton)
    
    #print(array_edges_skeleton)
    #print("Number of array_edges_skeleton: {0} \n".format(len(array_edges_skeleton)))
    

    
    # build directed graph from skeleton/structure data
    ####################################################################
    print("Building directed graph from 3D skeleton/structure ...\n")
    G_unordered = gt.Graph(directed = True)
    
    # assert directed graph
    #print(G.is_directed())
    
    nodes = G_unordered.add_vertex(num_vertex_skeleton)
    
    G_unordered.add_edge_list(array_edges_skeleton.tolist()) 
    
    #gt.graph_draw(G_unordered, vertex_text = G_unordered.vertex_index, output = current_path + "graph_view.pdf")
    
    
    # find all end vertices by fast iteration of all vertices
    end_vlist = []
    
    # for conecting all branches 
    end_vlist_offset = []
    
    start_vlist = []
    
    for v in G_unordered.iter_vertices():
        
        # get all start vertices
        if G_unordered.vertex(v).out_degree() == 1 and G_unordered.vertex(v).in_degree() == 0:
        
            start_vlist.append(v)
        
        # get all end vertices
        if G_unordered.vertex(v).out_degree() == 0 and G_unordered.vertex(v).in_degree() == 1:
        
            end_vlist.append(v)
        
            if (v+1) == num_vertex_skeleton:
                end_vlist_offset.append(v)
            else:
                end_vlist_offset.append(v+1)
                
                
    print("Number of start nodes {}, Number of End nodes in the graph: {} \n".format(len(start_vlist), len(end_vlist)))

    
    '''
    print("start_vlist = {} \n".format(start_vlist))
    print("end_vlist_offset = {} \n".format(end_vlist_offset))
    print("end_vlist = {} \n".format(end_vlist))
    '''

    

    #test angle calculation
    #vector1 = [0,0,1]
    # [1,0,0]
    #vector2 = [0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1]
    #print(dot_product_angle(vector1,vector2))
    #print(cart2sph(0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1)) 

    
    # parse all the sub branches edges and vetices, start vertices, end vertices
    #####################################################################################
    sub_branch_list = []
    
    sub_branch_length_rec = []
    sub_branch_angle_rec = []
    
    sub_branch_start_rec = []
    sub_branch_end_rec = []
    
    avg_node_distance_rec = []
    
    #sub_branch_projection_rec = []
    #sub_branch_radius_rec = []
    
    sub_branch_xs_rec = []
    sub_branch_ys_rec = []
    sub_branch_zs_rec = []
    
    #if len(end_vlist) == len(end_vlist_offset):
        
    for idx, v_end in enumerate(end_vlist):
        
        #print(idx, v_end)
        #construct list of vertices in sub branches
        if idx == 0:
            v_list = [*range(0, int(end_vlist[idx])+1)]
        else:
            v_list = [*range(int(end_vlist[idx-1])+1, int(end_vlist[idx])+1)]
            
        # change type to interger 
        int_v_list = [int(i) for i in v_list]
        
        
        
        # current sub branch length
        sub_branch_length = path_length(X_skeleton[int_v_list], Y_skeleton[int_v_list], Z_skeleton[int_v_list])
        
        avg_node_distance = sub_branch_length/(len(int_v_list)-1)
        
        #print("avg_node_distance = {}, sub_branch_length = {}, n_nodes = {}\n".format(avg_node_distance, sub_branch_length, len(int_v_list)))
        
        avg_node_distance_rec.append(avg_node_distance)
        
        # current sub branch start and end points 
        start_v = [X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]]
        
        end_v = [X_skeleton[int_v_list[0]] - X_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Y_skeleton[int_v_list[0]] - Y_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Z_skeleton[int_v_list[0]] - Z_skeleton[int_v_list[int(len(int_v_list)-1.0)]]]
        
        # angle of current branch vs Z direction
        angle_sub_branch = dot_product_angle(start_v, end_v)
        
        # projection radius of current branch length
        #p0 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]])
        
        #p1 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[-1]]])
        
        #projection_radius = np.linalg.norm(p0 - p1)
        
        #radius value from fitted point cloud contours
        #radius_edge = float(radius_vtx[int_v_list[0]])*1
        
        sub_branch_xs = X_skeleton[int_v_list[0]]
        sub_branch_ys = Y_skeleton[int_v_list[0]]
        sub_branch_zs = Z_skeleton[int_v_list[0]]
        
        # apply thersh length threshold
        #if sub_branch_length > 0: 
        
        if len(int_v_list) > len_ratio:
        #if sub_branch_length > thresh_length: 
        
            # save computed parameters for each branch
            sub_branch_list.append(int_v_list)
            
            sub_branch_length_rec.append(sub_branch_length)
            sub_branch_angle_rec.append(angle_sub_branch)
            
            sub_branch_start_rec.append(int_v_list[0])
            sub_branch_end_rec.append(int_v_list[-1])
            
            #sub_branch_projection_rec.append(projection_radius)
            #sub_branch_radius_rec.append(radius_edge)
            
            sub_branch_xs_rec.append(sub_branch_xs)
            sub_branch_ys_rec.append(sub_branch_ys)
            sub_branch_zs_rec.append(sub_branch_zs)
    
    #print(min(sub_branch_angle_rec))
    #print(max(sub_branch_angle_rec))
    
    
    ####################################################################
    
    
    
    
    
    '''
    # sort branches according to length feature in descending order
    ####################################################################
    sorted_idx_len = np.argsort(sub_branch_length_rec)
    
    #reverse the order from accending to descending
    sorted_idx_len_des = sorted_idx_len[::-1]
    
    #fliter smaller bracnhes with length ranking lower than 15% of average length
    # size desired
    k = int(len(sorted_idx_len_des)*(1.0 - len_ratio))
 
    sorted_idx_len_loc = sorted_idx_len_des[0 : k]


    print("sorted_idx_len = {} sorted_idx_len_loc = {}\n".format(len(sorted_idx_len_des), len(sorted_idx_len_loc)))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_len_loc]
     
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_len_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_len_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_len_loc]
    
    #sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_len_loc]
    #sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_len_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_len_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_len_loc]

    
    print("number of sub_branch_end_rec is: {} \n".format(len(sub_branch_end_rec)))
    '''
    
    
    
    # graph: find closest point pairs and connect graph nodes and edges, building a connected tree graph
    ####################################################################
    print("Converting skeleton to graph and connecting edges and vertices...\n")
    

    v_closest_pair_rec = []

    closest_pts = []
        
    idx_visited_start = []
    
    idx_visited_end = []
    
    anchor_point_rec = []
    
    dis_v_closest_pair_rec = []
    
    candidate_list = []
    

    
    
    
    #find closest point set and connect graph edges

    # start, stop, step
    #for idx in range(0, 1):
        
    #test_branch_idx = 3
    
    ####################################################################
    # find connection nodes along each branches
    
    for idx, (sub_branch, start_v, end_v) in enumerate(zip(sub_branch_list, sub_branch_start_rec, sub_branch_end_rec)):
    
    #for idx, (sub_branch, start_v, end_v) in enumerate(zip(sub_branch_list[test_branch_idx:(test_branch_idx+1)], sub_branch_start_rec[test_branch_idx:(test_branch_idx+1)], sub_branch_end_rec[test_branch_idx:(test_branch_idx+1)])):
        
      
        print("Processing branch ID : {}, contains {} nodes\n".format(idx, len(sub_branch)))
        
 
        # branch points in 3d 
        point_set = np.zeros((len(sub_branch), 3))
        
        point_set[:,0] = X_skeleton[sub_branch]
        point_set[:,1] = Y_skeleton[sub_branch]
        point_set[:,2] = Z_skeleton[sub_branch]
        
        # all start points of branches
        query_points = np.zeros((len(end_vlist_offset), 3))
        
        query_points[:,0] = X_skeleton[end_vlist_offset]
        query_points[:,1] = Y_skeleton[end_vlist_offset]
        query_points[:,2] = Z_skeleton[end_vlist_offset]
        
        # convex hull cannot be built if branches less than 3 nodes
        
        # distance threshold for connecting branches, number of nodes in the shortest root path
        if len(sub_branch) > len_ratio:
            
            # test whether start points of branches that lie in the given convex hull built from point_set,
            # It returns a boolean array where True values indicate points that lie in the given convex hull.
            #inside_idx = in_hull(query_points, point_set)

            #print(in_hull(query_points, point_set))
            #if len(inside_idx) > 0:
                #candidate_list = [end_vlist_offset[i] for i in range(len(end_vlist_offset)) if inside_idx[i]]
            
            candidate_list = end_vlist_offset

            # search not visited nodes
            #s = set(idx_visited_start)
            #candidate_list = [x for x in end_vlist_offset if x not in s]
            
            #print(candidate_list)
        
            for idx_c, candidate_v in enumerate(candidate_list):
                
                if start_v != candidate_v:

                    # start vertex of an edge
                    anchor_point = (X_skeleton[candidate_v], Y_skeleton[candidate_v], Z_skeleton[candidate_v])
                    
                    anchor_point_rec.append(candidate_v)
                    
                    (index_cp, value_cp) = closest_point(point_set, anchor_point)

                    v_closest_pair = [index_cp, candidate_v]

                    dis_v_closest_pair = path_length(X_skeleton[v_closest_pair], Y_skeleton[v_closest_pair], Z_skeleton[v_closest_pair])

                    #define distance threshold counted as connection nodes
                    #is_close = (dis_v_closest_pair < avg_node_distance_rec[idx]*dis_factor) 

                    #if is_close: 
                    
                    closest_pts.append(sub_branch[index_cp])
                    
                    dis_v_closest_pair_rec.append(dis_v_closest_pair)

                    v_closest_pair_rec.append(v_closest_pair)
                    
                    #print("closest point pair: {0}".format(v_closest_pair))
                    
                    idx_visited_start.append(candidate_v)
                    
                    idx_visited_end.append(end_v)
                    
                    #connect graph edges
                    G_unordered.add_edge(sub_branch[index_cp], candidate_v)
                        

    ################################################################
        
    #sort tha data in the order of accending 
    #sorted_idx_len = np.argsort(dis_v_closest_pair_rec)

    #sort all lists according 
    #dis_v_closest_pair_rec[:] = [dis_v_closest_pair_rec[i] for i in sorted_idx_len]

    #closest_pts[:] = [closest_pts[i] for i in sorted_idx_len]
    
    closest_pts_sorted = list(set(closest_pts))
    
    idx_visited_start_sorted = list(set(idx_visited_start))

    dis_v_closest_pair_rec_sorted = list(set(dis_v_closest_pair_rec))
    

    #print("closest_pts = {} \n".format(closest_pts))
    #print("visited_start = {} \n".format(idx_visited_start))
    #print("dis_v_closest_pair_rec = {} \n".format(dis_v_closest_pair_rec))

    print("Number of connection points = {}\n".format(len(closest_pts_sorted)))
        

    #find shortest path between start vertices and end vertices
    ####################################################################
    
    #define start and end vertex index
    start_v = 0
    
    vlist_path_rec = []
    
    avg_quaternion_path_rec = []
    composition_path_rec = []
    distance_path_rec = []
    diff_path_rec = []
    
    
    
    rotVec_rec_avg = []
    rotVec_rec_composition = []
    rotVec_rec_diff = []
    
    path_length_rec = []
    
    index_inv = []
    
    list_quaternion = []
    
    
    start_v_common = np.repeat(0, repeats = len(sub_branch_end_rec), axis = 0)
    
    
    #import time
    #t = time.process_time()
    
    # loop over all paths and compute quaternions values
    #for idx, (start_v, end_v) in enumerate(zip(sub_branch_start_rec, sub_branch_end_rec)):
    
    for idx, (start_v, end_v) in enumerate(zip(start_v_common, sub_branch_end_rec)):
        
       
        # Count the number of shortest paths from source to target.
        n_paths = gt.count_shortest_paths(G_unordered, start_v, end_v)
        

        if n_paths > 0:
            
            # Find the shortest path from source to target.
            vlist_path = short_path_finder(G_unordered, start_v, end_v)
            
 
            # List of vertices from source to target in the shortest path was not empty
            if len(vlist_path) > 0:
                
                vlist_path_rec.append(vlist_path)
    
    print("Found {} shortest path \n".format(len(vlist_path_rec)))
    
    
    #elapsed_time = time.process_time() - t
    #print("First loop time cost : {}\n".format(elapsed_time))
    
    
    for idx, vlist_path in enumerate(vlist_path_rec): 
    
        (avg_quaternion, q_composition, q_diff, cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized, path_length_N2N) = analyze_path_traits(vlist_path, Data_array_skeleton)
        
        # Average of quaternions
        ###############################################################
        # compute average of quaternions from Quaternion averaging functions from scikit-surgerycore, The quaternions input are arranged as (w,x,y,z),
        
        # compute average quaternion values from a list of quarternion along the path
        avg_quaternion_path_rec.append(avg_quaternion)
        
        # get the rotation vector
        rotVec_rec_avg.append(rotVec_from_quaternion(avg_quaternion))
        
        
        # Composition of quaternions
        #################################################################
        composition_path_rec.append(q_composition)

        # get the rotation vector
        rotVec_rec_composition.append(rotVec_from_quaternion(q_composition))


        # Differential of quaternions
        #################################################################
        diff_path_rec.append(q_diff)

        # get the rotation vector
        rotVec_rec_diff.append(rotVec_from_quaternion(q_diff))

        # Distance of quaternions
        #################################################################
        #(cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized) = quaternion_list_distance(list_quaternion)

        distance_path_rec.append([cumulative_Q_D_absolute, cumulative_Q_D_intrinsic, cumulative_Q_D_symmetrized])

        
        ####################################################################
        # record current path length
        path_length_rec.append(path_length_N2N)
    



    ####################################################################
    #find the dominant cluster of the average quaternion as 4 dimensional vectors 
    avg_quaternion_path_rec_list = []
    
    for vector in avg_quaternion_path_rec:
        avg_quaternion_path_rec_list.append(vector.reshape(4,1))

    avg_quaternion_path_rec_list_reshape = np.asarray(avg_quaternion_path_rec).reshape((len(vlist_path_rec),4))
    print("quaternion_path_rec_list_reshape: {}\n".format((avg_quaternion_path_rec_list_reshape.shape)))
    
    ####################################################################
    #find the dominant cluster of the composition quaternion as 4 dimensional vectors 
    composition_path_rec_list = []
    
    for vector in composition_path_rec:
        composition_path_rec_list.append(vector.reshape(4,1))

    composition_path_rec_list_reshape = np.asarray(composition_path_rec).reshape((len(vlist_path_rec),4))
    print("composition_path_rec_list_reshape: {}\n".format((composition_path_rec_list_reshape.shape)))
    
    ####################################################################
    #find the dominant cluster of the differential quaternion as 4 dimensional vectors 
    diff_path_rec_list = []
    
    for vector in diff_path_rec:
        diff_path_rec_list.append(vector.reshape(4,1))

    diff_path_rec_list_reshape = np.asarray(diff_path_rec).reshape((len(vlist_path_rec),4))
    print("diff_path_rec_list_reshape: {}\n".format((diff_path_rec_list_reshape.shape)))
    
    
    ####################################################################
    #find the dominant cluster of the distance quaternion as 3 dimensional vectors 
    distance_path_rec_list = []
    
    for vector in distance_path_rec_list:
        distance_path_rec_list.append(vector.reshape(3,1))

    distance_path_rec_list_reshape = np.asarray(distance_path_rec).reshape((len(vlist_path_rec),3))
    print("distance_path_rec_list_reshape: {}\n".format((distance_path_rec_list_reshape.shape)))
    
    

    
    ################################################################################
    #Keman cluster of quaternion values for all the paths

    #find the best number of clusters
    
    if args["n_cluster"] > 0:
        
        number_cluster = args["n_cluster"]
    else:
        
        number_cluster = optimize_n_clusters(avg_quaternion_path_rec_list_reshape)
    
    

    kmeans = KMeans(init = "k-means++", n_clusters = number_cluster)
    
    if type_quaternion == 0:
        s = kmeans.fit(avg_quaternion_path_rec_list_reshape)
    elif type_quaternion == 1:
        s = kmeans.fit(composition_path_rec_list_reshape)
    elif type_quaternion == 2:
        s = kmeans.fit(diff_path_rec_list_reshape)
    elif type_quaternion == 3:
        s = kmeans.fit(distance_path_rec_list_reshape)

        
        
    labels = kmeans.labels_
    
    labels = list(labels)
    
    #print((labels))
    
    # compute cluster center quaternion and related rotation vector
    centroid = kmeans.cluster_centers_
    
    #print("Centroid: {} \n".format(centroid))
    
    q_centroid_cluster = []
    rotVec_centroid_cluster = []
    for value in centroid:
        q_centroid_cluster.append(value)
        rotVec_centroid_cluster.append(rotVec_from_quaternion(value))
    
    #print("rotVec_centroid_cluster: {} \n".format(rotVec_centroid_cluster))
    
    # compute the ratio of each cluster
    percent = []
    for i in range(len(centroid)):
        j = labels.count(i)
        j = j/(len(labels))
        percent.append(j)
        
    #print(percent)
    

    #descending order sorting as per frequency count
    sorted_idx = np.argsort(percent)
    
    #reverse the order from accending to descending
    sorted_idx_percent = sorted_idx[::-1]
    
    percent_sorted = []
    for value in sorted_idx_percent:
        percent_sorted.append(percent[value])
    
    
    '''
    ################################################################################
    #Mixture of von Mises Fisher clustering (soft)
    
    
    vmf_soft = VonMisesFisherMixture(n_clusters = number_cluster, posterior_type = 'soft', n_init=20)
    
    
    if type_quaternion == 0:
        s = vmf_soft.fit(avg_quaternion_path_rec_list_reshape)
    elif type_quaternion == 1:
        s = vmf_soft.fit(composition_path_rec_list_reshape)
    elif type_quaternion == 2:
        s = vmf_soft.fit(diff_path_rec_list_reshape)
    elif type_quaternion == 3:
        s = vmf_soft.fit(distance_path_rec_list_reshape)

        
    labels = s.labels_
    
    labels = list(labels)
    
    
    #print((labels))
    
    # compute cluster center quaternion and related rotation vector
    centroid = s.centers
    
    #print("Centroid: {} \n".format(centroid))
    
    q_centroid_cluster = []
    rotVec_centroid_cluster = []
    for value in centroid:
        q_centroid_cluster.append(value)
        rotVec_centroid_cluster.append(rotVec_from_quaternion(value))
    
    #print("rotVec_centroid_cluster: {} \n".format(rotVec_centroid_cluster))
    
    # compute the ratio of each cluster
    percent = []
    for i in range(len(centroid)):
        j = labels.count(i)
        j = j/(len(labels))
        percent.append(j)
        
    #print(percent)
    

    #descending order sorting as per frequency count
    sorted_idx = np.argsort(percent)
    
    #reverse the order from accending to descending
    sorted_idx_percent = sorted_idx[::-1]
    
    percent_sorted = []
    for value in sorted_idx_percent:
        percent_sorted.append(percent[value])

    '''

    ##########################################################################################
    # compute the dominant cluster of quaternion vectors and related rotation vectors
    
    
    #quaternion_path_traits = ['average', 'composition', 'diff']
    
    q_average_cluster = []
    q_composition_cluster = []
    q_diff_cluster = []
    
    rotVec_average_cluster = []
    rotVec_composition_cluster = []
    rotVec_diff_cluster = []
    
    path_length_cluster = []
    distance_cluster = []
    

    
    #centroid
    
    for idx_value in sorted_idx_percent:
        
        index_selected = [index for index in range(len(labels))  if labels[index] == idx_value]
        
        average_path_q =  [avg_quaternion_path_rec[i] for i in index_selected]
        composition_path_q = [composition_path_rec[i] for i in index_selected]
        diff_path_q = [diff_path_rec[i] for i in index_selected]
        
        rotVec_path_average = [rotVec_rec_avg[i] for i in index_selected]
        rotVec_path_composition = [rotVec_rec_composition[i] for i in index_selected]
        rotVec_path_diff = [rotVec_rec_diff[i] for i in index_selected]
        
        distance_path_value = [distance_path_rec[i] for i in index_selected]

        path_length_value = [path_length_rec[i] for i in index_selected]
        
        q_average_cluster.append(average_path_q)
        q_composition_cluster.append(composition_path_q)
        q_diff_cluster.append(diff_path_q)
        
        rotVec_average_cluster.append(rotVec_path_average)
        rotVec_composition_cluster.append(rotVec_path_composition)
        rotVec_diff_cluster.append(rotVec_path_diff)

        path_length_cluster.append(path_length_value)
        
        distance_cluster.append(distance_path_value)
        
        

    
    #print("sorted_idx_percent = {} percent = {}".format(sorted_idx_percent, percent))
    
    for i in range(number_cluster):
        
        print("The {} cluster contains {} paths\n".format(i+1, len(path_length_cluster[i])))
    

    ############################################################################################
    color_array = np.repeat(np.array(percent).reshape(1, number_cluster), repeats = 4, axis = 0)

    #text = "{} {}".format(range(len(centroid)), percent)
    fig = plt.pie(percent, colors = np.transpose(color_array), labels = np.arange(len(centroid)))
    
    # create trait file using sub folder name
    pie_chart = (current_path + folder_name + '_pie.png')
    
    plt.savefig(pie_chart)
    
    plt.close()

    ####################################################################

    his_plot(path_length_cluster[0], current_path, folder_name)
    
    
    path_index = list(range(1,len(vlist_path_rec)+1))
    

    ###################################################################

    

        
    
    
    #Skeleton Visualization pipeline
    ####################################################################
    # The number of points per line
    
    if args["visualize_model"]:
        
        
        ############################################################################
        graph_vis = mlab.figure("Structure_graph", size = (800, 800), bgcolor = (0, 0, 0))
        
        graph_vis = mlab.clf()
        
        
        #1 . Visualize  *.ply 3D point cloud model
        if not (filename_pcloud is None):
            
            # load data coordinates
            model_pcloud = current_path + filename_pcloud
            
            print("Loading 3D point cloud {}...\n".format(filename_pcloud))
            
            model_pcloud_name_base = os.path.splitext(model_pcloud)[0]
            
            #load model file
            pcd = o3d.io.read_point_cloud(model_pcloud)
            
            Data_array_pcloud = np.asarray(pcd.points)
            
            #print(Data_array_pcloud.shape)
            
            #visualize point cloud model with color
            if pcd.has_colors():
                
                print("Render colored point cloud\n")
                
                pcd_color = np.asarray(pcd.colors)
                
                if len(pcd_color) > 0: 
                    
                    pcd_color = np.rint(pcd_color * 255.0)
                
                #pcd_color = tuple(map(tuple, pcd_color))
            else:
                
                print("Generate random color\n")
            
                pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
        
        
            
            x, y, z = Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2] 
            
            # visualize data coordinates
            #mlab.figure("Point cloud model", size = (800, 800), bgcolor = (0, 0, 0))
        
            #mlab.clf()
            
            graph_vis = mlab.points3d(x,y,z, mode = 'point')
            
            #pts = mlab.quiver3d(x,y,z)
            
            sc = tvtk.UnsignedCharArray()
            
            sc.from_array(pcd_color)

            graph_vis.mlab_source.dataset.point_data.scalars = sc
            
            graph_vis.mlab_source.dataset.modified()
        


        #2. visualize skeleton model, edge, nodes
        ####################################################################
        N = 2
        
        #graph_vis = mlab.figure("Structure_graph", size = (800, 800), bgcolor = (0, 0, 0))
        
        #graph_vis = mlab.clf()
        
        sf_value = 0.04

        '''
        #############################################################
        #visualize all edges in the graph
        x = list()
        y = list()
        z = list()
        s = list()
        connections = list()
        
        # The index of the current point in the total amount of points
        index = 0
        
        # Create each line one after the other in a loop
        for i in range(N_edges_skeleton):
        #for val in vlist_path:
            #if i in vertex_dominant:
            if True:
                #i = int(val)
                #print("Edges {0} has nodes {1}, {2}\n".format(i, array_edges[i][0], array_edges[i][1]))
              
                x.append(X_skeleton[array_edges_skeleton[i][0]])
                y.append(Y_skeleton[array_edges_skeleton[i][0]])
                z.append(Z_skeleton[array_edges_skeleton[i][0]])
                
                x.append(X_skeleton[array_edges_skeleton[i][1]])
                y.append(Y_skeleton[array_edges_skeleton[i][1]])
                z.append(Z_skeleton[array_edges_skeleton[i][1]])
                
                # The scalar parameter for each line
                s.append(array_edges_skeleton[i])
                
                # This is the tricky part: in a line, each point is connected
                # to the one following it. We have to express this with the indices
                # of the final set of points once all lines have been combined
                # together, this is why we need to keep track of the total number of
                # points already created (index)
                #connections.append(np.vstack(array_edges[i]).T)
                
                connections.append(np.vstack(
                               [np.arange(index,   index + N - 1.5),
                                np.arange(index + 1, index + N - .5)]
                                    ).T)
                index += 2
            
            
        # Now collapse all positions, scalars and connections in big arrays
        x = np.hstack(x)
        y = np.hstack(y)
        z = np.hstack(z)
        s = np.hstack(s)
        #connections = np.vstack(connections)

        # Create the points
        graph_vis = mlab.pipeline.scalar_scatter(x, y, z, s)
        
        # Connect them
        graph_vis.mlab_source.dataset.lines = connections
        
        #src.mlab_source.dataset.arrows = connections
        
        graph_vis.update()

        # display the set of lines
        graph_vis = mlab.pipeline.surface(graph_vis, colormap = 'Accent', line_width = 5, opacity = 0.7)
        '''
        
        
        
        ####################################################################
        # draw graph using connected arrows
        
        cmap = get_cmap(len(vlist_path_rec))

        for i, vlist_path in enumerate(vlist_path_rec):
            
            idx_pair_arr = np.vstack(pairwise(vlist_path))

            start_idx = idx_pair_arr[:,0]
            end_idx = idx_pair_arr[:,1]

            path_color = cmap(i)[:len(cmap(i))-1]
            
            #path_color = (1,1,1)
            
            graph_plot(X_skeleton, Y_skeleton, Z_skeleton, start_idx, end_idx, path_color)
            
            #graph_vis = mlab.points3d(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path], \
                    #color = path_color, mode = 'sphere', scale_factor = sf_value)
            
            # show the index of path
            #graph_vis = mlab.text3d(X_skeleton[vlist_path[-1]], Y_skeleton[vlist_path[-1]], Z_skeleton[vlist_path[-1]], \
                                    #str("{:.0f}".format(i)), color = (0,1,0), \
                                    #scale = (sf_value, sf_value, sf_value))

        # show the root tip point
        #graph_vis = mlab.points3d(X_skeleton[start_vlist[0]], Y_skeleton[start_vlist[0]], Z_skeleton[start_vlist[0]], \
                                    #color = (1,1,1), mode = 'sphere', scale_factor = sf_value*1.5)
        
        
        '''
        #################################################################################3
        # visualiztion all the nodes and vetices, debug purpose
        cmap = get_cmap(len(sub_branch_list))
        
        #draw all the sub branches in loop 
        for i, (sub_branch, start_v, end_v, end_v_offset) in enumerate(zip(sub_branch_list, sub_branch_start_rec, sub_branch_end_rec, end_vlist_offset)):

            #color_rgb = cmap(i)[:len(cmap(i))-1]
            
            color_rgb = (1,1,1)
            
            # show all the branches
            graph_vis = mlab.points3d(X_skeleton[sub_branch], Y_skeleton[sub_branch], Z_skeleton[sub_branch], \
                                color = color_rgb, mode = 'sphere', scale_factor = sf_value)
            
            
            # show the index of start and end node
            graph_vis = mlab.text3d(X_skeleton[start_v], Y_skeleton[start_v], Z_skeleton[start_v], \
                                    str("{:.0f}".format(start_v)), color = (1,1,0), \
                                    scale = (sf_value, sf_value, sf_value))
            
            graph_vis = mlab.text3d(X_skeleton[end_v], Y_skeleton[end_v], Z_skeleton[end_v], \
                                    str("{:.0f}".format(end_v)), color = (0,1,1), \
                                    scale = (sf_value, sf_value, sf_value))
            
            
            #graph_vis = mlab.text3d(X_skeleton[end_v_offset], Y_skeleton[end_v_offset], Z_skeleton[end_v_offset]-0.05, \
                                    #str("{:.0f}".format(end_v_offset)), color = (0,1,0), \
                                    #scale = (sf_value, sf_value, sf_value))
            
        

        '''
        ###################################################################################################
        '''
        # visualize all the start points
        #graph_vis = mlab.points3d(X_skeleton[start_vlist], Y_skeleton[start_vlist], Z_skeleton[start_vlist], \
                                    #color = (1,0,0), mode = 'sphere', scale_factor = sf_value*1.5)
        
        
        
        # visualize all the end points
        #graph_vis = mlab.points3d(X_skeleton[end_vlist], Y_skeleton[end_vlist], Z_skeleton[end_vlist], \
                                    #color = (1,0,0), mode = 'sphere', scale_factor = sf_value)
                                    
        
        
        
        graph_vis = mlab.points3d(X_skeleton[sub_branch_list[0]], Y_skeleton[sub_branch_list[0]], Z_skeleton[sub_branch_list[0]], \
                                    color = (0,1,1), mode = 'sphere', scale_factor = sf_value*0.5)
                                    
        #graph_vis = mlab.points3d(X_skeleton[sub_branch_list[test_branch_idx]], Y_skeleton[sub_branch_list[test_branch_idx]], Z_skeleton[sub_branch_list[test_branch_idx]], \
                                    #color = (0,0,1), mode = 'sphere', scale_factor = sf_value)
        
        #visualize all the connection points along the longest path
        #graph_vis = draw_nodes_index(X_skeleton, Y_skeleton, Z_skeleton, closest_pts_sorted, color_rgb_value = (1,0,0), scale_factor = sf_value*1.5)
        
        
        # show all the end_vlist_offset nodes
        #graph_vis = draw_nodes_index(X_skeleton, Y_skeleton, Z_skeleton, idx_visited_start, color_rgb_value = (0,1,0), scale_factor = sf_value)
        

        
        mlab.show()
        

        
        
        #3. visualize sphere and vectors
        ###############################################################################
        
        # Display a semi-transparent sphere, for the surface of the Earth
        mlab.figure("sphere_representation_rotation_vector", size = (800, 800), bgcolor = (0, 0, 0))
        mlab.clf()
        # We use a sphere Glyph, through the points3d mlab function, rather than
        # building the mesh ourselves, because it gives a better transparent
        # rendering.
        sphere = mlab.points3d(0, 0, 0, scale_mode='none',
                                scale_factor=2,
                                color=(0.67, 0.77, 0.93),
                                resolution=50,
                                opacity=0.7,
                                name='Earth')

        # These parameters, as well as the color, where tweaked through the GUI,
        # with the record mode to produce lines of code usable in a script.
        sphere.actor.property.specular = 0.45
        sphere.actor.property.specular_power = 5
        # Backface culling is necessary for more a beautiful transparent
        # rendering.
        sphere.actor.property.backface_culling = True

        #scalars = np.random.randint(1, size = (len(rotVec_rec),3))
        
        
        
        ################################################################
        #draw rotatin vectors
        
        # Key point: set an integer for each point
        #scalars = genotype_sub

        # Define color table (including alpha), which must be uint8 and [0,255]
        #colors = (np.random.random((N, 4))*255).astype(np.uint8)
        #colors[:,-1] = 255 # No transparency
        
        
        if type_quaternion == 0:
            rotVec_sel = np.asarray(rotVec_average_cluster, dtype = object)
            
        elif type_quaternion == 1:
            rotVec_sel = np.asarray(rotVec_composition_cluster, dtype = object)
            
        elif type_quaternion == 2:
            rotVec_sel = np.asarray(rotVec_diff_cluster, dtype = object)

        
        # draw all the rotation vectors in pipeline
        color_cluser = [(1, 0, 0), (0, 1, 0), (0, 0, 1)]
        
        line_width_value = 2.0
        
        pts = draw_rotation_vectors(rotVec_sel, color_cluser, line_width_value)
        
        
        #####################################################################33
        # draw cluster center rotation vector
        
        
        line_width_value = 6.0
        
        rotVec_centroid_list = [i for i in rotVec_centroid_cluster]
        
        #for i in range(len(rotVec_centroid_list)):
        for i, vectors in enumerate(rotVec_centroid_list):
                        
            #print(rotVec_centroid_list[i])
        
            pts = mlab.quiver3d(0, 0, 0, np.asarray(vectors)[0], np.asarray(vectors)[1], np.asarray(vectors)[2], color = color_cluser[i], mode = '2darrow', line_width = line_width_value)
        
        
        ###############################################################################
        
        # Plot the equator and the tropiques
        theta = np.linspace(0, 2 * np.pi, 100)
        for angle in (- np.pi / 6, 0, np.pi / 6):
            x = np.cos(theta) * np.cos(angle)
            y = np.sin(theta) * np.cos(angle)
            z = np.ones_like(theta) * np.sin(angle)

            pts = mlab.plot3d(x, y, z, color=(1, 1, 1), opacity=0.2, tube_radius=None)

        mlab.view(63.4, 73.8, 4, [-0.05, 0, 0])
    
        mlab.orientation_axes()
        '''
        

        #################################################################################
        
        
        mlab.orientation_axes()
        
        mlab.show()
        

    
    return number_cluster, percent_sorted, q_average_cluster, q_composition_cluster, q_diff_cluster,\
            rotVec_average_cluster, rotVec_composition_cluster, rotVec_diff_cluster, \
            distance_cluster, path_length_cluster
            #, rotVec_centroid_cluster, q_centroid_cluster

    




if __name__ == '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "path to *.ply model file")
    ap.add_argument("-m1", "--model_skeleton", required = True, help = "skeleton file name")
    ap.add_argument("-m2", "--model_pcloud", required = False, default = None, help = "point cloud model file name, same path with ply model")
    ap.add_argument("-n", "--n_cluster", required = False, type = int, default = 0, help = "Number of clusters to filter the small length paths")
    ap.add_argument("-r", "--len_ratio", required = False, type = int, default = 50, help = "length threshold to filter the roots, number of nodes in the shortest length path")
    ap.add_argument("-tq", "--type_quaternion", required = False, type = int, default = 0, help = "analyze quaternion type, average_quaternion=0, composition_quaternion=1, diff_quaternion=2, distance_quaternion=3")
    ap.add_argument("-v", "--visualize_model", required = False, type = int, default = 0, help = "Display model or not, deafult not display")
    args = vars(ap.parse_args())



    # setting path to model file 
    current_path = args["path"]
    filename_skeleton = args["model_skeleton"]
    model_skeleton_name_base = os.path.splitext(current_path + filename_skeleton)[0]
    
    #thresh_join = args["thresh_join"]
    #thresh_length = args["thresh_length"]
    
    len_ratio = args["len_ratio"]
    
    #number_cluster = args["n_cluster"]
    
    type_quaternion = args["type_quaternion"]
    
    if args["model_pcloud"] is None:
        filename_pcloud = None
    else:
        filename_pcloud = args["model_pcloud"]
    
    # analysis result path
    #print ("results_folder: " + current_path + "\n")
    
    
    # save folder construction
    abs_path = os.path.abspath(current_path)
    
    #print(abs_path)
    
    
    if type_quaternion == 0:
        mkpath = abs_path +'/average'
    elif type_quaternion == 1:
        mkpath = abs_path +'/composition'
    elif type_quaternion == 2:
        mkpath = abs_path +'/diff'
    elif type_quaternion == 3:
        mkpath = abs_path +'/distance'
        
    print(mkpath)

    mkdir(mkpath)
    save_path = mkpath + '/'
    print ("results_folder: " + save_path)
    
    
    
    
    result_list = []
    
    (number_cluster, percent_sorted, q_average_cluster, q_composition_cluster, q_diff_cluster,\
            rotVec_average_cluster, rotVec_composition_cluster, rotVec_diff_cluster, \
            distance_cluster, path_length_cluster) = analyze_skeleton(current_path, filename_skeleton, filename_pcloud)
    
    
    
    result_traits = []
    
    percent_arr_list = []
    
    #rotVec_centroid_arr_list = []
    #q_centroid_cluster_list = []
    
    path_len_arr_list = []
    
    q_average_arr_list = []
    q_composition_arr_list = []
    q_diff_arr_list = []
    
    q_distance_arr_list = []
    
    rotVec_avg_arr_list = []
    rotVec_composition_arr_list = []
    rotVec_diff_arr_list = []
    
    for i in range(number_cluster):
        
        traits_row = []
        
        
        
        q_average_arr = np.vstack(q_average_cluster[i])
        q_composition_arr = np.vstack(q_composition_cluster[i])
        q_diff_arr = np.vstack(q_diff_cluster[i])
        
        rotVec_avg_arr = np.vstack(rotVec_average_cluster[i])
        rotVec_composition_arr = np.vstack(rotVec_composition_cluster[i])
        rotVec_diff_arr = np.vstack(rotVec_diff_cluster[i])

        path_len_arr = np.vstack(path_length_cluster[i])
        q_distance_arr = np.vstack(distance_cluster[i])
        
        percent_arr = np.repeat(percent_sorted[i], repeats = len(q_average_arr), axis = 0)

        #rotVec_centroid_arr = np.tile(rotVec_centroid_cluster[i],(len(q_average_arr),1))
        
        #q_centroid_arr = np.tile(q_centroid_cluster[i],(len(q_average_arr),1))
        
        
        q_average_arr_list.append(q_average_arr)
        q_composition_arr_list.append(q_composition_arr)
        q_diff_arr_list.append(q_diff_arr)

        q_distance_arr_list.append(q_distance_arr)

        rotVec_avg_arr_list.append(rotVec_avg_arr)
        rotVec_composition_arr_list.append(rotVec_composition_arr)
        rotVec_diff_arr_list.append(rotVec_diff_arr)
        
        percent_arr_list.append(percent_arr)
        #rotVec_centroid_arr_list.append(rotVec_centroid_arr)
        #q_centroid_cluster_list.append(q_centroid_arr)
        path_len_arr_list.append(path_len_arr)
        
        #print("rotVec_centroid_arr: {} \n".format(rotVec_centroid_arr))
        
        for i, (v0, v1,v2,v3,v4, v5,v6,v7,v8, v9,v10,v11,v12, v13,v14,v15, v16,v17,v18, v19,v20,v21, v22,v23,v24, v25) in enumerate(zip(percent_arr, q_average_arr[:,0], q_average_arr[:,1], q_average_arr[:,2], q_average_arr[:,3],\
                                                        q_composition_arr[:,0], q_composition_arr[:,1], q_composition_arr[:,2], q_composition_arr[:,3],\
                                                        q_diff_arr[:,0], q_diff_arr[:,1], q_diff_arr[:,2], q_diff_arr[:,3],\
                                                        q_distance_arr[:,0], q_distance_arr[:,1], q_distance_arr[:,2],\
                                                        rotVec_avg_arr[:,0], rotVec_avg_arr[:,1], rotVec_avg_arr[:,2],\
                                                        rotVec_composition_arr[:,0], rotVec_composition_arr[:,1], rotVec_composition_arr[:,2],\
                                                        rotVec_diff_arr[:,0], rotVec_diff_arr[:,1], rotVec_diff_arr[:,2], path_len_arr[:,0])):
                                                        #rotVec_centroid_arr[:,0], rotVec_centroid_arr[:,1], rotVec_centroid_arr[:,2], \
                                                        #q_centroid_arr[:,0], q_centroid_arr[:,1], q_centroid_arr[:,2], q_centroid_arr[:,3])):
                                                            
            traits_row.append([v0, v1,v2,v3,v4, v5,v6,v7,v8, v9,v10,v11,v12, v13,v14,v15, v16,v17,v18, v19,v20,v21, v22,v23,v24, v25])
        
        result_traits.append(traits_row)

    #save reuslt file
    ####################################################################

    

    trait_path = os.path.dirname(save_path + filename_skeleton)
    
    folder_name = os.path.basename(trait_path)
    
    #print("current_path folder ={}".format(folder_name))
    
    # create trait file using sub folder name
    trait_file = (save_path + folder_name + '_quaternion.xlsx')
    
    #trait_file_csv = (save_path + folder_name + '_quaternion.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        
        
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet_quaternion_1 = wb['sheet_quaternion_1']
        sheet_quaternion_1.delete_rows(2, sheet_quaternion_1.max_row + 1) # for entire sheet
        
        #Get the current Active Sheet
        #sheet_quaternion_2 = wb['sheet_quaternion_2']
        #sheet_quaternion_2.delete_rows(2, sheet_quaternion_2.max_row + 1) # for entire sheet
        
        #Get the current Active Sheet
        #sheet_quaternion_3 = wb['sheet_quaternion_3']
        #sheet_quaternion_3.delete_rows(2, sheet_quaternion_3.max_row + 1) # for entire sheet
        
    else:
        # Keep presets
        # Keep presets
        wb = openpyxl.Workbook()
       
        
        sheet_quaternion_1 = wb.active
        sheet_quaternion_1.title = "sheet_quaternion_1"

        sheet_quaternion_1.cell(row = 1, column = 1).value = 'Ratio of cluster'
        
        sheet_quaternion_1.cell(row = 1, column = 2).value = 'quaternion_a'
        sheet_quaternion_1.cell(row = 1, column = 3).value = 'quaternion_b'
        sheet_quaternion_1.cell(row = 1, column = 4).value = 'quaternion_c'
        sheet_quaternion_1.cell(row = 1, column = 5).value = 'quaternion_d'
        
        sheet_quaternion_1.cell(row = 1, column = 6).value = 'composition_quaternion_a'
        sheet_quaternion_1.cell(row = 1, column = 7).value = 'composition_quaternion_b'
        sheet_quaternion_1.cell(row = 1, column = 8).value = 'composition_quaternion_c'
        sheet_quaternion_1.cell(row = 1, column = 9).value = 'composition_quaternion_d'
        
        sheet_quaternion_1.cell(row = 1, column = 10).value = 'diff_quaternion_a'
        sheet_quaternion_1.cell(row = 1, column = 11).value = 'diff_quaternion_b'
        sheet_quaternion_1.cell(row = 1, column = 12).value = 'diff_quaternion_c'
        sheet_quaternion_1.cell(row = 1, column = 13).value = 'diff_quaternion_d'
        
        sheet_quaternion_1.cell(row = 1, column = 14).value = 'distance_absolute'
        sheet_quaternion_1.cell(row = 1, column = 15).value = 'distance_intrinsic'
        sheet_quaternion_1.cell(row = 1, column = 16).value = 'distance_symmetrized'
        
        sheet_quaternion_1.cell(row = 1, column = 17).value = 'rotVec_avg_0'
        sheet_quaternion_1.cell(row = 1, column = 18).value = 'rotVec_avg_1'
        sheet_quaternion_1.cell(row = 1, column = 19).value = 'rotVec_avg_2'
        
        sheet_quaternion_1.cell(row = 1, column = 20).value = 'rotVec_composition_0'
        sheet_quaternion_1.cell(row = 1, column = 21).value = 'rotVec_composition_1'
        sheet_quaternion_1.cell(row = 1, column = 22).value = 'rotVec_composition_2'
        
        sheet_quaternion_1.cell(row = 1, column = 23).value = 'rotVec_diff_0'
        sheet_quaternion_1.cell(row = 1, column = 24).value = 'rotVec_diff_1'
        sheet_quaternion_1.cell(row = 1, column = 25).value = 'rotVec_diff_2'
        
        sheet_quaternion_1.cell(row = 1, column = 26).value = 'path_length'


        '''
        #####################################################################################
        sheet_quaternion_2 = wb.create_sheet()
        sheet_quaternion_2.title = "sheet_quaternion_2"

        sheet_quaternion_2.cell(row = 1, column = 1).value = 'Ratio of cluster'
        
        sheet_quaternion_2.cell(row = 1, column = 2).value = 'quaternion_a'
        sheet_quaternion_2.cell(row = 1, column = 3).value = 'quaternion_b'
        sheet_quaternion_2.cell(row = 1, column = 4).value = 'quaternion_c'
        sheet_quaternion_2.cell(row = 1, column = 5).value = 'quaternion_d'
        
        sheet_quaternion_2.cell(row = 1, column = 6).value = 'composition_quaternion_a'
        sheet_quaternion_2.cell(row = 1, column = 7).value = 'composition_quaternion_b'
        sheet_quaternion_2.cell(row = 1, column = 8).value = 'composition_quaternion_c'
        sheet_quaternion_2.cell(row = 1, column = 9).value = 'composition_quaternion_d'

        sheet_quaternion_2.cell(row = 1, column = 10).value = 'diff_quaternion_a'
        sheet_quaternion_2.cell(row = 1, column = 11).value = 'diff_quaternion_b'
        sheet_quaternion_2.cell(row = 1, column = 12).value = 'diff_quaternion_c'
        sheet_quaternion_2.cell(row = 1, column = 13).value = 'diff_quaternion_d'
        
        sheet_quaternion_2.cell(row = 1, column = 14).value = 'distance_absolute'
        sheet_quaternion_2.cell(row = 1, column = 15).value = 'distance_intrinsic'
        sheet_quaternion_2.cell(row = 1, column = 16).value = 'distance_symmetrized'
        
        sheet_quaternion_2.cell(row = 1, column = 17).value = 'rotVec_avg_0'
        sheet_quaternion_2.cell(row = 1, column = 18).value = 'rotVec_avg_1'
        sheet_quaternion_2.cell(row = 1, column = 19).value = 'rotVec_avg_2'
        
        sheet_quaternion_2.cell(row = 1, column = 20).value = 'rotVec_composition_0'
        sheet_quaternion_2.cell(row = 1, column = 21).value = 'rotVec_composition_1'
        sheet_quaternion_2.cell(row = 1, column = 22).value = 'rotVec_composition_2'
        
        sheet_quaternion_2.cell(row = 1, column = 23).value = 'rotVec_diff_0'
        sheet_quaternion_2.cell(row = 1, column = 24).value = 'rotVec_diff_1'
        sheet_quaternion_2.cell(row = 1, column = 25).value = 'rotVec_diff_2'
        
        sheet_quaternion_2.cell(row = 1, column = 26).value = 'path_length'

        
        #############################################################################
        sheet_quaternion_3 = wb.create_sheet()
        sheet_quaternion_3.title = "sheet_quaternion_3"

        sheet_quaternion_3.cell(row = 1, column = 1).value = 'Ratio of cluster'
        
        sheet_quaternion_3.cell(row = 1, column = 2).value = 'quaternion_a'
        sheet_quaternion_3.cell(row = 1, column = 3).value = 'quaternion_b'
        sheet_quaternion_3.cell(row = 1, column = 4).value = 'quaternion_c'
        sheet_quaternion_3.cell(row = 1, column = 5).value = 'quaternion_d'
        
        sheet_quaternion_3.cell(row = 1, column = 6).value = 'composition_quaternion_a'
        sheet_quaternion_3.cell(row = 1, column = 7).value = 'composition_quaternion_b'
        sheet_quaternion_3.cell(row = 1, column = 8).value = 'composition_quaternion_c'
        sheet_quaternion_3.cell(row = 1, column = 9).value = 'composition_quaternion_d'
        
        sheet_quaternion_3.cell(row = 1, column = 10).value = 'diff_quaternion_a'
        sheet_quaternion_3.cell(row = 1, column = 11).value = 'diff_quaternion_b'
        sheet_quaternion_3.cell(row = 1, column = 12).value = 'diff_quaternion_c'
        sheet_quaternion_3.cell(row = 1, column = 13).value = 'diff_quaternion_d'
        
        sheet_quaternion_3.cell(row = 1, column = 14).value = 'distance_absolute'
        sheet_quaternion_3.cell(row = 1, column = 15).value = 'distance_intrinsic'
        sheet_quaternion_3.cell(row = 1, column = 16).value = 'distance_symmetrized'
        
        sheet_quaternion_3.cell(row = 1, column = 17).value = 'rotVec_avg_0'
        sheet_quaternion_3.cell(row = 1, column = 18).value = 'rotVec_avg_1'
        sheet_quaternion_3.cell(row = 1, column = 19).value = 'rotVec_avg_2'
        
        sheet_quaternion_3.cell(row = 1, column = 20).value = 'rotVec_composition_0'
        sheet_quaternion_3.cell(row = 1, column = 21).value = 'rotVec_composition_1'
        sheet_quaternion_3.cell(row = 1, column = 22).value = 'rotVec_composition_2'
        
        sheet_quaternion_3.cell(row = 1, column = 23).value = 'rotVec_diff_0'
        sheet_quaternion_3.cell(row = 1, column = 24).value = 'rotVec_diff_1'
        sheet_quaternion_3.cell(row = 1, column = 25).value = 'rotVec_diff_2'
        
        sheet_quaternion_3.cell(row = 1, column = 26).value = 'path_length'
        '''

    for row in result_traits[0]:
        sheet_quaternion_1.append(row)
    
    '''
    for row in result_traits[1]:
        sheet_quaternion_2.append(row)
   
    for row in result_traits[2]:
        sheet_quaternion_3.append(row)
    '''
    
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error saving Result file\n")

  
    wb = openpyxl.load_workbook(trait_file)
    
    # get_active_sheet()
    sh = wb.active 
    

    '''
    ####################################################################
    #Multi-dimension plots in ploty
    
    percent_all = np.concatenate((percent_arr_list[0], percent_arr_list[1], percent_arr_list[2]), axis = 0)
    
    #percent_all = np.concatenate((percent_arr_list[0], percent_arr_list[1]), axis = 0)
    
    
    if type_quaternion == 0:
        
        q_all = np.concatenate((q_average_arr_list[0], q_average_arr_list[1], q_average_arr_list[2]), axis = 0)
        rotVec_all = np.concatenate((rotVec_avg_arr_list[0], rotVec_avg_arr_list[1], rotVec_avg_arr_list[2]), axis = 0)
        
        plot_file = (save_path + folder_name + '_avg_quaternion.html')
        plot_file_rotVec = (save_path + folder_name + '_rotVec_avg.html')
        
    elif type_quaternion == 1:
        
        q_all = np.concatenate((q_composition_arr_list[0], q_composition_arr_list[1], q_composition_arr_list[2]), axis = 0)
        rotVec_all = np.concatenate((rotVec_composition_arr_list[0], rotVec_composition_arr_list[1], rotVec_composition_arr_list[2]), axis = 0)
        
        plot_file = (save_path + folder_name + '_composition_quaternion.html')
        plot_file_rotVec = (save_path + folder_name + '_rotVec_composition.html')
        
    elif type_quaternion == 2:
        
        q_all = np.concatenate((q_diff_arr_list[0], q_diff_arr_list[1], q_diff_arr_list[2]), axis = 0)
        rotVec_all = np.concatenate((rotVec_diff_arr_list[0], rotVec_diff_arr_list[1], rotVec_diff_arr_list[2]), axis = 0)
        
        plot_file = (save_path + folder_name + '_diff_quaternion.html')
        plot_file_rotVec = (save_path + folder_name + '_rotVec_diff.html')
        
    elif type_quaternion == 3:
        q_average_all = np.concatenate((q_distance_arr_list[0], q_distance_arr_list[1], q_distance_arr_list[2]), axis = 0) 
    
    
    
    
    plot_quaternion_result(q_all, percent_all, plot_file, type_quaternion, 4)
    
    plot_quaternion_result(rotVec_all, percent_all, plot_file_rotVec, type_quaternion, 3)
    
    '''
    

    
