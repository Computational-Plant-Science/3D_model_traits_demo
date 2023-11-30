"""
Version: 1.0

Summary: visualization of two sets of rotation vectors

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE

    python3 compare_distribution.py -p ~/example/quaternion/species_comp/ -v 1 -tq 0


Input:
    *.xlsx

header:
    file_name    Ratio of cluster    quaternion_a    quaternion_b    quaternion_c    quaternion_d    rotVec_rec_0    rotVec_rec_1    rotVec_rec_2    genotype    genotype_label
    'quaternion_Mahalanobis','quaternion_p','rotVec_Mahalanobis', 'rotVec_p'


"""
import os
import glob
from pathlib import Path

from mayavi import mlab
from tvtk.api import tvtk

import argparse
import pandas as pd

import numpy as np
from numpy import arctan2, sqrt
import numexpr as ne
  
import plotly
import plotly.graph_objs as go
import plotly.express as px

import plotly as py
import plotly.tools as tls
import plotly.figure_factory as ff


import matplotlib
#matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.colors as mcolors

import math
import random
import itertools

from scipy import linalg 
from scipy.spatial.transform import Rotation as R
from scipy.stats import kstest
from scipy.spatial.distance import mahalanobis

from sklearn.preprocessing import normalize, StandardScaler
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE


from statistics import mean 

import seaborn as sns

from collections import OrderedDict

import openpyxl
import csv
from tabulate import tabulate


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        print (path+' path exists!')
        return False
        



#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)



def euler_to_rotMat(yaw, pitch, roll):
    Rz_yaw = np.array([
        [np.cos(yaw), -np.sin(yaw), 0],
        [np.sin(yaw),  np.cos(yaw), 0],
        [          0,            0, 1]])
    Ry_pitch = np.array([
        [ np.cos(pitch), 0, np.sin(pitch)],
        [             0, 1,             0],
        [-np.sin(pitch), 0, np.cos(pitch)]])
    Rx_roll = np.array([
        [1,            0,             0],
        [0, np.cos(roll), -np.sin(roll)],
        [0, np.sin(roll),  np.cos(roll)]])
    # R = RzRyRx
    rotMat = np.dot(Rz_yaw, np.dot(Ry_pitch, Rx_roll))
    
    return rotMat


# RPY/Euler angles to Rotation Vector
def euler_to_rotVec(yaw, pitch, roll):

    # compute the rotation matrix
    Rmat = euler_to_rotMat(yaw, pitch, roll)
    
    theta = math.acos(((Rmat[0, 0] + Rmat[1, 1] + Rmat[2, 2]) - 1) / 2)
    sin_theta = math.sin(theta)
    if sin_theta  ==  0:
        rx, ry, rz = 0.0, 0.0, 0.0
    else:
        multi = 1 / (2 * math.sin(theta))
        rx = multi * (Rmat[2, 1] - Rmat[1, 2]) * theta
        ry = multi * (Rmat[0, 2] - Rmat[2, 0]) * theta
        rz = multi * (Rmat[1, 0] - Rmat[0, 1]) * theta
    return rx, ry, rz



#get rotation matrix between two vectors using scipy
def get_rotation_matrix(vec2, vec1):
    """ Find the rotation matrix that aligns vec1 to vec2
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    
    
    vec1 = np.reshape(vec1, (1, -1))
    
    vec2 = np.reshape(vec2, (1, -1))
    
    r = R.align_vectors(vec2, vec1)
        
    return r[0].as_matrix()
    
    '''
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return numpy.eye(3) #cross of all zeros only occurs on identical directions
    '''


    

# Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
# The quaternions are arranged as (w,x,y,z), with w being the scalar
# The result will be the average quaternion of the input. Note that the signs
# of the output quaternion can be reversed, since q and -q describe the same orientation
def averageQuaternions(Q):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape = (4,4))

    for i in range(0,M):
        q = Q[i,:]
        # multiply q with its transposed version q' and add A
        A = np.outer(q,q) + A

    # scale
    A = (1.0/M)*A
    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)
    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]
    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0])
    return np.ravel(eigenVectors[:,0])
    

def averageVectors(avg_quaternion):
    
    avg_quaternion = avg_quaternion.flatten()

    # get Rotation matrix from quaternion
    rot = R.from_quat(avg_quaternion)

    # get the rotation vector
    avg_rotVec = rot.as_rotvec()
    
    return avg_rotVec



def cMap(x):
    #whatever logic you want for colors
    return [random.random() for i in x]
    


def cart2sph(x,y,z, ceval = ne.evaluate):
    """ x, y, z :  ndarray coordinates
        ceval: backend to use: 
              - eval :  pure Numpy
              - numexpr.evaluate:  Numexpr """
    azimuth = ceval('arctan2(y,x)')
    
    xy2 = ceval('x**2 + y**2')
    
    elevation = ceval('arctan2(z, sqrt(xy2))')
    
    r = eval('sqrt(xy2 + z**2)')
    
    return azimuth, elevation, r




def plot_rotated_axes(ax, r, name = None, offset = (0, 0, 0), scale = 1):

    colors = ("#FF6666", "#005533", "#1199EE")  # Colorblind-safe RGB

    loc = np.array([offset, offset])

    for i, (axis, c) in enumerate(zip((ax.xaxis, ax.yaxis, ax.zaxis), colors)):

        axlabel = axis.axis_name

        axis.set_label_text(axlabel)

        axis.label.set_color(c)

        axis.line.set_color(c)

        axis.set_tick_params(colors = c)

        line = np.zeros((2, 3))

        line[1, i] = scale

        line_rot = r.apply(line)

        line_plot = line_rot + loc

        ax.plot(line_plot[:, 0], line_plot[:, 1], line_plot[:, 2], c)

        text_loc = line[1]*1.2

        text_loc_rot = r.apply(text_loc)

        text_plot = text_loc_rot + loc[0]

        ax.text(*text_plot, axlabel.upper(), color = c, va = "center", ha = "center")

    ax.text(*offset, name, color = "k", va = "center", ha = "center", bbox = {"fc": "w", "alpha": 0.8, "boxstyle": "circle"})



# compute Mahalanobis Distance between the point x1 and the distribution X
def mahalanobis_p2cluster(X, x1):
    
    # Calculate the mean vector and covariance matrix of the dataset
    mu = np.mean(X, axis = 0)
    sigma = np.cov(X.T)

    # Calculate the Mahalanobis Distance between two points
    dist_mahalanobis = mahalanobis(x1, mu, np.linalg.inv(sigma))

    return dist_mahalanobis



  
def analyze_mahalanobis(rotVec_rec, data_q_arr, genotype_label, genotype_unique):
    
  
    #####################################################################
    #group quaternion values and rotation vectors by genotypes
    avg_rotVec_list = []
    
    #genotype_label_unique = np.unique(genotype_label)
    
    indexes = np.unique(genotype_label, return_index = True)[1]
    
    genotype_label_unique = [genotype_label[index] for index in (indexes)]
        
    print("genotype_name = {}, genoype_ID = {}".format(genotype_unique, genotype_label_unique))

   
    #print((rotVec_rec.shape))
    ###################################################################
    #normalize rotation vectors
    rotVec_rec_norm = normalize(rotVec_rec, axis = 1, norm = 'l1')
    
    #print((normed_rotVec.shape))
    

    
    rotVec_list = []
    
    for idx, genoype_value  in enumerate(genotype_label_unique):
        
        #print("genotype_ID = {}, genoype_value = {}".format(idx, genoype_value))
        
        index_sel = np.where(genotype_label  ==  genotype_label_unique[idx])[0]
    
        #print(data_q_arr[index_sel])
        
       
        rotVec_list.append(rotVec_rec[index_sel])
        
        # use eigenvalues to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
        avg_quaternion = averageQuaternions(data_q_arr[index_sel])

        # use components averaging to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
        #avg_quaternion = ((sum_quaternion.sum(axis = 0))/len(vlist_path)).flatten()

        #the signs of the output quaternion can be reversed, since q and -q describe the same orientation
        #avg_quaternion = np.absolute(avg_quaternion)

        avg_quaternion = avg_quaternion.flatten()
        
        # transform quaternion value to rotation vector
        avg_rotVec = averageVectors(avg_quaternion)
        
        normalized_avg_rotVec = avg_rotVec/np.linalg.norm(avg_rotVec)
        
        #print("normalized_avg_rotVec = {}\n".format(normalized_avg_rotVec))
        
        avg_rotVec_list.append(normalized_avg_rotVec)
    
        
    
    
    
    ###############################################################################
    #print(rotVec_list)

    ###############################################################################
    # Generate all possible two-element combinations 
    # Convert the resulting iterator to a list
    #combinations = list(itertools.combinations(genotype_label_unique, 2))
    
    combinations = list(itertools.permutations(genotype_label_unique, 2)) 

    # Print the list of combinations to the console
    print("There are {} Genotype combinations in total\n".format(len(combinations)))

    
    
    
    genotype_pair = []
    genotype_id_pair = []
    rot_mat_pair = []
    euler_r_pair = []
    quaternion_r_pair = []
    
    #dis_mahalanobis_avg = []
    dis_mahalanobis_cluster = []
    
    # loop over all adjacent vector pairs 
    for i, value in enumerate(combinations):
        
        vec1_idx =  list(value)[0]
        vec2_idx = list(value)[1]
        
        

        # compoute rotation matrix between adjacent directed vectors
        mat_r = get_rotation_matrix(vec1  = avg_rotVec_list[vec1_idx], vec2 = avg_rotVec_list[vec2_idx])

        # compoute quaternion between adjacent directed vectors
        #The returned quaternion value is in scalar-last (x, y, z, w) format.
        quaternion_r = R.from_matrix(mat_r).as_quat()

        #compute rotation vector between adjacent directed vectors
        #rotVec_r = R.from_matrix(mat).as_rotvec()

        # change the order of the quaternion_r value from (x, y, z, w)  to (w, x, y, z)
        quaternion_r_rearanged = np.hstack((quaternion_r[3], quaternion_r[0], quaternion_r[1], quaternion_r[2]))
        
         #compute rotation angles in euler coordinates
        euler_r = R.from_matrix(mat_r).as_euler('xyz', degrees = True)
           
        
        #################################################################

        
        # compute Mahalanobis Distance between the average vector and the paired distribution in current combination
        dis_mahalanobis = mahalanobis_p2cluster(rotVec_list[vec1_idx], avg_rotVec_list[vec2_idx])
        
        dis_mahalanobis_cluster.append(dis_mahalanobis)
        
        ###################################################################
        genotype_pair.append((genotype_unique[vec1_idx], genotype_unique[vec2_idx]))
        genotype_id_pair.append((vec1_idx, vec2_idx))
        rot_mat_pair.append(mat_r)
        euler_r_pair.append(euler_r)
        quaternion_r_pair.append((quaternion_r_rearanged))

        #print("genotype_id_pair = {} {}, genoype_name_pair = {} {}, dis_mahalanobis = {}".format(vec1_idx, vec2_idx, genotype_unique[vec1_idx], genotype_unique[vec2_idx], dis_mahalanobis))
        
        print("genotype_id_pair = {} {}, genoype_name_pair = {}, dis_mahalanobis = {}".format(vec1_idx, vec2_idx, genotype_pair[i], dis_mahalanobis_cluster[i]))



    #print("dis_mahalanobis_cluster = {}\n".format((dis_mahalanobis_cluster)))
    

    #################################################################################

    ########################################################################################
    '''
    r0 = R.identity()

    ax = plt.figure().add_subplot(projection = "3d", proj_type = "ortho")
    
    plot_rotated_axes(ax, r0, name = "r0", offset = (0, 0, 0))

    for i, (euler_r, genotype_pair_name) in enumerate(zip(euler_r_pair, genotype_pair)):
        
        if i < 2:
            
            r = R.from_euler("XYZ", euler_r, degrees = True)
            
            name = str("r{}".format(i+1))
            
            plot_rotated_axes(ax, r, name, offset = ((i+1)*3, 0, 0))
            
    _ = ax.annotate("r0: Identity Rotation\n"
                    "\n"
                    "r1: Euler Rotation between" + str("{}".format(genotype_pair[0])) + " (XYZ)\n",
                    #"\n"
                    #"r1: Euler Rotation between" + str("{}".format(genotype_pair[1])) + " (XYZ)\n", 
                    xy = (0.6, 0.7), xycoords = "axes fraction", ha = "left")


    #ax.set(xlim = (-1.25, 7.25), ylim = (-1.25, 1.25), zlim = (-1.25, 1.25))

    #ax.set(xticks = range(-1, 8), yticks = [-1, 0, 1], zticks = [-1, 0, 1])

    ax.set_aspect("equal", adjustable = "box")

    ax.figure.set_size_inches(12, 10)

    plt.tight_layout()
    
    result_path = current_path + 'vector_rotation.png'
    print("result file was saved as {}\n".format(result_path))
    plt.savefig(result_path, bbox_inches = 'tight', dpi = 1000)
    plt.close()
    #plt.show()
    '''
    ##########################################################################################
    
    if args['visualize']  ==  1:
        ###############################################################################
        # Display a semi-transparent sphere

        mlab.figure("Rotation_vector_in_sphere", size = (800, 800), bgcolor = (0, 0, 0))
        
        # use a sphere Glyph, through the points3d mlab function, rather than
        # building the mesh ourselves, because it gives a better transparent
        # rendering.
        sphere = mlab.points3d(0, 0, 0, scale_mode = 'none',
                                scale_factor = 2,
                                color = (0.67, 0.77, 0.93),
                                resolution = 50,
                                opacity = 0.7,
                                name = 'Sphere')

        # These parameters, as well as the color, where tweaked through the GUI,
        # with the record mode to produce lines of code usable in a script.
        sphere.actor.property.specular = 0.45
        sphere.actor.property.specular_power = 5
        
        # Backface culling is necessary for more a beautiful transparent rendering.
        sphere.actor.property.backface_culling = True
        
        #cmap = matplotlib.cm.get_cmap('viridis')
        

        
        '''
        ###########################################################################
        # Visualzie vectors by genotypes, colored by different genotypes
        
        #rotVec_rec_norm = rotVec_rec
        
        cm = plt.get_cmap('turbo')
        
        # Primitives
        #N = len(np.unique(genotype_label)) 
        
        N = len(rotVec_rec_norm)
        
        #print(N)

        # Key point: set an integer for each point
        scalars = genotype_label

        # Define color table (including alpha), which must be uint8 and [0,255]
        colors = (np.random.random((N, 4))*255).astype(np.uint8)
        colors[:,-1] = 255 # No transparency

        zeros = np.zeros(N)

        # draw all the rotation vectors in pipeline
        #mlab.quiver3d( 0,0,0, Vec_arr[0], Vec_arr[1], Vec_arr[2], color = current_color)
        sphere = mlab.quiver3d(zeros,zeros,zeros, rotVec_rec_norm[:,0], rotVec_rec_norm[:,1], rotVec_rec_norm[:,2], scalars = scalars, mode = '2ddash')
            
        # Color by scalar
        sphere.glyph.color_mode = 'color_by_scalar' 

        # Set look-up table and redraw
        sphere.module_manager.scalar_lut_manager.lut.table = colors
        '''

        
        ###################################################################################################################
        # visualize average rotation vector from average quaterunion
        
        print("{} genotypes in total, representative rotation vectors  = {}\n".format(len(genotype_label_unique), avg_rotVec_list))
        
        cmap = get_cmap(len(avg_rotVec_list))
        
        #color_cluser = [(1, 0, 0), (0, 1, 0), (0, 0, 1)]
        
        sf_value = 0.02
        
        
        for idx, avg_rotVec  in enumerate(avg_rotVec_list):
            
            vec_color = cmap(idx)[:len(cmap(idx))-1]
            
            sphere = mlab.quiver3d(0,0,0, avg_rotVec[0], avg_rotVec[1], avg_rotVec[2], color = vec_color, mode = '2darrow', line_width = 10)
                                
            sphere = mlab.points3d(avg_rotVec[0], avg_rotVec[1], avg_rotVec[2], color = vec_color, mode = 'sphere', scale_factor = sf_value*2)
            
            sphere = mlab.text3d(avg_rotVec[0], avg_rotVec[1], avg_rotVec[2], str("{}".format(genotype_unique[idx])), color = vec_color, scale = (sf_value, sf_value, sf_value))
            

        mlab.show()
    
    
    
    # return values
        
    return genotype_id_pair, genotype_pair, quaternion_r_pair, euler_r_pair, dis_mahalanobis_cluster
    





if __name__  ==  '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "path to *.ply model file")
    ap.add_argument("-ft", "--filetype", required = False, default = 'xlsx', help = "file type")
    ap.add_argument("-v", "--visualize", required = False, type =  int, default = 0, help = "Visualize rotation vector or not")
    ap.add_argument("-tq", "--type_quaternion", required = False, type = int, default = 0, help = "analyze quaternion type, average_quaternion = 0, composition_quaternion = 1, diff_quaternion = 2, distance_quaternion = 3")
    ap.add_argument("-r", "--result", required = False,    help="result path")

    args = vars(ap.parse_args())

    
    ###################################################################
    
    current_path = args["path"]
    
    file_path = current_path + '*.' + args['filetype']
    
    type_quaternion = args["type_quaternion"]

    # get the absolute path of all Excel files 
    ExcelFiles_list = glob.glob(file_path)
    
    
    #test_Mahalanobis(0, 0, 0)

    
    ####################################################################
    # loop over the list of excel files
    data_q = []
    
    data_v = []
    
    genotype_label = []
    
    #sample_rate = 100
    
    for f in ExcelFiles_list:
        
        filename = Path(f).name
        
        base_name = filename.replace(".xlsx", "")
       
        print("Processing file '{}'...\n".format(filename))
        
        # read the csv file
        df = pd.read_excel(f)
        

        
        ###############################################################
        #get all path related rotation vectors
        #rotVec = np.vstack((data['rotVec_rec_0'],data['rotVec_rec_1'],data['rotVec_rec_2'])).T
        
        if type_quaternion  ==  0:
            cols_vec = ['rotVec_avg_0','rotVec_avg_1','rotVec_avg_2']
        elif type_quaternion  ==  1:
            cols_vec = ['rotVec_composition_0','rotVec_composition_1','rotVec_composition_2']
        elif type_quaternion  ==  2:
            cols_vec = ['rotVec_diff_0','rotVec_diff_1','rotVec_diff_2']
        elif type_quaternion  ==  3:
            cols_vec = ['rotVec_avg_0','rotVec_avg_1','rotVec_avg_2']
        

        data_v = df[cols_vec].values.tolist()
        
        
        data_v_arr = np.asarray(data_v)


       
        ################################################################
        #get quarterunion values
        
        if type_quaternion  ==  0:
            cols_q = ['quaternion_a','quaternion_b','quaternion_c', 'quaternion_d']
        elif type_quaternion  ==  1:
            cols_q = ['composition_quaternion_a','composition_quaternion_b','composition_quaternion_c', 'composition_quaternion_d']
        elif type_quaternion  ==  2:
            cols_q = ['diff_quaternion_a','diff_quaternion_b','diff_quaternion_c', 'diff_quaternion_d']
        elif type_quaternion  ==  3:
            cols_q = ['distance_absolute','distance_intrinsic', 'distance_symmetrized']
        
        data_q = df[cols_q].values.tolist()
        
        
        data_q_arr = np.asarray(data_q)
        
    
    
        #################################################################################

        #fig = px.scatter(df, x = "quaternion_Mahalanobis", y = "rotVec_Mahalanobis", color = "genotype", symbol = "genotype")
        #plotly.offline.plot(fig, auto_open = False, filename = current_path + 'Mahalanobis_q_rotVec.html')
        
        
        #PCA Visualize all the original dimensions
        features = cols_q

        fig = px.scatter_matrix(df, dimensions = features, color = "genotype")
        fig.update_traces(diagonal_visible = False)
        plotly.offline.plot(fig, auto_open = False, filename = current_path + 'PCA_quaternion.html')
        
        
        #Visualize all the principal components
        pca = PCA()
        
        components = pca.fit_transform(df[features])
        
        labels = {
        str(i): f"PC {i+1} ({var:.1f}%)"
        for i, var in enumerate(pca.explained_variance_ratio_ * 100)
        }

        fig = px.scatter_matrix(components, labels = labels, dimensions = range(4), color = df["genotype"])
        
        fig.update_traces(diagonal_visible = False)
        
        plotly.offline.plot(fig, auto_open = False, filename = current_path + 'PCA_quaternion_components.html')
        
        
        # 2D PCA Scatter Plot and Loadings
        X = df[features]

        pca = PCA(n_components = 2)
        
        components = pca.fit_transform(X)

        loadings = pca.components_.T * np.sqrt(pca.explained_variance_)

        fig = px.scatter(components, x = 0, y = 1, color = df['genotype'])

        for i, feature in enumerate(features):
            fig.add_annotation(
                ax = 0, ay = 0,
                axref = "x", ayref = "y",
                x = loadings[i, 0],
                y = loadings[i, 1],
                showarrow = True,
                arrowsize = 2,
                arrowhead = 2,
                xanchor = "right",
                yanchor = "top"
            )
            fig.add_annotation(
                x = loadings[i, 0],
                y = loadings[i, 1],
                ax = 0, ay = 0,
                xanchor = "center",
                yanchor = "bottom",
                text = feature,
                yshift = 5,
            )
        
        plotly.offline.plot(fig, auto_open = False, filename = current_path + 'PCA_Loadings.html')
        

        #histogram
        #fig = px.histogram(df, x = "quaternion_Mahalanobis", y = "rotVec_Mahalanobis", color = "genotype",  marginal = "box", hover_data = df.columns)# or violin, rug
                       
        fig = px.histogram(df, x = "quaternion_Mahalanobis", color = "genotype", marginal = "box", hover_data = df.columns)
        
        plotly.offline.plot(fig, auto_open = False, filename = current_path + 'Mahalanobis_histogram.html')
        
        
        # Group data together
        #x1 = df.loc[df.genotype == genotype_unique[0], 'quaternion_Mahalanobis'].values
        #x2 = df.loc[df.genotype == genotype_unique[1], 'quaternion_Mahalanobis'].values
        #hist_data = [x1, x2]
        #group_labels = ['bean', 'maize']

        # Create distplot with custom bin_size
        #fig = ff.create_distplot(hist_data, group_labels, bin_size = .2)
        #plotly.offline.plot(fig, auto_open = False, filename = current_path + 'Mahalanobis_histogram_group.html')
        
        
        
        #Project data into 2D with t-SNE and px.scatter
        #feature_col = ['quaternion_Mahalanobis','quaternion_p','rotVec_Mahalanobis', 'rotVec_p']
        
        
        feature_col = ['quaternion_Mahalanobis', 'rotVec_Mahalanobis']
        
        features_data = df[feature_col]

        tsne = TSNE(n_components = 2, random_state = 0)
        
        projections = tsne.fit_transform(features_data)

        fig = px.scatter(projections, x = 0, y = 1, color = df.genotype, labels = {'color': 'genotype'})
        
        plotly.offline.plot(fig, auto_open = False, filename = current_path + 'Mahalanobis_TSNE.html')
        
    
    
        ################################################################
        #get genotype and label values
        genotype_label = df['genotype_label'].values.tolist()
        
        genotype_label_arr = np.asarray(genotype_label)
        
        #print("genotype_label are {} \n".format(genotype_label))


        ################################################################
        genotype_type = df['genotype'].values.tolist()
        
        genotype_name_arr = np.asarray(genotype_type)
        
        #genotype_unique = list(set(genotype_name_arr))
        

        # not sorted unique 
        genotype_unique = sorted(list(OrderedDict.fromkeys(genotype_name_arr)))
        
        #genotype_label_unique = sorted(genotype_label)
        
        print("Genotypes are {} \n".format(genotype_unique))
        
        #print("genotype_name = {}, genoype_ID = {}".format(genotype_unique, genotype_label_unique))
        
        
        ################################################################
        #get Mahalanobis distance values
        cols_ma = ['quaternion_Mahalanobis','quaternion_p','rotVec_Mahalanobis', 'rotVec_p']
        data_ma = df[cols_ma].values.tolist()
        
        data_ma_arr = np.asarray(data_ma)
        
       
        
        #The center of the box represents the median while the borders represent the first (Q1) and third quartile (Q3), respectively. 
        #sns.boxplot(data = data, x = 'quaternion_Mahalanobis', y = 'genotype')
        #plt.title("Boxplot")

        #histogram groups the data into equally wide bins and plots the number of observations within each bin
        #sns.histplot(data = df, y = 'quaternion_Mahalanobis', hue = 'genotype', bins = 50)
        #plt.title("Histogram")


        
        #Density Histogram
        fig = plt.plot(figsize  = (10, 7), tight_layout = True)
        sns.histplot(data = df, x = 'quaternion_Mahalanobis', hue = 'genotype', bins = 50, stat = 'density', common_norm = False)
        plt.title("Density Histogram")
        
        result_path = current_path + base_name + '_Density_his.png'
        print("result file was saved as {}\n".format(result_path))
        plt.savefig(result_path, bbox_inches = 'tight', dpi = 1000)
        plt.close()
        
        
        #Kernel Density
        fig = plt.plot(figsize  = (10, 7), tight_layout = True)
        sns.kdeplot(x = 'quaternion_Mahalanobis', data = df, hue = 'genotype', common_norm = False)
        plt.title("Kernel Density Function")
        result_path = current_path + base_name + '_Kernel_Density.png'
        print("result file was saved as {}\n".format(result_path))
        plt.savefig(result_path, bbox_inches = 'tight', dpi = 1000)
        plt.close()
        
        #sns.histplot(x = 'quaternion_Mahalanobis', data = df, hue = 'genotype', bins = len(df), stat = "density", element = "step", fill = False, cumulative = True, common_norm = False);
        #plt.title("Cumulative distribution function")


        
        ###########################################################################################
        q_m = df['quaternion_Mahalanobis'].values
        
        q_type_1 = df.loc[df.genotype == genotype_unique[0], 'quaternion_Mahalanobis'].values
        q_type_2 = df.loc[df.genotype == genotype_unique[1], 'quaternion_Mahalanobis'].values
        
        stat, p_value = kstest(q_type_1, q_type_2)
        print(f" Kolmogorov-Smirnov Test: statistic = {stat:.4f}, p-value = {p_value:.4f}")

        #################################################################
        #qq plot
        #q stands for quantile. The Q-Q plot plots the quantiles of the two distributions against each other. 
        #If the distributions are the same, we should get a 45-degree line.
        
        df_pct = pd.DataFrame()
        df_pct[genotype_unique[0]] = np.percentile(q_type_1, range(100))
        df_pct[genotype_unique[1]] = np.percentile(q_type_2, range(100))
        
        
        fig = plt.plot(figsize  = (10, 7), tight_layout = True)
        plt.scatter(x = genotype_unique[0], y = genotype_unique[1], data = df_pct, label = 'Actual fit');
        sns.lineplot(x = genotype_unique[0], y = genotype_unique[0], data = df_pct, color = 'r', label = 'Line of perfect fit')
        plt.xlabel('Quantile of quaternion_Mahalanobis' + genotype_unique[0])
        plt.ylabel('Quantile of quaternion_Mahalanobis' + genotype_unique[1])
        plt.legend()
        plt.title("QQ plot")
        
        result_path = current_path + base_name + '_QQ_plot.png'
        print("result file was saved as {}\n".format(result_path))
        plt.savefig(result_path, bbox_inches = 'tight', dpi = 1000)
        plt.close()
        
        ###############################################################
        #Kolmogorov-Smirnov Test
        # If the p-value is below 5%: we reject the null hypothesis that the two distributions are the same, with 95% confidence.
        
        df_ks = pd.DataFrame()
        
        df_ks['quaternion_Mahalanobis'] = np.sort(df['quaternion_Mahalanobis'].unique())
        
        
        df_ks[genotype_unique[0]] = df_ks['quaternion_Mahalanobis'].apply(lambda x: np.mean(q_type_1<= x))
        
        df_ks[genotype_unique[1]] = df_ks['quaternion_Mahalanobis'].apply(lambda x: np.mean(q_type_2<= x))
        
        print(df_ks.head())
        
        
        k = np.argmax( np.abs(df_ks[genotype_unique[0]] - df_ks[genotype_unique[1]]))
        ks_stat = np.abs(df_ks[genotype_unique[1]][k] - df_ks[genotype_unique[0]][k])
        
        y = (df_ks[genotype_unique[1]][k] + df_ks[genotype_unique[0]][k])/2
        
        
        # Creating histogram
        fig = plt.plot(figsize  = (10, 7), tight_layout = True)
        
        plt.plot('quaternion_Mahalanobis', genotype_unique[0], data = df_ks, label = genotype_unique[0])
        plt.plot('quaternion_Mahalanobis', genotype_unique[1], data = df_ks, label = genotype_unique[1])
        plt.errorbar(x = df_ks['quaternion_Mahalanobis'][k], y = y, yerr = ks_stat/2, color = 'k', capsize = 5, mew = 3, label = f"Test statistic: {ks_stat:.4f}, p_value {p_value:.4f}")
        plt.legend(loc = 'center right')
        plt.title("Kolmogorov-Smirnov Test")
        
        result_path = current_path + base_name + '_Kolmogorov_Smirnov_Test.png'
        
        print("result file was saved as {}\n".format(result_path))
        
        plt.savefig(result_path, bbox_inches = 'tight', dpi = 1000)
        plt.close()
        
        
    

        
    
    ###########################################################################
    #indices_HighN = np.where(genotype_label  ==  'HighN')
    
    #visualize rotation vectors
    #visualization_rotation_vector(np.asarray(normalized_data_v), genotype_label)
    
    #result_list = []
        
    (genotype_id_pair, genotype_pair, quaternion_r_pair, euler_r_pair, dis_mahalanobis_cluster) = analyze_mahalanobis(np.asarray(data_v_arr), np.asarray(data_q_arr), genotype_label_arr, genotype_unique)
    

    #print(genotype_pair)
    
    ###########################################################################
    #write excel output
    result_traits = []
    
    genotype_pair_list = []
    quaternion_pair_list = []
    euler_pair_list = []
    dis_mahalanobis_list = []
    
    for i in range(len(genotype_pair)):
        
        genotype_pair_arr = np.asarray(genotype_pair[i])
        quaternion_pair_arr = np.asarray(quaternion_r_pair[i])
        euler_pair_arr = np.asarray(euler_r_pair[i])
        dis_mahalanobis_cluster_arr = dis_mahalanobis_cluster[i]
        
         
        #print(genotype_pair_arr)
        #print(quaternion_pair_arr)
        #print(euler_pair_arr)
        #print(dis_mahalanobis_cluster_arr)
        
        #print(genotype_pair_arr.shape)

        
        result_traits.append([genotype_pair_arr[0], genotype_pair_arr[1], \
                            quaternion_pair_arr[0], quaternion_pair_arr[1], quaternion_pair_arr[2], quaternion_pair_arr[3],\
                            euler_pair_arr[0], euler_pair_arr[1], euler_pair_arr[2],\
                            dis_mahalanobis_cluster_arr])
    
    
    
    ################################################################################################
    # Break a list into chunks of size N (N = 11, genotypes = 12, combination number = 11*12 = 132)
    dis_mahalanobis_split = list(np.array_split(np.asarray(dis_mahalanobis_cluster), len(genotype_unique)))

    #print(len(dis_mahalanobis_split[0]))
    
    result_matrix = []
    
    if len(genotype_unique) == 12:
    
        #B101	B112	DKIB014	LH123HT	Pa762	PHZ51	DKPB80_3IIH6	H96_3IIH6	LH59_PHG29	Pa762_3IIH6	PHG50_PHG47	PHZ51_LH59
        # 0      1      2        5      10      8       3               4           6           11          7           9       
        for i, value in enumerate(dis_mahalanobis_split):
            

            genotype_row = [genotype_pair[0], genotype_pair[1], genotype_pair[4], genotype_pair[9], genotype_pair[7], genotype_pair[2], genotype_pair[3], genotype_pair[5], genotype_pair[10], genotype_pair[6], genotype_pair[8]]
            
            #print(genotype_row)
            
            matrix_row = [value[0], value[1], value[4], value[9], value[7], value[2], value[3], value[5], value[10], value[6], value[8]]

            matrix_row.insert(int(i), 0)
            
            #print(matrix_row)
            
            result_matrix.append(matrix_row)
        

    ###############################################################################
    #output in command window in a sum table
    #table = tabulate(result_traits, headers = ['genotype_pair_element_A', 'genotype_pair_element_B', 'quaternion_a', 'quaternion_b', 'quaternion_c' ,'quaternion_d', 'Euler_rotation_roll', 'Euler_rotation_pitch', 'Euler_rotation_yaw', 'mahalanobis_distance_between_elements'], tablefmt = 'orgtbl')

    #print(table + "\n")
    
    

    
    ################################################################################
    # save excel results
    
    
    mkpath = os.path.dirname(current_path) +'/results'
    mkdir(mkpath)
    save_path = mkpath + '/'
    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        #trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (save_path + 'trait.xlsx')
        #trait_file_csv = (save_path + 'trait.csv')


    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row + 1) # for entire sheet
        
        # create sheet for genetic distance comparison
        sheet_matrix = wb.create_sheet()
        

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        #########################################################################
        sheet.cell(row = 1, column = 1).value = 'Genotype_pair_element_A'
        sheet.cell(row = 1, column = 2).value = 'Genotype_pair_element_B'
        sheet.cell(row = 1, column = 3).value = 'Quaternion_a'
        sheet.cell(row = 1, column = 4).value = 'Quaternion_b'
        sheet.cell(row = 1, column = 5).value = 'Quaternion_c'
        sheet.cell(row = 1, column = 6).value = 'Quaternion_d'
        sheet.cell(row = 1, column = 7).value = 'Euler_rotation_roll'
        sheet.cell(row = 1, column = 8).value = 'Euler_rotation_pitch'
        sheet.cell(row = 1, column = 9).value = 'Euler_rotation_yaw'
        sheet.cell(row = 1, column = 10).value = 'Mahalanobis_distance_between_elements'


        ########################################################################
        sheet_matrix = wb.create_sheet()

        sheet_matrix.cell(row = 1, column = 1).value = 'Individual1'
        sheet_matrix.cell(row = 1, column = 2).value = 'B101'
        sheet_matrix.cell(row = 1, column = 3).value = 'B112'
        sheet_matrix.cell(row = 1, column = 4).value = 'DKIB014'
        sheet_matrix.cell(row = 1, column = 5).value = 'LH123HT'
        sheet_matrix.cell(row = 1, column = 6).value = 'Pa762'
        sheet_matrix.cell(row = 1, column = 7).value = 'PHZ51'
        sheet_matrix.cell(row = 1, column = 8).value = 'DKPB80_3IIH6'
        sheet_matrix.cell(row = 1, column = 9).value = 'H96_3IIH6'
        sheet_matrix.cell(row = 1, column = 10).value = 'LH59_PHG29'
        sheet_matrix.cell(row = 1, column = 11).value = 'Pa762_3IIH6'
        sheet_matrix.cell(row = 1, column = 12).value = 'PHG50_PHG47'
        sheet_matrix.cell(row = 1, column = 12).value = 'PHZ51_LH59'

        
    for row in result_traits:
        sheet.append(row)

    for row in result_matrix:
        sheet_matrix.append(row)
        
    #save the excel file
    wb.save(trait_file)
    
    
    
    
    
    

    ###########################################################################################3
    #print(ExcelFiles_list[0])
    # merge all excel files read all Excel files at once
    df = pd.concat(pd.read_excel(excelFile) for excelFile in ExcelFiles_list)
    
    n_colors = len(np.unique(genotype_label))
    
    if n_colors < 2:
        markercolor = px.colors.sample_colorscale("turbo", [n/(n_colors) for n in range(n_colors)])
    else:
        markercolor = px.colors.sample_colorscale("turbo", [n/(n_colors -1) for n in range(n_colors)])
    
    
    if type_quaternion  ==  0:

        X_col_q = 'quaternion_b'
        Y_col_q = 'quaternion_c'
        Z_col_q = 'quaternion_d'
        
        X_col_v = 'rotVec_avg_0'
        Y_col_v = 'rotVec_avg_1'
        Z_col_v = 'rotVec_avg_2'
        
        quaternion_4D = (current_path + 'avg_quaternion_4D.html')
        
    elif type_quaternion  ==  1:
        
        X_col_q = 'composition_quaternion_b'
        Y_col_q = 'composition_quaternion_c'
        Z_col_q = 'composition_quaternion_d'
        
        X_col_v = 'rotVec_composition_0'
        Y_col_v = 'rotVec_composition_1'
        Z_col_v = 'rotVec_composition_2'
        
        quaternion_4D = (current_path + 'composition_quaternion_4D.html')
        
    elif type_quaternion  ==  2:
        
        X_col_q = 'diff_quaternion_b'
        Y_col_q = 'diff_quaternion_c'
        Z_col_q = 'diff_quaternion_d'

        X_col_v = 'rotVec_diff_0'
        Y_col_v = 'rotVec_diff_1'
        Z_col_v = 'rotVec_diff_2'
        
        quaternion_4D = (current_path + 'diff_quaternion_4D.html')
        

        

    
    fig = px.scatter_3d(df, x = X_col_q, y = Y_col_q, z = Z_col_q, color = 'genotype',   size_max = 20, opacity = 1.0)

    #fig = px.scatter_3d(df, x = 'quaternion_b', y = 'quaternion_c', z = 'quaternion_d', color = 'genotype',  color_discrete_sequence = markercolor,  symbol = 'genotype',  size = 'label', size_max = 20, opacity = 1.0)

    fig.update_traces(marker_size = 4)
    
    
    #fig.add_scatter3d(x = data2['quaternion_b'], y = data2['quaternion_c'], z = data2['quaternion_d'])

    #fig.add_scatter3d(df, x = 'quaternion_b', y = 'quaternion_c', z = 'quaternion_d', color = 'genotype',  color_discrete_sequence = markercolor)

    #fig.update_traces(marker_size = 10)
    
   
    plotly.offline.plot(fig, auto_open = False, filename = quaternion_4D)

    
    
    
    ######################################################################
    fig = px.scatter_3d(df, x = X_col_v, y = Y_col_v, z = Z_col_v, color = 'genotype', size_max = 20, opacity = 1.0)
    
    fig.update_traces(marker_size = 4)
    
    
    
    plotly.offline.plot(fig, auto_open = False, filename = quaternion_4D)
    



    

