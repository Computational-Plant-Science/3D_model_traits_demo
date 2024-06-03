"""
Version: 1.5

Summary: CDF visualization

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 average_quarterunion.py -p ~/cluster_test/syngenta_data/results/


argument:
("-p", "--path", required = True,    help = "path to image file")

"""

#!/usr/bin/python
# Standard Libraries



import glob
from pathlib import Path
import os,fnmatch,os.path
import argparse

import shutil
import pandas as pd
import numpy as np 

from scipy.spatial.transform import Rotation as R
import math

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import csv


def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False


def euler_to_rotMat(yaw, pitch, roll):
    Rz_yaw = np.array([
        [np.cos(yaw), -np.sin(yaw), 0],
        [np.sin(yaw),  np.cos(yaw), 0],
        [          0,            0, 1]])
    Ry_pitch = np.array([
        [ np.cos(pitch), 0, np.sin(pitch)],
        [             0, 1,             0],
        [-np.sin(pitch), 0, np.cos(pitch)]])
    Rx_roll = np.array([
        [1,            0,             0],
        [0, np.cos(roll), -np.sin(roll)],
        [0, np.sin(roll),  np.cos(roll)]])
    # R = RzRyRx
    rotMat = np.dot(Rz_yaw, np.dot(Ry_pitch, Rx_roll))
    
    return rotMat
    
    

# RPY/Euler angles to Rotation Vector
def euler_to_rotVec(yaw, pitch, roll):

    # compute the rotation matrix
    Rmat = euler_to_rotMat(yaw, pitch, roll)
    
    theta = math.acos(((Rmat[0, 0] + Rmat[1, 1] + Rmat[2, 2]) - 1) / 2)
    sin_theta = math.sin(theta)
    if sin_theta == 0:
        rx, ry, rz = 0.0, 0.0, 0.0
    else:
        multi = 1 / (2 * math.sin(theta))
        rx = multi * (Rmat[2, 1] - Rmat[1, 2]) * theta
        ry = multi * (Rmat[0, 2] - Rmat[2, 0]) * theta
        rz = multi * (Rmat[1, 0] - Rmat[0, 1]) * theta
    return rx, ry, rz
    

# Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
# The quaternions are arranged as (w,x,y,z), with w being the scalar
# The result will be the average quaternion of the input. Note that the signs
# of the output quaternion can be reversed, since q and -q describe the same orientation
def averageQuaternions(Q):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))

    for i in range(0,M):
        q = Q[i,:]
        # multiply q with its transposed version q' and add A
        A = np.outer(q,q) + A

    # scale
    A = (1.0/M)*A
    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)
    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]
    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0])
    return np.ravel(eigenVectors[:,0])
    
    

if __name__ == '__main__':

    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to excel file")
    #ap.add_argument("-f", "--filename", required = True, help = "data file name")
    args = vars(ap.parse_args())

    ###################################################################
    
    current_path = args["path"]
    
    file_path = current_path + "*.xlsx"

    # get the absolute path of all Excel files 
    ExcelFiles_list = glob.glob(file_path)



    mkpath = current_path + '/avg_quarterunion/'
    mkdir(mkpath)
    save_path = mkpath + '/'

    ####################################################################
    # add filename to the first column of all excel files
    # loop over the list of excel files
    
    for f in ExcelFiles_list:
        
        # extarct path and name of the image file
        abs_path = os.path.abspath(f)

        filename, file_extension = os.path.splitext(abs_path)

        # extract the base name 
        base_name = os.path.splitext(os.path.basename(filename))[0]
        
        # read the csv file
        data = pd.read_excel(f)

        data.to_excel(f)
        
        ####################################################################
        
        cols = ['quaternion_a','quaternion_b','quaternion_c','quaternion_d']
        
        sum_quaternion = data[cols].values
        
        #print(sum_quaternion)
        
        # use eigenvalues to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
        avg_quaternion = averageQuaternions(sum_quaternion)

        # use components averaging to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
        #avg_quaternion = ((sum_quaternion.sum(axis=0))/len(vlist_path)).flatten()
        
        #the signs of the output quaternion can be reversed, since q and -q describe the same orientation
        avg_quaternion = np.absolute(avg_quaternion)

        avg_quaternion = avg_quaternion.flatten()

        rot = R.from_quat(avg_quaternion)
        
        avg_euler = rot.as_euler('xyz')
        
        #avg_euler = np.mean(sum_euler, axis = 0)
        
        avg_rotVec = euler_to_rotVec(avg_euler[0], avg_euler[1], avg_euler[2])

        #print((avg_quaternion[0]))
        
        print(avg_quaternion, avg_rotVec)

        ####################################################################

        # create trait file using sub folder name
        trait_file = (save_path + base_name + '_avg_q.xlsx')
        
        result_avg = []
        
        result_avg.append([avg_quaternion[0], avg_quaternion[1], avg_quaternion[2], avg_quaternion[3], avg_rotVec[0], avg_rotVec[1], avg_rotVec[2]])
        

        #Open an xlsx for reading
        wb = openpyxl.Workbook()

        #Get the current Active Sheet
        sheet_quaternion_avg = wb.active
        sheet_quaternion_avg.title = "Average_quaternion"

        sheet_quaternion_avg.cell(row = 1, column = 1).value = 'quaternion_a'
        sheet_quaternion_avg.cell(row = 1, column = 2).value = 'quaternion_b'
        sheet_quaternion_avg.cell(row = 1, column = 3).value = 'quaternion_c'
        sheet_quaternion_avg.cell(row = 1, column = 4).value = 'quaternion_d'
        sheet_quaternion_avg.cell(row = 1, column = 5).value = 'rotVec_rec_0'
        sheet_quaternion_avg.cell(row = 1, column = 6).value = 'rotVec_rec_1'
        sheet_quaternion_avg.cell(row = 1, column = 7).value = 'rotVec_rec_2'

        for row in result_avg:
            
            sheet_quaternion_avg.append(row)
        
            #save the csv file
        wb.save(trait_file)
        
        if os.path.exists(trait_file):
            
            print("Result file was saved at {}\n".format(trait_file))
        else:
            print("Error saving Result file\n")
