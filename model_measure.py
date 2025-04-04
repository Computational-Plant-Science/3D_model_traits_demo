"""
Version: 1.5

Summary: compute the whole root traits based on 3D Sorghum model

Author: Suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 model_measure.py -i ~/example/test.ply  -o ~/example/ --n_slices 200 --slicing_factor 0.3
    
INPUT:

    3D point cloud model, cleaned and aligned along Z direction in 3D coordinates.

OUTPUT:

    Excel result file contains traits computation 

PARAMETERS:

    ("-i", "--input", dest="input", required=True, type=str, help="full path to 3D model file")
    ("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    ("--nb_neighbors", required = False, type = int, default = 20, help = "nb_neighbors")
    ("--std_ratio", required = False, type = float, default = 5.0, help = "outlier remove ratio, small number = aggresive")
    ("--black_filter", required = False, type = int, default = 0, help = "Apply black points removal filter or not, 0 = not, 1 = Apply")
    ("--black_threshold", required = False, type = float, default = 0.2, help = "threshold for black points removal")
    ("--n_slices", dest = "n_slices", type = int, required = False, default = 10,  help = "Number of planes to segment the 3d model along Z direction")
    ("--slicing_factor", dest = "slicing_factor", type = float, required = False, default = 0.7,  help = "Slicing adjust factor")
    ("--visualize", dest = "visualize", required = False, type = int, default = 0, help = "Display model or not, default not display")

"""
#!/usr/bin/env python

import subprocess, os, glob
import numpy as np
import sys
import pathlib
import argparse
from pathlib import Path


import open3d as o3d
import copy

import openpyxl

import math

from findpeaks import findpeaks

from sklearn.cluster import KMeans

#from pyransac3d import Cylinder

from plyfile import PlyData, PlyElement

import graph_tool.all as gt
import statistics


import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt

# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)



# generate folder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False


# execute script inside program
def execute_script(cmd_line):
    
    try:
        #print(cmd_line)
        #os.system(cmd_line)

        process = subprocess.getoutput(cmd_line)
        
        print(process)
        
        #process = subprocess.Popen(cmd_line, shell = True, stdout = subprocess.PIPE)
        #process.wait()
        #print (process.communicate())
        
    except OSError:
        
        print("Failed ...!\n")


# Performs statistical outlier removal on a point cloud.
def statistical_outlier_removal(pcd, nb_neighbors, std_ratio):
    """
    Performs statistical outlier removal on a point cloud.

    Args:
        pcd (open3d.geometry.PointCloud): The input point cloud.
        nb_neighbors (int): Number of neighbors to consider for each point.
        std_ratio (float): Standard deviation ratio threshold.

    Returns:
        open3d.geometry.PointCloud: The filtered point cloud.
    """
    
    #print("Statistical outlier removal\n")
    
    pcd_np = np.asarray(pcd.points)

    # Compute distances to neighbors
    pcd_tree = o3d.geometry.KDTreeFlann(pcd)
    distances = []
    for i in range(len(pcd_np)):
        _, indices, _ = pcd_tree.search_knn_vector_3d(pcd_np[i], nb_neighbors + 1)
        distances.append(np.mean(np.linalg.norm(pcd_np[indices[1:]] - pcd_np[i], axis=1)))

    # Compute mean and standard deviation of distances
    mean_dist = np.mean(distances)
    std_dist = np.std(distances)

    # Filter out outliers
    inliers = np.where(np.abs(distances - mean_dist) < std_ratio * std_dist)[0]
    filtered_pcd = pcd.select_by_index(inliers)

    return filtered_pcd



# compute parameters of a 3d point cloud by slicing it into n_slices segments
def get_pt_sel_parameter(Data_array_pt, n_slices):
    
    ####################################################################
    
    # load skeleton coordinates and radius 
    Z_pt_sorted = np.sort(Data_array_pt[:,2])
    
    pt_plane = []
    
    
    # initialize paramters
    pt_plane_center = []
    
    pt_plane_diameter_max = []
    
    pt_plane_diameter_min = []
    
    pt_plane_diameter_avg = []
    

    
    filter_plane_center = []
    
    filter_plane_volume = []
    
    filter_plane_eccentricity = []
    
    filter_plane_density = []
    
    
    # slicing models based number of planes along Z axis
    for idx, x in enumerate(range(n_slices)):
        
        ratio_s = idx/n_slices
        ratio_e = (idx+1)/n_slices
        
        print("ratio_s = {} ratio_e = {}\n".format(ratio_s, ratio_e))
        
        # index of end plane 
        idx_sel_e = int(len(Z_pt_sorted)*ratio_e) 
    
        Z_e = Z_pt_sorted[idx_sel_e]  if idx_sel_e < len(Data_array_pt) else (len(Data_array_pt) - 1)
        
        # index of start plane
        idx_sel_s = int(len(Z_pt_sorted)*ratio_s) 
    
        Z_s = Z_pt_sorted[idx_sel_s]

        # mask between the start and end plane
        Z_mask = (Data_array_pt[:,2] <= Z_e) & (Data_array_pt[:,2] >= Z_s) 
        
        Z_pt_sel = Data_array_pt[Z_mask]
        
        
        #print(Z_pt_sel.shape)
        
        # initialize the o3d object
        pcd_Z_mask = o3d.geometry.PointCloud()
    
        pcd_Z_mask.points = o3d.utility.Vector3dVector(Z_pt_sel)
        
        ###########################
        # Parameters of statistical_outlier filter for 3d point cloud
        nb_neighbors = 20
        std_ratio = 5.0
        
        # Apply noise filter
        pcd_Z_mask = statistical_outlier_removal(pcd_Z_mask, nb_neighbors, std_ratio)
        ##############################
        
        # get the diameter of the sliced model 
        (pt_diameter_max, pt_diameter_min, pt_diameter_avg, pt_length, pt_volume, pt_ob_volume) = get_pt_parameter(pcd_Z_mask)
        
        #print("Current slice diameter_max = {}, diameter_min = {}, diameter_avg = {}\n".format(pt_diameter_max, pt_diameter_min, pt_diameter_avg))
        
        # get the model center position
        model_center = pcd_Z_mask.get_center()

        pt_plane.append(pcd_Z_mask)
        
        pt_plane_center.append(model_center)
        
        #pt_plane_diameter.append(pt_diameter)
        
        filter_plane_eccentricity.append(pt_diameter_min/pt_diameter_max)
        
        pt_plane_diameter_max.append(pt_diameter_max)
        
        pt_plane_diameter_min.append(pt_diameter_min)
        
        pt_plane_diameter_avg.append(pt_diameter_avg)
        
        #filter_plane_bushiness.append(pt_volume/pt_ob_volume)
        
        # filter sliced models using sphere with radius and compute parameters
        ################################################################
        # copy current sliced model
        pt_sel_filter = copy.deepcopy(pcd_Z_mask)
        
        # get 3d points
        points = np.asarray(pt_sel_filter.points)

        # Sphere center and radius
        radius = pt_diameter_avg*0.5
        
        #print("radius =  {} \n".format(radius))

        # Calculate distances to center, set new points
        distances = np.linalg.norm(points - model_center, axis=1)
        
        pt_sel_filter.points = o3d.utility.Vector3dVector(points[distances <= radius])
        
        # filter sliced model
        (filter_diameter_max, filter_diameter_min, filter_diameter, filter_length, filter_volume, filter_density) = get_pt_parameter(pcd_Z_mask)
        
        
        filter_plane_center.append(pt_sel_filter.get_center())
        
        filter_plane_volume.append(filter_volume)
                
        #########################################################################
        # compute eccentricity using oriented bounding box axis
        
        # get OrientedBoundingBox
        obb = pt_sel_filter.get_oriented_bounding_box()

        # assign color for OrientedBoundingBox
        obb.color = (0, 0, 1)

        # get the eight points that define the bounding box.
        pcd_coord = obb.get_box_points()

        #print(obb.get_box_points())

        #pcd_coord.color = (1, 0, 0)

        # From Open3D to numpy array
        np_points = np.asarray(pcd_coord)

        # create Open3D format for points 
        #pcd_coord = o3d.geometry.PointCloud()
        #pcd_coord.points = o3d.utility.Vector3dVector(np_points)
    
        # check the length of the joint 3 vector in the bounding box to estimate the orientation of model
        list_dis = [np.linalg.norm(np_points[0] - np_points[1]), np.linalg.norm(np_points[0] - np_points[2]), np.linalg.norm(np_points[0] - np_points[3])]
        
        #print("list_dis =  {} \n".format(list_dis))
        
        filter_plane_density.append(min(list_dis[0],list_dis[1])/max(list_dis[0],list_dis[1]))
        
        #print("filter_plane_eccentricity =  {} \n".format(filter_plane_eccentricity))
        
        # get rotation matrix
        #rotation_array = obb.R.tolist()
        # get the eight points that define the bounding box.
        #pcd_coord = obb.get_box_points()
        #print("obb.R =  {} \n".format(obb.R))
        #rotation_array = obb.R.tolist()
        #r = R.from_matrix(rotation_array)
        #orientation_angle = r.as_euler('xyz', degrees=True)
        #print("orientation_angle =  {} \n".format(orientation_angle))
        
        #visualize the convex hull as a red LineSet
        #o3d.visualization.draw_geometries([pt_sel_filter, obb])

        
        ################################################################
        
        #pt_plane_volume.append(pt_volume)
        

    return pt_plane, pt_plane_center, pt_plane_diameter_max, pt_plane_diameter_min, pt_plane_diameter_avg, filter_plane_center, filter_plane_volume, filter_plane_eccentricity, filter_plane_density
    

    

# compute dimensions of point cloud
def get_pt_parameter(pcd):
    
    # get convex hull of a point cloud is the smallest convex set that contains all points.
    hull, idx = pcd.compute_convex_hull()
    
    #hull_ls = o3d.geometry.LineSet.create_from_triangle_mesh(hull)
    #hull_ls.paint_uniform_color((1, 0, 0))
    
    #o3d.visualization.draw_geometries([pcd, hull_ls])
    
    # get AxisAlignedBoundingBox
    aabb = pcd.get_axis_aligned_bounding_box()
    aabb.color = (0, 1, 0)
    
    #Get the extent/length of the bounding box in x, y, and z dimension.
    aabb_extent = aabb.get_extent()
    
    aabb_extent_half = aabb.get_half_extent()
    
    # get OrientedBoundingBox
    obb = pcd.get_oriented_bounding_box()
    
    obb.color = (0, 0, 1)

    #visualize the convex hull as a red LineSet
    #o3d.visualization.draw_geometries([pcd, aabb, obb, hull_ls])
    
    
    # compute parameters
    #pt_diameter_max = max(aabb_extent[0], aabb_extent[1])
    
    pt_diameter_max = (math.sqrt(pow(aabb_extent[0],2) + pow(aabb_extent[1],2)) + max(aabb_extent[0], aabb_extent[1])) / 2.0
    
    pt_diameter_min = min(aabb_extent_half[0], aabb_extent_half[1])
    
    pt_diameter_avg = (pt_diameter_max + pt_diameter_min)*0.5

    pt_length = (aabb_extent[2])

    # compute as cylinder
    #pt_volume = np.pi * ((pt_diameter_avg*0.5) ** 2) * pt_length
    
    #print("hull.get_volume() = {}\n".format(hull.get_volume()))

    
    # compute volume
    if hull.get_volume() is None:
        
        print("Mesh is not watertight, using boudning box volume...\n")
        
        pt_volume = obb.get_volume()
        
    else:
        #print("Mesh is watertight, using convex_hull volume...\n")
        # compute as convexhull volume
        pt_volume = hull.get_volume()
    
    
    
    # oriented bounding box volume
    pt_ob_volume = pcd.get_oriented_bounding_box().volume()

    return pt_diameter_max, pt_diameter_min, pt_diameter_avg, pt_length, pt_volume, pt_ob_volume
    
    
    
    

#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)




# save point cloud data from numpy array as ply file, compatible format with open3d library
def write_ply(path, data_numpy_array):
    
    #data_range = 100
    
    #Normalize data range for generate cross section level set scan
    #min_max_scaler = preprocessing.MinMaxScaler(feature_range = (0, data_range))

    #point_normalized = min_max_scaler.fit_transform(data_numpy_array)
    
    #initialize pcd object for open3d 
    pcd = o3d.geometry.PointCloud()
     
    pcd.points = o3d.utility.Vector3dVector(data_numpy_array)
    
    #pcd.colors = o3d.utility.Vector3dVector(Z_pt_color)
    
    # get the model center position
    #model_center = pcd.get_center()
    
    # geometry points are translated directly to the model_center position
    #pcd.translate(-1*(model_center))
    
    #write out point cloud file
    o3d.io.write_point_cloud(path, pcd, write_ascii = True)
    
    
    # check saved file
    if os.path.exists(path):
        print("Converted 3d model was saved at {0}\n".format(path))
        return True
    else:
        return False
        print("Model file converter failed !")
        #sys.exit(0)


# compute diameter from area
def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    
    #note: return diameter instead of radius
    return 2*radius 



#compute angle
def angle(directions):
    """Return the angle between vectors"""
    vec2 = directions[1:]
    vec1 = directions[:-1]

    norm1 = np.sqrt((vec1 ** 2).sum(axis=1))
    norm2 = np.sqrt((vec2 ** 2).sum(axis=1))
    cos = (vec1 * vec2).sum(axis=1) / (norm1 * norm2)   
    return np.arccos(cos)


#first derivative function
def first_derivative(x) :

    return x[2:] - x[0:-2]


#second derivative function
def second_derivative(x) :
    
    return x[2:] - 2 * x[1:-1] + x[:-2]


#compute curvature
def curvature(x, y) :

    x_1 = first_derivative(x)
    x_2 = second_derivative(x)
    y_1 = first_derivative(y)
    y_2 = second_derivative(y)
    return np.abs(x_1 * y_2 - y_1 * x_2) / np.sqrt((x_1**2 + y_1**2)**3)


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        return (90 - angle)
        
    '''
    if angle > 0 and angle < 45:
        return (90 - angle)
    elif angle < 90:
        return angle
    else:
        return (180- angle)
    '''


# smooth 1d list
def smooth_data_convolve_average(arr, span):
    re = np.convolve(arr, np.ones(span * 2 + 1) / (span * 2 + 1), mode="same")

    # The "my_average" part: shrinks the averaging window on the side that 
    # reaches beyond the data, keeps the other side the same size as given 
    # by "span"
    re[0] = np.average(arr[:span])
    for i in range(1, span + 1):
        re[i] = np.average(arr[:i + span])
        re[-i] = np.average(arr[-i - span:])
    return re    




#cluster 1D list using Kmeans
def cluster_list(list_array, n_clusters):
    
    data = np.array(list_array)
    
    if data.ndim == 1:
        
        data = data.reshape(-1,1)
    
    #kmeans = KMeans(n_clusters).fit(data.reshape(-1,1))
        
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(data)
    
    #kmeans = KMeans(n_clusters).fit(data)
    
    labels = kmeans.labels_
    
    centers = kmeans.cluster_centers_
    
    center_labels = kmeans.predict(centers)
    
    #print(kmeans.cluster_centers_)

    ######################################################
    
    
    return labels, centers, center_labels



# Removing black points from a point cloud by filtering color values of the points based on the intensity
def remove_black_points(pcd, black_threshold):
    
    # Access the point cloud colors
    colors = np.asarray(pcd.colors)

    # Define a threshold for black points
    #black_threshold = 0.2  # Adjust as needed

    # Create a mask for black points
    black_mask = np.all(colors <= black_threshold, axis=1)

    # Remove black points
    pcd = pcd.select_by_index(np.where(black_mask == False)[0])

    return pcd, black_mask




# remove the noise in 3d point cloud 
def model_clean(pcd):
    
    # Apply noise filter to the input open3d pcd object
    pcd_filtered = statistical_outlier_removal(pcd, nb_neighbors, std_ratio)
    
    # get color values of the point cloud 
    #color_array = np.asarray(pcd_filter.colors)
    

    # black points removal
    if black_filter == 0:

        pcd_sel = pcd_filtered
        print("No color based noise filter was applied!\n")
        
    else:
        # Remove black points
        (pcd_sel, black_mask) = remove_black_points(pcd_filtered, black_threshold)
        
    
    #o3d.visualization.draw_geometries([pcd])
    #o3d.visualization.draw_geometries([pcd_sel])
    
    #print("Showing outliers (red) and inliers (gray): ")
    #pcd_sel.paint_uniform_color([1, 0, 0])
    #o3d.visualization.draw_geometries([pcd, pcd_sel])


    # copy original point cloud for rotation
    pcd_cleaned = copy.deepcopy(pcd_sel)

    # get the model center postion
    model_center = pcd_cleaned.get_center()
    
    # geometry points are translated directly to the model_center position
    pcd_cleaned.translate(-1*(model_center))


    return pcd_cleaned


# slice point cloud model
def slice_pt(Data_array_pcloud, ratio_s, ratio_e):

    
    # load skeleton coordinates and radius 
    Z_pt_sorted = np.sort(Data_array_pcloud[:,2])
    
    # index of end plane 
    idx_sel_e = int(len(Z_pt_sorted)*ratio_e) 

    Z_e = Z_pt_sorted[idx_sel_e]  if idx_sel_e < len(Data_array_pcloud) else (len(Data_array_pcloud) - 1)
    
    # index of start plane
    idx_sel_s = int(len(Z_pt_sorted)*ratio_s) 

    Z_s = Z_pt_sorted[idx_sel_s]

    # mask between the start and end plane
    Z_mask = (Data_array_pcloud[:,2] <= Z_e) & (Data_array_pcloud[:,2] >= Z_s) 
    
    Z_pt_sel = Data_array_pcloud[Z_mask]
    
    
    return Z_pt_sel, Z_mask



# detect peaks in a curved line
def peak_detection(X):
    
    # peak detection
    
    # Input Data
    #X = List_N_seg_smooth
    
    # Initialize
    fp = findpeaks(method='peakdetect', lookahead = 1)

    # return dictionary object
    results = fp.fit(X)
    
    # Plot and save figure
    fp.plot("Depth v.s. diameter")
    
    result_file = result_path + 'N_seg.png'
    
    plt.savefig(result_file)
    
    plt.close()
    
    #####################################################################

    # parse the dictionary object and get the pd.DataFrame
    df = results.get("df")
 
    # print df names
    # The output contains multiple variables
    print(list(df.columns))


    # Use boolean indexing to extract peak locations to a new pd frame
    df_valley= df.loc[df['valley'] == True]
    
    label_x = df['labx'].tolist()

    #print(label_x)
    
    unique_label_x = list(set(label_x))

    print("unique_label_x = {}\n".format(len(unique_label_x)))
    
    # convert x, y values into list
    #valley_y = df_valley['y'].tolist()
    #valley_x = df_valley['x'].tolist()
    
    
    #print("valley_y = {}\n".format(len(valley_y)))
    
    
    return unique_label_x


#calculate length of a 3D path or curve
def path_length(X, Y, Z):

    n = len(X)
     
    lv = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1,n)]
    
    L = sum(lv)
    
    return L


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        #return angle
        
        
        if angle > 0 and angle < 45:
            return (90 - angle)
        elif angle < 90:
            return angle
        else:
            return (180- angle)
        

#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    
    hxy = np.hypot(x, y)
    
    r = np.hypot(hxy, z)
    
    elevation = np.arctan2(z, hxy)*180/math.pi
    
    azimuth = np.arctan2(y, x)*180/math.pi
    
    return r[2], azimuth[2], elevation[2]
    '''
    if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
    '''


# compute the dimension of a skeleton part
def dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, select):
    
    loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(loc_start), min(loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(loc_end), min(loc_end)))
    
    max_dimension_length = abs(max(max(loc_start), max(loc_end) - min(min(loc_start), min(loc_end))))

    min_dimension_length = abs(min(min(loc_start), min(loc_end) - max(max(loc_start), max(loc_end))))
    
    #print("max_length = {}\n".format(max_length))
    
    if select == 1:
        return max_dimension_length
    else:
        return min_dimension_length


# compute the WCSS function, within cluster sum of squares
#The objective function in k-means is the WCSS (within cluster sum of squares). 
def calculate_wcss(data):
    wcss = []
    for n in range(2, 10):
        kmeans = KMeans(n_clusters=n)
        kmeans.fit(X=data)
        wcss.append(kmeans.inertia_)

    return wcss


# compute the optimal number for clustering
def optimal_number_of_clusters(wcss):
    x1, y1 = 2, wcss[0]
    x2, y2 = 20, wcss[len(wcss)-1]

    distances = []
    for i in range(len(wcss)):
        x0 = i+2
        y0 = wcss[i]
        numerator = abs((y2-y1)*x0 - (x2-x1)*y0 + x2*y1 - y2*x1)
        denominator = math.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        distances.append(numerator/denominator)

    return distances.index(max(distances)) + 0


# compute parameters for current level
def level_branch(sub_branch_level, length_level, angle_level, radius_level, projection_level, N_level):
    
    N_curr = []
    avg_length_curr = []
    avg_angle_curr = []
    avg_diameter_curr = []
    avg_projection_curr = []
            
    for i in N_level:
        
        N_curr.append(int(len(sub_branch_level[i])))
        avg_length_curr.append(np.mean(length_level[i]))
        avg_angle_curr.append(np.mean(angle_level[i]))
        avg_diameter_curr.append(np.mean(radius_level[i]))
        avg_projection_curr.append(np.mean(projection_level[i]))
            

    N_curr_avg = sum(N_curr)
    avg_length_level = statistics.mean(avg_length_curr)
    avg_angle_level = statistics.mean(avg_angle_curr)
    avg_diameter_level = statistics.mean(avg_diameter_curr)
    avg_projection_level = statistics.mean(avg_projection_curr)
    
    
    return N_curr_avg, avg_length_level, avg_angle_level, avg_diameter_level, avg_projection_level



# visualization of point cloud and skeleton
def visualize_pt(Data_array_skeleton, array_edges_skeleton, pcd_pt):
    
    # visualization 
    ####################################################################

    pcd_skeleton = o3d.geometry.PointCloud()
    
    pcd_skeleton.points = o3d.utility.Vector3dVector(Data_array_skeleton)
    
    #pcd_skeleton.points = o3d.utility.Vector3dVector(sel_points)
    
    pcd_skeleton.paint_uniform_color([0, 0, 1])
    
    #o3d.visualization.draw_geometries([pcd])
    
    points = Data_array_skeleton
    
    lines = array_edges_skeleton
    
    colors = [[1, 0, 0] for i in range(len(lines))]
    
    line_set = o3d.geometry.LineSet(
        points=o3d.utility.Vector3dVector(points),
        lines=o3d.utility.Vector2iVector(lines),
    )
    
    line_set.colors = o3d.utility.Vector3dVector(colors)
    
    
    #o3d.visualization.draw_geometries([line_set])
    
    
    vis = o3d.visualization.Visualizer()
    vis.create_window()
    vis.add_geometry(line_set)
    vis.add_geometry(pcd_skeleton)
    vis.add_geometry(pcd_pt)
    vis.get_render_option().line_width = 1
    vis.get_render_option().point_size = 1
    vis.get_render_option().background_color = (0, 0, 0)
    vis.get_render_option().show_coordinate_frame = True
    
    vis.run()




# Traits analysis for the input 3D model
def analyze_pt(pt_file):
    
    
    ###################################################################
    #load point cloud input in *.ply format
    print("Loading 3D point cloud {}...\n".format(pt_file))

    pcd_input = o3d.io.read_point_cloud(pt_file)
    
    
    ###################################################################
    # model clean
    pcd = model_clean(pcd_input)
    
    #write out point cloud file
    #pcd_cleaned = (result_path + 'pcd_cleaned.ply')
    #o3d.io.write_point_cloud(pcd_cleaned, pcd, write_ascii = True)
    
    ###################################################################
    # load point cloud as np array 
    # get 3d points
    Data_array_pcloud = np.asarray(pcd.points)
    
    # check points data structure 
    print(Data_array_pcloud.shape)
    

    # get 3d points coordinates 
    X = Data_array_pcloud[:,0] 
    Y = Data_array_pcloud[:,1] 
    Z = Data_array_pcloud[:,2] 
    
    # assign colors to points
    if pcd.has_colors():
        
        print("Render colored point cloud\n")
        
        pcd_color = np.asarray(pcd.colors)
        
        # change format
        if len(pcd_color) > 0: 
            
            pcd_color = np.rint(pcd_color * 255.0)
        
        #pcd_color = tuple(map(tuple, pcd_color))
    else:
        
        print("Generate random color\n")
    
        pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
    
    
    #########################################################################################################
    #initialize parameters
    pt_diameter_max = pt_diameter_min = pt_length = pt_diameter = pt_eccentricity = pt_density = pt_angle = pt_angle_max = pt_angle_min = sum_volume = pt_stem_diameter = 0
    
    
    #initialize parameters
    #pt_diameter_max = pt_diameter_min = pt_length = pt_diameter = pt_eccentricity = pt_stem_diameter = pt_density = 0
    

    #compute dimensions of point cloud data
    (pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume, pt_ob_volume) = get_pt_parameter(pcd)

    print("pt_diameter_max = {} pt_diameter_min = {} pt_diameter = {} pt_length = {} pt_volume = {}\n".format(pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume))
    

    ########################################################################################################
    # slicing models using n_slices
    print("Using {} planes to scan the model along Z axis...".format(n_slices))
    
    (pt_plane, pt_plane_center, pt_plane_diameter_max, pt_plane_diameter_min, pt_plane_diameter_avg, filter_plane_center, filter_plane_volume, filter_plane_eccentricity, filter_plane_density) = get_pt_sel_parameter(Data_array_pcloud, n_slices)
    
    #o3d.visualization.draw_geometries(pt_plane)
    
    pt_center_arr = np.vstack(pt_plane_center)
    
    print("pt_plane_diameter_avg = {}\n".format(pt_plane_diameter_avg))
    
    # argmax returns the max element's index
    idx_max = np.argmax(pt_plane_diameter_avg)
    
    print("pt_plane_diameter_max = {}, len(pt_plane_diameter_avg) = {}, index_max = {}\n".format(max(pt_plane_diameter_avg), len(pt_plane_diameter_avg), idx_max))
    
    # slicing the list only keep the ascending order 
    pt_plane_diameter_slice = pt_plane_diameter_avg[0:idx_max+1]
    
    
    
    ###################################################################
    # smooth the list with ascending order 
    span = 5
  
    List_N_seg_smooth = smooth_data_convolve_average(np.array(pt_plane_diameter_slice), span)
    
    print("List_N_seg_smooth = {}\n".format(len(List_N_seg_smooth)))
    
    
    #List_N_seg_smooth = List_N_seg

    ####################################################################
    # peak detection 
    
    unique_label_x = peak_detection(List_N_seg_smooth)

    # set number of clusters 
    if len(unique_label_x) > 4:
        N_cluster = 4
    elif len(unique_label_x) > 2:
        N_cluster = 3
    else:
        N_cluster = 2

    print("N_cluster = {}\n".format(N_cluster))
    
    
    ##################################################################################
    # cluster the list to find different parts
    
    #(labels, centers, center_labels) = cluster_list(valley_y, n_clusters = N_cluster)
    
    (labels, centers, center_labels) = cluster_list(List_N_seg_smooth, n_clusters = N_cluster)
    
    sorted_idx = np.argsort(centers[:,0])
    
    centers = centers[sorted_idx]
    
    print("centers = {}\n".format(centers))
    
    print("sorted_idx = {}\n".format(sorted_idx))
    
    print("labels = {}\n".format(labels))
    
    print("labels_len = {}\n".format(len(labels)))
    
    #print("len(centers[0]) = {}, len(centers[1]) = {}, len(centers[2]) = {}\n".format(len(centers[0])))
    
    
    # Get indices
    indices_first = list(filter(lambda i: labels[i] == sorted_idx[0], range(len(labels))))
    
    #indices_first = list(filter(lambda i: labels[i] == sorted_idx[1], range(len(labels))))

    print("indices_first_len = {}\n".format(len(indices_first)))
    

    #######################################################################################
    # slicing start and end ratios
    ratio_s = slicing_factor*len(indices_first)/len(pt_plane_diameter_avg)

    ratio_e = 1
    
    print("ratio_s = {}; ratio_e = {}\n".format(ratio_s, ratio_e))
    
    (Z_pt_sel, Z_pt_mask) = slice_pt(Data_array_pcloud, ratio_s, ratio_e)
    
    # save pcd file
    pcd_maskfile = (result_path + 'pcd.xyz')

    write_ply(pcd_maskfile, Z_pt_sel)
    
    
    ########################################################################################################
    # compute stem diameter

    slice_index = round(ratio_s*len(pt_plane_diameter_avg))

    print("slice_index = {}\n".format(slice_index))
    
    # slicing the list only keep the ascending order 
    pt_plane_diameter_stem = pt_plane_diameter_avg[0:slice_index+1]
    

    # compute stem diameter
    pt_stem_diameter = statistics.mean(pt_plane_diameter_stem)
    
    print("pt_stem_diameter = {}\n".format(pt_stem_diameter))
    
    '''
    # using cylinder fitting method
    # slice stem
    ratio_s = 0
    ratio_e = factor_part*len(indices_first)/len(pt_plane_diameter_avg)
    
    (pt_stem,pt_stem_mask) = slice_pt(Data_array_pcloud, ratio_s, ratio_e)
    
    # save pcd file
    pcd_maskfile = (result_path + 'pcd_stem.xyz')

    write_ply(pcd_maskfile, pt_stem)
    '''
    
    
    #######################################################################################
    # compute ratios for each part
    
    part_ratio = []
    
    sum_part_ratio = []
    
    # get each ratio from depth and diameter list
    for idx in range(len(sorted_idx)):
        
        indices_current = list(filter(lambda i: labels[i] == sorted_idx[idx], range(len(labels))))

        ratio_current = len(indices_current)/len(pt_plane_diameter_avg)
        
        part_ratio.append(ratio_current)
        
        sum_part_ratio.append(sum(part_ratio))
    
    # construct ratio list in range (0,1)
    sum_part_ratio.insert(0,0)
    
    sum_part_ratio.append(1)
    
    print(part_ratio)
    
    print(sum_part_ratio)
    
    '''
    # save slice parts
    for idx, (ratio_value) in enumerate(sum_part_ratio):
        
        if idx < len(sum_part_ratio) - 1:
            
            ratio_s_curr = ratio_value
            ratio_e_curr = sum_part_ratio[idx + 1]
            
            #print("ratio_s = {}; ratio_e = {}\n".format(ratio_s, ratio_e))
        
            (Z_pt_slice, slice_mask) = slice_pt(Data_array_pcloud, ratio_s_curr, ratio_e_curr)

            # save pcd file
            
            pcd_slice = (result_path + 'pcd_' + str(idx) + '.xyz')

            write_ply(pcd_slice, Z_pt_slice)
    '''
    
    
    # compute the R_1 and R_2 ratios
    if len(part_ratio) > 2:
        R_1 = pt_length * part_ratio[1] * slicing_factor
        R_2 = pt_length * part_ratio[2] * slicing_factor
        N_w = 3
    else:
        R_1 = pt_length * part_ratio[1] * slicing_factor
        R_2 = 0
        N_w = 2
    
    print("pt_length = {}, R_1 = {}, R_2 = {}\n".format(pt_length, R_1, R_2))
    
    


    ########################################################################################
    # compute parameters of whole point cloud level
    filter_plane_angle = []
    
    
    # define unit vector
    v_x = [1,0,0]
    v_y = [0,1,0]
    v_z = [0,0,1]


    # compute side angles for each sliced model
    for idx, f_center in enumerate(filter_plane_center):
        
        if idx > 0:
            
            #print(idx, f_center)
            
            center_vector = [f_center[0] - filter_plane_center[idx-1][0], f_center[1] - filter_plane_center[idx-1][1], f_center[2] - filter_plane_center[idx-1][2]]
        
            norm_center_vector = center_vector / np.linalg.norm(center_vector)
        
            cur_angle = dot_product_angle(norm_center_vector, v_z)
            
            #print("cur_angle = {} ...".format(cur_angle))
    
            filter_plane_angle.append(cur_angle)
    
    
    pt_angle = np.mean(filter_plane_angle)
    
    pt_angle_max = max(filter_plane_angle)
    
    pt_angle_min = min(filter_plane_angle)
    
    print("pt_angle = {}, pt_angle_max = {}, pt_angle_min = {}\n".format(pt_angle, pt_angle_max, pt_angle_min))
    

    # Sum of all volume for each sliced model 
    sum_volume = sum(filter_plane_volume)

    # average of eccentricity
    avg_eccentricity = np.mean(filter_plane_eccentricity)
    
    # average of bushiness
    avg_density = np.mean(filter_plane_density)
    

    ######################################################################################################
    # compute skeleton
    print("Compute structure and skeleton from point cloud model ...\n")
    
    # docker test
    #skeleton_compute = "/opt/code/compiled/Release/bin/AdTree " + result_path + basename + ".xyz " + result_path + " -s"
    
    pcd_file = Path(pcd_maskfile)
    
    # check input file exists or not
    if pcd_file.is_file():

        # local test using local path in repo
        skeleton_compute = "/home/suxing/3D_model_traits_demo/compiled/Release/bin/AdTree " + pcd_maskfile + " " + result_path + " -s"
        
        # docker version 
        #skeleton_compute = "/opt/code/compiled/Release/bin/AdTree " + pcd_maskfile + " " + result_path + " -s"
        
        #print(skeleton_compute)
        
        execute_script(skeleton_compute)
        
        print("Skeleton analysis finsihed...\n")
    
    else:
        
        print("The pcd_file is missing or not readable!\n")

        print("Exiting the program...")

        sys.exit(0)
    
    
    
    ##################################################################################################
    print("Compute traits and analyze skeleton...\n")
    
    #Skeleton analysis
    #def analyze_skeleton(current_path, filename_skeleton, filename_ptcloud, imgList):

    model_skeleton = (result_path + 'pcd_skeleton.ply')
    
    
    # check skeleton file exists or not
    if Path(model_skeleton).is_file():

        print("Skeleton file was found...\n")

    
    else:
        
        print("Skeleton file was not found!\n")

        print("Exiting the program...")

        sys.exit(0)
    
    
    
    # load skeleton file
    print("Loading 3D skeleton file {}...\n".format(model_skeleton))
    
    #model_skeleton_basename = os.path.splitext(model_skeleton)[0]
    
    #model_basename = Path(ske_file).stem
    
    
    #print(model_basename)

    
    #load the ply format skeleton file 
    try:
        with open(model_skeleton, 'rb') as f:
            plydata_skeleton = PlyData.read(f)
            num_vertex_skeleton = plydata_skeleton.elements[0].count
            N_edges_skeleton = len(plydata_skeleton['edge'].data['vertex_indices'])
            array_edges_skeleton = plydata_skeleton['edge'].data['vertex_indices']
            
            #print("Ply data structure: \n")
            #print(plydata_skeleton)
            #print("\n")
            
            print("Number of 3D points in skeleton model: {0} \n".format(num_vertex_skeleton))
            print("Number of edges: {0} \n".format(N_edges_skeleton))

        
    except:
        sys.exit("Model skeleton file does not exist!")
    
    
    #Parse ply format skeleton file and extract the data
    Data_array_skeleton = np.zeros((num_vertex_skeleton, 3))
    
    # load skeleton coordinates
    Data_array_skeleton[:,0] = plydata_skeleton['vertex'].data['x']
    Data_array_skeleton[:,1] = plydata_skeleton['vertex'].data['y']
    Data_array_skeleton[:,2] = plydata_skeleton['vertex'].data['z']
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    
    # get radius of each edges
    radius_vtx = plydata_skeleton['vertex'].data['radius']
    
    
    #print(Data_array_skeleton.shape)
    #print(radius_vtx.shape)
    #print(array_edges_skeleton)
    
    ####################################################################
    # Visualize the skeleton and point cloud together
    
    if args["visualize"] == 1:
        
        #initialize pcd object for open3d 
        pcd_pt = o3d.geometry.PointCloud()
         
        pcd_pt.points = o3d.utility.Vector3dVector(Z_pt_sel)
        
        pcd_pt.colors = o3d.utility.Vector3dVector(np.asarray(pcd.colors)[Z_pt_mask])

        visualize_pt(Data_array_skeleton, array_edges_skeleton, pcd_pt)
    
    
   
    
    ####################################################################
    # build directed graph from skeleton/structure data
    
    print("Building directed graph from 3D skeleton/structure ...\n")
    
    G_unordered = gt.Graph(directed = True)
    
    # assert directed graph
    #print(G.is_directed())
    
    nodes = G_unordered.add_vertex(num_vertex_skeleton)
    
    G_unordered.add_edge_list(array_edges_skeleton.tolist()) 
    
    #gt.graph_draw(G_unordered, vertex_text = G_unordered.vertex_index, output = current_path + "graph_view.pdf")
    
    
    # find all end vertices by fast iteration of all vertices
    end_vlist = []
    
    end_vlist_offset = []
    
    for v in G_unordered.iter_vertices():
        
        #print(G.vertex(v).out_degree(), G.vertex(v).in_degree())
        
        if G_unordered.vertex(v).out_degree() == 0 and G_unordered.vertex(v).in_degree() == 1:
        
            end_vlist.append(v)
            
            if (v+1) == num_vertex_skeleton:
                end_vlist_offset.append(v)
            else:
                end_vlist_offset.append(v+1)
            
    #print("end_vlist = {} \n".format(end_vlist))
    #print("end_vlist_offset = {} \n".format(end_vlist_offset))
    
    
    # save graph
    ##########################################
    #graph_file = (result_path + 'graph.gt.gz')
    #G_unordered.save(graph_file)

    #test angle calculation
    #vector1 = [0,0,1]
    # [1,0,0]
    #vector2 = [0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1]
    #print(dot_product_angle(vector1,vector2))
    #print(cart2sph(0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1)) 
    
    
    
    #####################################################################################
    # parse all the edges and vetices, start, end vetices for sub branches 
    
    
    sub_branch_list = []
    sub_branch_length_rec = []
    sub_branch_angle_rec = []
    sub_branch_start_rec = []
    sub_branch_end_rec = []
    sub_branch_projection_rec = []
    sub_branch_radius_rec = []
    
    sub_branch_xs_rec = []
    sub_branch_ys_rec = []
    sub_branch_zs_rec = []
    
    #factor = 0.77
    
    #if len(end_vlist) == len(end_vlist_offset):
        
    for idx, v_end in enumerate(end_vlist):
        
        #print(idx, v_end)
        #construct list of vertices in sub branches
        if idx == 0:
            v_list = [*range(0, int(end_vlist[idx])+1)]
        else:
            v_list = [*range(int(end_vlist[idx-1])+1, int(end_vlist[idx])+1)]
            
        # change type to interger 
        int_v_list = [int(i) for i in v_list]
        
        # current sub branch length
        sub_branch_length = path_length(X_skeleton[int_v_list], Y_skeleton[int_v_list], Z_skeleton[int_v_list])
        
        # current sub branch start and end points 
        start_v = [X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]]
        
        end_v = [X_skeleton[int_v_list[0]] - X_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Y_skeleton[int_v_list[0]] - Y_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Z_skeleton[int_v_list[0]] - Z_skeleton[int_v_list[int(len(int_v_list)-1.0)]]]
        
        # angle of current branch vs Z direction
        angle_sub_branch = dot_product_angle(start_v, end_v)
        
        # projection radius of current branch length
        p0 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]])
        
        p1 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[-1]]])
        
        projection_radius = np.linalg.norm(p0 - p1)
        
        #radius value from fitted point cloud contours
        radius_edge = float(radius_vtx[int_v_list[0]])*1
        #radius_edge = 0
        
        
        sub_branch_xs = X_skeleton[int_v_list[0]]
        sub_branch_ys = Y_skeleton[int_v_list[0]]
        sub_branch_zs = Z_skeleton[int_v_list[0]]
        
        
        # save computed parameters for each branch
        sub_branch_list.append(v_list)
        sub_branch_length_rec.append(sub_branch_length)
        sub_branch_angle_rec.append(angle_sub_branch)
        sub_branch_start_rec.append(int_v_list[0])
        sub_branch_end_rec.append(int_v_list[-1])
        sub_branch_projection_rec.append(projection_radius)
        sub_branch_radius_rec.append(radius_edge)
        
        sub_branch_xs_rec.append(sub_branch_xs)
        sub_branch_ys_rec.append(sub_branch_ys)
        sub_branch_zs_rec.append(sub_branch_zs)
    
    #print(min(sub_branch_angle_rec))
    #print(max(sub_branch_angle_rec))
    
    

    
    ####################################################################
    # sort branches according to length feature in descending order
    
    sorted_idx_len = np.argsort(sub_branch_length_rec)
    
    #reverse the order from accending to descending
    sorted_idx_len_loc = sorted_idx_len[::-1]

    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_len_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_len_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_len_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_len_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_len_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_len_loc]
    sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_len_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_len_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_len_loc]

    ####################################################################
    # compute the dimensions of skeletons
    
    max_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    
    min_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    
    print("max_length = {} {} {}\n".format(max_length_x, max_length_y, max_length_z))
    
    print("min_length = {} {} {}\n".format(min_length_x, min_length_y, min_length_z))
    
    
    

    
    
    ####################################################################
    # construct sub branches with length and radius feature 
    
    combined_list = np.array(list(zip(sub_branch_length_rec, sub_branch_radius_rec))).reshape(len(sub_branch_length_rec), 2)
    
    # calculating the within clusters sum-of-squares 
    sum_of_squares = calculate_wcss(combined_list)
    
    # calculating the optimal number of clusters
    n_optimal = optimal_number_of_clusters(sum_of_squares)
    
    print("optimal_number_of_clusters = {}\n".format(n_optimal))
    
    n_optimal = 4
    
    
    # find sub branches cluster with length and radius feature 
    ####################################################################
    cluster_number = n_optimal + 4
    
    (labels, centers, center_labels) = cluster_list(combined_list, n_clusters = cluster_number)
    
    sorted_idx = np.argsort(centers[:,0])[::-1]

    print("sorted_idx = {}\n".format(sorted_idx))
    
    
    indices_level = []
    sub_branch_level = []
    sub_branch_start_level = []
    sub_branch_startZ_level = []
    sub_branch_startY_level = []
    sub_branch_startX_level = []
    
    radius_level = []
    length_level = []
    angle_level = []
    projection_level = []
    
    for idx, (idx_value) in enumerate(sorted_idx):
        
        #print(labels_length_rec.tolist().index(idx_value))
        
        #print("cluster {}, center value {}".format(idx, idx_value))
        indices = [i for i, x in enumerate(labels.tolist()) if x == idx_value]
        
        #print(indices)
        
        sub_branch_start_rec_selected = [sub_branch_start_rec[i] for i in indices]
        Z_loc = [Z_skeleton[index] for index in sub_branch_start_rec_selected]
        Y_loc = [Y_skeleton[index] for index in sub_branch_start_rec_selected]
        X_loc = [X_skeleton[index] for index in sub_branch_start_rec_selected]
        
        
        sub_loc = [sub_branch_list[index] for index in indices]
        radius_loc = [sub_branch_radius_rec[index] for index in indices]
        length_loc = [sub_branch_length_rec[index] for index in indices]
        angle_loc = [sub_branch_angle_rec[index] for index in indices]
        projection_loc = [sub_branch_projection_rec[index] for index in indices]
        
        
        indices_level.append(indices)
        sub_branch_level.append(sub_loc)
        sub_branch_start_level.append(sub_branch_start_rec_selected)
        sub_branch_startZ_level.append(Z_loc)
        sub_branch_startY_level.append(Y_loc)
        sub_branch_startX_level.append(X_loc)
        
        radius_level.append(radius_loc)
        length_level.append(length_loc)
        angle_level.append(angle_loc)
        projection_level.append(projection_loc)

    
    ###################################################################
    # output sub branches levels
    for idx in range(cluster_number):
        
        print("sub_branch_level[{}] = {}\n".format(idx, len(sub_branch_level[idx])))
    

    
    #compute paramters
    if len(sub_branch_level) > 4:
        
        n_level = 3
        
        N_level = list(range(0, n_level))
        
        # compute average parameters
        (N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1) = level_branch(sub_branch_level, length_level, angle_level, radius_level, projection_level, N_level)

        print(N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1)
    
        N_level = list(range(3, n_level+2))
        
        (N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2) = level_branch(sub_branch_level, length_level, angle_level, radius_level, projection_level, N_level)
        
        print(N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2)
        
    
    avg_diameter_N3 = np.mean(radius_level[3])

    ####################################################################

    
    # sum_volume or pt_volume
    return pt_diameter_max, pt_diameter_min, pt_diameter, pt_stem_diameter, pt_length, avg_eccentricity, avg_density, sum_volume, \
            N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1, \
            N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2, \
            avg_diameter_N3, N_w, R_1, R_2
            



# get file information from the file path using python3
def get_file_info(file_full_path):
    
    p = pathlib.Path(file_full_path)
    filename = p.name
    basename = p.stem

    file_path = p.parent.absolute()
    file_path = os.path.join(file_path, '')

    return file_path, filename, basename



# save results as excel file
def write_output(trait_file, trait_sum):
    
    if os.path.isfile(trait_file):
        # update values

        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'RC diameter max'
        sheet.cell(row = 1, column = 2).value = 'RC diameter min'
        sheet.cell(row = 1, column = 3).value = 'RC diameter'
        sheet.cell(row = 1, column = 4).value = 'stem diamerter'
        sheet.cell(row = 1, column = 5).value = 'RC length'
        sheet.cell(row = 1, column = 6).value = 'RC eccentricity'
        sheet.cell(row = 1, column = 7).value = 'RC density'
        sheet.cell(row = 1, column = 8).value = 'RC volume'
        
        sheet.cell(row = 1, column = 9).value = 'number of brace roots'
        sheet.cell(row = 1, column = 10).value = 'brace root length'
        sheet.cell(row = 1, column = 11).value = 'brace root angle'
        sheet.cell(row = 1, column = 12).value = 'brace root diameter'
        sheet.cell(row = 1, column = 13).value = 'brace root projection radius'
        
        sheet.cell(row = 1, column = 14).value = 'number of crown roots'
        sheet.cell(row = 1, column = 15).value = 'crown root length'
        sheet.cell(row = 1, column = 16).value = 'crown root angle'
        sheet.cell(row = 1, column = 17).value = 'crown root diameter'
        sheet.cell(row = 1, column = 18).value = 'crown root projection radius'
        
        sheet.cell(row = 1, column = 19).value = 'lateral root radius'
        sheet.cell(row = 1, column = 20).value = 'number of whorls'
        sheet.cell(row = 1, column = 21).value = 'whorl distance 1'
        sheet.cell(row = 1, column = 22).value = 'whorl distance 2'
        

    for row in trait_sum:
        sheet.append(row)
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
        
    else:
        print("Error in saving Result file\n")




if __name__ == '__main__':
    

    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-i", "--input", dest="input", required=True, type=str, help="full path to 3D model file")
    ap.add_argument("-o", "--output_path", dest = "output_path", type = str, required = False, help = "result path")
    ap.add_argument("--nb_neighbors", required = False, type = int, default = 20, help = "nb_neighbors")
    ap.add_argument("--std_ratio", required = False, type = float, default = 5.0, help = "outlier remove ratio, small number = aggresive")
    ap.add_argument("--black_filter", required = False, type = int, default = 0, help = "Apply black points removal filter or not, 0 = not, 1 = Apply")
    ap.add_argument("--black_threshold", required = False, type = float, default = 0.2, help = "threshold for black points removal")
    ap.add_argument("--n_slices", dest = "n_slices", type = int, required = False, default = 10,  help = "Number of planes to segment the 3d model along Z direction")
    ap.add_argument("--slicing_factor", dest = "slicing_factor", type = float, required = False, default = 0.7,  help = "Slicing adjust factor")
    ap.add_argument("--visualize", dest = "visualize", required = False, type = int, default = 0, help = "Display model or not, default not display")
    args = vars(ap.parse_args())


    
    if os.path.isfile(args["input"]):

        input_file = args["input"]

        (file_path, filename, basename) = get_file_info(input_file)

        print("Processing 3d model point cloud file '{} {} {}'...\n".format(file_path, filename, basename))

        # result path
        result_path = args["output_path"] if args["output_path"] is not None else file_path

        result_path = os.path.join(result_path, '')

        # print out result path
        print("results_folder: {}\n".format(result_path))
        
        
        nb_neighbors = args["nb_neighbors"]

        std_ratio = args["std_ratio"]
        
        black_filter = args["black_filter"]
    
        black_threshold = args["black_threshold"]

        slicing_factor = args["slicing_factor"] 
        
        
        # number of slices for cross section
        n_slices = args['n_slices']

        visualize = args["visualize"]

        # start pipeline
        ########################################################################################3
        # compute parameters
        (pt_diameter_max, pt_diameter_min, pt_diameter, pt_stem_diameter, pt_length, avg_eccentricity, avg_density, sum_volume, \
            N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1, \
            N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2, \
            avg_diameter_N3, N_w, R_1, R_2) = analyze_pt(input_file)
        
        # save result as an excel file
        trait_sum = []

        trait_sum.append([pt_diameter_max, pt_diameter_min, pt_diameter, pt_stem_diameter, pt_length, avg_eccentricity, avg_density, sum_volume, \
            N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1, \
            N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2, \
            avg_diameter_N3, N_w, R_1, R_2])

        trait_file = (result_path + basename + '_trait.xlsx')

        write_output(trait_file, trait_sum)


    else:

        print("The input file is missing or not readable!\n")

        print("Exiting the program...")

        sys.exit(0)


    
    '''
    # loop multiple files for batch processing 
    # python3 model_measurement.py -p ~/example/ -ft ply -o ~/example/ -n 5 -v 0
    #######################################################################################
    
    # path to model file 
    file_path = args["path"]
    
    ext = args['filetype']
    
    files = file_path + '*.' + ext
    
    n_slices = args['n_slices']

    visualize = args["visualize"]
    
    
    # obtain image file list
    fileList = sorted(glob.glob(files))


    

    for input_file in fileList:
        
        if os.path.isfile(input_file):
            
            (file_path, filename, basename) = get_file_info(input_file)

            print("Processing 3d model point cloud file '{} {} {}'...\n".format(file_path, filename, basename))

            # result path
            result_path = args["output_path"] if args["output_path"] is not None else file_path

            result_path = os.path.join(result_path, '')

            # print out result path
            nb_neighbors = args["nb_neighbors"]

            std_ratio = args["std_ratio"]
            
            black_filter = args["black_filter"]
        
            black_threshold = args["black_threshold"]

            slicing_factor = args["slicing_factor"] 

            # number of slices for cross section
            n_slices = args['n_slices']

            visualize = args["visualize"]

            # start pipeline
            ########################################################################################3
            # compute parameters
            (pt_diameter_max, pt_diameter_min, pt_diameter, pt_stem_diameter, pt_length, avg_eccentricity, avg_density, sum_volume, \
                N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1, \
                N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2, \
                avg_diameter_N3, N_w, R_1, R_2) = analyze_pt(input_file)
            
            # save result as an excel file
            trait_sum = []

            trait_sum.append([pt_diameter_max, pt_diameter_min, pt_diameter, pt_stem_diameter, pt_length, avg_eccentricity, avg_density, sum_volume, \
                N_1, avg_length_N1, avg_angle_N1, avg_diameter_N1, avg_projection_N1, \
                N_2, avg_length_N2, avg_angle_N2, avg_diameter_N2, avg_projection_N2, \
                avg_diameter_N3, N_w, R_1, R_2])

            trait_file = (result_path + basename + '_trait.xlsx')

            write_output(trait_file, trait_sum)
                


        else:
        
            print("The input file is missing or not readable!\n")
            
            print("Exiting the program...")
            
            sys.exit(0)
    '''


    
