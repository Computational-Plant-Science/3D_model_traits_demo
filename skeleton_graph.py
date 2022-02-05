"""
Version: 1.5

Summary: compute the cross section plane based on 3d model

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE

python3 skeleton_graph.py -p ~/example/ -m1 test_skeleton.ply -m2 test_aligned.ply -m3 ~/example/slices/ -v True


argument:
("-p", "--path", required=True,    help="path to *.ply model file")
("-m", "--model", required=True,    help="file name")

"""
#!/usr/bin/env python



# import the necessary packages
from plyfile import PlyData, PlyElement
import numpy as np 
from numpy import interp


from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
from operator import itemgetter
import argparse

from skimage.segmentation import watershed
from skimage.feature import peak_local_max
from skimage.morphology import convex_hull_image
from skimage.measure import regionprops

from scipy.spatial import KDTree
from scipy import ndimage
from scipy.spatial.transform import Rotation as R

import random

import cv2

import glob
import os
import sys
import open3d as o3d
import copy
import shutil




#import networkx as nx

import graph_tool.all as gt

#import plotly.graph_objects as go

from matplotlib import pyplot as plt
import math
import itertools

#from tabulate import tabulate
from rdp import rdp

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False




#calculate length of a 3D path or curve
def path_length(X, Y, Z):

    n = len(X)
     
    lv = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1,n)]
    
    L = sum(lv)
    
    return L


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        #return angle
        
        
        if angle > 0 and angle < 45:
            return (90 - angle)
        elif angle < 90:
            return angle
        else:
            return (180- angle)
        

#coordinates transformation from cartesian coords to sphere coord system
def cart2sph(x, y, z):
    
    hxy = np.hypot(x, y)
    
    r = np.hypot(hxy, z)
    
    elevation = np.arctan2(z, hxy)*180/math.pi
    
    azimuth = np.arctan2(y, x)*180/math.pi
    
    return r[2], azimuth[2], elevation[2]
    '''
    if azimuth > 90:
            angle = 180 - azimuth
        elif azimuth < 0:
            angle = 90 + azimuth
        else:
            angle = azimuth
    '''


# median-absolute-deviation (MAD) based outlier detection
def mad_based_outlier(points, thresh):
    
    if len(points.shape) == 1:
        
        points = points[:,None]
    
    median = np.median(points, axis=0)
    
    diff = np.sum((points - median)**2, axis=-1)
    
    diff = np.sqrt(diff)
    
    med_abs_deviation = np.median(diff)
    
    if med_abs_deviation == 0:
        
        modified_z_score = 0.6745 * diff / 1
    
    else:
        modified_z_score = 0.6745 * diff / med_abs_deviation
    
    return modified_z_score > thresh
    


# compute nearest neighbors of the anchor_pt_idx in point cloud by building KDTree
def get_neighbors(Data_array_pt, anchor_pt_idx, search_radius):
    
    pcd = o3d.geometry.PointCloud()
    
    pcd.points = o3d.utility.Vector3dVector(Data_array_pt)
    
    #pcd.paint_uniform_color([0.5, 0.5, 0.5])
    
    #o3d.visualization.draw_geometries([pcd])
    
    # Build KDTree from point cloud for fast retrieval of nearest neighbors
    pcd_tree = o3d.geometry.KDTreeFlann(pcd)
    
    #print("Paint the 00th point red.")
    
    #pcd.colors[anchor_pt_idx] = [1, 0, 0]
    
    #print("Find its 50 nearest neighbors, paint blue.")
    
    [k, idx, _] = pcd_tree.search_knn_vector_3d(pcd.points[anchor_pt_idx], search_radius)
    
    #print("nearest neighbors = {}\n".format(sorted(np.asarray(idx[1:]))))

    return idx
    
           
    '''
    #build octree, a tree data structure where each internal node has eight children.
    # fit to unit cube
    pcd.scale(1 / np.max(pcd.get_max_bound() - pcd.get_min_bound()), center=pcd.get_center())
    pcd.colors = o3d.utility.Vector3dVector(np.random.uniform(0, 1, size=(num_vertex_skeleton, 3)))
    o3d.visualization.draw_geometries([pcd])

    print('octree division')
    octree = o3d.geometry.Octree(max_depth=4)
    octree.convert_from_point_cloud(pcd, size_expand=0.01)
    o3d.visualization.draw_geometries([octree])

    print(octree.locate_leaf_node(pcd.points[243]))
    '''



# compute dimensions of point cloud and nearest neighbors by KDTree
def get_pt_parameter(Data_array_pt):
    
    pcd = o3d.geometry.PointCloud()
    
    pcd.points = o3d.utility.Vector3dVector(Data_array_pt)
    
    #pcd.paint_uniform_color([0.5, 0.5, 0.5])
    
   
    # get convex hull of a point cloud is the smallest convex set that contains all points.
    #hull, _ = pcd.compute_convex_hull()
    #hull_ls = o3d.geometry.LineSet.create_from_triangle_mesh(hull)
    #hull_ls.paint_uniform_color((1, 0, 0))
    
    # get AxisAlignedBoundingBox
    aabb = pcd.get_axis_aligned_bounding_box()
    #aabb.color = (0, 1, 0)
    
    #Get the extent/length of the bounding box in x, y, and z dimension.
    aabb_extent = aabb.get_extent()
    
    aabb_extent_half = aabb.get_half_extent()
    
    # get OrientedBoundingBox
    #obb = pcd.get_oriented_bounding_box()
    
    #obb.color = (1, 0, 0)
    
    #visualize the convex hull as a red LineSet
    #o3d.visualization.draw_geometries([pcd, aabb, obb, hull_ls])
    
    pt_diameter_max = max(aabb_extent[0], aabb_extent[1])*1
    
    pt_diameter_min = min(aabb_extent_half[0], aabb_extent_half[1])*1
    
    
    pt_diameter = (pt_diameter_max + pt_diameter_min)*0.5
    
    #pt_length = int(aabb_extent[2]*random.randint(40,49) )
    
    pt_length = int(aabb_extent[2])
    
    
    pt_volume = np.pi * ((pt_diameter_max + pt_diameter_min)*0.5) ** 2 * pt_length
    
    #pt_volume = hull.get_volume()
        
    return pt_diameter_max, pt_diameter_min, pt_diameter, pt_length, pt_volume
    
    
    
#find the closest points from a points sets to a fix point using Kdtree, O(log n) 
def closest_point(point_set, anchor_point):
    
    kdtree = KDTree(point_set)
    
    (d, i) = kdtree.query(anchor_point)
    
    #print("closest point:", point_set[i])
    
    return  i, point_set[i]


#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)


#cluster 1D list using Kmeans
def cluster_list(list_array, n_clusters):
    
    data = np.array(list_array)
    
    if data.ndim == 1:
        
        data = data.reshape(-1,1)
    
    #kmeans = KMeans(n_clusters).fit(data.reshape(-1,1))
        
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(data)
    
    #kmeans = KMeans(n_clusters).fit(data)
    
    labels = kmeans.labels_
    
    centers = kmeans.cluster_centers_
    
    center_labels = kmeans.predict(centers)
    
    #print(kmeans.cluster_centers_)
    '''
    #visualzie data clustering 
    ######################################################
    y_kmeans = kmeans.predict(data)
    
    plt.scatter(data[:, 0], data[:, 1], c=y_kmeans, s=50, cmap='viridis')

    centers = kmeans.cluster_centers_
    
    plt.scatter(centers[:, 0], centers[:, 1], c='black', s=200, alpha=0.5)
    
    #plt.legend()

    plt.show()
    
    '''
    
    
    '''
    from sklearn.metrics import silhouette_score
    
    range_n_clusters = [2, 3, 4, 5, 6, 7, 8]
    silhouette_avg = []
    for num_clusters in range_n_clusters:
     
         # initialize kmeans
         kmeans = KMeans(n_clusters=num_clusters)
         kmeans.fit(data)
         cluster_labels = kmeans.labels_
         
         # silhouette score
         silhouette_avg.append(silhouette_score(data, cluster_labels))
    
    plt.plot(range_n_clusters,silhouette_avg,'bx-')
    plt.xlabel('Values of K')
    plt.ylabel('Silhouette score')
    plt.title('Silhouette analysis For Optimal k')
    plt.show()
    
    
    Sum_of_squared_distances = []
    K = range(2,8)
    for num_clusters in K :
        kmeans = KMeans(n_clusters=num_clusters)
        kmeans.fit(data)
        Sum_of_squared_distances.append(kmeans.inertia_)
        
    plt.plot(K,Sum_of_squared_distances,'bx-')
    plt.xlabel('Values of K') 
    plt.ylabel('Sum of squared distances/Inertia') 
    plt.title('Elbow Method For Optimal k')
    plt.show()
    '''
    ######################################################
    
    
    return labels, centers, center_labels


#remove outliers
def outlier_remove(data_list):

    #find index of k biggest elements in list
    ####################################################
    
    k = int(len(data_list) * 0.8)
    
    #print(k)
    
    #k biggest
    idx_dominant = np.argsort(data_list)[-k:]
    
    #k smallest
    #idx_dominant_dis_closest_pts = np.argsort(dis_closest_pts)[:k]
    
    #print("idx_dominant_dis_closest_pts = {}".format(idx_dominant_dis_closest_pts))
    
    #print(idx_dominant_dis_closest_pts)
    
    outlier_remove_list = [data_list[index] for index in idx_dominant] 
    
    #print("outlier_remove_list = {}".format(outlier_remove_list))
    
    return outlier_remove_list, idx_dominant
    ####################################################


# save point cloud data from numpy array as ply file, open3d compatiable format
def write_ply(path, data_numpy_array):
    
    #data_range = 100
    
    #Normalize data range for generate cross section level set scan
    #min_max_scaler = preprocessing.MinMaxScaler(feature_range = (0, data_range))

    #point_normalized = min_max_scaler.fit_transform(data_numpy_array)
    
    #initialize pcd object for open3d 
    pcd = o3d.geometry.PointCloud()
     
    pcd.points = o3d.utility.Vector3dVector(data_numpy_array)
    
    # get the model center postion
    model_center = pcd.get_center()
    
    # geometry points are translated directly to the model_center position
    pcd.translate(-1*(model_center))
    
    #write out point cloud file
    o3d.io.write_point_cloud(path, pcd, write_ascii = True)
    
    
    # check saved file
    if os.path.exists(path):
        print("Converted 3d model was saved at {0}".format(path))
        return True
    else:
        return False
        print("Model file converter failed !")
        #sys.exit(0)


# compute diameter from area
def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    
    #note: return diameter instead of radius
    return 2*radius 



# segmentation of overlapping components 
def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    #localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
    
    localMax = peak_local_max(D, min_distance = min_distance_value,  indices = False, labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    #print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels
    


# analyze cross section paramters
def crosssection_analysis(image_file):
    
    path, filename = os.path.split(image_file)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    #print("processing image : {0} \n".format(str(filename)))
    
    # load the image 
    imgcolor = cv2.imread(image_file)
    
    # if cross scan images are white background and black foreground
    #imgcolor = ~imgcolor
    
    # accquire image dimensions 
    height, width, channels = imgcolor.shape
    #shifted = cv2.pyrMeanShiftFiltering(image, 5, 5)

    #Image binarization by apltying otsu threshold
    img = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2GRAY)
    
    # Convert BGR to GRAY
    img_lab = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2LAB)
    
    gray = cv2.cvtColor(img_lab, cv2.COLOR_BGR2GRAY)
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    
    
    #find contours and fill contours 
    ####################################################################
    #container version
    contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #local version
    #_, contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    contours_img = []
    
    #define image morphology operation kernel
    #kernel = cv2.getStructuringElement(cv2.MORPH_CROSS,(3,3))
    
    #draw all the filled contours
    for c in contours:
        
        #fill the connected contours
        contours_img = cv2.drawContours(binary, [c], -1, (255, 255, 255), cv2.FILLED)
        
        #contours_img = cv2.erode(contours_img, kernel, iterations = 5)

    #Obtain the threshold image using OTSU adaptive filter
    thresh_filled = cv2.threshold(contours_img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary_filled = cv2.threshold(contours_img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    
    
    # process filled contour images to extract connected Components
    ####################################################################
    
    contours, hier = cv2.findContours(binary_filled, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #local version
    #_, contours, hier = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #draw all the filled contours
    for c in contours:
        
        #fill the connected contours
        contours_img = cv2.drawContours(binary, [c], -1, (255, 255, 255), cv2.FILLED)

    # define kernel
    connectivity = 8
    
    #find connected components 
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(contours_img, connectivity , cv2.CV_32S)
    
    #find the component with largest area 
    largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA]) 
    
    #collection of all component area
    areas = [s[4] for s in stats]
    
    # average of component area
    area_avg = sum(areas)/len(np.unique(labels))
    
    #area_avg = sum(areas)
    
    ###################################################################
    # segment overlapping components
    #make backup image
    orig = imgcolor.copy()

    min_distance_value = 5
    
    #watershed based segmentaiton 
    labels = watershed_seg(contours_img, thresh_filled, min_distance_value)
    
    #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))
    #label_hue[labels == largest_label] = np.uint8(15)
    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
    
    # save label results
    result_file = (label_path + base_name + '_label.png')
    
    #print(result_file)

    cv2.imwrite(result_file, labeled_img)

    ####################################################################

    #Convert the mean shift image to grayscale, then apply Otsu's thresholding
    convexhull = convex_hull_image(gray)
    
    img_convexhull = np.uint8(convexhull)*255
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh_hull = cv2.threshold(img_convexhull, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #find contours and get the external one
    #1ocal version
    #image_result, contours, hier = cv2.findContours(img_convexhull, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #container version
    contours, hier = cv2.findContours(img_convexhull, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    #print("len(contours)")
    #print(len(contours))
    
    #label image regions
    #label_image_convexhull = label(convexhull)
    
    #Measure properties 
    regions = regionprops(img_convexhull)
    
    if regions[0].area > 0:
        
        density = area_avg/regions[0].area
    else:
        density = sum(areas)/len((labels))+1
    
       
    ####################################################################
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    area_rec = []

    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
        
        # apply individual object mask
        masked = cv2.bitwise_and(contours_img, contours_img, mask = mask)
        
        #define result path 
        #result_img_path = (label_path + 'component_' + str(label) + '.png')
        #cv2.imwrite(result_img_path, masked)
        
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key = cv2.contourArea)
        
        area_rec.append(cv2.contourArea(c))
        
        
    radius_rec = [area_radius(area_val) for area_val in area_rec]
    
    #print(area_rec)
    #print(radius_rec)
    
    #area_avg = sum(area_rec)/len(area_rec)
    
    radius_avg = np.mean(radius_rec)
    
    return radius_avg, area_avg, density


#compute angle
def angle(directions):
    """Return the angle between vectors"""
    vec2 = directions[1:]
    vec1 = directions[:-1]

    norm1 = np.sqrt((vec1 ** 2).sum(axis=1))
    norm2 = np.sqrt((vec2 ** 2).sum(axis=1))
    cos = (vec1 * vec2).sum(axis=1) / (norm1 * norm2)   
    return np.arccos(cos)


#first derivative function
def first_derivative(x) :

    return x[2:] - x[0:-2]


#second derivative function
def second_derivative(x) :
    
    return x[2:] - 2 * x[1:-1] + x[:-2]


#compute curvature
def curvature(x, y) :

    x_1 = first_derivative(x)
    x_2 = second_derivative(x)
    y_1 = first_derivative(y)
    y_2 = second_derivative(y)
    return np.abs(x_1 * y_2 - y_1 * x_2) / np.sqrt((x_1**2 + y_1**2)**3)


#define angle computation for turing points detection
def turning_points(x, y, turning_points, smoothing_radius,cluster_radius):

    
    if smoothing_radius:
        weights = np.ones(2 * smoothing_radius + 1)
        new_x = ndimage.convolve1d(x, weights, mode='constant', cval=0.0)
        new_x = new_x[smoothing_radius:-smoothing_radius] / np.sum(weights)
        new_y = ndimage.convolve1d(y, weights, mode='constant', cval=0.0)
        new_y = new_y[smoothing_radius:-smoothing_radius] / np.sum(weights)
    else :
        new_x, new_y = x, y
        
    k = curvature(new_x, new_y)
    turn_point_idx = np.argsort(k)[::-1]
    t_points = []
    
    while len(t_points) < turning_points and len(turn_point_idx) > 0:
        t_points += [turn_point_idx[0]]
        idx = np.abs(turn_point_idx - turn_point_idx[0]) > cluster_radius
        turn_point_idx = turn_point_idx[idx]
        
    t_points = np.array(t_points)
    t_points += smoothing_radius + 1
    
    return t_points.astype(int)
    

# visualize CDF curve
def CDF_visualization(radius_avg_rec):
    
    trait_file = (label_path + '/CDF.xlsx')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active
    
    for row in enumerate(radius_avg_rec):
    
        sheet.append(row)

    #save the csv file
    wb.save(trait_file)
    
    num_bins = 10
    
    #counts, bin_edges = np.histogram(list(zip(*result)[0]), bins = num_bins, normed = True)
    counts, bin_edges = np.histogram(radius_avg_rec, bins = num_bins)
    
    # compute CDF curve
    cdf = np.cumsum(counts)
    
    #cdf = cdf / cdf[-1] #normalize
    
    x = bin_edges[1:]
    y = cdf
    
    # assembly points of CDF curve 
    trajectory = np.vstack((x, y)).T

    index_turning_pt = turning_points(x, y, turning_points = 4, smoothing_radius = 2, cluster_radius = 2)
    

    #Ramer-Douglas-Peucker Algorithm
    #simplify points et using rdp library 
    simplified_trajectory = rdp(trajectory, epsilon = 0.00200)
    #simplified_trajectory = rdp(trajectory)
    sx, sy = simplified_trajectory.T
    
    #print(sx)
    #print(sy)

    #compute plateau in curve
    dis_sy = [j-i for i, j in zip(sy[:-1], sy[1:])]
    
    #get index of plateau location
    index_sy = [i for i in range(len(dis_sy)) if dis_sy[i] <= 1.3]
    
    dis_index_sy = [j-i for i, j in zip(index_sy[:-1], index_sy[1:])]
    
    for idx, value in enumerate(dis_index_sy):
        
        if idx < len(index_sy)-2:
        
            if value == dis_index_sy[idx+1]:
            
                index_sy.remove(index_sy[idx+1])
    
    # Define a minimum angle to treat change in direction
    # as significant (valuable turning point).
    #min_angle = np.pi / 36.0
    min_angle = np.pi / 180.0
    #min_angle = np.pi /1800.0
    
    # Compute the direction vectors on the simplified_trajectory.
    directions = np.diff(simplified_trajectory, axis = 0)
    theta = angle(directions)

    # Select the index of the points with the greatest theta.
    # Large theta is associated with greatest change in direction.
    idx = np.where(theta > min_angle)[0] + 1
    
    index_turning_pt = sorted(idx)
    
    Turing_points =  np.unique(sy[idx].astype(int))
    
    #max_idx = max(max_idx)
    #print("Turing points: {0} \n".format(Turing_points))
    
    # plot CDF 
    #fig = plt.plot(bin_edges[1:], cdf, '-r', label = 'CDF')
    fig = plt.figure(1)
    #ax = fig.add_subplot(111)
    plt.grid(True)
    #plt.legend(loc = 'right')
    plt.title('CDF curve')
    plt.xlabel('Root area, unit:pixel')
    plt.ylabel('Depth of level-set, unit:pixel')
    
    plt.plot(sx, sy, 'gx-', label = 'simplified trajectory')
    plt.plot(bin_edges[1:], cdf, '-b', label = 'CDF')
    #plt.plot(sx[idx], sy[idx], 'ro', markersize = 7, label='turning points')
    
    plt.plot(sx[index_sy], sy[index_sy], 'ro', markersize = 7, label = 'plateau points')
    
    #plt.plot(sx[index_turning_pt], sy[index_turning_pt], 'bo', markersize = 7, label='turning points')
    #plt.vlines(sx[index_turning_pt], sy[index_turning_pt]-100, sy[index_turning_pt]+100, color='b', linewidth = 2, alpha = 0.3)
    #plt.legend(loc='best')
    
    result_file_CDF = label_path + '/'  + 'cdf.png'
    plt.savefig(result_file_CDF)
    plt.close()
    
    return sy


# compute number of whorls
def wholr_number_count(imgList):
    
    area_avg_rec = []
    
    density_rec = []
    
    for img in imgList:

        #(area_avg, area_sum, n_unique_labels) = root_area_label(img)
        
        (radius_avg, area_avg, density) = crosssection_analysis(img)
        
        area_avg_rec.append(area_avg)
        
        density_rec.append(density)
    
    #visualzie the CDF graph of first return value 
    list_thresh = sorted(CDF_visualization(area_avg_rec))

    #compute plateau in curve
    dis_array = [j-i for i, j in zip(list_thresh[:-1], list_thresh[1:])]
    
    #get index of plateau location
    index = [i for i in range(len(dis_array)) if dis_array[i] <= 1.3]
    
    dis_index = [j-i for i, j in zip(index[:-1], index[1:])]
    
    for idx, value in enumerate(dis_index):
        
        if idx < len(index)-2:
        
            if value == dis_index[idx+1]:
            
                index.remove(index[idx+1])
    
    
    reverse_index = sorted(index, reverse = True)
    
    #count = sum(1 for x in dis_array if float(x) <= 1.3)
    #get whorl number count 
    
    count_wholrs = int(math.ceil(len(index)))

    
    #compute wholr location
     #compute whorl location
    whorl_dis = []
    whorl_loc = []
    
    for idx, value in enumerate(reverse_index, start=1):
        
        #dis_value = list_thresh[value+1] - list_thresh[value-1]
        loc_value = int(len(imgList) - list_thresh[value+1])
        whorl_loc.append(loc_value)
        #print("adding value : {0} \n".format(str(loc_value)))
    
    
    #compute whorl distance
    whorl_dis_array = [j-i for i, j in zip(whorl_loc[:-1], whorl_loc[1:])]
    whorl_loc.extend([0, len(imgList)])
    whorl_loc = list(dict.fromkeys(whorl_loc))
    whorl_loc_ex = sorted(whorl_loc)
    
    #print("list_thresh : {0} \n".format(str(list_thresh)))
    
    
    return count_wholrs, whorl_loc_ex, sum(density_rec)/len(density_rec)




# compute average from cross section scan
def crosssection_analysis_range(start_idx, end_idx):

    radius_avg_rec = []
    
    for img in imgList[int(start_idx): int(end_idx)]:

        (radius_avg, area_avg, density) = crosssection_analysis(img)
        
        radius_avg_rec.append(radius_avg)
        
    #print(radius_avg_rec)
    
    k = int(len(radius_avg_rec) * 0.8)
    
    #print(k)
    
    #k smallest
    idx_dominant = np.argsort(radius_avg_rec)[:k]
    
    outlier_remove_list = [radius_avg_rec[index] for index in idx_dominant] 
    
    #print(outlier_remove_list)
    
    return np.mean(outlier_remove_list)


#compute muti-dimension distance between two points
def multiDimenDist(point1,point2):
   #find the difference between the two points, its really the same as below
   deltaVals = [point2[dimension]-point1[dimension] for dimension in range(len(point1))]
   runningSquared = 0
   #because the pythagarom theorm works for any dimension we can just use that
   for coOrd in deltaVals:
       runningSquared += coOrd**2
   return runningSquared**(1/2)
   

#compute two vectors form two points in muti-dimension space
def findVec(point1,point2,unitSphere = False):
  #setting unitSphere to True will make the vector scaled down to a sphere with a radius one, instead of it's orginal length
  finalVector = [0 for coOrd in point1]
  for dimension, coOrd in enumerate(point1):
      #finding total differnce for that co-ordinate(x,y,z...)
      deltaCoOrd = point2[dimension]-coOrd
      #adding total difference
      finalVector[dimension] = deltaCoOrd
  if unitSphere:
      totalDist = multiDimenDist(point1,point2)
      unitVector =[]
      for dimen in finalVector:
          unitVector.append( dimen/totalDist)
      return unitVector
  else:
      return finalVector



#get rotation matrix between two vectors using scipy
def get_rotation_matrix(vec2, vec1=np.array([1, 0, 0])):

    vec1 = np.reshape(vec1, (1, -1))
    
    vec2 = np.reshape(vec2, (1, -1))
    
    r = R.align_vectors(vec2, vec1)
    
    #from scipy.spatial.transform import Rotation as R

    return r[0].as_matrix()


#Find the rotation matrix that aligns vec1 to vec2
def rotation_matrix_from_vectors(vec1, vec2):
    """ 
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return np.eye(3) #cross of all zeros only occurs on identical directions


# Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
# The quaternions are arranged as (w,x,y,z), with w being the scalar
# The result will be the average quaternion of the input. Note that the signs
# of the output quaternion can be reversed, since q and -q describe the same orientation
def averageQuaternions(Q):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))

    for i in range(0,M):
        q = Q[i,:]
        # multiply q with its transposed version q' and add A
        A = np.outer(q,q) + A

    # scale
    A = (1.0/M)*A
    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)
    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]
    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0])
    return np.ravel(eigenVectors[:,0])


# Average multiple quaternions with specific weights
# The weight vector w must be of the same length as the number of rows in the
# quaternion maxtrix Q
def weightedAverageQuaternions(Q, w):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))
    weightSum = 0

    for i in range(0,M):
        q = Q[i,:]
        A = w[i] * np.outer(q,q) + A
        weightSum += w[i]

    # scale
    A = (1.0/weightSum) * A

    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)

    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]

    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0].A1)
    return np.ravel(eigenVectors[:,0])


def euler_to_rotMat(yaw, pitch, roll):
    Rz_yaw = np.array([
        [np.cos(yaw), -np.sin(yaw), 0],
        [np.sin(yaw),  np.cos(yaw), 0],
        [          0,            0, 1]])
    Ry_pitch = np.array([
        [ np.cos(pitch), 0, np.sin(pitch)],
        [             0, 1,             0],
        [-np.sin(pitch), 0, np.cos(pitch)]])
    Rx_roll = np.array([
        [1,            0,             0],
        [0, np.cos(roll), -np.sin(roll)],
        [0, np.sin(roll),  np.cos(roll)]])
    # R = RzRyRx
    rotMat = np.dot(Rz_yaw, np.dot(Ry_pitch, Rx_roll))
    
    return rotMat


# RPY/Euler angles to Rotation Vector
def euler_to_rotVec(yaw, pitch, roll):

    # compute the rotation matrix
    Rmat = euler_to_rotMat(yaw, pitch, roll)
    
    theta = math.acos(((Rmat[0, 0] + Rmat[1, 1] + Rmat[2, 2]) - 1) / 2)
    sin_theta = math.sin(theta)
    if sin_theta == 0:
        rx, ry, rz = 0.0, 0.0, 0.0
    else:
        multi = 1 / (2 * math.sin(theta))
        rx = multi * (Rmat[2, 1] - Rmat[1, 2]) * theta
        ry = multi * (Rmat[0, 2] - Rmat[2, 0]) * theta
        rz = multi * (Rmat[1, 0] - Rmat[0, 1]) * theta
    return rx, ry, rz



#find shortest path between start and end vertex
def short_path_finder(G_unordered, start_v, end_v):
    
    #find shortest path between start and end vertex
    ####################################################################
    
    #define start and end vertex index
    #start_v = 0
    #end_v = 1559
    #end_v = 608
    
    
    #print(X_skeleton[start_v], Y_skeleton[start_v], Z_skeleton[start_v])
    
    # find shortest path in the graph between start and end vertices 
    vlist, elist = gt.shortest_path(G_unordered, G_unordered.vertex(start_v), G_unordered.vertex(end_v))
    
    vlist_path = [str(v) for v in vlist]
    
    elist_path = [str(e) for e in elist]
    
    #change format form str to int
    int_vlist_path = [int(i) for i in vlist_path]
    
    #print(int_vlist_path)
    '''
    if len(vlist_path) > 0: 
        
        print("Shortest path found in graph! \n")
        
        print("vlist_path = {} \n".format(int_vlist_path))
    
        #curve_length = path_length(X_skeleton[int_vlist_path], Y_skeleton[int_vlist_path], Z_skeleton[int_vlist_path])
    
        #print("curve_length = {} \n".format(curve_length))
        
    else:
        print("No shortest path found in graph...\n")
    '''
    return int_vlist_path


def dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, select):
    
    loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(loc_start), min(loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(loc_end), min(loc_end)))
    
    max_dimension_length = abs(max(max(loc_start), max(loc_end) - min(min(loc_start), min(loc_end))))

    min_dimension_length = abs(min(min(loc_start), min(loc_end) - max(max(loc_start), max(loc_end))))
    
    #print("max_length = {}\n".format(max_length))
    
    if select == 1:
        return max_dimension_length
    else:
        return min_dimension_length



def calculate_wcss(data):
    wcss = []
    for n in range(2, 10):
        kmeans = KMeans(n_clusters=n)
        kmeans.fit(X=data)
        wcss.append(kmeans.inertia_)

    return wcss



def optimal_number_of_clusters(wcss):
    x1, y1 = 2, wcss[0]
    x2, y2 = 20, wcss[len(wcss)-1]

    distances = []
    for i in range(len(wcss)):
        x0 = i+2
        y0 = wcss[i]
        numerator = abs((y2-y1)*x0 - (x2-x1)*y0 + x2*y1 - y2*x1)
        denominator = math.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        distances.append(numerator/denominator)

    return distances.index(max(distances)) + 0


# Skeleton analysis
def analyze_skeleton(current_path, filename_skeleton, filename_pcloud):
    
    model_skeleton = current_path + filename_skeleton
    print("Loading 3D skeleton file {}...\n".format(filename_skeleton))
    model_skeleton_name_base = os.path.splitext(model_skeleton)[0]
    
    #load the ply format skeleton file 
    try:
        with open(model_skeleton, 'rb') as f:
            plydata_skeleton = PlyData.read(f)
            num_vertex_skeleton = plydata_skeleton.elements[0].count
            N_edges_skeleton = len(plydata_skeleton['edge'].data['vertex_indices'])
            array_edges_skeleton = plydata_skeleton['edge'].data['vertex_indices']
            
            print("Ply data structure: \n")
            #print(plydata_skeleton)
            #print("\n")
            print("Number of 3D points in skeleton model: {0} \n".format(num_vertex_skeleton))
            print("Number of edges: {0} \n".format(N_edges_skeleton))

        
    except:
        sys.exit("Model skeleton file does not exist!")
    
    
    #Parse ply format skeleton file and Extract the data
    Data_array_skeleton = np.zeros((num_vertex_skeleton, 3))
    
    Data_array_skeleton[:,0] = plydata_skeleton['vertex'].data['x']
    Data_array_skeleton[:,1] = plydata_skeleton['vertex'].data['y']
    Data_array_skeleton[:,2] = plydata_skeleton['vertex'].data['z']
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    
    
    #load radius values
    ####################################################################
    #radius_skeleton = current_path + filename_skeleton
    
    base_name = os.path.splitext(os.path.basename(model_skeleton_name_base))[0]
    txt_base_name = base_name.replace("_skeleton", "_avr.txt")
    radius_file = current_path + txt_base_name

    print("Loading 3D skeleton radius txt file {}...\n".format(radius_file))
    
    #check file exits
    if os.path.isfile(radius_file):
        
        with open(radius_file) as file:
            lines = file.readlines()
            radius_vtx = [line.rstrip() for line in lines]
    else:
        
        sys.exit("Could not load 3D skeleton radius txt file")  
    
    
    # build directed graph from skeleton/structure data
    ####################################################################
    print("Building directed graph from 3D skeleton/structure ...\n")
    G_unordered = gt.Graph(directed = True)
    
    # assert directed graph
    #print(G.is_directed())
    
    nodes = G_unordered.add_vertex(num_vertex_skeleton)
    
    G_unordered.add_edge_list(array_edges_skeleton.tolist()) 
    
    #gt.graph_draw(G_unordered, vertex_text = G_unordered.vertex_index, output = current_path + "graph_view.pdf")
    
    
    # find all end vertices by fast iteration of all vertices
    end_vlist = []
    
    end_vlist_offset = []
    
    for v in G_unordered.iter_vertices():
        
        #print(G.vertex(v).out_degree(), G.vertex(v).in_degree())
        
        if G_unordered.vertex(v).out_degree() == 0 and G_unordered.vertex(v).in_degree() == 1:
        
            end_vlist.append(v)
            
            if (v+1) == num_vertex_skeleton:
                end_vlist_offset.append(v)
            else:
                end_vlist_offset.append(v+1)
            
    #print("end_vlist = {} \n".format(end_vlist))
    #print("end_vlist_offset = {} \n".format(end_vlist_offset))
    

    #test angle calculation
    #vector1 = [0,0,1]
    # [1,0,0]
    #vector2 = [0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1]
    #print(dot_product_angle(vector1,vector2))
    #print(cart2sph(0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1)) 

    
    # parse all the sub branches edges and vetices, start, end vetices
    #####################################################################################
    sub_branch_list = []
    sub_branch_length_rec = []
    sub_branch_angle_rec = []
    sub_branch_start_rec = []
    sub_branch_end_rec = []
    sub_branch_projection_rec = []
    sub_branch_radius_rec = []
    
    sub_branch_xs_rec = []
    sub_branch_ys_rec = []
    sub_branch_zs_rec = []
    
    #if len(end_vlist) == len(end_vlist_offset):
        
    for idx, v_end in enumerate(end_vlist):
        
        #print(idx, v_end)
        #construct list of vertices in sub branches
        if idx == 0:
            v_list = [*range(0, int(end_vlist[idx])+1)]
        else:
            v_list = [*range(int(end_vlist[idx-1])+1, int(end_vlist[idx])+1)]
            
        # change type to interger 
        int_v_list = [int(i) for i in v_list]
        
        # current sub branch length
        sub_branch_length = path_length(X_skeleton[int_v_list], Y_skeleton[int_v_list], Z_skeleton[int_v_list])
        
        # current sub branch start and end points 
        start_v = [X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]]
        
        end_v = [X_skeleton[int_v_list[0]] - X_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Y_skeleton[int_v_list[0]] - Y_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Z_skeleton[int_v_list[0]] - Z_skeleton[int_v_list[int(len(int_v_list)-1.0)]]]
        
        # angle of current branch vs Z direction
        angle_sub_branch = dot_product_angle(start_v, end_v)
        
        # projection radius of current branch length
        p0 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]])
        
        p1 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[-1]]])
        
        projection_radius = np.linalg.norm(p0 - p1)
        
        #radius value from fitted point cloud contours
        radius_edge = float(radius_vtx[int_v_list[0]])*1
        
        sub_branch_xs = X_skeleton[int_v_list[0]]
        sub_branch_ys = Y_skeleton[int_v_list[0]]
        sub_branch_zs = Z_skeleton[int_v_list[0]]
        
        
        # save computed parameters for each branch
        sub_branch_list.append(v_list)
        sub_branch_length_rec.append(sub_branch_length)
        sub_branch_angle_rec.append(angle_sub_branch)
        sub_branch_start_rec.append(int_v_list[0])
        sub_branch_end_rec.append(int_v_list[-1])
        sub_branch_projection_rec.append(projection_radius)
        sub_branch_radius_rec.append(radius_edge)
        
        sub_branch_xs_rec.append(sub_branch_xs)
        sub_branch_ys_rec.append(sub_branch_ys)
        sub_branch_zs_rec.append(sub_branch_zs)
    
    #print(min(sub_branch_angle_rec))
    #print(max(sub_branch_angle_rec))
    

    
    '''
    # sort branches according to the start vertex location(Z value) 
    ####################################################################
    Z_loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    
    sorted_idx_Z_loc = np.argsort(Z_loc_value)
    
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_Z_loc]
    
    
    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_Z_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_Z_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_Z_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_Z_loc]

    #print("sub_branch_length_rec = {}\n".format(sub_branch_length_rec[0:20]))
    
    '''
    # sort branches according to length feature in descending order
    ####################################################################
    sorted_idx_len = np.argsort(sub_branch_length_rec)
    
    #reverse the order from accending to descending
    sorted_idx_len_loc = sorted_idx_len[::-1]

    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_len_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_len_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_len_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_len_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_len_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_len_loc]
    sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_len_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_len_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_len_loc]

    
    ####################################################################
    #(count_wholrs, whorl_loc_ex, avg_density) = wholr_number_count(imgList)
    
    #print("number of whorls is: {} whorl_loc_ex : {} avg_density = {}\n".format(count_wholrs, str(whorl_loc_ex), avg_density))
    
    '''
    Z_loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    Z_loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(Z_loc_start), min(Z_loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(Z_loc_end), min(Z_loc_end)))
    
    max_length = abs(max(max(Z_loc_start), max(Z_loc_end) - min(min(Z_loc_start), min(Z_loc_end))))
    '''
    '''
    max_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    max_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 1)
    
    min_length_x = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_y = dimension_size(Y_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    min_length_z = dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, 0)
    
    print("max_length = {} {} {}\n".format(max_length_x, max_length_y, max_length_z))
    
    print("min_length = {} {} {}\n".format(min_length_x, min_length_y, min_length_z))
    
    s_diameter_max = max(max_length_x, max_length_y)
    
    s_diameter_min = min(min_length_x, min_length_y)
    
    #s_diameter = (s_diameter_max + s_diameter_min)*0.5
    
    s_length = max_length_z
    '''
            
    '''
    # construct sub branches with length and radius feature 
    ####################################################################
    combined_list = np.array(list(zip(sub_branch_length_rec, sub_branch_radius_rec))).reshape(len(sub_branch_length_rec), 2)
    
    # calculating the within clusters sum-of-squares 
    sum_of_squares = calculate_wcss(combined_list)
    
    # calculating the optimal number of clusters
    n_optimal = optimal_number_of_clusters(sum_of_squares)
    
    print("optimal_number_of_clusters = {}\n".format(n_optimal))
    
   
    # find sub branches cluster with length and radius feature 
    ####################################################################
    cluster_number = n_optimal + 4
    
    (labels, centers, center_labels) = cluster_list(combined_list, n_clusters = cluster_number)
    
    sorted_idx = np.argsort(centers[:,0])[::-1]

    print("sorted_idx = {}\n".format(sorted_idx))
    
    
    indices_level = []
    sub_branch_level = []
    sub_branch_start_level = []
    sub_branch_startZ_level = []
    radius_level = []
    length_level = []
    angle_level = []
    projection_level = []
    
    for idx, (idx_value) in enumerate(sorted_idx):
        
        #print(labels_length_rec.tolist().index(idx_value))
        
        #print("cluster {}, center value {}".format(idx, idx_value))
        indices = [i for i, x in enumerate(labels.tolist()) if x == idx_value]
        
        #print(indices)
        
        sub_branch_start_rec_selected = [sub_branch_start_rec[i] for i in indices]
        Z_loc = [Z_skeleton[index] for index in sub_branch_start_rec_selected]
        
        sub_loc = [sub_branch_list[index] for index in indices]
        radius_loc = [sub_branch_radius_rec[index] for index in indices]
        length_loc = [sub_branch_length_rec[index] for index in indices]
        angle_loc = [sub_branch_angle_rec[index] for index in indices]
        projection_loc = [sub_branch_projection_rec[index] for index in indices]
        
        
        print("max = {} min = {} ".format(max(sub_branch_start_rec_selected), min(sub_branch_start_rec_selected)))
        print("max_Z = {} min_Z = {} average = {}".format(max(Z_loc), min(Z_loc), np.mean(Z_loc)))
        print("max_radius = {} min_radius = {} average = {}".format(max(radius_loc), min(radius_loc), np.mean(radius_loc)))
        print("max_length = {} min_length = {} average = {}".format(max(length_loc), min(length_loc), np.mean(length_loc)))
        print("max_angle = {} min_angle = {} average = {}".format(max(angle_loc), min(angle_loc), np.mean(angle_loc)))
        print("max_projection = {} min_projection = {}".format(max(projection_loc), min(projection_loc), np.mean(projection_loc)))
        print("number of roots = {} {} {}\n".format(len(indices), len(Z_loc), len(radius_loc)))

        indices_level.append(indices)
        sub_branch_level.append(sub_loc)
        sub_branch_start_level.append(sub_branch_start_rec_selected)
        sub_branch_startZ_level.append(Z_loc)
        radius_level.append(radius_loc)
        length_level.append(length_loc)
        angle_level.append(angle_loc)
        projection_level.append(projection_loc)
        
    
    #compute paramters
    #avg_radius_stem = max(radius_level[0])*2
    avg_radius_stem = np.mean(radius_level[0])*2
    

    
    num_brace = len(indices_level[0]) + len(indices_level[1])
    avg_brace_length = np.mean(length_level[1])
    avg_brace_angle = np.mean(angle_level[1])
    avg_radius_brace = np.mean(radius_level[1])*2
    avg_brace_projection = np.mean(projection_level[1])
    

    
    num_crown = len(indices_level[2]) - len(indices_level[1]) - len(indices_level[0])
    avg_crown_length = np.mean(length_level[2])
    avg_crown_angle = np.mean(angle_level[2])
    avg_radius_crown = np.mean(radius_level[2])*2
    avg_crown_projection = np.mean(projection_level[2])
    
    avg_radius_lateral = np.mean(radius_level[3])

    

    
    if num_brace ==0:
        num_crown = 18

    if num_crown < 10:
        num_crown = num_crown*2 + int(num_crown*0.7)

    if num_brace < 10 and num_brace > 0:
        num_brace = num_brace*2 + int(num_brace*0.7)
    elif num_brace >20:
        num_brace = round(interp(num_brace,[1,num_brace*2],[15,20]))

    if num_crown < 10 and num_crown > 0:
        num_crown = num_crown*2 + int(num_crown*0.7)
    elif num_crown >30:
        num_crown = round(interp(num_crown,[1,num_crown*2],[18,26]))
    elif num_crown ==0 and num_brace > 12: 
        num_crown = 35
    elif num_crown ==0 or num_crown < 0:
        num_crown = num_brace + 10
    
    
    whorl_dis_1 = abs(np.mean(sub_branch_startZ_level[0]) - np.mean(sub_branch_startZ_level[1]))
    whorl_dis_2 = abs(np.mean(sub_branch_startZ_level[1]) - np.mean(sub_branch_startZ_level[2]))


    
    if num_brace < 25 and num_crown < 27:
        n_whorl = count_wholrs + 2
    else:
        n_whorl = count_wholrs + 3
    '''
    ################################################################################################################################

    
    
    # graph: find closest point pairs and connect close graph edges
    ####################################################################
    print("Converting skeleton to graph and connecting edges and vertices...\n")
    
    v_closest_pair_rec = []
    
    closest_pts = []
    
    #find closest point set and connect graph edges
    for idx, (sub_branch, anchor_point) in enumerate(zip(sub_branch_list, end_vlist_offset)):
        
        # start vertex of an edge
        anchor_point = (X_skeleton[end_vlist_offset[idx]], Y_skeleton[end_vlist_offset[idx]], Z_skeleton[end_vlist_offset[idx]])

        # curve of the edge in 3D
        point_set = np.zeros((len(sub_branch_list[0]), 3))
        
        point_set[:,0] = X_skeleton[sub_branch_list[0]]
        point_set[:,1] = Y_skeleton[sub_branch_list[0]]
        point_set[:,2] = Z_skeleton[sub_branch_list[0]]
        
        (index_cp, value_cp) = closest_point(point_set, anchor_point)

        v_closest_pair = [index_cp, end_vlist_offset[idx]]

        dis_v_closest_pair = path_length(X_skeleton[v_closest_pair], Y_skeleton[v_closest_pair], Z_skeleton[v_closest_pair])
        
        #small threshold indicating close pair vetices
        if dis_v_closest_pair < 0.11:
            
            closest_pts.append(index_cp)
            
            #print("dis_v_closest_pair = {}".format(dis_v_closest_pair))
            v_closest_pair_rec.append(v_closest_pair)
            
            #print("closest point pair: {0}".format(v_closest_pair))
            
            #connect graph edges
            G_unordered.add_edge(index_cp, end_vlist_offset[idx])
            
    #print("v_closest_pair_rec = {}\n".format(v_closest_pair_rec))
    
    #get the unique values from the list
    #closest_pts_unique = list(set(closest_pts))
    
    #keep repeat values for correct indexing order
    
    print("closest_pts = {}\n".format(closest_pts))
    
    closest_pts_unique = list((closest_pts))
    
    closest_pts_unique_sorted = sorted(closest_pts_unique)
    
    print("closest_pts_unique_sorted = {}\n".format(closest_pts_unique_sorted))
    '''
    
    #sort and combine adjacent connecting vertices in closest_pts  
    ####################################################################
    X = X_skeleton[closest_pts_unique_sorted]
    Y = Y_skeleton[closest_pts_unique_sorted]
    Z = Z_skeleton[closest_pts_unique_sorted]
    
    # compute distance between adjacent vertices in closest_pts_unique_sorted
    dis_closest_pts = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1, len(X))]
    
    #print("distance between closest_pts_unique = {}\n".format(dis_closest_pts))
    '''
    
    
    '''
    #find outlier of closest points based on its distance list, then merge close points
    ####################################################################
    index_outlier = mad_based_outlier(np.asarray(dis_closest_pts),3.5)
    
    #print("index_outlier = {}".format(index_outlier))
    
    index_outlier_loc = [i for i, x in enumerate(index_outlier) if x]
    
    closest_pts_unique_sorted_combined = [closest_pts_unique_sorted[index] for index in index_outlier_loc]
    
    #print("index_outlier = {}\n".format(index_outlier_loc))
    
    
    
    if len(closest_pts_unique_sorted_combined) < 1:
        
        closest_pts_unique_sorted_combined = list(set(closest_pts))
    
        print("Adjusted closest_pts_unique_sorted_combined = {}\n".format(closest_pts_unique_sorted_combined))
    
    else:
        
        print("closest_pts_unique_sorted_combined = {}\n".format(closest_pts_unique_sorted_combined))
    
    

    #find Z locations of each part
    Z_range_stem = (Z_skeleton[0], Z_skeleton[closest_pts_unique_sorted_combined[-1]])
    #Z_range_crown = (Z_skeleton[closest_pts_unique_sorted_combined[0]], sub_branch_start_Z[-1])
    if len(sub_branch_start_Z) < 1:
        Z_range_brace = (Z_skeleton[closest_pts_unique_sorted_combined[0]], sub_branch_start_Z[0])
    else:
        Z_range_brace = (sub_branch_start_Z[-1], sub_branch_start_Z[0])
        
    Z_range_crown = (sub_branch_start_Z[-1], sub_branch_end_Z[0])
    
    #####################################################################
    '''

    
    '''
    #Search skeleton graph
    ####################################################################
    search_radius = 150
    
    neighbors_idx_rec = []
    
    # search neighbors of every vertex in closest_pts_unique_sorted_combined to find sub branches
    for idx, val in enumerate(closest_pts_unique_sorted_combined):
        
        anchor_pt_idx = int(val)
        
        idx = get_neighbors(Data_array_skeleton, anchor_pt_idx, search_radius)
    
        neighbors_idx = sorted(list(np.asarray(idx)))
        
        #find branches within near neighbors search range
        #print("neighbors_idx = {}\n".format(neighbors_idx))

        #find branches within near neighbors 
        neighbors_match = sorted(list(set(sub_branch_start_rec).intersection(set(neighbors_idx))))
        
        print("Found {} matches, neighbors_match = {}\n".format(len(neighbors_match), neighbors_match))
        
        neighbors_idx_rec.append(neighbors_match)
        
    ####################################################################
    
    
    v_closest_pair_rec_selected = [v_closest_pair_rec[index] for index in index_outlier_loc] 
    
    v_closest_start_selected = [v_closest_pair_rec[index][1] for index in index_outlier_loc]
    
    #print("v_closest_pair_rec_selected = {}\n".format(v_closest_pair_rec_selected))
    
    #print("v_closest_start_selected = {}\n".format(v_closest_start_selected))
    
    
    #sub_branch_selected = [sub_branch_list[index+1] for index in index_outlier_loc]
    
    index_level_selected = [int(index+1) for index in index_outlier_loc]
    
    print("index_level_selected = {}\n".format(index_level_selected))
    
    
    level_range_set = []
    
    for idx, val in enumerate(index_level_selected):
        
        if (idx+1) < len(index_level_selected): 
            
            range_idx = range(index_level_selected[idx], index_level_selected[idx+1])
        
            #print([*range_idx])
            
            level_range_set.append([*range_idx])


    #choose level set depth
    combined_level_range_set = level_range_set[0:2]
    
    combined_level_range_set = [item for sublist in combined_level_range_set for item in sublist]
    
    print("combined_level_range_set = {}\n".format(combined_level_range_set))
    
    #sub_branch_selected = [sub_branch_list[index] for index in combined_level_range_set]
    '''
   
    
    #convert skeleton data to KDTree using Open3D to search nearest neighbors
    #find branches within near neighbors search range
    ####################################################################
    '''
    anchor_pt_idx = 30
    
    search_radius = 150
    
    idx = get_neighbors(Data_array_skeleton, anchor_pt_idx, search_radius)
    
    neighbors_idx = sorted(list(np.asarray(idx)))
    
    print("neighbors_idx = {}\n".format(neighbors_idx))
    
    #find branches within near neighbors 
    neighbors_match = sorted(list(set(sub_branch_start_rec).intersection(set(neighbors_idx))))
    
    print("neighbors_match = {}\n".format(neighbors_match))
    
    
    
    level = 1
    
    neighbors_match_idx = [i for i, item in enumerate(sub_branch_start_rec) if item in neighbors_idx_rec[level]]
    
    #neighbors_match_idx = [int(i) for i in neighbors_match_idx]
    
    sub_branch_selected = [sub_branch_list[index] for index in sorted(neighbors_match_idx)]
    
    #print("neighbors_match_idx = {}\n".format(neighbors_match_idx))
    #print("sub_branch_selected = {}\n".format(len(sub_branch_selected)))
    
    num_1_order = len(sub_branch_selected)
    
    angle_1_order = [sub_branch_angle_rec[index] for index in sorted(neighbors_match_idx)]
    
    length_1_order = [sub_branch_length_rec[index] for index in sorted(neighbors_match_idx)]
    
    print("num_1_order = {0}\n  angle_1_order = {1}\n length_1_order = {2}\n".format(num_1_order, angle_1_order, length_1_order))
    '''

    #find shortest path between start and end vertex
    ####################################################################
    
    #define start and end vertex index
    start_v = 0
    #end_v = 1559
    #end_v = 608
    
    #int_vlist_path = short_path_finder(G_unordered, 0, 608)
    
    
    vlist_path_rec = []
    
    quaternion_path_rec = []
    
    rotVec_rec = []
    
    #for idx, end_v in enumerate(sub_branch_end_rec):
    for idx, end_v in enumerate(sub_branch_end_rec[0:1000]):
        
        #print("start_v = {} end_v = {} \n".format(start_v, end_v))
   
        vlist_path = short_path_finder(G_unordered, start_v, end_v)
        
        if len(vlist_path) > 0:
        
            vlist_path_rec.append(vlist_path)
            
            sum_quaternion = np.zeros([len(vlist_path), 4])
            
            #sum_euler = np.zeros([len(vlist_path), 3])
            
            for i, v_path in enumerate(vlist_path):
        
                if i + 2 < len(vlist_path):
                    
                    vector1 = [X_skeleton[vlist_path[i]], Y_skeleton[vlist_path[i]], Z_skeleton[vlist_path[i]]]
                    vector2 = [X_skeleton[vlist_path[i + 1]], Y_skeleton[vlist_path[i + 1]], Z_skeleton[vlist_path[i + 1]]]
                    vector3 = [X_skeleton[vlist_path[i + 2]], Y_skeleton[vlist_path[i + 2]], Z_skeleton[vlist_path[i + 2]]]
            
                    vector_12 = findVec(vector1,vector2)
                    vector_23 = findVec(vector2,vector3)
                    
                    mat = get_rotation_matrix(vec1 = vector_12, vec2 = vector_23)

                    quaternion_r = R.from_matrix(mat).as_quat()
                    
                    #euler_r = R.from_matrix(mat).as_euler('xyz', degrees = True)
                    
                    sum_quaternion[i,:] = quaternion_r
                    
                    #sum_euler[i,:] = euler_r
                    
                    #print("vlist_path = {} quaternion_r = {}".format(idx, quaternion_r))
            
            avg_quaternion = averageQuaternions(sum_quaternion)
            
            rot = R.from_quat(avg_quaternion)
            
            avg_euler = rot.as_euler('xyz')
            
            #avg_euler = np.mean(sum_euler, axis = 0)
            
            rotVec = euler_to_rotVec(avg_euler[0], avg_euler[1], avg_euler[2])
            
            print("vlist_path = {} avg_quaternion = {} avg_euler = {} rotVec = {}".format(idx, avg_quaternion, avg_euler, rotVec))
            
            rotVec_rec.append(rotVec)
            
    print("Found {} shortest path \n".format(len(vlist_path_rec)))
    
    #print(rotVec_rec)
    
    '''
    for idx, v_path in enumerate(vlist_path_rec[0]):
        
        if idx + 2 < len(vlist_path):
            
            vector1 = [X_skeleton[vlist_path_rec[0][idx]], Y_skeleton[vlist_path_rec[0][idx]], Z_skeleton[vlist_path_rec[0][idx]]]
            vector2 = [X_skeleton[vlist_path_rec[0][idx + 1]], Y_skeleton[vlist_path_rec[0][idx + 1]], Z_skeleton[vlist_path_rec[0][idx + 1]]]
            vector3 = [X_skeleton[vlist_path_rec[0][idx + 2]], Y_skeleton[vlist_path_rec[0][idx + 2]], Z_skeleton[vlist_path_rec[0][idx + 2]]]
    
            vector_12 = findVec(vector1,vector2)
            vector_23 = findVec(vector2,vector3)
            
            mat = get_rotation_matrix(vec1 = vector_12, vec2 = vector_23)
            
            print(mat)

            quaternion_r = R.from_matrix(mat).as_quat()
            
            print(quaternion_r)
    '''
    
    '''
    print(vlist_path_rec[0][10])
    
    vec1 = [X_skeleton[vlist_path_rec[0][10]], Y_skeleton[vlist_path_rec[0][10]], Z_skeleton[vlist_path_rec[0][10]]]
    
    vec2 = [X_skeleton[vlist_path_rec[0][11]], Y_skeleton[vlist_path_rec[0][11]], Z_skeleton[vlist_path_rec[0][11]]]
    
    #vectors = findVec([0,0,0],[3,6,8])
    
    vectors = findVec(vec1,vec2)
    
    print(vectors)
    
    #vec1 = np.array([1, 0, 0])
    #vec2 = np.array([0, 1, 0])
    
    vec1 = [2, 3, 2.5]
    vec2 = [-3, 1, -3.4]

    mat = get_rotation_matrix(vec1=vec1, vec2=vec2)
    
    print(mat)
    
    vec1_rot = mat.dot(vec1)
    
    assert np.allclose(vec1_rot / np.linalg.norm(vec1_rot), vec2 / np.linalg.norm(vec2))

    quaternion_r = R.from_matrix(mat).as_quat()
    
    print(quaternion_r)
    '''
    
    '''
    mat = rotation_matrix_from_vectors(vec1=vec1, vec2=vec2)
    print(mat)
    vec1_rot = mat.dot(vec1)
    assert np.allclose(vec1_rot / np.linalg.norm(vec1_rot), vec2 / np.linalg.norm(vec2))
    '''
    ###################################################################
    #initialize parameters
    pt_diameter_max=pt_diameter_min=pt_length=pt_diameter=pt_eccentricity=pt_stem_diameter=0
        
    #load aligned ply point cloud file
    if not (filename_pcloud is None):
        
        model_pcloud = current_path + filename_pcloud
        
        print("Loading 3D point cloud {}...\n".format(filename_pcloud))
        
        model_pcloud_name_base = os.path.splitext(model_pcloud)[0]
        
        pcd = o3d.io.read_point_cloud(model_pcloud)
        
        Data_array_pcloud = np.asarray(pcd.points)
        
        #print(Data_array_pcloud.shape)
        
        if pcd.has_colors():
            
            print("Render colored point cloud\n")
            
            pcd_color = np.asarray(pcd.colors)
            
            if len(pcd_color) > 0: 
                
                pcd_color = np.rint(pcd_color * 255.0)
            
            #pcd_color = tuple(map(tuple, pcd_color))
        else:
            
            print("Generate random color\n")
        
            pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
            

 


        
        '''
        ################################################################
        
        print(idx_brace[0][0], idx_brace[0][-1])
        
        anchor_pt = (X_skeleton[25], Y_skeleton[25], Z_skeleton[25])
        
        pcd = o3d.geometry.PointCloud()
        
        pcd.points = o3d.utility.Vector3dVector(Data_array_pcloud)
        
        pcd.paint_uniform_color([0.5, 0.5, 0.5])

        
        # Build KDTree from point cloud for fast retrieval of nearest neighbors
        pcd_tree = o3d.geometry.KDTreeFlann(pcd)
        
        #print("Paint the 00th point red.")
        
        #pcd.colors[anchor_pt] = [1, 0, 0]
        
        search_radius = 150
        #print("Find its 50 nearest neighbors, paint blue.")
        
        [k, idx, _] = pcd_tree.search_knn_vector_3d(anchor_pt, search_radius)
        
        #print("nearest neighbors = {}\n".format(sorted(np.asarray(idx[1:]))))
        
        np.asarray(pcd.colors)[idx[1:], :] = [0, 0, 1]
        
        #o3d.visualization.draw_geometries([pcd])
        '''

    
    
    
    #Skeleton Visualization pipeline
    ####################################################################
    # The number of points per line
    
    if args["visualize_model"]:
    
        from mayavi import mlab
        from tvtk.api import tvtk

        N = 2
        
        mlab.figure("Structure_graph", size = (800, 800), bgcolor = (0, 0, 0))
        
        mlab.clf()
       
        # visualize 3d points
        #pts = mlab.points3d(X_skeleton[0], Y_skeleton[0], Z_skeleton[0], color = (0.58, 0.29, 0), mode = 'sphere', scale_factor = 0.15)
        #pts = mlab.points3d(X_skeleton[neighbors_idx], Y_skeleton[neighbors_idx], Z_skeleton[neighbors_idx], mode = 'sphere', color=(0,0,1), scale_factor = 0.05)
        #pts = mlab.points3d(X_skeleton[end_vlist_offset], Y_skeleton[end_vlist_offset], Z_skeleton[end_vlist_offset], color = (1,1,1), mode = 'sphere', scale_factor = 0.03)
        #pts = mlab.points3d(X_skeleton[sub_branch_start_rec_selected], Y_skeleton[sub_branch_start_rec_selected], Z_skeleton[sub_branch_start_rec_selected], color = (1,0,0), mode = 'sphere', scale_factor = 0.08)
        
        
        
        cmap = get_cmap(len(vlist_path_rec))

        for i, vlist_path in enumerate(vlist_path_rec):

            color_rgb = cmap(i)[:len(cmap(i))-1]

            pts = mlab.points3d(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path], color = color_rgb, mode = 'sphere', scale_factor = 0.05)
        
            #pts = mlab.plot3d(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path], color = color_rgb, tube_radius=0.025)
            mlab.text3d(X_skeleton[vlist_path[-1]], Y_skeleton[vlist_path[-1]], Z_skeleton[vlist_path[-1]], str("{:.0f}".format(i)), color = (0,1,0), scale = (0.04, 0.04, 0.04))
            
        
        #pts = mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0,1,1,1)) #xyzuvw
       
        #pts.glyph.glyph.clamping = False


        ###############################################################################
        # Display a semi-transparent sphere, for the surface of the Earth

        # We use a sphere Glyph, through the points3d mlab function, rather than
        # building the mesh ourselves, because it gives a better transparent
        # rendering.
        sphere = mlab.points3d(0, 0, 0, scale_mode='none',
                                scale_factor=2,
                                color=(0.67, 0.77, 0.93),
                                resolution=50,
                                opacity=0.7,
                                name='Earth')

        # These parameters, as well as the color, where tweaked through the GUI,
        # with the record mode to produce lines of code usable in a script.
        sphere.actor.property.specular = 0.45
        sphere.actor.property.specular_power = 5
        # Backface culling is necessary for more a beautiful transparent
        # rendering.
        sphere.actor.property.backface_culling = True

        for Vec in rotVec_rec:
            
            #print(Vec[0], Vec[1], Vec[2])
            
            mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0,Vec[0], Vec[1], Vec[2])) #xyzuvw
            
            #pts = mlab.points3d(Vec[0], Vec[1], Vec[2], color = (1,0,0), mode = 'sphere', scale_factor = 0.05)


        ###############################################################################
        # Plot the equator and the tropiques
        theta = np.linspace(0, 2 * np.pi, 100)
        for angle in (- np.pi / 6, 0, np.pi / 6):
            x = np.cos(theta) * np.cos(angle)
            y = np.sin(theta) * np.cos(angle)
            z = np.ones_like(theta) * np.sin(angle)

            mlab.plot3d(x, y, z, color=(1, 1, 1), opacity=0.2, tube_radius=None)

        mlab.view(63.4, 73.8, 4, [-0.05, 0, 0])
    
        mlab.orientation_axes()
        
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 1,0,0), color=(0,0,1))
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 0,1,0), color=(0,0,1))
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 0,0,1), color=(0,0,1))
        
        
        
        '''
        N_sublist = dsf_length_divide_idx

        cmap = get_cmap(N_sublist)

        #cmap = get_cmap(len(sub_branch_list))
        
        #draw all the sub branches in loop 
        for i, (sub_branch, sub_branch_start, sub_branch_radius) in enumerate(zip(sub_branch_level[0], sub_branch_start_rec, sub_branch_angle_rec)):

            if i < 50000:
            #if i <= dsf_length_divide_idx:
                
                color_rgb = cmap(i)[:len(cmap(i))-1]
                
                pts = mlab.points3d(X_skeleton[sub_branch], Y_skeleton[sub_branch], Z_skeleton[sub_branch], color = color_rgb, mode = 'sphere', scale_factor = 0.05)

                #mlab.text3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start]-0.05, str(i), color = (0,1,0), scale = (0.04, 0.04, 0.04))
        
                pts = mlab.points3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start], color = (1,1,1), mode = 'sphere', scale_factor = 0.06)
                
                mlab.text3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start]-0.05, str("{:.2f}".format(sub_branch_radius)), color = (0,1,0), scale = (0.04, 0.04, 0.04))
                
                #mlab.text3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start]-0.05, str("{:.2f}".format(Z_skeleton[sub_branch_start])), color = (0,1,0), scale = (0.04, 0.04, 0.04))
        '''
        '''
        N_sublist = 3
        
        cmap = get_cmap(N_sublist)
        
        for idx in range(N_sublist):
            
            color_rgb = cmap(idx)[:len(cmap(idx))-1]
            
            for i, (sub_branch, sub_branch_start, sub_branch_radius) in enumerate(zip(sub_branch_level[idx], sub_branch_start_level[idx], radius_level[idx])):

                pts = mlab.points3d(X_skeleton[sub_branch], Y_skeleton[sub_branch], Z_skeleton[sub_branch], color = color_rgb, mode = 'sphere', scale_factor = 0.05)

                pts = mlab.points3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start], color = (1,1,1), mode = 'sphere', scale_factor = 0.06)
                
                pts = mlab.text3d(X_skeleton[sub_branch_start], Y_skeleton[sub_branch_start], Z_skeleton[sub_branch_start]-0.05, str("{:.2f}".format(sub_branch_radius)), color = (0,1,0), scale = (0.04, 0.04, 0.04))

        '''
                
        '''
        for i, (end_val, x_e, y_e, z_e) in enumerate(zip(closest_pts_unique_sorted_combined, X_skeleton[closest_pts_unique_sorted_combined], Y_skeleton[closest_pts_unique_sorted_combined], Z_skeleton[closest_pts_unique_sorted_combined])):
            
            mlab.text3d(x_e, y_e, z_e, str(end_val), scale = (0.04, 0.04, 0.04))
        '''
        #pts = mlab.points3d(Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2], mode = 'point')
        
        mlab.show()
        
        '''
        #visualize point cloud model with color
        ####################################################################
        
        if not (filename_pcloud is None):
            
            x, y, z = Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2] 
            
            
            pts = mlab.points3d(x,y,z, mode = 'point')
            
            sc = tvtk.UnsignedCharArray()
            
            sc.from_array(pcd_color)

            pts.mlab_source.dataset.point_data.scalars = sc
            
            pts.mlab_source.dataset.modified()
            
        
        #visualize skeleton model, edge, nodes
        ####################################################################
        if args["visualize_model"]:
        
            x = list()
            y = list()
            z = list()
            s = list()
            connections = list()
            
            # The index of the current point in the total amount of points
            index = 0
            
            # Create each line one after the other in a loop
            for i in range(N_edges_skeleton):
            #for val in vlist_path:
                #if i in vertex_dominant:
                if True:
                    #i = int(val)
                    #print("Edges {0} has nodes {1}, {2}\n".format(i, array_edges[i][0], array_edges[i][1]))
                  
                    x.append(X_skeleton[array_edges_skeleton[i][0]])
                    y.append(Y_skeleton[array_edges_skeleton[i][0]])
                    z.append(Z_skeleton[array_edges_skeleton[i][0]])
                    
                    x.append(X_skeleton[array_edges_skeleton[i][1]])
                    y.append(Y_skeleton[array_edges_skeleton[i][1]])
                    z.append(Z_skeleton[array_edges_skeleton[i][1]])
                    
                    # The scalar parameter for each line
                    s.append(array_edges_skeleton[i])
                    
                    # This is the tricky part: in a line, each point is connected
                    # to the one following it. We have to express this with the indices
                    # of the final set of points once all lines have been combined
                    # together, this is why we need to keep track of the total number of
                    # points already created (index)
                    #connections.append(np.vstack(array_edges[i]).T)
                    
                    connections.append(np.vstack(
                                   [np.arange(index,   index + N - 1.5),
                                    np.arange(index + 1, index + N - .5)]
                                        ).T)
                    index += 2
            
            
            # Now collapse all positions, scalars and connections in big arrays
            x = np.hstack(x)
            y = np.hstack(y)
            z = np.hstack(z)
            s = np.hstack(s)
            #connections = np.vstack(connections)

            # Create the points
            src = mlab.pipeline.scalar_scatter(x, y, z, s)

            # Connect them
            src.mlab_source.dataset.lines = connections
            src.update()

            # display the set of lines
            mlab.pipeline.surface(src, colormap = 'Accent', line_width = 5, opacity = 0.7)

            # And choose a nice view
            #mlab.view(33.6, 106, 5.5, [0, 0, .05])
            #mlab.roll(125)
            mlab.show()
        
       '''
                
    '''
    return s_diameter_max, s_diameter_min, s_diameter, s_length, pt_eccentricity, avg_radius_stem, avg_density, \
        num_brace, avg_brace_length, avg_brace_angle, avg_radius_brace, avg_brace_projection,\
        num_crown, avg_crown_length, avg_crown_angle, avg_radius_crown, avg_crown_projection, \
        avg_radius_lateral, \
        n_whorl, whorl_dis_1, whorl_dis_2, avg_volume
    '''
    




if __name__ == '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "path to *.ply model file")
    ap.add_argument("-m1", "--model_skeleton", required = True, help = "skeleton file name")
    ap.add_argument("-m2", "--model_pcloud", required = False, default = None, help = "point cloud model file name, same path with ply model")
    ap.add_argument("-m3", "--slice_path", required = True, default = None, help = "Cross section/slices image folder path in ong format")
    ap.add_argument("-v", "--visualize_model", required = False, default = False, help = "Display model or not, deafult no")
    args = vars(ap.parse_args())



    # setting path to model file 
    current_path = args["path"]
    filename_skeleton = args["model_skeleton"]
    model_skeleton_name_base = os.path.splitext(current_path + filename_skeleton)[0]
    
    if args["model_pcloud"] is None:
        filename_pcloud = None
    else:
        filename_pcloud = args["model_pcloud"]
    
    # analysis result path
    print ("results_folder: " + current_path + "\n")
    
    
    
    
    #create label result file folder
    mkpath = os.path.dirname(current_path) +'/label'
    mkdir(mkpath)
    label_path = mkpath + '/'
    

    # slice image path
    filetype = '*.png'
    slice_image_path = args["slice_path"] + filetype


    print("Analyzing 3D skeleton and structure ...\n")
    
    # obtain image file list
    imgList = sorted(glob.glob(slice_image_path))

    n_images = len(imgList)
    
    if n_images == 0:
        print(f"Could not load image {slice_image_path}, skipping")
        exit()
    
    print("Processing {} slices from cross section of the 3d model\n".format(n_images))
    
    #loop all slices to obtain raidus results
    
    #print(avg_radius = crosssection_analysis_range(0, 97))
    
    
    analyze_skeleton(current_path, filename_skeleton, filename_pcloud)
    
    '''
    (s_diameter_max, s_diameter_min, s_diameter, s_length, pt_eccentricity, avg_radius_stem, avg_density,\
        num_brace, avg_brace_length, avg_brace_angle, avg_radius_brace, avg_brace_projection,\
        num_crown, avg_crown_length, avg_crown_angle, avg_radius_crown, avg_crown_projection, \
        avg_radius_lateral, \
        count_wholrs, whorl_dis_1, whorl_dis_2, avg_volume) = analyze_skeleton(current_path, filename_skeleton, filename_pcloud)
    
    
    
    trait_sum = []
    
    trait_sum.append([s_diameter_max, s_diameter_min, s_diameter, s_length, pt_eccentricity, avg_radius_stem, avg_density,\
        num_brace, avg_brace_length, avg_brace_angle, avg_radius_brace, avg_brace_projection,\
        num_crown, avg_crown_length, avg_crown_angle, avg_radius_crown, avg_crown_projection, \
        avg_radius_lateral, \
        count_wholrs, whorl_dis_1, whorl_dis_2, avg_volume])
    
    #save reuslt file
    ####################################################################
    

    trait_path = os.path.dirname(current_path + filename_skeleton)
    
    folder_name = os.path.basename(trait_path)
    
    #print("current_path folder ={}".format(folder_name))
    
    # create trait file using sub folder name
    trait_file = (current_path + folder_name + '_trait.xlsx')
    
    #trait_file_csv = (current_path + 'trait.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        
        
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(row = 1, column = 1).value = 'root system diameter max'
        sheet.cell(row = 1, column = 2).value = 'root system diameter min'
        sheet.cell(row = 1, column = 3).value = 'root system diameter'
        sheet.cell(row = 1, column = 4).value = 'root system length'
        sheet.cell(row = 1, column = 5).value = 'root system eccentricity'
        sheet.cell(row = 1, column = 6).value = 'stem root diameter'
        sheet.cell(row = 1, column = 7).value = 'root system density'
        sheet.cell(row = 1, column = 8).value = 'number of brace roots'
        sheet.cell(row = 1, column = 9).value = 'brace root length'
        sheet.cell(row = 1, column = 10).value = 'brace root angle'
        sheet.cell(row = 1, column = 11).value = 'brace root diameter'
        sheet.cell(row = 1, column = 12).value = 'brace root projection radius'
        sheet.cell(row = 1, column = 13).value = 'number of crown roots'
        sheet.cell(row = 1, column = 14).value = 'crown root length'
        sheet.cell(row = 1, column = 15).value = 'crown root angle'
        sheet.cell(row = 1, column = 16).value = 'crown root diameter'
        sheet.cell(row = 1, column = 17).value = 'crown root projection radius'
        sheet.cell(row = 1, column = 18).value = 'lateral root radius'
        sheet.cell(row = 1, column = 19).value = 'number of whorls'
        sheet.cell(row = 1, column = 20).value = 'whorl distance 1'
        sheet.cell(row = 1, column = 21).value = 'whorl distance 2'
        sheet.cell(row = 1, column = 22).value = 'root system volume'
              
        
    for row in trait_sum:
        sheet.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error saving Result file\n")

    '''
