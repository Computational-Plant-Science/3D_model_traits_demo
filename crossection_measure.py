"""
Version: 1.0
Summary: root object detection and simplification, parallel processing
Author: suxing liu
Author-email: suxingliu@gmail.com

USAGE

python3 crossection_measure.py -p ~/example/ -ext png -t 1


"""

# import necessary packages
from skimage.measure import regionprops
from skimage.morphology import convex_hull_image

import numpy as np
import argparse
import cv2
import sys, traceback
import os
import glob

import math

import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from openpyxl import load_workbook
from openpyxl import Workbook


def mkdir(path):
    # import module
    #import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end 
    path=path.rstrip("\\")
  
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        print (path + ' folder constructed!')
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False


# get average of a list 
def Average(lst): 
    return sum(lst) / len(lst)


def area_radius(area_of_circle):
    radius = ((area_of_circle/ math.pi)** 0.5)
    return radius 




def comp_external_contour(orig,thresh):
    
    img_height, img_width, img_channels = orig.shape
    
    #Convert image to grayscale, then apply Otsu's thresholding
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    convexhull = convex_hull_image(thresh)
    
    img_convexhull = np.uint8(convexhull)*255
    
    #Obtain the threshold image using OTSU adaptive filter
    thresh_hull = cv2.threshold(img_convexhull, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
   #find contours and get the external one
    image_result, contours, hier = cv2.findContours(img_convexhull, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    print("len(contours)")
    print(len(contours))
    
     # Measure properties 
    regions = regionprops(img_convexhull)

    #center location of region
    y_cvh, x_cvh = regions[0].centroid
    print("Convexhull center of root system: {0}, {1} \n".format(int(x_cvh),int(y_cvh)))
    
    convexhull_diameter = regions[0].equivalent_diameter 
    
    return img_convexhull,convexhull_diameter, y_cvh, x_cvh
    




# detect root convexhull area based on regionprops method
def root_area_label(image_file):
    
    path, filename = os.path.split(image_file)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    print("processing image : {0} \n".format(str(filename)))
    
    #result_img_path = file_path + str(filename[0:-4]) + '_lab.png'
       
    #print(result_img_path)
    
     # load the image and perform pyramid mean shift filtering to aid the thresholding step
    imgcolor = cv2.imread(image_file)
    
    # accquire image dimensions 
    height, width, channels = imgcolor.shape
    #shifted = cv2.pyrMeanShiftFiltering(image, 5, 5)

    #Image binarization by apltying otsu threshold
    img = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2GRAY)
    
    # Convert BGR to GRAY
    img_lab = cv2.cvtColor(imgcolor, cv2.COLOR_BGR2LAB)
    
    gray = cv2.cvtColor(img_lab, cv2.COLOR_BGR2GRAY)
    
     #Obtain the threshold image using OTSU adaptive filter
    thresh = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #Compute the gemetrical shape in convexhull
    (img_convexhull, convexhull_diameter, y_cvh, x_cvh) = comp_external_contour(imgcolor.copy(),thresh)
    
    #print("convexhull_diameter: {0} \n".format(convexhull_diameter))
    
    
    #Obtain the threshold image using OTSU adaptive filter
    ret, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    connectivity = 8
    
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary , connectivity , cv2.CV_32S)
   
    largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA]) 
    
    areas = [s[4] for s in stats]
    
    areas.remove(max(areas))
    
    radius = [area_radius(x) for x in areas]
    
    min_radius = 1
    
    radius = [x for x in radius if x > min_radius]
    
    #print(radius)
    #print(type(radius))
    
    sorted_idx_as = np.argsort(radius)
    
    sorted_idx_de = sorted_idx_as[::-1]
    
    #print(sorted(radius, reverse=True))
    
    radius_max = radius[sorted_idx_de[0]]

    
    if pattern_id == 1:
        
        radius = [x for x in radius if x == radius_max]
        
        num_primary_root = 1
        num_lateral_root = 0
        num_fine_root = 0
    
    elif (pattern_id == 2):
        
        num_primary_root = 1
        num_lateral_root = len(radius) - 1
        num_fine_root = 0

    elif (pattern_id == 3):
        
        #radius = [x for x in radius if x != radius_max]
        
        num_primary_root = 0
        num_lateral_root = 0
        num_fine_root = len(radius)

    
    print("num_primary_root is : {0}\n".format(str(num_primary_root)))
    print("num_lateral_root is : {0}\n".format(str(num_lateral_root)))
    print("num_fine_root is : {0}\n".format(str(num_fine_root)))
    
    return base_name, convexhull_diameter, len(radius), num_primary_root, num_lateral_root, num_fine_root, sorted(radius, reverse=True)
     
  
def parallel_root_area_label(images,save_path):
    
    # parallel processing 
    agents = multiprocessing.cpu_count()
    chunksize = 3
    
    with closing(Pool(processes = agents)) as pool:
        
        result = pool.map(root_area_label, images, chunksize)
        
        pool.terminate()
        
       
    base_name_rec = list(zip(*result)[0])
    convexhull_diameter_rec = list(zip(*result)[1])
    len_radius_rec = list(zip(*result)[2])
    num_primary_root_rec = list(zip(*result)[3])
    num_lateral_root_rec = list(zip(*result)[4])
    num_fine_root_rec = list(zip(*result)[5])
    radius_rec = list(zip(*result)[6])
    
   
    #print(radius_rec[0], len(radius_rec[0]))
    avr_len = Average(len_radius_rec)
    
    #print(avr_len)
    
    #print(radius_rec)
    #print(type(radius_rec))
    
    #write measured parameters as excel file 
    trait_file = (file_path + 'root_trait_' + str(pattern_id) + '.xlsx')
    
        
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'Root system diameter'
        sheet.cell(row = 1, column = 3).value = 'Number of roots'
        sheet.cell(row = 1, column = 4).value = 'Number of primary root'
        sheet.cell(row = 1, column = 5).value = 'Number of lateral root'
        sheet.cell(row = 1, column = 6).value = 'Number of fine root'
        sheet.cell(row = 1, column = 7).value = 'Each root radius'
    
    
    
    for idx, row in enumerate(radius_rec):
        
        row.insert(0, base_name_rec[idx])
        row.insert(1, convexhull_diameter_rec[idx])
        row.insert(2, len_radius_rec[idx])
        row.insert(3, num_primary_root_rec[idx])
        row.insert(4, num_lateral_root_rec[idx])
        row.insert(5, num_fine_root_rec[idx])
        
        sheet.append(row)

    #save the csv file
    wb.save(trait_file)
    
    


if __name__ == '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help="path to image file")
    ap.add_argument("-ext", "--extension", required = False, default = 'png', help = "extension name. default is 'png'.")  
    ap.add_argument("-t", "--type", required = True, type = int, help = "type")
    args = vars(ap.parse_args())


    # setting path to model file
    global file_path,pattern_id
    file_path = args["path"]
    ext = args['extension']
    pattern_id = args["type"]
    
    
    if not os.path.exists(file_path):
        #ignore if no such file is present.
        print("File does not exist!!!")
        #raise

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype

        
    images = sorted(glob.glob(image_file_path))
    
        
    parallel_root_area_label(images,file_path)
    

    
    
    
   

    


