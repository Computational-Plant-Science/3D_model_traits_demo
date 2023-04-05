"""
Version: 1.5

Summary: compute the Quaternions representation of 3d model graph

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE

#default parameter: python3 skeleton_graph.py -p ~/example/test/ -m1 test_skeleton.ply -m2 test_aligned.ply -v 1

#customized parameter: python3 skeleton_graph.py -p ~/example/test/ -m1 test_skeleton.ply -m2 test_aligned.ply -th 0.21 -v 1

#customized parameter: python3 skeleton_graph.py -p ~/example/pt_cloud/tiny/ -m1 tiny_skeleton.ply -v 1


argument:
("-p", "--path", required=True,    help="path to *.ply model file")
("-m", "--model", required=True,    help="file name")

"""
#!/usr/bin/env python

from mayavi import mlab
from tvtk.api import tvtk

# import the necessary packages
from plyfile import PlyData, PlyElement
import numpy as np 
from numpy import interp


from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
from sklearn.cluster import KMeans
from operator import itemgetter
import argparse
import kmeans1d

from skimage.segmentation import watershed
from skimage.feature import peak_local_max
from skimage.morphology import convex_hull_image
from skimage.measure import regionprops

from scipy.spatial import KDTree
from scipy import ndimage
from scipy.spatial.transform import Rotation as R

import random

import cv2

import glob
import os
import sys
import open3d as o3d
import copy
import shutil

import graph_tool.all as gt

from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import math
import itertools

#from tabulate import tabulate
from rdp import rdp

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import csv


import pandas as pd
import plotly
import plotly.graph_objs as go

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
from matplotlib import colors
from matplotlib.ticker import PercentFormatter


from pyquaternion import Quaternion
from pathlib import Path

'''
# import warnings filter
from warnings import simplefilter
# ignore all future warnings
simplefilter(action='ignore', category=FutureWarning)
'''

# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        #shutil.rmtree(path)
        #os.makedirs(path)
        return False




#calculate length of a 3D path or curve
def path_length(X, Y, Z):

    n = len(X)
     
    lv = [math.sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])**2) for i in range (1,n)]
    
    L = sum(lv)
    
    return L


#compute angle between two vectors(works for n-dimensional vector),
def dot_product_angle(v1,v2):

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        
        print("Zero magnitude vector!")
        
        return 0
        
    else:
        vector_dot_product = np.dot(v1,v2)
        
        arccos = np.arccos(vector_dot_product / (np.linalg.norm(v1) * np.linalg.norm(v2)))
        
        angle = np.degrees(arccos)
        
        #return angle
        
        
        if angle > 0 and angle < 45:
            return (90 - angle)
        elif angle < 90:
            return angle
        else:
            return (180- angle)
        
    
#find the closest points from a points sets to a fix point using Kdtree, O(log n) 
def closest_point(point_set, anchor_point):
    
    kdtree = KDTree(point_set)
    
    (d, i) = kdtree.query(anchor_point)
    
    #print("closest point:", point_set[i])
    
    return  i, point_set[i]


#colormap mapping
def get_cmap(n, name = 'hsv'):
    """get the color mapping""" 
    #viridis, BrBG, hsv, copper, Spectral
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    return plt.cm.get_cmap(name,n+1)


#compute muti-dimension distance between two points
def multiDimenDist(point1,point2):
   #find the difference between the two points, its really the same as below
   deltaVals = [point2[dimension]-point1[dimension] for dimension in range(len(point1))]
   runningSquared = 0
   #because the pythagarom theorm works for any dimension we can just use that
   for coOrd in deltaVals:
       runningSquared += coOrd**2
   return runningSquared**(1/2)
   

#compute two vectors form two points in muti-dimension space
def findVec(point1,point2,unitSphere = False):
  #setting unitSphere to True will make the vector scaled down to a sphere with a radius one, instead of it's orginal length
  finalVector = [0 for coOrd in point1]
  for dimension, coOrd in enumerate(point1):
      #finding total differnce for that co-ordinate(x,y,z...)
      deltaCoOrd = point2[dimension]-coOrd
      #adding total difference
      finalVector[dimension] = deltaCoOrd
  if unitSphere:
      totalDist = multiDimenDist(point1,point2)
      unitVector =[]
      for dimen in finalVector:
          unitVector.append( dimen/totalDist)
      return unitVector
  else:
      return finalVector


# normalize a list of data
def normalize(lst):
    
    #normlized_list = [float(i)/sum(lst) for i in lst]
    
    normlized_list = [float(i)/max(lst) for i in lst]
    
    return normlized_list
    

#get rotation matrix between two vectors using scipy
def get_rotation_matrix(vec2, vec1):
    
    vec1 = np.reshape(vec1, (1, -1))
    
    vec2 = np.reshape(vec2, (1, -1))
    
    r = R.align_vectors(vec2, vec1)
        
    return r[0].as_matrix()
    '''
    """ Find the rotation matrix that aligns vec1 to vec2
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return numpy.eye(3) #cross of all zeros only occurs on identical directions
    '''

    


#Find the rotation matrix that aligns vec1 to vec2
def rotation_matrix_from_vectors(vec1, vec2):
    """ 
    :param vec1: A 3d "source" vector
    :param vec2: A 3d "destination" vector
    :return mat: A transform matrix (3x3) which when applied to vec1, aligns it with vec2.
    """
    a, b = (vec1 / np.linalg.norm(vec1)).reshape(3), (vec2 / np.linalg.norm(vec2)).reshape(3)
    v = np.cross(a, b)
    if any(v): #if not all zeros then 
        c = np.dot(a, b)
        s = np.linalg.norm(v)
        kmat = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
        return np.eye(3) + kmat + kmat.dot(kmat) * ((1 - c) / (s ** 2))

    else:
        return np.eye(3) #cross of all zeros only occurs on identical directions


# Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
# The quaternions are arranged as (w,x,y,z), with w being the scalar
# The result will be the average quaternion of the input. Note that the signs
# of the output quaternion can be reversed, since q and -q describe the same orientation
def averageQuaternions(Q):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))

    for i in range(0,M):
        q = Q[i,:]
        # multiply q with its transposed version q' and add A
        A = np.outer(q,q) + A

    # scale
    A = (1.0/M)*A
    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)
    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]
    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0])
    return np.ravel(eigenVectors[:,0])


# Average multiple quaternions with specific weights
# The weight vector w must be of the same length as the number of rows in the
# quaternion maxtrix Q
def weightedAverageQuaternions(Q, w):
    # Number of quaternions to average
    M = Q.shape[0]
    A = np.zeros(shape=(4,4))
    weightSum = 0

    for i in range(0,M):
        q = Q[i,:]
        A = w[i] * np.outer(q,q) + A
        weightSum += w[i]

    # scale
    A = (1.0/weightSum) * A

    # compute eigenvalues and -vectors
    eigenValues, eigenVectors = np.linalg.eig(A)

    # Sort by largest eigenvalue
    eigenVectors = eigenVectors[:,eigenValues.argsort()[::-1]]

    # return the real part of the largest eigenvector (has only real part)
    #return np.real(eigenVectors[:,0].A1)
    return np.ravel(eigenVectors[:,0])


def euler_to_rotMat(yaw, pitch, roll):
    Rz_yaw = np.array([
        [np.cos(yaw), -np.sin(yaw), 0],
        [np.sin(yaw),  np.cos(yaw), 0],
        [          0,            0, 1]])
    Ry_pitch = np.array([
        [ np.cos(pitch), 0, np.sin(pitch)],
        [             0, 1,             0],
        [-np.sin(pitch), 0, np.cos(pitch)]])
    Rx_roll = np.array([
        [1,            0,             0],
        [0, np.cos(roll), -np.sin(roll)],
        [0, np.sin(roll),  np.cos(roll)]])
    # R = RzRyRx
    rotMat = np.dot(Rz_yaw, np.dot(Ry_pitch, Rx_roll))
    
    return rotMat


# RPY/Euler angles to Rotation Vector
def euler_to_rotVec(yaw, pitch, roll):

    # compute the rotation matrix
    Rmat = euler_to_rotMat(yaw, pitch, roll)
    
    theta = math.acos(((Rmat[0, 0] + Rmat[1, 1] + Rmat[2, 2]) - 1) / 2)
    sin_theta = math.sin(theta)
    if sin_theta == 0:
        rx, ry, rz = 0.0, 0.0, 0.0
    else:
        multi = 1 / (2 * math.sin(theta))
        rx = multi * (Rmat[2, 1] - Rmat[1, 2]) * theta
        ry = multi * (Rmat[0, 2] - Rmat[2, 0]) * theta
        rz = multi * (Rmat[1, 0] - Rmat[0, 1]) * theta
    return rx, ry, rz



#find shortest path between start and end vertex
def short_path_finder(G_unordered, start_v, end_v):
    
    #find shortest path between start and end vertex
    ####################################################################
    
    #define start and end vertex index
    #start_v = 0
    #end_v = 1559
    #end_v = 608
    
    
    #print(X_skeleton[start_v], Y_skeleton[start_v], Z_skeleton[start_v])
    
    # find shortest path in the graph between start and end vertices 
    
    #vlist, elist = gt.all_shortest_paths(G_unordered, G_unordered.vertex(start_v), G_unordered.vertex(end_v))
    
    vlist, elist = gt.shortest_path(G_unordered, G_unordered.vertex(start_v), G_unordered.vertex(end_v))
    
    vlist_path = [str(v) for v in vlist]
    
    elist_path = [str(e) for e in elist]
    
    #change format form str to int
    int_vlist_path = [int(i) for i in vlist_path]
    
    #print(int_vlist_path)
    
    if len(vlist_path) > 0: 
        
        return int_vlist_path
    else:
        print("No shortest path found in graph...\n")
        
        return 0


def dimension_size(Z_skeleton, sub_branch_start_rec, sub_branch_end_rec, select):
    
    loc_start = [Z_skeleton[index] for index in sub_branch_start_rec]
    loc_end = [Z_skeleton[index] for index in sub_branch_end_rec]
    
    print("Z_loc_start max = {} min = {}".format(max(loc_start), min(loc_start)))
    print("Z_loc_end max = {} min = {}\n".format(max(loc_end), min(loc_end)))
    
    max_dimension_length = abs(max(max(loc_start), max(loc_end) - min(min(loc_start), min(loc_end))))

    min_dimension_length = abs(min(min(loc_start), min(loc_end) - max(max(loc_start), max(loc_end))))
    
    #print("max_length = {}\n".format(max_length))
    
    if select == 1:
        return max_dimension_length
    else:
        return min_dimension_length



def calculate_wcss(data):
    wcss = []
    for n in range(2, 10):
        kmeans = KMeans(n_clusters=n)
        kmeans.fit(X=data)
        wcss.append(kmeans.inertia_)

    return wcss



def optimal_number_of_clusters(wcss):
    x1, y1 = 2, wcss[0]
    x2, y2 = 20, wcss[len(wcss)-1]

    distances = []
    for i in range(len(wcss)):
        x0 = i+2
        y0 = wcss[i]
        numerator = abs((y2-y1)*x0 - (x2-x1)*y0 + x2*y1 - y2*x1)
        denominator = math.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        distances.append(numerator/denominator)

    return distances.index(max(distances)) + 0



def his_plot(path_length_rec, current_path, filename_skeleton):
    
  
    legend = ['Path length histogram distribution']

    N_points = len(path_length_rec)
    x = path_length_rec
    n_bins = 20
    
    
    # Creating histogram
    fig, axs = plt.subplots(1, 1,
                        figsize =(10, 7),
                        tight_layout = True)


    # Remove axes splines
    for s in ['top', 'bottom', 'left', 'right']:
        axs.spines[s].set_visible(False)

    # Remove x, y ticks
    axs.xaxis.set_ticks_position('none')
    axs.yaxis.set_ticks_position('none')

    # Add padding between axes and labels
    axs.xaxis.set_tick_params(pad = 5)
    axs.yaxis.set_tick_params(pad = 10)

    # Add x, y gridlines
    axs.grid(visible = True, color ='grey',
        linestyle ='-.', linewidth = 0.5,
        alpha = 0.6)

    bin_size = 0.1
    min_edge = 0.0
    max_edge = 1.0
    Nplus1 = (max_edge-min_edge)/bin_size + 1
    bin_list = np.linspace(min_edge, max_edge, int(Nplus1))


    #fixed_bins = [0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0]
    
    # Creating histogram
    #N, bins, patches = axs.hist(x, bins = n_bins)
    
    N, bins, patches = axs.hist(x, bins = bin_list)
    
    axs.set_ylim([0, 60000])

    # Setting color
    fracs = ((N**(1 / 5)) / N.max())
    norm = colors.Normalize(fracs.min(), fracs.max())

    for thisfrac, thispatch in zip(fracs, patches):
        color = plt.cm.viridis(norm(thisfrac))
        thispatch.set_facecolor(color)

    # Adding extra features   
    plt.xlabel("Path length in 3D model space")
    plt.ylabel("Counts")
    plt.legend(legend)
    plt.title('Path length distribution')
    
    
    trait_path = os.path.dirname(current_path + filename_skeleton)
    folder_name = os.path.basename(trait_path)
    
    # create trait file using sub folder name
    path_histogram = (current_path + folder_name + '_his.png')
    
    plt.savefig(path_histogram)



#cluster 1D list using Kmeans
def cluster_list(list_array, n_clusters):
    
    data = np.array(list_array)
    
    if data.ndim == 1:
        
        data = data.reshape(-1,1)
    
    #kmeans = KMeans(n_clusters).fit(data.reshape(-1,1))
        
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(data)
    
    #kmeans = KMeans(n_clusters).fit(data)
    
    labels = kmeans.labels_
    
    centers = kmeans.cluster_centers_
    
    center_labels = kmeans.predict(centers)
    
    #print(kmeans.cluster_centers_)

    return labels, centers, center_labels




# Skeleton analysis
def analyze_skeleton(current_path, filename_skeleton, filename_pcloud):
    
    model_skeleton = current_path + filename_skeleton
    print("Loading 3D skeleton file {}...\n".format(filename_skeleton))
    model_skeleton_name_base = os.path.splitext(model_skeleton)[0]
    
    #load the ply format skeleton file 
    try:
        with open(model_skeleton, 'rb') as f:
            plydata_skeleton = PlyData.read(f)
            num_vertex_skeleton = plydata_skeleton.elements[0].count
            N_edges_skeleton = len(plydata_skeleton['edge'].data['vertex_indices'])
            array_edges_skeleton = plydata_skeleton['edge'].data['vertex_indices']
            
            print("Ply data structure: \n")
            #print(plydata_skeleton)
            #print("\n")
            print("Number of 3D points in skeleton model: {0} \n".format(num_vertex_skeleton))
            print("Number of edges: {0} \n".format(N_edges_skeleton))

        
    except:
        sys.exit("Model skeleton file does not exist!")
    
    
    #Parse ply format skeleton file and Extract the data
    Data_array_skeleton = np.zeros((num_vertex_skeleton, 3))
    
    Data_array_skeleton[:,0] = plydata_skeleton['vertex'].data['x']
    Data_array_skeleton[:,1] = plydata_skeleton['vertex'].data['y']
    Data_array_skeleton[:,2] = plydata_skeleton['vertex'].data['z']
    
    X_skeleton = Data_array_skeleton[:,0]
    Y_skeleton = Data_array_skeleton[:,1]
    Z_skeleton = Data_array_skeleton[:,2]
    
    '''
    #load radius values
    ####################################################################
    #radius_skeleton = current_path + filename_skeleton
    
    base_name = os.path.splitext(os.path.basename(model_skeleton_name_base))[0]
    txt_base_name = base_name.replace("_skeleton", "_avr.txt")
    radius_file = current_path + txt_base_name

    print("Loading 3D skeleton radius txt file {}...\n".format(radius_file))
    
    #check file exits
    if os.path.isfile(radius_file):
        
        with open(radius_file) as file:
            lines = file.readlines()
            radius_vtx = [line.rstrip() for line in lines]
    else:
        
        sys.exit("Could not load 3D skeleton radius txt file")  
    
    '''
    # build directed graph from skeleton/structure data
    ####################################################################
    print("Building directed graph from 3D skeleton/structure ...\n")
    G_unordered = gt.Graph(directed = True)
    
    # assert directed graph
    #print(G.is_directed())
    
    nodes = G_unordered.add_vertex(num_vertex_skeleton)
    
    G_unordered.add_edge_list(array_edges_skeleton.tolist()) 
    
    #gt.graph_draw(G_unordered, vertex_text = G_unordered.vertex_index, output = current_path + "graph_view.pdf")
    
    
    # find all end vertices by fast iteration of all vertices
    end_vlist = []
    
    end_vlist_offset = []
    
    for v in G_unordered.iter_vertices():
        
        #print(G.vertex(v).out_degree(), G.vertex(v).in_degree())
        
        if G_unordered.vertex(v).out_degree() == 0 and G_unordered.vertex(v).in_degree() == 1:
        
            end_vlist.append(v)
            
            if (v+1) == num_vertex_skeleton:
                end_vlist_offset.append(v)
            else:
                end_vlist_offset.append(v+1)
            
    #print("end_vlist = {} \n".format(end_vlist))
    #print("end_vlist_offset = {} \n".format(end_vlist_offset))
    

    #test angle calculation
    #vector1 = [0,0,1]
    # [1,0,0]
    #vector2 = [0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1]
    #print(dot_product_angle(vector1,vector2))
    #print(cart2sph(0-math.sqrt(2)/2, 0-math.sqrt(2)/2, 1)) 

    
    # parse all the sub branches edges and vetices, start, end vetices
    #####################################################################################
    sub_branch_list = []
    sub_branch_length_rec = []
    sub_branch_angle_rec = []
    sub_branch_start_rec = []
    sub_branch_end_rec = []
    sub_branch_projection_rec = []
    sub_branch_radius_rec = []
    
    sub_branch_xs_rec = []
    sub_branch_ys_rec = []
    sub_branch_zs_rec = []
    
    #if len(end_vlist) == len(end_vlist_offset):
        
    for idx, v_end in enumerate(end_vlist):
        
        #print(idx, v_end)
        #construct list of vertices in sub branches
        if idx == 0:
            v_list = [*range(0, int(end_vlist[idx])+1)]
        else:
            v_list = [*range(int(end_vlist[idx-1])+1, int(end_vlist[idx])+1)]
            
        # change type to interger 
        int_v_list = [int(i) for i in v_list]
        
        # current sub branch length
        sub_branch_length = path_length(X_skeleton[int_v_list], Y_skeleton[int_v_list], Z_skeleton[int_v_list])
        
        # current sub branch start and end points 
        start_v = [X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]]
        
        end_v = [X_skeleton[int_v_list[0]] - X_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Y_skeleton[int_v_list[0]] - Y_skeleton[int_v_list[int(len(int_v_list)-1.0)]], Z_skeleton[int_v_list[0]] - Z_skeleton[int_v_list[int(len(int_v_list)-1.0)]]]
        
        # angle of current branch vs Z direction
        angle_sub_branch = dot_product_angle(start_v, end_v)
        
        # projection radius of current branch length
        p0 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[0]]])
        
        p1 = np.array([X_skeleton[int_v_list[0]], Y_skeleton[int_v_list[0]], Z_skeleton[int_v_list[-1]]])
        
        projection_radius = np.linalg.norm(p0 - p1)
        
        #radius value from fitted point cloud contours
        #radius_edge = float(radius_vtx[int_v_list[0]])*1
        
        sub_branch_xs = X_skeleton[int_v_list[0]]
        sub_branch_ys = Y_skeleton[int_v_list[0]]
        sub_branch_zs = Z_skeleton[int_v_list[0]]
        
        
        # save computed parameters for each branch
        sub_branch_list.append(v_list)
        sub_branch_length_rec.append(sub_branch_length)
        sub_branch_angle_rec.append(angle_sub_branch)
        sub_branch_start_rec.append(int_v_list[0])
        sub_branch_end_rec.append(int_v_list[-1])
        sub_branch_projection_rec.append(projection_radius)
        #sub_branch_radius_rec.append(radius_edge)
        
        sub_branch_xs_rec.append(sub_branch_xs)
        sub_branch_ys_rec.append(sub_branch_ys)
        sub_branch_zs_rec.append(sub_branch_zs)
    
    #print(min(sub_branch_angle_rec))
    #print(max(sub_branch_angle_rec))
    

    # sort branches according to length feature in descending order
    ####################################################################
    sorted_idx_len = np.argsort(sub_branch_length_rec)
    
    #reverse the order from accending to descending
    sorted_idx_len_loc = sorted_idx_len[::-1]

    #print("Z_loc = {}\n".format(sorted_idx_Z_loc))
    
    #sort all lists according to sorted_idx_Z_loc order
    sub_branch_list[:] = [sub_branch_list[i] for i in sorted_idx_len_loc] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in sorted_idx_len_loc]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in sorted_idx_len_loc]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in sorted_idx_len_loc]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in sorted_idx_len_loc]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in sorted_idx_len_loc]
    #sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in sorted_idx_len_loc]
    
    sub_branch_xs_rec[:] = [sub_branch_xs_rec[i] for i in sorted_idx_len_loc]
    sub_branch_ys_rec[:] = [sub_branch_ys_rec[i] for i in sorted_idx_len_loc]
    sub_branch_zs_rec[:] = [sub_branch_zs_rec[i] for i in sorted_idx_len_loc]

    
    print("number of sub_branch_end_rec is: {} \n".format(len(sub_branch_end_rec)))
    
    
    
    
    # filter sub branches with dominant length threshold
    ####################################################################
    cluster_number = 4
    
    #(labels, centers, center_labels) = cluster_list(sub_branch_length_rec, n_clusters = cluster_number)
    
    (clusters, centroids) = kmeans1d.cluster(sub_branch_length_rec, cluster_number)

    print(sub_branch_length_rec)
    print(clusters)   # [1, 1, 1, 0, 3, 3, 3, 2, 2, 2]
    print(centroids)  # [-50.0, 4.1, 94.0, 200.5]

    
    
    #sorted_idx = np.argsort(centers[:,0])[::-1]
    
    #print(sorted_idx)
    
    max_cluster_index = centroids.index(max(centroids))
    value_keep_index =  [idx for idx, value in enumerate(clusters) if value == max_cluster_index]

    #print(value_keep_index)
    
    print(len(sub_branch_list))

    sub_branch_list[:] = [sub_branch_list[i] for i in value_keep_index] 
    sub_branch_length_rec[:] = [sub_branch_length_rec[i] for i in value_keep_index]
    sub_branch_angle_rec[:] = [sub_branch_angle_rec[i] for i in value_keep_index]
    sub_branch_start_rec[:] = [sub_branch_start_rec[i] for i in value_keep_index]
    sub_branch_end_rec[:] = [sub_branch_end_rec[i] for i in value_keep_index]
    sub_branch_projection_rec[:] = [sub_branch_projection_rec[i] for i in value_keep_index]
    #sub_branch_radius_rec[:] = [sub_branch_radius_rec[i] for i in value_keep_index]
    
    ################################################################################################################################
    print(len(sub_branch_list))
    
    
    # graph: find closest point pairs and connect close graph edges
    ####################################################################
    print("Converting skeleton to graph and connecting edges and vertices...\n")
    
    v_closest_pair_rec = []
    
    closest_pts = []
    
    #find closest point set and connect graph edges
    for idx, (sub_branch, anchor_point) in enumerate(zip(sub_branch_list, end_vlist_offset)):
        
        # start vertex of an edge
        anchor_point = (X_skeleton[end_vlist_offset[idx]], Y_skeleton[end_vlist_offset[idx]], Z_skeleton[end_vlist_offset[idx]])

        # curve of the edge in 3D
        point_set = np.zeros((len(sub_branch_list[0]), 3))
        
        point_set[:,0] = X_skeleton[sub_branch_list[0]]
        point_set[:,1] = Y_skeleton[sub_branch_list[0]]
        point_set[:,2] = Z_skeleton[sub_branch_list[0]]
        
        (index_cp, value_cp) = closest_point(point_set, anchor_point)

        v_closest_pair = [index_cp, end_vlist_offset[idx]]

        dis_v_closest_pair = path_length(X_skeleton[v_closest_pair], Y_skeleton[v_closest_pair], Z_skeleton[v_closest_pair])
        
        #small threshold indicating close pair vetices
        if dis_v_closest_pair < thresh_join:
            
            closest_pts.append(index_cp)
            
            #print("dis_v_closest_pair = {}".format(dis_v_closest_pair))
            v_closest_pair_rec.append(v_closest_pair)
            
            #print("closest point pair: {0}".format(v_closest_pair))
            
            #connect graph edges
            G_unordered.add_edge(index_cp, end_vlist_offset[idx])
            
    #print("v_closest_pair_rec = {}\n".format(v_closest_pair_rec))
    
    #get the unique values from the list
    #closest_pts_unique = list(set(closest_pts))
    
    #keep repeat values for correct indexing order
    
    #print("closest_pts = {}\n".format(closest_pts))
    
    closest_pts_unique = list((closest_pts))
    
    closest_pts_unique_sorted = sorted(closest_pts_unique)
    
    #print("closest_pts_unique_sorted = {}\n".format(closest_pts_unique_sorted))
 
    #find shortest path between start and end vertex
    ####################################################################
    
    #define start and end vertex index
    start_v = 0
    
    vlist_path_rec = []
    
    quaternion_path_rec = []
    
    rotVec_rec = []
    
    path_length_rec = []
    

    

    # loop over all paths and compute quaternions values
    for idx, end_v in enumerate(sub_branch_end_rec):
        
        #print("start_v = {} end_v = {} \n".format(start_v, end_v))
   
        vlist_path = short_path_finder(G_unordered, start_v, end_v)
        
        if len(vlist_path) > 0:
        
            #path_count+=1
            #print("Shortest path found in graph! \n")
            
            vlist_path_rec.append(vlist_path)
            
            # record current path length
            path_length_N2N = path_length(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path])
            
            path_length_rec.append(path_length_N2N)
            
            # Q is a Nx4 numpy matrix and contains the quaternions to average in the rows.
            # The quaternions are arranged as (w,x,y,z), with w being the scalar
            sum_quaternion = np.zeros([len(vlist_path), 4])
            
            #sum_euler = np.zeros([len(vlist_path), 3])
            
            for i, v_path in enumerate(vlist_path):
                
                #print("closest_pts = {}\n".format(vlist_path[i]))
        
                if i + 2 < len(vlist_path):
                    
                    # get adjacent vector coordinates
                    vector1 = [X_skeleton[vlist_path[i]], Y_skeleton[vlist_path[i]], Z_skeleton[vlist_path[i]]]
                    vector2 = [X_skeleton[vlist_path[i + 1]], Y_skeleton[vlist_path[i + 1]], Z_skeleton[vlist_path[i + 1]]]
                    vector3 = [X_skeleton[vlist_path[i + 2]], Y_skeleton[vlist_path[i + 2]], Z_skeleton[vlist_path[i + 2]]]
            
                    # get adjacent directed vectors
                    vector_12 = findVec(vector1,vector2)
                    vector_23 = findVec(vector2,vector3)
                    
                    # compoute rotation matrix between adjacent directed vectors
                    mat = get_rotation_matrix(vec1 = vector_12, vec2 = vector_23)
                    
                    # compoute quaternion between adjacent directed vectors
                    #The returned quaternion value is in scalar-last (x, y, z, w) format.
                    quaternion_r = R.from_matrix(mat).as_quat()
                    
                    # change the order of the quaternion_r value from (x, y, z, w)  to (w, x, y, z)
                    quaternion_r_rearanged = np.hstack((quaternion_r[3], quaternion_r[0], quaternion_r[1], quaternion_r[2]))
                    
                    #euler_r = R.from_matrix(mat).as_euler('xyz', degrees = True)
                                       
                    sum_quaternion[i,:] = quaternion_r_rearanged
                    
                    #sum_euler[i,:] = euler_r
                    
                    #print("vlist_path = {} quaternion_r = {}".format(idx, quaternion_r))
                    
            
            # use eigenvalues to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
            #avg_quaternion = averageQuaternions(sum_quaternion)

            # use components averaging to compute average of quaternions, The quaternions input are arranged as (w,x,y,z),
            avg_quaternion = ((sum_quaternion.sum(axis=0))/len(vlist_path)).flatten()
            #avg_quaternion = avg_quaternion.flatten()

            rot = R.from_quat(avg_quaternion)
            
            avg_euler = rot.as_euler('xyz')
            
            #avg_euler = np.mean(sum_euler, axis = 0)
            
            rotVec = euler_to_rotVec(avg_euler[0], avg_euler[1], avg_euler[2])
            
            rotVec_rec.append(rotVec)

            quaternion_path_rec.append(avg_quaternion)
            
            #print("vlist_path = {} avg_quaternion = {} avg_euler = {} rotVec = {}\n".format(idx, avg_quaternion, avg_euler, rotVec))
                
            '''
            if rotVec[2] > 0:
                
                rotVec_rec.append(rotVec)
                
                quaternion_path_rec.append(avg_quaternion)
            '''

            
    print("Found {} shortest path \n".format(len(vlist_path_rec)))
    
    #print("Path length are: {}\n".format(path_length_rec)) 
    
    
   
    

    
    ####################################################
    
    
    path_length_rec = sub_branch_length_rec
    
    #normalize path length againste the max value
    #path_length_rec_normalize = normalize(path_length_rec)
    
    path_length_rec_normalize = path_length_rec
    
    # generate histogram
    his_plot(path_length_rec_normalize, current_path, filename_skeleton)
    
    
    path_index = list(range(1,len(vlist_path_rec)+1))
    

    ###################################################################
    #initialize parameters
    pt_diameter_max=pt_diameter_min=pt_length=pt_diameter=pt_eccentricity=pt_stem_diameter=0
        
    #load aligned ply point cloud file
    if not (filename_pcloud is None):
        
        model_pcloud = current_path + filename_pcloud
        
        print("Loading 3D point cloud {}...\n".format(filename_pcloud))
        
        model_pcloud_name_base = os.path.splitext(model_pcloud)[0]
        
        
        pcd = o3d.io.read_point_cloud(model_pcloud)
        
        Data_array_pcloud = np.asarray(pcd.points)
        
        #print(Data_array_pcloud.shape)
        
        if pcd.has_colors():
            
            print("Render colored point cloud\n")
            
            pcd_color = np.asarray(pcd.colors)
            
            if len(pcd_color) > 0: 
                
                pcd_color = np.rint(pcd_color * 255.0)
            
            #pcd_color = tuple(map(tuple, pcd_color))
        else:
            
            print("Generate random color\n")
        
            pcd_color = np.random.randint(256, size = (len(Data_array_pcloud),3))
        
    
    
    #Skeleton Visualization pipeline
    ####################################################################
    # The number of points per line
    
    if args["visualize_model"]:
    
        N = 2
        
        mlab.figure("Structure_graph", size = (800, 800), bgcolor = (0, 0, 0))
        
        mlab.clf()
        
        
        #visualize paths found between root node and end notes
        ############################################################################
        
        # visualize 3d points
        
        cmap = get_cmap(len(vlist_path_rec))

        for i, vlist_path in enumerate(vlist_path_rec):

            color_rgb = cmap(i)[:len(cmap(i))-1]
            
            if i == 0: 
                scale_factor_value = 0.075
            else:
                scale_factor_value = 0.01
            
            pts = mlab.points3d(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path], color = color_rgb, mode = 'sphere', scale_factor = scale_factor_value)
            

             #pts = mlab.plot3d(X_skeleton[vlist_path], Y_skeleton[vlist_path], Z_skeleton[vlist_path], color = color_rgb, tube_radius=0.025)
            #mlab.text3d(X_skeleton[vlist_path[-1]], Y_skeleton[vlist_path[-1]], Z_skeleton[vlist_path[-1]], str("{:.0f}".format(i)), color = (0,1,0), scale = (0.04, 0.04, 0.04))
            
        
        
        
        
        #pts = mlab.points3d(Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2], mode = 'point')
        ################################################################################################################

        
        
        #visualize point cloud model with color
        ####################################################################
        
        if not (filename_pcloud is None):
            
            x, y, z = Data_array_pcloud[:,0], Data_array_pcloud[:,1], Data_array_pcloud[:,2] 
            
            
            pts = mlab.points3d(x,y,z, mode = 'point')
            
            sc = tvtk.UnsignedCharArray()
            
            sc.from_array(pcd_color)

            pts.mlab_source.dataset.point_data.scalars = sc
            
            pts.mlab_source.dataset.modified()
            
        
        #visualize skeleton model, edge, nodes
        ####################################################################
        if args["visualize_model"]:
        
            x = list()
            y = list()
            z = list()
            s = list()
            connections = list()
            
            # The index of the current point in the total amount of points
            index = 0
            
            # Create each line one after the other in a loop
            for i in range(N_edges_skeleton):
            #for val in vlist_path:
                #if i in vertex_dominant:
                if True:
                    #i = int(val)
                    #print("Edges {0} has nodes {1}, {2}\n".format(i, array_edges[i][0], array_edges[i][1]))
                  
                    x.append(X_skeleton[array_edges_skeleton[i][0]])
                    y.append(Y_skeleton[array_edges_skeleton[i][0]])
                    z.append(Z_skeleton[array_edges_skeleton[i][0]])
                    
                    x.append(X_skeleton[array_edges_skeleton[i][1]])
                    y.append(Y_skeleton[array_edges_skeleton[i][1]])
                    z.append(Z_skeleton[array_edges_skeleton[i][1]])
                    
                    # The scalar parameter for each line
                    s.append(array_edges_skeleton[i])
                    
                    # This is the tricky part: in a line, each point is connected
                    # to the one following it. We have to express this with the indices
                    # of the final set of points once all lines have been combined
                    # together, this is why we need to keep track of the total number of
                    # points already created (index)
                    #connections.append(np.vstack(array_edges[i]).T)
                    
                    connections.append(np.vstack(
                                   [np.arange(index,   index + N - 1.5),
                                    np.arange(index + 1, index + N - .5)]
                                        ).T)
                    index += 2
            
            
            # Now collapse all positions, scalars and connections in big arrays
            x = np.hstack(x)
            y = np.hstack(y)
            z = np.hstack(z)
            s = np.hstack(s)
            #connections = np.vstack(connections)

            # Create the points
            src = mlab.pipeline.scalar_scatter(x, y, z, s)

            # Connect them
            src.mlab_source.dataset.lines = connections
            src.update()

            # display the set of lines
            mlab.pipeline.surface(src, colormap = 'Accent', line_width = 5, opacity = 0.7)

            # And choose a nice view
            #mlab.view(33.6, 106, 5.5, [0, 0, .05])
            #mlab.roll(125)
            #mlab.show()
        
        
       
        ###############################################################################
        # Display a semi-transparent sphere, for the surface of the Earth
        
        mlab.figure("sphere_representation_rotation_vector", size = (800, 800), bgcolor = (0, 0, 0))
        
        # We use a sphere Glyph, through the points3d mlab function, rather than
        # building the mesh ourselves, because it gives a better transparent
        # rendering.
        sphere = mlab.points3d(0, 0, 0, scale_mode='none',
                                scale_factor=2,
                                color=(0.67, 0.77, 0.93),
                                resolution=50,
                                opacity=0.7,
                                name='Earth')

        # These parameters, as well as the color, where tweaked through the GUI,
        # with the record mode to produce lines of code usable in a script.
        sphere.actor.property.specular = 0.45
        sphere.actor.property.specular_power = 5
        # Backface culling is necessary for more a beautiful transparent
        # rendering.
        sphere.actor.property.backface_culling = True

        #scalars = np.random.randint(1, size = (len(rotVec_rec),3))
        
        for idx, Vec in enumerate(rotVec_rec):
            
            #print(Vec[0], Vec[1], Vec[2])
            
            #mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, Vec[0], Vec[1], Vec[2], )) #xyz
            
            mlab.quiver3d(0,0,0, Vec[0], Vec[1], Vec[2], color=(1, 0, 0)) #xyz
            
            #pts = mlab.points3d(Vec[0], Vec[1], Vec[2], color = (1,0,0), mode = 'sphere', scale_factor = 0.05)
            
           

        
        ###############################################################################
        # Plot the equator and the tropiques
        theta = np.linspace(0, 2 * np.pi, 100)
        for angle in (- np.pi / 6, 0, np.pi / 6):
            x = np.cos(theta) * np.cos(angle)
            y = np.sin(theta) * np.cos(angle)
            z = np.ones_like(theta) * np.sin(angle)

            pts = mlab.plot3d(x, y, z, color=(1, 1, 1), opacity=0.2, tube_radius=None)

        mlab.view(63.4, 73.8, 4, [-0.05, 0, 0])
    
        mlab.orientation_axes()
        
        
        '''
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 1,0,0), color=(0,0,1))
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 0,1,0), color=(0,0,1))
        mlab.pipeline.vectors(mlab.pipeline.vector_scatter(0,0,0, 0,0,1), color=(0,0,1))
        '''
        #################################################################################
        
        
        mlab.show()
                

    
    return path_index, quaternion_path_rec, rotVec_rec
    
    
    




if __name__ == '__main__':
    
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help = "path to *.ply model file")
    ap.add_argument("-m1", "--model_skeleton", required = True, help = "skeleton file name")
    ap.add_argument("-m2", "--model_pcloud", required = False, default = None, help = "point cloud model file name, same path with ply model")
    ap.add_argument("-th", "--thresh_join", required = False, type = float, default = 3.21, help = "threshhold value to join all disconnected graph nodes")
    ap.add_argument("-v", "--visualize_model", required = False, type = int, default = 0, help = "Display model or not, deafult not display")
    args = vars(ap.parse_args())



    # setting path to model file 
    current_path = args["path"]
    filename_skeleton = args["model_skeleton"]
    model_skeleton_name_base = os.path.splitext(current_path + filename_skeleton)[0]
    
    thresh_join = args["thresh_join"]
    
    if args["model_pcloud"] is None:
        filename_pcloud = None
    else:
        filename_pcloud = args["model_pcloud"]
    
    # analysis result path
    print ("results_folder: " + current_path + "\n")
    
    
    
    result_list = []
    
    (path_index, quaternion_path_rec, rotVec_rec) = analyze_skeleton(current_path, filename_skeleton, filename_pcloud)
    
    rotVec_rec_arr = np.vstack(rotVec_rec)
    quaternion_path_arr = np.vstack(quaternion_path_rec)
    
    #print(rotVec_rec)
    
    #print((quaternion_path_arr.shape))
    
    for i, (v0,v1,v2,v3,v4,v5,v6,v7) in enumerate(zip(path_index, quaternion_path_arr[:,0], quaternion_path_arr[:,1], quaternion_path_arr[:,2], quaternion_path_arr[:,3], rotVec_rec_arr[:,0], rotVec_rec_arr[:,1], rotVec_rec_arr[:,2])):

        result_list.append([v0,v1,v2,v3,v4,v5,v6,v7])
    
    '''
    #save reuslt file
    ####################################################################
    

    trait_path = os.path.dirname(current_path + filename_skeleton)
    
    folder_name = os.path.basename(trait_path)
    
    #print("current_path folder ={}".format(folder_name))
    
    # create trait file using sub folder name
    trait_file = (current_path + folder_name + '_quaternion.xlsx')
    
    trait_file_csv = (current_path + folder_name + '_quaternion.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        
        
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
         #Get the current Active Sheet
        sheet_quaternion = wb['sheet_quaternion']
        
        sheet_quaternion.delete_rows(2, sheet_quaternion.max_row+1) # for entire sheet
        
    else:
        # Keep presets
        # Keep presets
        wb = openpyxl.Workbook()
        
        #sheet = wb.active
        
        sheet_quaternion = wb.active
        sheet_quaternion.title = "sheet_quaternion"
        

        sheet_quaternion.cell(row = 1, column = 1).value = 'graph path index'
        sheet_quaternion.cell(row = 1, column = 2).value = 'quaternion_a'
        sheet_quaternion.cell(row = 1, column = 3).value = 'quaternion_b'
        sheet_quaternion.cell(row = 1, column = 4).value = 'quaternion_c'
        sheet_quaternion.cell(row = 1, column = 5).value = 'quaternion_d'
        sheet_quaternion.cell(row = 1, column = 6).value = 'rotVec_rec_0'
        sheet_quaternion.cell(row = 1, column = 7).value = 'rotVec_rec_1'
        sheet_quaternion.cell(row = 1, column = 8).value = 'rotVec_rec_2'
              
        
    for row in result_list:
        sheet_quaternion.append(row)
   
   
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Result file was saved at {}\n".format(trait_file))
    else:
        print("Error saving Result file\n")

  
    wb = openpyxl.load_workbook(trait_file)
    
    # get_active_sheet()
    sh = wb.active 
    
    
    # save excel file as csv format
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows: 
            c.writerow([cell.value for cell in r])
    

    
    ###################################################################
    #visualize quaternion values a + b*i + c*j + d*k
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    
    
    x = quaternion_path_arr[:,1]
    y = quaternion_path_arr[:,2]
    z = quaternion_path_arr[:,3]
    c = quaternion_path_arr[:,0]

    img = ax.scatter(x, y, z, c=c, cmap=plt.hot())
    #img = ax.scatter(x, y, z, c=c, cmap=plt.hot())
    fig.colorbar(img)
    #plt.show()
    
    #define result path
    trait_path = os.path.dirname(current_path + filename_skeleton)
    folder_name = os.path.basename(trait_path)
    
    # create trait file using sub folder name
    quaternion_scatter = (current_path + folder_name + '_quaternion_scatter.png')
    
    plt.savefig(quaternion_scatter)
    
    
    
    ####################################################################
    #Multi-dimension plots in ploty, color represents quaternion_a

    #Read cars data from csv
    data = pd.read_csv(trait_file_csv)

    #Set marker properties
    markercolor = data['quaternion_a']

    #Make Plotly figure
    fig1 = go.Scatter3d(x=data['quaternion_b'],
                    y=data['quaternion_c'],
                    z=data['quaternion_d'],
                    marker=dict(color=markercolor,
                                opacity=1,
                                reversescale=True,
                                colorscale='Viridis',
                                colorbar=dict(thickness=10),
                                size=5),
                    line=dict (width=0.02),
                    mode='markers')

    #Make Plot.ly Layout
    mylayout = go.Layout(scene=dict(xaxis=dict( title="quaternion_b"),
                                yaxis=dict( title="quaternion_c"),
                                zaxis=dict(title="quaternion_d")),)
    

    quaternion_4D = (current_path + folder_name + '_quaternion_4D.html')
    
    #Plot and save html
    plotly.offline.plot({"data": [fig1],
                     "layout": mylayout},
                     auto_open=True,
                     filename=quaternion_4D)
    '''
    
