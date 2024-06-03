"""
Version: 1.5

Summary: Analyze and visualzie tracked traces

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

python3 trace_load_connect.py -p /home/suxingliu/ply_data/trace_track/ -dt 50.5 -ma 0.1 -dr 4.8 


argument:
("-p", "--path", required = True, help="path to trace file")

default file format: *.txt  

"""

# Import python libraries
import numpy as np
from numpy import arctan2, sqrt
import numexpr as ne

import matplotlib as mpl
import matplotlib.cm as cm
import matplotlib.pyplot as plt

import glob
import fnmatch
import os, os.path
import math

import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
        
from scipy.spatial import distance


import itertools

import warnings
#warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore")

def mkdir(path):
    """Create result folder"""
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False

def trace_angle(x, y, z):
    """compute the angle of each trace in 3D space"""
    '''
    cx = x[0] - x[len(x)-1]
    cy = y[0] - y[len(y)-1]
    cz = z[0] - z[len(z)-1]

    (r,theta,phi) = asSpherical(cx, cy, cz)
    '''
    (r,theta,phi) = asSpherical(x, y, z)
    
    return r, theta, phi

def unit_vector(vector):
    """ Returns the unit vector of the vector.  """
    return vector / np.linalg.norm(vector)

def angle_between(v1, v2):
    """ Returns the angle in radians between vectors 'v1' and 'v2':
    """
    v1_u = unit_vector(v1)
    v2_u = unit_vector(v2)
    return np.arccos(np.clip(np.dot(v1_u, v2_u), -1.0, 1.0))*180/math.pi



def angle_vector(v1, v2, acute):
    """compute the angle between two vectors"""
    # v1 is firsr vector
    # v2 is second vector
    angle = np.arccos(np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2)))
    
    if (acute == True):
        return angle
    else:
        return 2 * np.pi - angle

def get_cmap(n, name='hsv'):
    """get the color mapping"""
    
    #Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    #RGB color; the keyword argument name must be a standard mpl colormap name
    
    return plt.cm.get_cmap(name, n)


def asCartesian(r, theta, phi):
    """coordinates transormation from sphere coords to cartesian coord system"""
    
    #takes list rthetaphi (single coord)
    r       = r
    theta   = theta* math.pi/180 # to radian
    phi     = phi* math.pi/180
    x = r * sin( theta ) * cos( phi )
    y = r * sin( theta ) * sin( phi )
    z = r * cos( theta )
    return x,y,z


def asSpherical(x, y, z):
    """coordinates transormation from cartesian coords to sphere coord system"""

    r = math.sqrt(x*x + y*y + z*z)
    
    elevation = math.acos(z/r)*180/math.pi #to degrees
    
    azimuth = arctan2(y,x)*180/math.pi
    
    return r, elevation, azimuth



    

def Line_length(X, Y, Z):
	
    """compute 3D line length """
    n = len(X)
    
    lv = [sqrt((X[i]-X[i-1])**2 + (Y[i]-Y[i-1])**2 + (Z[i]-Z[i-1])) for i in range (1,n)]
    
    L = sum(lv)
    
    # calculate length of line
    #l = np.sqrt( np.diff(X)**2 + np.diff(Y)**2 + np.diff(Z)**2 )
    
    return L

def Distance(x1, y1, x2, y2):
    """compute distance between two 2D points"""
    
    dist = math.hypot(x2 - x1, y2 - y1)
    return dist
    

def line_num_for_phrase_in_file(filename, phrase):
    """extract trace size index"""
    
    num_loc = []
    
    for num, line in enumerate(filename, 1):
        
        if (phrase in line):
            
            #print line, num
            
            num_loc.append(num)
    
    return num_loc
    


# read data from files
def read_data(file_path):
    """load data from files"""
    
    with open(file_path, 'r') as f:

        num_loc = line_num_for_phrase_in_file(f, '#')
        
        #print("Number of tracked traces: {0} \n".format(len(num_loc)))

        f.close

    # load data matrix only
    data  = np.loadtxt(file_path)

    return data, num_loc


def lsq_line_fit(data):
    """fitting 3D pints into best fit line using lsq method"""
     
    # compute data range
    data_range  = (data.max() - data.min())/2
    
    # Perturb with some Gaussian noise
    data += np.random.normal(size=data.shape) * 0.4

    # Calculate the mean of the points, i.e. the 'center' of the cloud
    datamean = data.mean(axis=0)

    # Do an SVD on the mean-centered data.
    uu, dd, vv = np.linalg.svd(data - datamean)

    # Now vv[0] contains the first principal component, i.e. the direction
    # vector of the 'best fit' line in the least squares sense.
    # data_range of the data can be accquired by min and max values
    # and it should have mean 0. it's a straight line, so we only need 2 points.
    linepts = vv[0] * np.mgrid[-data_range:data_range:2j][:, np.newaxis]

    # shift by the mean to get the line in the right place
    linepts += datamean
    
    return linepts
   

def fit_line(x,y,z):
    """fit the trace points as a two points line """
    
    data_fit = np.concatenate((x[:, np.newaxis], y[:, np.newaxis], z[:, np.newaxis]), axis=1)

    if len(data_fit) > 0:
        
        linepts = lsq_line_fit(data_fit)
        
        cx = linepts[0,0] - linepts[1,0]
        cy = linepts[0,1] - linepts[1,1]
        cz = linepts[0,2] - linepts[1,2]

        (fit_line_angle_r,fit_line_angle_theta,fit_line_angle_phi) = asSpherical(cx, cy, cz)
        
        xx = linepts[:,0]
        yy = linepts[:,1]
        zz = linepts[:,2]
    else:
        
        fit_line_angle_theta = 0
        print("fit_line_angle_theta is zero\n")
    
    return xx, yy, zz, fit_line_angle_theta, fit_line_angle_r



def trace_angle(x, y, z):
    """compute the angle of each trace in 3D space"""
    
    cx = x[0] - x[len(x)-1]
    cy = y[0] - y[len(y)-1]
    cz = z[0] - z[len(z)-1]

    (r,theta,phi) = asSpherical(cx, cy, cz)
    
    return r, theta, phi


# load trace data and connect adjacent traces 
def trace_measure(fname, trace_rec, file_idx):
    """load trace data and connect adjacent traces"""
    
    #trace file path
    trace_file = args["path"] + str(fname)

    # read data from trace file, row by row
    (data, num_loc) = read_data(trace_file)
    
    print("Loading {0} , it contains {1} trace lines...\n".format(fname,len(num_loc)))
    

    # loop over individual trace point set
    for index in range(len(num_loc)-1):
        
        # Extract the line data by deviding mark '#'
        line_start = num_loc[index] - index -1
        line_end = line_start + (num_loc[index+1] - num_loc[index] -1) 
            
        #extract one trace line data
        data_array = data[line_start:line_end, :]
        
        #extract xyz coordinates
        x = data_array[:,0]
        y = data_array[:,1]
        z = data_array[:,2]
        
        #define min distance tracking index
        min_idx = -1
        min_dst = 10000
        
        (r_data, theta_data, phi_data) = trace_angle(x, y, z)

        # compare distance between pointset and the start point of new trace line,
        # connect lines if under threhold, otherwise create and adding new trace lines
        if file_idx > 0:
        
            #new_start_point =  data_array[0,[0,1,2]].reshape(1,3)
            
            z_min_data_array = np.where(z==np.amin(z))
            
            new_start_point = data_array[z_min_data_array [0][0],[0,1,2]].reshape(1,3)
          
            #new_start_point =  np.asarray([x.mean(), y.mean(), z.mean()]).reshape(1,3)
            
            for tr_idx, trace_coord in enumerate(trace_rec):
                
                # compute the euclidean between 3d point and pointset
                dst = distance.cdist(trace_rec[tr_idx][:,[0,1,2]], new_start_point, 'euclidean')

                #dst = distance.cdist(np.asarray([trace_rec[tr_idx][:,0].mean(), trace_rec[tr_idx][:,1].mean(), trace_rec[tr_idx][:,2].mean()]) .reshape(1,3), new_start_point, 'euclidean')
               
                #min_idx = np.argmin(dst) 
                if np.min(dst) < min_dst:
                    min_dst = np.min(dst)
                    min_idx = tr_idx

            # connect two lines if close enough
            if (min_dst < dis_tracking):
                
                x_trace = trace_rec[min_idx][:,0]
                y_trace = trace_rec[min_idx][:,1]
                z_trace = trace_rec[min_idx][:,2]
                
                #(xx, yy, zz, fit_line_angle_theta_trace, r)  = fit_line(x_trace, y_trace, z_trace)
                #print("fit_line_angle_theta {0} \n".format(abs(fit_line_angle_theta - fit_line_angle_theta_trace)))
                
                #xy_dist = math.hypot(x_trace[len(x_trace)-1]-x[0], y_trace[len(y_trace)-1]-y[0])
                
                (r_trace, theta_trace, phi_trace) = trace_angle(x_trace, y_trace, z_trace)
                
                #if xy_dist < xy_dist_thresh:
                #if abs(fit_line_angle_theta - fit_line_angle_theta_trace) < min_angle:
                #if abs(theta_data - theta_trace) < min_angle:
                
                trace_rec[min_idx] = np.concatenate((trace_rec[min_idx], data_array), axis = 0)
                
               
            else:
                # creat new trace line and add it to collection
         
                    trace_rec.append(data_array)
            
        else:
            #print("First trace file loaded!")
            trace_rec.append(data_array)
    
    return trace_rec



def trace_angle_connect(trace_rec):
    """connect the trace based on the similarity of trace direction angles in 3D space"""
    
    #initialize parameters
    trace_rec_update = trace_rec
    trace_rec_fitting_angle = []
    trace_rec_center = []
    
    # compute all trace direction angles 
    for idx, trace_pointset in enumerate(trace_rec):

        x = trace_rec[idx][:,0]
        y = trace_rec[idx][:,1]
        z = trace_rec[idx][:,2]
        
        (xx, yy, zz, theta, r) = fit_line(x, y, z)
        
        #b = np.array([[np.mean(xx, axis=0)], [y]])
         
        center = np.array([np.mean(xx, axis=0), np.mean(yy, axis=0), np.mean(zz, axis=0)])
        
        trace_rec_fitting_angle.append(theta)
        
        trace_rec_center.append(center)
   

    # sort trace according to angle
    sorted_idx = sorted(range(len(trace_rec_fitting_angle)), key=lambda k: trace_rec_fitting_angle[k])
    

    adjacent_value_rec = []
    
    # connect trace with similar direction angles 
    for idx, value in enumerate(sorted_idx):
        
        if idx < len(sorted_idx) - 1:
            
            adjacent_value = sorted_idx[idx + 1]
            
            adjacent_dst = distance.euclidean(trace_rec_center[value], trace_rec_center[adjacent_value])
            
            #print adjacent_dst

            if abs(trace_rec_fitting_angle[adjacent_value] - trace_rec_fitting_angle[value]) < min_angle:
            #or adjacent_dst < dis_thresh*dist_ratio: 
                
                trace_rec_update[idx] = np.concatenate((trace_rec_update[idx], trace_rec[adjacent_value]), axis = 0)
                
                #print(trace_rec[value], trace_rec[adjacent_value])
                
                #print value, adjacent_value
                
                adjacent_value_rec.append(adjacent_value)
            
  
    #print len(trace_rec_update)
    #print adjacent_value_rec
    #trace_rec = []
    
    # delete traces which was joined with adjacent traces
    trace_new = [v for i,v in enumerate(trace_rec_update) if i not in adjacent_value_rec]
    
    return trace_new
    

def trait_measure(trace_rec):
    """measure trace paramters """
    
    #initialize parameters
    index_rec = []
    length_rec = []
    angle_rec = []
    diameter_rec = []
    projection_radius = []
    
    #connect the trace with similar direction angles 
    trace_new = trace_angle_connect(trace_rec)
    
    for idx, trace_pointset in enumerate(trace_new):
        
        # extract x y z coordinates
        x = trace_new[idx][:,0]
        y = trace_new[idx][:,1]
        z = trace_new[idx][:,2]
        
        diameter = np.mean(trace_new[idx][:,3],axis=0)*2
        #print("average diameter = {}\n".format(diameter))
        
        '''
        #calculate line length
        l = Line_length(x, y, z)
        
        #print("line length is:")
        #print(np.sum(l))

        #line_length = np.average(l)
        line_length = np.sum(l)
        '''
        #accquire max and min value of z and their indexes
        z_maxid = np.where(z==np.amax(z))
        z_minid = np.where(z==np.amin(z))
        
        p_s = trace_new[idx][z_minid[0][0],[0,1,2]]
        p_e = trace_new[idx][z_maxid[0][0],[0,1,2]]
        
        #compute line length
        line_length = distance.euclidean(p_s, p_e)
        
        
        theta_offest = np.zeros(4)
        r_offest = np.zeros(4)
        
        #calculate angles
        for offest in range(0,4):
            
            interval = int(((offest+1)*0.25)*line_length)  
            
            if interval >= len(x):
                interval = len(x)-1
            
            cx = x[interval] - x[0]
            cy = y[interval] - y[0]  
            cz = z[interval] - z[0]
            
            (r,theta,phi) = asSpherical(cx, cy, cz)
            
            if theta > 90:
                theta = 180 -theta

            theta_offest[offest] = theta

            r_offest[offest] = r
       
        # record all parameters
        index_rec.append(idx)
        length_rec.append(line_length)
        angle_rec.append(theta_offest[2])
        diameter_rec.append(diameter)
        projection_radius.append(r_offest[2])
        
    return index_rec, length_rec, angle_rec, diameter_rec, projection_radius, trace_new


# visualize the trace in 2D and apply color coding for each trace
def visualize_trace_mayavi(file_path, trace_new):
    
    from mayavi import mlab
    
    #mlab.figure(size = (400,300))
    f = mlab.figure(bgcolor = (1,1,1), fgcolor = (0.5, 0.5, 0.5), size = (600,400))
    
    cmap = get_cmap(len(trace_rec))
    
    #print("{0} Traces were successfuly tracked...\n".format(len(trace_rec)))
    
    for idx, trace_pointset in enumerate(trace_new):
        
        #print("Coordinate of Trace{0}: \n".format(idx))
        #print(trace_rec[idx])

        X = trace_new[idx][:,0]
        Y = trace_new[idx][:,1]
        Z = trace_new[idx][:,2]/3
        (xx, yy, zz, angle, r) = fit_line(X,Y,Z)

        radius = np.average(trace_new[idx][:,3])
        
        #generate random different colors for different traces
        color_rgb = cmap(idx)[:len(cmap(idx))-1]
        
        #Draw 3d points and lines 
        #pts = mlab.points3d(X, Y, Z, color = color_rgb, scale_factor = radius)
        
        pts = mlab.points3d(X, Y, Z, color = color_rgb, mode = 'point')
        
        #pts = mlab.points3d(X.mean(), Y.mean(), Z.mean(), color = color_rgb, scale_factor = 4.5, mode = 'sphere')

        #pts = mlab.plot3d(xx, yy, zz, color = color_rgb, tube_radius = 1.1, line_width = 2)
        
        #pts = mlab.plot3d(X, Y, Z, color = color_rgb, tube_radius = 0.1, extent = [0,1,0,1,0,1])

        pts.actor.property.set(point_size = 2.5)
    
    
    #Save as ply models, optional
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    base_foler = os.path.basename(file_path[:-1])
    
    model_file = (parent_path + '/' + base_foler + '_root_trace_model' + '.obj')

    mlab.savefig(model_file)
    
    #show model
    mlab.show()


if __name__ == '__main__':
    
    # construct the argument and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True, help="path to trace file")
    ap.add_argument("-dt", "--dis_tracking", required = False, default = '50.5', type = float, help = "dis_tracking")
    ap.add_argument("-ma", "--min_angle", required = False, default = '0.1', type = float, help = "min_angle")
    ap.add_argument("-dr", "--dist_ratio", required = False, default = '4.8', type = float, help = "dist_ratio")
    args = vars(ap.parse_args())

    #extract file list in specified path
    filetype = '*.txt' 
    file_path = args["path"]
    
    #accquire image file list

    file_list = sorted(fnmatch.filter(os.listdir(args["path"]), filetype))
      
    #file_list.sort(key=lambda f: int(filter(str.isdigit, f)))
    
    print(file_list)
    
    global trace_rec, dis_tracking, min_angle, dist_ratio
    
    trace_rec = []
    
    #define min distance tracking threshold
    dis_tracking = args["dis_tracking"]
    
    min_angle = args["min_angle"]
    
    dist_ratio = args["dist_ratio"]

    
    # make the folder to store the results
    parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    mkpath = parent_path + '/' + str('analysis_result')
    mkdir(mkpath)
    save_path_result = mkpath + '/'
  
    
    #loop to all tracked trace files
    for file_idx, fname in enumerate(file_list):

        trace_rec = trace_measure(fname, trace_rec, file_idx)
    
    (index_rec, length_rec, angle_rec, diameter_rec, projection_radius, trace_new) = trait_measure(trace_rec)
    
    print("{0} Traces were successfuly tracked...\n".format(len(trace_new)))
 
    #disaply trace in 3D space
    #visualize_trace_mayavi(file_path, trace_new)
    
    
    
    
    print ("results_folder: " + save_path_result) 
    ##################################################################
    #Start of writing measured parameters as excel file 
    
    #parent_path = os.path.abspath(os.path.join(file_path, os.pardir))
    
    #base_folder = os.path.basename(file_path[:-1])
    
    #trait_file = (parent_path + '/' + base_folder + 'root_trace_measure' + '.xlsx')
    
    #trait_file_csv = (parent_path + '/' + base_folder + 'root_trace_measure' + '.csv')
    
    trait_file = (save_path_result + 'root_trace_measure' + '.xlsx')
    
    trait_file_csv = (save_path_result + 'root_trace_measure' + '.csv')
    
    if os.path.exists(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file, read_only = False)
        sheet = wb.active

        os.remove(trait_file)
        
    else:
        # Keep presents
        wb = Workbook()
        sheet = wb.active
        
        sheet.cell(row = 1, column = 1).value = 'Root trace index'
        sheet.cell(row = 1, column = 2).value = 'Root trace length'
        sheet.cell(row = 1, column = 3).value = 'Root trace angle'
        sheet.cell(row = 1, column = 4).value = 'Root trace diameter'
        sheet.cell(row = 1, column = 5).value = 'Root trace projection radius'
        sheet.cell(row = 1, column = 6).value = 'Root number in total'
    
    sheet = wb.active

    for row in zip(index_rec, length_rec, angle_rec, diameter_rec, projection_radius):
        sheet.append(row)
    
    sheet.cell(row = 2, column = 6).value = len(trace_new)
    
    # create new sheet for coordinates storage
    wb.create_sheet('Trace coordinates')
    sheet_coord = wb["Trace coordinates"]
    
    sheet_coord.cell(row = 1, column = 1).value = 'Root trace index'
    sheet_coord.cell(row = 1, column = 2).value = 'Root trace coordinate X'
    sheet_coord.cell(row = 1, column = 3).value = 'Root trace coordinate Y'
    sheet_coord.cell(row = 1, column = 4).value = 'Root trace coordinate Z'
        
    # write coordinates for each root trace
    for idx, row in enumerate(trace_new):

        for idy, ind_trace in enumerate(trace_new[idx]):
            
            coords = np.hstack([index_rec[idx],ind_trace])
            
            sheet_coord.append(coords.tolist())
                
    
    #save the csv file
    wb.save(trait_file)
    
    if os.path.exists(trait_file):
        
        print("Trait result was saved!\n")
    
    import openpyxl
    import csv

    wb = load_workbook(trait_file)
    sh = wb.get_active_sheet()
    
    #with open(trait_file_csv, 'wb') as f:
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
    
    
    ##################################################################
    #End of writing measured parameters as excel file 
    
    
    






